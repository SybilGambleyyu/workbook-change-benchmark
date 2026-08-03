"""Deterministic generation of the WCAB fixture tree.

The fixtures are intentionally generated rather than copied from real business
workbooks.  That makes every assertion inspectable and the redistribution
rights unambiguous.
"""

from __future__ import annotations

import base64
import io
import json
import re
import struct
from collections.abc import Callable
from datetime import date, datetime, timezone
from hashlib import sha256
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Protection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.table import Table, TableStyleInfo

from .manifest import write_manifest

FIXTURE_SCHEMA_VERSION = 3
_FIXED_TIMESTAMP = datetime(2024, 1, 1, tzinfo=timezone.utc)
_CORE_MODIFIED = re.compile(rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)")
WorkbookFactory = Callable[[], Workbook]
WorkbookMutator = Callable[[Workbook], None]
ArchiveMutator = Callable[[dict[str, bytes]], None]
_SPREADSHEETML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_PACKAGE_RELATIONSHIPS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_DOCUMENT_RELATIONSHIPS_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
_XML_DIGITAL_SIGNATURE_NS = "http://www.w3.org/2000/09/xmldsig#"
_OFFICE_2013_SPREADSHEET_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
_DYNAMIC_ARRAY_NS = "http://schemas.microsoft.com/office/spreadsheetml/2017/dynamicarray"
_EXTERNAL_WORKBOOK_LINK_FORMULA = "='[WCABSource.xlsx]Inputs'!$B$2"
_EXTERNAL_WORKBOOK_LINK_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/externalLink"
_EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/externalLinkPath"
_EXTERNAL_WORKBOOK_LINK_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"
)
_EXTERNAL_WORKBOOK_LINK_MEMBER = "xl/externalLinks/externalLink1.xml"
_EXTERNAL_WORKBOOK_LINK_RELATIONSHIPS_MEMBER = "xl/externalLinks/_rels/externalLink1.xml.rels"
_EXTERNAL_WORKBOOK_LINK_WORKBOOK_RELATIONSHIP_ID = "rIdWCABExternalLink"
_EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP_ID = "rIdWCABExternalLinkPath"
_EXTERNAL_WORKBOOK_LINK_HOST_SHEET = "LinkedModel"
_EXTERNAL_WORKBOOK_LINK_HOST_CELL = "B2"
_EXTERNAL_WORKBOOK_LINK_EXTERNAL_SHEET = "Inputs"
_EXTERNAL_WORKBOOK_LINK_DASHBOARD_SHEET = "Dashboard"
_EXTERNAL_WORKBOOK_LINK_DASHBOARD_CELL = "B4"
_EXTERNAL_WORKBOOK_LINK_DASHBOARD_FORMULA = "=LinkedModel!$B$2"
_EXTERNAL_WORKBOOK_LINK_BASELINE_TARGET = (
    "https://approved.example.invalid/wcab-external-workbook/WCABSource.xlsx"
)
_EXTERNAL_WORKBOOK_LINK_CANDIDATE_TARGET = (
    "https://review.example.invalid/wcab-external-workbook/WCABSource.xlsx"
)
_EXTERNAL_DEFINED_NAME_SOURCE_NAME = "ScenarioRate"
_EXTERNAL_DEFINED_NAME_SOURCE_BASELINE_REFERS_TO = "'[WCABApprovedSource.xlsx]Inputs'!$B$2"
_EXTERNAL_DEFINED_NAME_SOURCE_CANDIDATE_REFERS_TO = "'[WCABReviewSource.xlsx]Inputs'!$B$2"
_EXTERNAL_DEFINED_NAME_SOURCE_MODEL_SHEET = "Model"
_EXTERNAL_DEFINED_NAME_SOURCE_MODEL_CELL = "B2"
_EXTERNAL_DEFINED_NAME_SOURCE_MODEL_FORMULA = "=ScenarioRate*2"
_EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_SHEET = "Dashboard"
_EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_CELL = "B4"
_EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_FORMULA = "=Model!$B$2"
_NAMED_LAMBDA_NAME = "ScenarioValue"
_NAMED_LAMBDA_BASELINE_REFERS_TO = "=LAMBDA(rate,amount,rate*amount)"
_NAMED_LAMBDA_CANDIDATE_REFERS_TO = "=LAMBDA(rate,amount,rate*(amount+10))"
_NAMED_LAMBDA_INPUT_SHEET = "Inputs"
_NAMED_LAMBDA_RATE_CELL = "B2"
_NAMED_LAMBDA_RATE_VALUE = 0.08
_NAMED_LAMBDA_AMOUNT_CELL = "B3"
_NAMED_LAMBDA_AMOUNT_VALUE = 100
_NAMED_LAMBDA_MODEL_SHEET = "Model"
_NAMED_LAMBDA_MODEL_CELL = "B2"
_NAMED_LAMBDA_MODEL_FORMULA = "=ScenarioValue(Inputs!B2,Inputs!B3)"
_NAMED_LAMBDA_DASHBOARD_SHEET = "Dashboard"
_NAMED_LAMBDA_DASHBOARD_CELL = "B4"
_NAMED_LAMBDA_DASHBOARD_FORMULA = "=Model!$B$2"
_ITERATIVE_CALCULATION_FORMULA = "=(B2+Inputs!$B$2)/2"
_ITERATION_COUNT = 100
_ITERATION_DELTA = 0.001
_PRECISION_AS_DISPLAYED_INPUT = 10.005
_PRECISION_AS_DISPLAYED_NUMBER_FORMAT = "0.00"
_PRECISION_AS_DISPLAYED_FORMULA = "=Inputs!$B$2*2"
_FORMULA_CACHED_RESULT_INPUT = 10
_FORMULA_CACHED_RESULT_FORMULA = "=Inputs!$B$2*2"
_FORMULA_CACHED_RESULT_BASELINE = 20
_FORMULA_CACHED_RESULT_CANDIDATE = 25
_WORKBOOK_DATE_SYSTEM_SERIAL = 45292
_WORKBOOK_DATE_SYSTEM_NUMBER_FORMAT = "yyyy-mm-dd"
_WORKBOOK_DATE_SYSTEM_FORMULA = "=Inputs!$B$2+30"
_WORKBOOK_DATE_SYSTEM_DASHBOARD_FORMULA = "=Model!$B$2"
_AUTO_FILTER_REF = "A1:B5"
_AUTO_FILTER_COLUMN_ID = 0
_AUTO_FILTER_BASELINE_VALUE = "North"
_AUTO_FILTER_CANDIDATE_VALUE = "South"
_AUTO_FILTER_SUBTOTAL_FORMULA = "=SUBTOTAL(109,B2:B5)"
_AUTO_FILTER_DASHBOARD_FORMULA = "=Report!$D$2"
_OFFICE_2014_REVISION_NS = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
_NAMED_SHEET_VIEW_NS = "http://schemas.microsoft.com/office/spreadsheetml/2019/namedsheetviews"
_NAMED_SHEET_VIEW_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2019/04/relationships/namedSheetView"
)
_NAMED_SHEET_VIEW_CONTENT_TYPE = "application/vnd.ms-excel.namedsheetviews+xml"
_NAMED_SHEET_VIEW_REPORT_SHEET = "Report"
_NAMED_SHEET_VIEW_DASHBOARD_SHEET = "Dashboard"
_NAMED_SHEET_VIEW_FILTER_REF = "A1:B5"
_NAMED_SHEET_VIEW_FILTER_COLUMN_ID = 0
_NAMED_SHEET_VIEW_BASELINE_VALUE = "North"
_NAMED_SHEET_VIEW_CANDIDATE_VALUE = "South"
_NAMED_SHEET_VIEW_SUBTOTAL_CELL = "D2"
_NAMED_SHEET_VIEW_SUBTOTAL_FORMULA = "=SUBTOTAL(109,B2:B5)"
_NAMED_SHEET_VIEW_DASHBOARD_CELL = "B4"
_NAMED_SHEET_VIEW_DASHBOARD_FORMULA = "=Report!$D$2"
_NAMED_SHEET_VIEW_MEMBER = "xl/namedSheetViews/namedSheetView1.xml"
_NAMED_SHEET_VIEW_RELATIONSHIP_MEMBER = "xl/worksheets/_rels/sheet1.xml.rels"
_NAMED_SHEET_VIEW_RELATIONSHIP_ID = "rIdWCABNamedSheetView"
_NAMED_SHEET_VIEW_FILTER_ID = "{00000000-0001-0000-0000-000000000000}"
_NAMED_SHEET_VIEW_ID = "{11111111-1111-1111-1111-111111111111}"
_NAMED_SHEET_VIEW_COLUMN_ID = "{22222222-2222-2222-2222-222222222222}"
_NAMED_SHEET_VIEW_NAME = "WCAB regional review"
_XML_SCHEMA_NS = "http://www.w3.org/2001/XMLSchema"
_XML_MAP_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/xmlMaps"
_TABLE_SINGLE_CELLS_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/tableSingleCells"
_XML_MAP_EXPORT_SHEET = "Export"
_XML_MAP_DASHBOARD_SHEET = "Dashboard"
_XML_MAP_TABLE_NAME = "InvoiceLines"
_XML_MAP_TABLE_REF = "A1:B3"
_XML_MAP_TABLE_MEMBER = "xl/tables/table1.xml"
_XML_MAP_MEMBER = "xl/xmlMaps.xml"
_XML_MAP_SINGLE_CELL_MEMBER = "xl/singleCellTables/singleCellTable1.xml"
_XML_MAP_WORKBOOK_RELATIONSHIP_ID = "rIdWCABXmlMaps"
_XML_MAP_SINGLE_CELL_RELATIONSHIP_ID = "rIdWCABXmlMappingSingleCells"
_XML_MAP_CONNECTION_ID = 7
_XML_MAP_ID = 1
_XML_MAP_SCHEMA_ID = "WCAB-INVOICE-EXPORT"
_XML_MAP_NAMESPACE = "urn:wcab:invoice-export"
_XML_MAP_ROOT_ELEMENT = "Invoice"
_XML_MAP_TABLE_COLUMN_ID = 2
_XML_MAP_TABLE_COLUMN_NAME = "Net amount"
_XML_MAP_BASELINE_XPATH = "/wcab:Invoice/wcab:Line/wcab:NetAmount"
_XML_MAP_CANDIDATE_XPATH = "/wcab:Invoice/wcab:Line/wcab:TaxAmount"
_XML_MAP_SINGLE_CELL = "E2"
_XML_MAP_SINGLE_CELL_XPATH = "/wcab:Invoice/wcab:Header/wcab:AsOf"
_XML_MAP_TOTAL_CELL = "D2"
_XML_MAP_TOTAL_FORMULA = "=SUM(InvoiceLines[Net amount])"
_XML_MAP_DASHBOARD_CELL = "B4"
_XML_MAP_DASHBOARD_FORMULA = "=Export!$D$2"
_WEB_EXTENSION_TASKPANES_NS = "http://schemas.microsoft.com/office/webextensions/taskpanes/2010/11"
_WEB_EXTENSION_NS = "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
_WEB_EXTENSION_TASKPANES_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2011/relationships/webextensiontaskpanes"
)
_WEB_EXTENSION_RELATIONSHIP = "http://schemas.microsoft.com/office/2011/relationships/webextension"
_WEB_EXTENSION_TASKPANES_CONTENT_TYPE = "application/vnd.ms-office.webextensiontaskpanes+xml"
_WEB_EXTENSION_CONTENT_TYPE = "application/vnd.ms-office.webextension+xml"
_OFFICE_WEB_ADDIN_TASKPANES_MEMBER = "xl/webextensions/taskpanes.xml"
_OFFICE_WEB_ADDIN_EXTENSION_MEMBER = "xl/webextensions/webextension1.xml"
_OFFICE_WEB_ADDIN_TASKPANES_RELATIONSHIPS_MEMBER = "xl/webextensions/_rels/taskpanes.xml.rels"
_OFFICE_WEB_ADDIN_WORKBOOK_RELATIONSHIP_ID = "rIdWCABOfficeWebAddinTaskpanes"
_OFFICE_WEB_ADDIN_EXTENSION_RELATIONSHIP_ID = "rIdWCABOfficeWebAddin"
_OFFICE_WEB_ADDIN_ID = "{33333333-3333-3333-3333-333333333333}"
_OFFICE_WEB_ADDIN_REFERENCE_ID = "{44444444-4444-4444-4444-444444444444}"
_OFFICE_WEB_ADDIN_REFERENCE_VERSION = "1.0.0.0"
_OFFICE_WEB_ADDIN_STORE = "wcab-review-assistant.xml"
_OFFICE_WEB_ADDIN_STORE_TYPE = "Filesystem"
_OFFICE_WEB_ADDIN_AUTO_SHOW_PROPERTY = "Office.AutoShowTaskpaneWithDocument"
_OFFICE_WEB_ADDIN_BASELINE_AUTO_SHOW = False
_OFFICE_WEB_ADDIN_CANDIDATE_AUTO_SHOW = True
_OFFICE_WEB_ADDIN_INPUT_SHEET = "Inputs"
_OFFICE_WEB_ADDIN_INPUT_CELL = "B2"
_OFFICE_WEB_ADDIN_INPUT_VALUE = 10
_OFFICE_WEB_ADDIN_MODEL_SHEET = "Model"
_OFFICE_WEB_ADDIN_MODEL_CELL = "B2"
_OFFICE_WEB_ADDIN_MODEL_FORMULA = "=Inputs!$B$2*2"
_OFFICE_WEB_ADDIN_DASHBOARD_SHEET = "Dashboard"
_OFFICE_WEB_ADDIN_DASHBOARD_CELL = "B4"
_OFFICE_WEB_ADDIN_DASHBOARD_FORMULA = "=Model!$B$2"
_OLE_OBJECT_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/oleObject"
_OLE_OBJECT_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.oleObject"
_OLE_OBJECT_WORKSHEET_MEMBER = "xl/worksheets/sheet1.xml"
_OLE_OBJECT_RELATIONSHIPS_MEMBER = "xl/worksheets/_rels/sheet1.xml.rels"
_OLE_OBJECT_RELATIONSHIP_ID = "rIdWCABEmbeddedOle"
_OLE_OBJECT_MEMBER = "xl/embeddings/wcab-review-embedded-object.bin"
_OLE_OBJECT_PROG_ID = "WCAB.Review.Embedded.Object"
_OLE_OBJECT_DV_ASPECT = "DVASPECT_CONTENT"
_OLE_OBJECT_SHAPE_ID = 1026
_OLE_OBJECT_BASELINE_AUTO_LOAD = False
_OLE_OBJECT_CANDIDATE_AUTO_LOAD = True
_OLE_OBJECT_INPUT_SHEET = "Inputs"
_OLE_OBJECT_INPUT_CELL = "B2"
_OLE_OBJECT_INPUT_VALUE = 10
_OLE_OBJECT_MODEL_SHEET = "Model"
_OLE_OBJECT_MODEL_CELL = "B2"
_OLE_OBJECT_MODEL_FORMULA = "=Inputs!$B$2*2"
_OLE_OBJECT_DASHBOARD_SHEET = "Dashboard"
_OLE_OBJECT_DASHBOARD_CELL = "B4"
_OLE_OBJECT_DASHBOARD_FORMULA = "=Model!$B$2"
_OLE_OBJECT_PAYLOAD = (
    b"WCAB opaque synthetic embedded-object fixture bytes; never deserialized or opened."
)
_EXTERNAL_DATA_CONNECTIONS_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/connections"
_EXTERNAL_DATA_CONNECTIONS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml"
)
_EXTERNAL_DATA_CONNECTION_MEMBER = "xl/connections.xml"
_EXTERNAL_DATA_CONNECTION_WORKBOOK_RELATIONSHIP_ID = "rIdWCABExternalDataConnection"
_EXTERNAL_DATA_CONNECTION_ID = 1
_EXTERNAL_DATA_CONNECTION_NAME = "WCAB synthetic external-data connection"
_EXTERNAL_DATA_CONNECTION_REFRESH_URL = "https://example.invalid/wcab-external-data-refresh"
_EXTERNAL_DATA_CONNECTION_SOURCE_BASELINE_URL = (
    "https://approved.example.invalid/wcab-external-data-source"
)
_EXTERNAL_DATA_CONNECTION_SOURCE_CANDIDATE_URL = (
    "https://review.example.invalid/wcab-external-data-source"
)
_PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP = f"{_PACKAGE_RELATIONSHIPS_NS}/digital-signature/origin"
_PACKAGE_SIGNATURE_SIGNATURE_RELATIONSHIP = (
    f"{_PACKAGE_RELATIONSHIPS_NS}/digital-signature/signature"
)
_PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/officeDocument"
_PACKAGE_SIGNATURE_RELATIONSHIP_TRANSFORM = (
    "http://schemas.openxmlformats.org/package/2006/digital-signature/RelationshipTransform"
)
_PACKAGE_SIGNATURE_C14N_ALGORITHM = "http://www.w3.org/TR/2001/REC-xml-c14n-20010315"
_PACKAGE_SIGNATURE_ORIGIN_CONTENT_TYPE = (
    "application/vnd.openxmlformats-package.digital-signature-origin"
)
_PACKAGE_SIGNATURE_XML_CONTENT_TYPE = (
    "application/vnd.openxmlformats-package.digital-signature-xmlsignature+xml"
)
_PACKAGE_SIGNATURE_RELATIONSHIPS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-package.relationships+xml"
)
_PACKAGE_SIGNATURE_ORIGIN_MEMBER = "_xmlsignatures/origin.sigs"
_PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIPS_MEMBER = "_xmlsignatures/_rels/origin.sigs.rels"
_PACKAGE_SIGNATURE_MEMBER = "_xmlsignatures/sig1.xml"
_PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP_ID = "rIdWCABPackageSignatureOrigin"
_PACKAGE_SIGNATURE_XML_RELATIONSHIP_ID = "rIdWCABPackageXmlSignature"
_PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP_ID = "rIdWCABOfficeDocument"
_PACKAGE_SIGNATURE_OBJECT_ID = "idWCABPackageObject"
_PACKAGE_SIGNATURE_WORKBOOK_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
)
_PACKAGE_SIGNATURE_WORKSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)
_PACKAGE_SIGNATURE_BASELINE_MANIFEST_URI = (
    "/xl/workbook.xml?ContentType=" + _PACKAGE_SIGNATURE_WORKBOOK_CONTENT_TYPE
)
_PACKAGE_SIGNATURE_CANDIDATE_MANIFEST_URI = (
    "/xl/worksheets/sheet1.xml?ContentType=" + _PACKAGE_SIGNATURE_WORKSHEET_CONTENT_TYPE
)
_PACKAGE_SIGNATURE_ROOT_RELATIONSHIPS_MANIFEST_URI = (
    "/_rels/.rels?ContentType=" + _PACKAGE_SIGNATURE_RELATIONSHIPS_CONTENT_TYPE
)
_PACKAGE_SIGNATURE_BASELINE_SELECTOR_SOURCE_ID = _PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP_ID
_PACKAGE_SIGNATURE_CANDIDATE_SELECTOR_SOURCE_ID = _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP_ID
_THREADED_COMMENT_NS = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
_THREADED_COMMENT_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2017/10/relationships/threadedComment"
)
_THREADED_COMMENT_PERSON_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2017/10/relationships/person"
)
_THREADED_COMMENT_CONTENT_TYPE = "application/vnd.ms-excel.threadedcomments+xml"
_THREADED_COMMENT_PERSON_CONTENT_TYPE = "application/vnd.ms-excel.person+xml"
_THREADED_COMMENT_SHEET = "Controls"
_THREADED_COMMENT_WORKSHEET_MEMBER = "xl/worksheets/sheet1.xml"
_THREADED_COMMENT_WORKSHEET_RELATIONSHIPS_MEMBER = "xl/worksheets/_rels/sheet1.xml.rels"
_THREADED_COMMENT_MEMBER = "xl/threadedComments/threadedComment1.xml"
_THREADED_COMMENT_PERSON_MEMBER = "xl/persons/person.xml"
_THREADED_COMMENT_WORKSHEET_RELATIONSHIP_ID = "rIdWCABThreadedComment"
_THREADED_COMMENT_PERSON_WORKBOOK_RELATIONSHIP_ID = "rIdWCABThreadedPerson"
_THREADED_COMMENT_PERSON_ID = "{66666666-6666-6666-6666-666666666666}"
_THREADED_COMMENT_ROOT_ID = "{77777777-7777-7777-7777-777777777777}"
_THREADED_COMMENT_CELL = "B10"
_THREADED_COMMENT_TIMESTAMP = "2024-01-01T00:00:00Z"
_THREADED_COMMENT_TEXT = "WCAB synthetic review thread"
_THREADED_COMMENT_PERSON_NAME = "WCAB Reviewer"
_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/revisionHeaders"
_SHARED_WORKBOOK_REVISION_LOG_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/revisionLog"
_SHARED_WORKBOOK_REVISION_HEADERS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.revisionHeaders+xml"
)
_SHARED_WORKBOOK_REVISION_LOG_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.revisionLog+xml"
)
_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER = "xl/revisions/revisionHeaders.xml"
_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER = (
    "xl/revisions/_rels/revisionHeaders.xml.rels"
)
_SHARED_WORKBOOK_REVISION_LOG_MEMBER = "xl/revisions/revisionLog1.xml"
_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIP_ID = "rIdWCABRevisionHeaders"
_SHARED_WORKBOOK_REVISION_LOG_RELATIONSHIP_ID = "rIdWCABRevisionLog"
_SHARED_WORKBOOK_REVISION_HEADER_GUID = "{88888888-8888-8888-8888-888888888888}"
_SHARED_WORKBOOK_REVISION_HEADER_TIMESTAMP = "2024-01-02T03:04:05Z"
_SHARED_WORKBOOK_REVISION_HEADER_AUTHOR = "WCAB Revision Author"
_SHARED_WORKBOOK_REVISION_BASELINE_HISTORIC_VALUE = "WCAB historic approved value"
_SHARED_WORKBOOK_REVISION_CANDIDATE_HISTORIC_VALUE = "WCAB historic candidate value"
_SHARED_WORKBOOK_REVISION_RECORDED_VALUE = "WCAB historic recorded value"
_QUERY_TABLE_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/queryTable"
_QUERY_TABLE_CONNECTIONS_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/connections"
_QUERY_TABLE_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.queryTable+xml"
)
_QUERY_TABLE_CONNECTIONS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml"
)
_QUERY_TABLE_SHEET = "ImportedData"
_QUERY_TABLE_WORKSHEET_MEMBER = "xl/worksheets/sheet1.xml"
_QUERY_TABLE_WORKSHEET_RELATIONSHIPS_MEMBER = "xl/worksheets/_rels/sheet1.xml.rels"
_QUERY_TABLE_MEMBER = "xl/queryTables/queryTable1.xml"
_QUERY_TABLE_CONNECTION_MEMBER = "xl/connections.xml"
_QUERY_TABLE_WORKBOOK_RELATIONSHIP_ID = "rIdWCABQueryTableConnections"
_QUERY_TABLE_WORKSHEET_RELATIONSHIP_ID = "rIdWCABQueryTable"
_QUERY_TABLE_CONNECTION_ID = 1
_QUERY_TABLE_NAME = "WCAB synthetic query table"
_QUERY_TABLE_CONNECTION_NAME = "WCAB synthetic query-table connection"
_QUERY_TABLE_SOURCE_URL = "https://example.invalid/wcab-query-table-refresh"
_QUERY_TABLE_BASELINE_REFRESH_ON_LOAD = False
_QUERY_TABLE_CANDIDATE_REFRESH_ON_LOAD = True
_QUERY_TABLE_BACKGROUND_REFRESH = False
_QUERY_TABLE_REFRESH_DISABLED = False
_QUERY_TABLE_REMOVE_DATA_ON_SAVE = False
_QUERY_TABLE_FILL_FORMULAS = False
_QUERY_TABLE_CONNECTION_EDIT_DISABLED = True
_QUERY_TABLE_GROWTH_BEHAVIOR = "insertClear"
_QUERY_TABLE_SAVED_VALUE_CELL = "B2"
_QUERY_TABLE_SAVED_VALUE = 100
_QUERY_TABLE_SUMMARY_SHEET = "Summary"
_QUERY_TABLE_SUMMARY_CELL = "B2"
_QUERY_TABLE_SUMMARY_FORMULA = "=ImportedData!$B$2"
_QUERY_TABLE_DASHBOARD_SHEET = "Dashboard"
_QUERY_TABLE_DASHBOARD_CELL = "B4"
_QUERY_TABLE_DASHBOARD_FORMULA = "=Summary!$B$2"
_CELL_HYPERLINK_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/hyperlink"
_CELL_HYPERLINK_SHEET = "Inputs"
_CELL_HYPERLINK_CELL = "B2"
_CELL_HYPERLINK_VALUE = "Open vendor portal"
_CELL_HYPERLINK_WORKSHEET_MEMBER = "xl/worksheets/sheet1.xml"
_CELL_HYPERLINK_WORKSHEET_RELATIONSHIPS_MEMBER = "xl/worksheets/_rels/sheet1.xml.rels"
_CELL_HYPERLINK_RELATIONSHIP_ID = "rIdWCABVendorPortal"
_CELL_HYPERLINK_BASELINE_TARGET = "https://approved.example.invalid/wcab-vendor-portal"
_CELL_HYPERLINK_CANDIDATE_TARGET = "https://review.example.invalid/wcab-vendor-portal"
_CELL_HYPERLINK_SUMMARY_SHEET = "Summary"
_CELL_HYPERLINK_SUMMARY_CELL = "B2"
_CELL_HYPERLINK_SUMMARY_FORMULA = "=Inputs!$B$2"
_CELL_HYPERLINK_DASHBOARD_SHEET = "Dashboard"
_CELL_HYPERLINK_DASHBOARD_CELL = "B4"
_CELL_HYPERLINK_DASHBOARD_FORMULA = "=Summary!$B$2"
_PIVOT_CACHE_ID = 1
_PIVOT_CACHE_SOURCE_SHEET = "Source"
_PIVOT_CACHE_SOURCE_REF = "A1:B5"
_PIVOT_REPORT_SHEET = "Report"
_PIVOT_REPORT_REF = "A1:B2"
_PIVOT_CACHE_DASHBOARD_FORMULA = "=Report!$B$2"
_PIVOT_DATA_FIELD_SOURCE_INDEX = 1
_PIVOT_DATA_FIELD_BASELINE_SUBTOTAL = "sum"
_PIVOT_DATA_FIELD_CANDIDATE_SUBTOTAL = "average"
_OFFICE_2010_SPREADSHEET_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
_PIVOT_SLICER_CACHE_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2007/relationships/slicerCache"
)
_PIVOT_SLICER_CACHE_CONTENT_TYPE = "application/vnd.ms-excel.slicerCache+xml"
_PIVOT_SLICER_CACHE_EXTENSION_URI = "{BBE1A952-AA13-448E-AADC-164F8A28A991}"
_PIVOT_SLICER_PIVOT_CACHE_EXTENSION_URI = "{725AE2AE-9491-48BE-B2B4-4EB974FC3084}"
_PIVOT_SLICER_NAME = "WCAB Region slicer"
_PIVOT_SLICER_SOURCE_NAME = "Region"
_PIVOT_SLICER_PIVOT_TABLE_NAME = "WCAB Pivot Report"
_PIVOT_SLICER_PIVOT_TAB_ID = 2
_PIVOT_SLICER_ITEM_COUNT = 2
_PIVOT_SLICER_BASELINE_SELECTED_INDEX = 0
_PIVOT_SLICER_CANDIDATE_SELECTED_INDEX = 1
_DATA_MASHUP_NS = "http://schemas.microsoft.com/DataMashup"
_POWER_QUERY_CUSTOM_XML_MEMBER = "customXml/item1.xml"
_POWER_QUERY_ROOT_RELATIONSHIP_ID = "rIdWCABPowerQuery"
_POWER_QUERY_SOURCE_SHEET = "Source"
_POWER_QUERY_SOURCE_TABLE = "SourceData"
_POWER_QUERY_SOURCE_REF = "A1:B5"
_POWER_QUERY_SECTION = "Section1"
_POWER_QUERY_NAME = "RegionQuery"
_POWER_QUERY_FILTER_COLUMN = "Region"
_POWER_QUERY_BASELINE_FILTER_VALUE = "North"
_POWER_QUERY_CANDIDATE_FILTER_VALUE = "South"
_SCENARIO_MANAGER_SHEET = "Inputs"
_SCENARIO_MANAGER_DASHBOARD_SHEET = "Dashboard"
_SCENARIO_MANAGER_CHANGING_CELL = "B2"
_SCENARIO_MANAGER_STABLE_INPUT_CELL = "B3"
_SCENARIO_MANAGER_RESULT_CELL = "D2"
_SCENARIO_MANAGER_DASHBOARD_CELL = "B4"
_SCENARIO_MANAGER_WORKSHEET_INPUT_VALUE = 0.1
_SCENARIO_MANAGER_WORKSHEET_STABLE_INPUT_VALUE = 125
_SCENARIO_MANAGER_FORMULA = "=B2*B3"
_SCENARIO_MANAGER_DASHBOARD_FORMULA = "=Inputs!$D$2"
_SCENARIO_MANAGER_NAME = "WCAB downside"
_SCENARIO_MANAGER_COMMENT = "Synthetic downside assumption set"
_SCENARIO_MANAGER_USER = "WCAB"
_SCENARIO_MANAGER_BASELINE_STORED_VALUE = "0.08"
_SCENARIO_MANAGER_CANDIDATE_STORED_VALUE = "0.16"
_SCENARIO_MANAGER_STABLE_STORED_VALUE = "125"
_SCENARIO_MANAGER_INPUT_NUMBER_FORMAT_ID = 10
_WHAT_IF_DATA_TABLE_SHEET = "Sensitivity"
_WHAT_IF_DATA_TABLE_MODEL_SHEET = "Model"
_WHAT_IF_DATA_TABLE_DASHBOARD_SHEET = "Dashboard"
_WHAT_IF_DATA_TABLE_MASTER_CELL = "D3"
_WHAT_IF_DATA_TABLE_OUTPUT_RANGE = "D3:D5"
_WHAT_IF_DATA_TABLE_PRIMARY_INPUT_CELL = "B2"
_WHAT_IF_DATA_TABLE_ALTERNATE_INPUT_CELL = "B3"
_WHAT_IF_DATA_TABLE_SCALE_CELL = "B4"
_WHAT_IF_DATA_TABLE_INPUT_VALUE_RANGE = "C3:C5"
_WHAT_IF_DATA_TABLE_PRIMARY_INPUT_VALUE = 0.08
_WHAT_IF_DATA_TABLE_ALTERNATE_INPUT_VALUE = 0.12
_WHAT_IF_DATA_TABLE_SCALE_VALUE = 100
_WHAT_IF_DATA_TABLE_GRID_VALUES = (0.04, 0.08, 0.12)
_WHAT_IF_DATA_TABLE_OUTPUT_FORMULA_CELL = "D2"
_WHAT_IF_DATA_TABLE_OUTPUT_FORMULA = "=Model!$B$2"
_WHAT_IF_DATA_TABLE_MODEL_CELL = "B2"
_WHAT_IF_DATA_TABLE_MODEL_FORMULA = "=Sensitivity!$B$2*Sensitivity!$B$3*Sensitivity!$B$4"
_WHAT_IF_DATA_TABLE_DASHBOARD_CELL = "B4"
_WHAT_IF_DATA_TABLE_DASHBOARD_FORMULA = "=Model!$B$2"
_DATA_VALIDATION_LIST_SHEET = "Inputs"
_DATA_VALIDATION_LIST_SOURCE_SHEET = "Lists"
_DATA_VALIDATION_LIST_MODEL_SHEET = "Model"
_DATA_VALIDATION_LIST_DASHBOARD_SHEET = "Dashboard"
_DATA_VALIDATION_LIST_TARGET_RANGE = "B2"
_DATA_VALIDATION_LIST_BASELINE_SOURCE_FORMULA = "=Lists!$A$2:$A$4"
_DATA_VALIDATION_LIST_CANDIDATE_SOURCE_FORMULA = "=Lists!$B$2:$B$4"
_DATA_VALIDATION_LIST_BASELINE_SOURCE_RANGE = "A2:A4"
_DATA_VALIDATION_LIST_CANDIDATE_SOURCE_RANGE = "B2:B4"
_DATA_VALIDATION_LIST_BASELINE_SOURCE_VALUES = ("Draft", "Review", "Approved")
_DATA_VALIDATION_LIST_CANDIDATE_SOURCE_VALUES = ("Draft", "Suspended", "Rejected")
_DATA_VALIDATION_LIST_INPUT_VALUE = "Draft"
_DATA_VALIDATION_LIST_MODEL_CELL = "B2"
_DATA_VALIDATION_LIST_MODEL_FORMULA = "=Inputs!$B$2"
_DATA_VALIDATION_LIST_DASHBOARD_CELL = "B4"
_DATA_VALIDATION_LIST_DASHBOARD_FORMULA = "=Model!$B$2"
_DATA_VALIDATION_LIST_ERROR_STYLE = "stop"
_DATA_VALIDATION_LIST_ERROR_TITLE = "Invalid status"
_DATA_VALIDATION_LIST_ERROR = "Choose an approved status."
_DATA_VALIDATION_LIST_PROMPT_TITLE = "Approved status"
_DATA_VALIDATION_LIST_PROMPT = "Choose a documented status."
_CONDITIONAL_FORMATTING_THRESHOLD_SHEET = "Operations"
_CONDITIONAL_FORMATTING_THRESHOLD_RANGE = "B2:B4"
_CONDITIONAL_FORMATTING_THRESHOLD_BASELINE_FORMULA = "100"
_CONDITIONAL_FORMATTING_THRESHOLD_CANDIDATE_FORMULA = "50"
_CONDITIONAL_FORMATTING_THRESHOLD_VALUES = (10, 75, 120)
_CONDITIONAL_FORMATTING_THRESHOLD_FILL_RGB = "FFFFC7CE"
_NUMBER_FORMAT_VISIBILITY_SHEET = "Operations"
_NUMBER_FORMAT_VISIBILITY_CELL = "B2"
_NUMBER_FORMAT_VISIBILITY_VALUE = 0.125
_NUMBER_FORMAT_VISIBILITY_BASELINE_FORMAT = "0.0%;[Red](0.0%);-"
_NUMBER_FORMAT_VISIBILITY_CANDIDATE_FORMAT = ";;;"
_NUMBER_FORMAT_VISIBILITY_CUSTOM_ID = 164
_NUMBER_FORMAT_VISIBILITY_FORMULA_CELL = "B3"
_NUMBER_FORMAT_VISIBILITY_FORMULA = "=B2"
_IGNORED_ERROR_SUPPRESSION_SHEET = "Operations"
_IGNORED_ERROR_SUPPRESSION_TARGET_RANGE = "B5"
_IGNORED_ERROR_SUPPRESSION_FLAG = "formulaRange"
_IGNORED_ERROR_SUPPRESSION_FORMULA = "=SUM(B2:B3)"
_IGNORED_ERROR_SUPPRESSION_ADJACENT_CELL = "B4"
_IGNORED_ERROR_SUPPRESSION_ADJACENT_VALUE = 30
_IGNORED_ERROR_SUPPRESSION_DOWNSTREAM_CELL = "C5"
_IGNORED_ERROR_SUPPRESSION_DOWNSTREAM_FORMULA = "=B5"
_WORKBOOK_STRUCTURE_PROTECTION_INPUT_SHEET = "Inputs"
_WORKBOOK_STRUCTURE_PROTECTION_HIDDEN_SHEET = "ReviewControls"
_WORKBOOK_STRUCTURE_PROTECTION_FORMULA_CELL = "D2"
_WORKBOOK_STRUCTURE_PROTECTION_FORMULA = "=B2*C2"
_SHEET_PROTECTION_SORT_SHEET = "Controls"
_SHEET_PROTECTION_SORT_FORMULA_CELL = "D2"
_SHEET_PROTECTION_SORT_FORMULA = "=B2*C2"
_SHEET_PROTECTION_SORT_DASHBOARD_SHEET = "Dashboard"
_SHEET_PROTECTION_SORT_DASHBOARD_CELL = "B4"
_SHEET_PROTECTION_SORT_DASHBOARD_FORMULA = "=Controls!$D$2"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_SHEET = "Controls"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_WORKSHEET_MEMBER = "xl/worksheets/sheet1.xml"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_INPUT_CELL = "B2"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_INPUT_VALUE = 12
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_FORMULA_CELL = "D2"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_FORMULA = "=B2*C2"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_SHEET = "Dashboard"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_CELL = "B4"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_FORMULA = "=Controls!$D$2"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_RANGE_NAME = "WCAB controlled input"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_RANGE_REF = "B2:B2"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_LEGACY_VERIFIER = "A1B2"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_BASELINE = "approved-editor@wcab.invalid"
_PROTECTED_RANGE_SECURITY_DESCRIPTOR_CANDIDATE = "review-editor@wcab.invalid"
_TABLE_CALCULATED_COLUMN_SHEET = "Ledger"
_TABLE_CALCULATED_COLUMN_NAME = "ScenarioLedger"
_TABLE_CALCULATED_COLUMN_REF = "A1:C4"
_TABLE_CALCULATED_COLUMN_MEMBER = "xl/tables/table1.xml"
_TABLE_CALCULATED_COLUMN_ID = 3
_TABLE_CALCULATED_COLUMN_HEADER = "Calculated amount"
_TABLE_CALCULATED_COLUMN_BASELINE_FORMULA = "A2*B2"
_TABLE_CALCULATED_COLUMN_CANDIDATE_FORMULA = "A2*(B2+1)"
_TABLE_CALCULATED_COLUMN_STABLE_FORMULA_CELLS = ("C2", "C3", "C4")
_TABLE_CALCULATED_COLUMN_DASHBOARD_SHEET = "Dashboard"
_TABLE_CALCULATED_COLUMN_DASHBOARD_CELL = "B4"
_TABLE_CALCULATED_COLUMN_DASHBOARD_FORMULA = "=SUM(ScenarioLedger[Calculated amount])"
_POWER_PIVOT_DATA_WORKBOOK_MEMBER = "xl/workbook.xml"
_POWER_PIVOT_DATA_WORKBOOK_RELATIONSHIPS_MEMBER = "xl/_rels/workbook.xml.rels"
_POWER_PIVOT_DATA_MEMBER = "xl/model/item.data"
_POWER_PIVOT_DATA_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/powerPivotData"
_POWER_PIVOT_DATA_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.model+data"
_POWER_PIVOT_DATA_WORKBOOK_RELATIONSHIP_ID = "rIdWCABPowerPivotData"
_POWER_PIVOT_DATA_EXTENSION_URI = "{FCE2AD5D-F65C-4FA6-A056-5C36A1767C68}"
_POWER_PIVOT_DATA_MIN_VERSION_LOAD = "5"
_POWER_PIVOT_DATA_FROM_TABLE = "SalesModel"
_POWER_PIVOT_DATA_FROM_COLUMN = "CalendarKey"
_POWER_PIVOT_DATA_TO_TABLE = "CalendarModel"
_POWER_PIVOT_DATA_BASELINE_TO_COLUMN = "DateKey"
_POWER_PIVOT_DATA_CANDIDATE_TO_COLUMN = "FiscalDateKey"
_POWER_PIVOT_DATA_MODEL_TABLES = (
    {
        "id": "SalesModel_{11111111-1111-1111-1111-111111111111}",
        "name": _POWER_PIVOT_DATA_FROM_TABLE,
        "connection": "SalesModel",
    },
    {
        "id": "CalendarModel_{22222222-2222-2222-2222-222222222222}",
        "name": _POWER_PIVOT_DATA_TO_TABLE,
        "connection": "CalendarModel",
    },
)
_POWER_PIVOT_DATA_PAYLOAD = (
    b"WCAB opaque synthetic Power Pivot Data Model payload v1; never deserialized or opened."
)
_XLM_AUTO_OPEN_WORKBOOK_MEMBER = "xl/workbook.xml"
_XLM_AUTO_OPEN_WORKBOOK_RELATIONSHIPS_MEMBER = "xl/_rels/workbook.xml.rels"
_XLM_AUTO_OPEN_MACRO_SHEET_MEMBER = "xl/macrosheets/sheet1.xml"
_XLM_AUTO_OPEN_MACRO_SHEET_NS = "http://schemas.microsoft.com/office/excel/2006/main"
_XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2006/relationships/xlMacrosheet"
)
_XLM_AUTO_OPEN_MACRO_SHEET_CONTENT_TYPE = "application/vnd.ms-excel.macrosheet+xml"
_XLM_AUTO_OPEN_WORKBOOK_CONTENT_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
_XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP_ID = "rIdWCABXlmMacroSheet"
_XLM_AUTO_OPEN_MACRO_SHEET_NAME = "Macro Automation"
_XLM_AUTO_OPEN_MACRO_SHEET_ID = "4"
_XLM_AUTO_OPEN_MACRO_SHEET_STATE = "veryHidden"
_XLM_AUTO_OPEN_DEFINED_NAME = "_xlnm.Auto_Open"
_XLM_AUTO_OPEN_EVENT = "Auto_Open"
_XLM_AUTO_OPEN_BASELINE_TARGET = "'Macro Automation'!$A$1"
_XLM_AUTO_OPEN_CANDIDATE_TARGET = "'Macro Automation'!$A$2"
_XLM_AUTO_OPEN_MACRO_FORMULA = "HALT()"
_XLM_AUTO_OPEN_MACRO_SHEET_PAYLOAD = (
    b'<?xml version="1.0" encoding="UTF-8"?>'
    b'<macrosheet xmlns="http://schemas.microsoft.com/office/excel/2006/main">'
    b'<sheetData><row r="1"><c r="A1"><f>HALT()</f></c></row>'
    b'<row r="2"><c r="A2"><f>HALT()</f></c></row></sheetData></macrosheet>'
)
_XLM_AUTO_OPEN_INPUT_SHEET = "Inputs"
_XLM_AUTO_OPEN_INPUT_CELL = "B2"
_XLM_AUTO_OPEN_INPUT_VALUE = 10
_XLM_AUTO_OPEN_MODEL_SHEET = "Model"
_XLM_AUTO_OPEN_MODEL_CELL = "B2"
_XLM_AUTO_OPEN_MODEL_FORMULA = "=Inputs!$B$2*2"
_XLM_AUTO_OPEN_DASHBOARD_SHEET = "Dashboard"
_XLM_AUTO_OPEN_DASHBOARD_CELL = "B4"
_XLM_AUTO_OPEN_DASHBOARD_FORMULA = "=Model!$B$2"
_CHART_SERIES_SOURCE_SHEET = "Source"
_CHART_SERIES_DASHBOARD_SHEET = "Dashboard"
_CHART_SERIES_ANCHOR = "D2"
_CHART_SERIES_TITLE_REFERENCE = "'Source'!B1"
_CHART_SERIES_CATEGORY_REFERENCE = "'Source'!$A$2:$A$4"
_CHART_SERIES_BASELINE_VALUE_REFERENCE = "'Source'!$B$2:$B$4"
_CHART_SERIES_CANDIDATE_VALUE_REFERENCE = "'Source'!$C$2:$C$4"
_DRAWINGML_CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"


def _configure_workbook(workbook: Workbook, *, title: str) -> None:
    workbook.properties.creator = "WCAB"
    workbook.properties.lastModifiedBy = "WCAB"
    workbook.properties.title = title
    workbook.properties.subject = "Synthetic workbook change-assurance fixture"
    workbook.properties.created = _FIXED_TIMESTAMP
    workbook.properties.modified = _FIXED_TIMESTAMP
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False


def _write_canonical_xlsx_members(path: Path, members: dict[str, bytes]) -> None:
    """Write OOXML package members with stable ordering and metadata."""

    core_properties = members.get("docProps/core.xml")
    if core_properties is not None:
        members["docProps/core.xml"] = _CORE_MODIFIED.sub(
            rb"\g<1>2024-01-01T00:00:00Z\g<2>", core_properties
        )
    with NamedTemporaryFile(dir=path.parent, prefix=f".{path.name}.", delete=False) as handle:
        temporary = Path(handle.name)
    try:
        with ZipFile(temporary, "w", compression=ZIP_DEFLATED, compresslevel=9) as target:
            for name in sorted(members):
                info = ZipInfo(filename=name, date_time=(1980, 1, 1, 0, 0, 0))
                info.compress_type = ZIP_DEFLATED
                info.create_system = 3
                info.external_attr = 0o100644 << 16
                target.writestr(info, members[name])
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)


def _canonicalize_xlsx(path: Path) -> None:
    """Rewrite a generated XLSX with stable ZIP order and timestamps.

    Openpyxl's XML generation is deterministic for the supported dependency
    versions, but ZIP member timestamps otherwise make byte-for-byte fixture
    reproduction impossible.
    """

    with ZipFile(path, "r") as source:
        members = {info.filename: source.read(info.filename) for info in source.infolist()}
    _write_canonical_xlsx_members(path, members)


def _rewrite_xlsx_parts(path: Path, mutate: ArchiveMutator) -> None:
    """Apply a deterministic, raw-OOXML fixture mutation to an XLSX package."""

    with ZipFile(path, "r") as source:
        members = {info.filename: source.read(info.filename) for info in source.infolist()}
    mutate(members)
    _write_canonical_xlsx_members(path, members)


def _save_workbook(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    _canonicalize_xlsx(path)


def _write_json(path: Path, value: dict[str, Any]) -> None:
    path.write_text(json.dumps(value, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _truth(
    *,
    case_id: str,
    title: str,
    family: str,
    review_expectation: str,
    facts: list[dict[str, Any]],
    must_reach: list[dict[str, Any]] | None = None,
    coverage: list[str] | None = None,
    coverage_expectations: list[dict[str, Any]] | None = None,
    topology: str = "pair",
) -> dict[str, Any]:
    return {
        "schema_version": FIXTURE_SCHEMA_VERSION,
        "id": case_id,
        "title": title,
        "family": family,
        "topology": topology,
        "review_expectation": review_expectation,
        "facts": facts,
        "must_reach": must_reach or [],
        "coverage": coverage or [],
        "coverage_expectations": coverage_expectations or [],
    }


def _write_pair(
    directory: Path,
    factory: WorkbookFactory,
    mutate_candidate: WorkbookMutator,
    truth: dict[str, Any],
) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    _save_workbook(factory(), directory / "baseline.xlsx")
    candidate = factory()
    mutate_candidate(candidate)
    _save_workbook(candidate, directory / "candidate.xlsx")
    _write_json(directory / "truth.json", truth)


def _finance_workbook() -> Workbook:
    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB finance forecast")
    assumptions = workbook.active
    assumptions.title = "Assumptions"
    assumptions.append(["Driver", "2024", "2025", "2026"])
    assumptions.append(["Units", 100, 115, 130])
    assumptions.append(["Price", 8.5, 8.7, 9.0])
    assumptions.append(["Cost per unit", 3.2, 3.4, 3.5])
    assumptions["B10"] = "Discount rate"
    assumptions["E10"] = 0.10
    assumptions["B11"] = "Stress discount rate"
    assumptions["E11"] = 0.12

    revenue = workbook.create_sheet("Revenue")
    revenue.append(["Metric", "2024", "2025", "2026"])
    for column, assumption_column in zip(("B", "C", "D"), ("B", "C", "D"), strict=True):
        revenue[f"{column}5"] = f"=Assumptions!{assumption_column}2"
        revenue[f"{column}6"] = f"=Assumptions!{assumption_column}3"
        revenue[f"{column}8"] = f"={column}5*{column}6"
        revenue[f"{column}9"] = f"={column}5*Assumptions!{assumption_column}4"
        revenue[f"{column}10"] = f"={column}8-{column}9"
    revenue["A5"] = "Units"
    revenue["A6"] = "Price"
    revenue["A8"] = "Revenue"
    revenue["A9"] = "Cost of goods"
    revenue["A10"] = "Gross profit"

    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "2024", "2025", "2026"])
    for column in ("B", "C", "D"):
        summary[f"{column}5"] = f"=Revenue!{column}8"
        summary[f"{column}6"] = f"=Revenue!{column}10"
        summary[f"{column}8"] = f"=Assumptions!{column}2"
        summary[f"{column}10"] = f"={column}6*(1-DiscountRate)"
    summary["A5"] = "Revenue"
    summary["A6"] = "Gross profit"
    summary["A8"] = "Units source"
    summary["A10"] = "Discounted gross profit"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Board dashboard"
    dashboard["A4"] = "2025 gross profit"
    dashboard["B4"] = "=Summary!C6"
    dashboard["A5"] = "2025 discounted gross profit"
    dashboard["B5"] = "=Summary!C10"

    workbook.defined_names.add(
        DefinedName("DiscountRate", attr_text="Assumptions!$E$10", comment="WCAB synthetic driver")
    )
    return workbook


def _operations_workbook() -> Workbook:
    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB operations tracker")
    sheet = workbook.active
    sheet.title = "Operations"
    sheet.append(
        ["Region", "SKU", "Units", "Unit cost", "Unit price", "Margin", "Status", "Regional sales"]
    )
    rows = [
        ("North", "A-100", 12, 4, 9),
        ("North", "B-200", 10, 5, 11),
        ("South", "A-100", 9, 4, 9),
        ("South", "B-200", 16, 5, 11),
    ]
    for row_number, (region, sku, units, cost, price) in enumerate(rows, start=6):
        sheet.cell(row_number, 1, region)
        sheet.cell(row_number, 2, sku)
        sheet.cell(row_number, 3, units)
        sheet.cell(row_number, 4, cost)
        sheet.cell(row_number, 5, price)
        sheet.cell(row_number, 6, f"=C{row_number}*(E{row_number}-D{row_number})")
        sheet.cell(row_number, 8, f"=SUMIFS($F$6:$F$9,$A$6:$A$9,A{row_number})")

    validation = DataValidation(type="whole", operator="between", formula1="0", formula2="1000")
    validation.promptTitle = "Units"
    validation.prompt = "Enter a whole number between 0 and 1000."
    sheet.add_data_validation(validation)
    validation.add("C6:C9")
    high_margin = PatternFill(fill_type="solid", fgColor="C6EFCE")
    sheet.conditional_formatting.add(
        "F6:F9", CellIsRule(operator="greaterThan", formula=["50"], fill=high_margin)
    )
    return workbook


def _data_validation_list_source_workbook() -> Workbook:
    """Build a list-controlled input with two stable local source lists.

    The candidate is later changed at the raw SpreadsheetML ``formula1``
    declaration only. The source lists, current input, and ordinary formula
    path remain unchanged so this fixture records a future entry-control
    boundary rather than a calculated model result.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB list validation source fixture")
    inputs = workbook.active
    inputs.title = _DATA_VALIDATION_LIST_SHEET
    inputs["A1"] = "Status input"
    inputs["A2"] = "Current status"
    inputs[_DATA_VALIDATION_LIST_TARGET_RANGE] = _DATA_VALIDATION_LIST_INPUT_VALUE
    validation = DataValidation(
        type="list",
        formula1=_DATA_VALIDATION_LIST_BASELINE_SOURCE_FORMULA,
        allow_blank=False,
        showErrorMessage=True,
        errorStyle=_DATA_VALIDATION_LIST_ERROR_STYLE,
        errorTitle=_DATA_VALIDATION_LIST_ERROR_TITLE,
        error=_DATA_VALIDATION_LIST_ERROR,
        showInputMessage=False,
        promptTitle=_DATA_VALIDATION_LIST_PROMPT_TITLE,
        prompt=_DATA_VALIDATION_LIST_PROMPT,
    )
    validation.showDropDown = False
    validation.add(_DATA_VALIDATION_LIST_TARGET_RANGE)
    inputs.add_data_validation(validation)

    lists = workbook.create_sheet(_DATA_VALIDATION_LIST_SOURCE_SHEET)
    lists["A1"] = "Approved statuses"
    lists["B1"] = "Restricted statuses"
    for row, (baseline_value, candidate_value) in enumerate(
        zip(
            _DATA_VALIDATION_LIST_BASELINE_SOURCE_VALUES,
            _DATA_VALIDATION_LIST_CANDIDATE_SOURCE_VALUES,
            strict=True,
        ),
        start=2,
    ):
        lists.cell(row, 1, baseline_value)
        lists.cell(row, 2, candidate_value)

    model = workbook.create_sheet(_DATA_VALIDATION_LIST_MODEL_SHEET)
    model["A1"] = "Selected status"
    model[_DATA_VALIDATION_LIST_MODEL_CELL] = _DATA_VALIDATION_LIST_MODEL_FORMULA

    dashboard = workbook.create_sheet(_DATA_VALIDATION_LIST_DASHBOARD_SHEET)
    dashboard["A1"] = "Board status"
    dashboard["A4"] = "Selected status"
    dashboard[_DATA_VALIDATION_LIST_DASHBOARD_CELL] = _DATA_VALIDATION_LIST_DASHBOARD_FORMULA
    return workbook


def _conditional_formatting_threshold_workbook() -> Workbook:
    """Build one stable exception-highlight rule with stored metric values.

    Conditional-formatting formulas live alongside worksheet controls rather
    than as ordinary cell formulas. The candidate is later changed only at the
    rule's raw threshold formula, which lets the fixture measure a visual
    control transition without asking a spreadsheet client to render it.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB conditional-formatting threshold fixture")
    operations = workbook.active
    operations.title = _CONDITIONAL_FORMATTING_THRESHOLD_SHEET
    operations["A1"] = "Metric"
    operations["B1"] = "Exception review metric"
    for row, value in enumerate(_CONDITIONAL_FORMATTING_THRESHOLD_VALUES, start=2):
        operations.cell(row, 1, f"Period {row - 1}")
        operations.cell(row, 2, value)
    exception_fill = PatternFill(
        fill_type="solid", fgColor=_CONDITIONAL_FORMATTING_THRESHOLD_FILL_RGB
    )
    operations.conditional_formatting.add(
        _CONDITIONAL_FORMATTING_THRESHOLD_RANGE,
        CellIsRule(
            operator="greaterThan",
            formula=[_CONDITIONAL_FORMATTING_THRESHOLD_BASELINE_FORMULA],
            fill=exception_fill,
        ),
    )
    return workbook


def _number_format_visibility_workbook() -> Workbook:
    """Build one reported metric with a custom display format.

    The candidate is later changed only in the raw custom numFmt declaration in
    styles.xml. Its target cell, style index, numeric value, and neighboring
    formula therefore remain stable; the fixture records stored display
    metadata without asking a spreadsheet client to render it.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB custom number-format fixture")
    operations = workbook.active
    operations.title = _NUMBER_FORMAT_VISIBILITY_SHEET
    operations["A1"] = "Metric"
    operations["B1"] = "Reported margin"
    operations["A2"] = "Base case"
    operations[_NUMBER_FORMAT_VISIBILITY_CELL] = _NUMBER_FORMAT_VISIBILITY_VALUE
    operations[
        _NUMBER_FORMAT_VISIBILITY_CELL
    ].number_format = _NUMBER_FORMAT_VISIBILITY_BASELINE_FORMAT
    operations["A3"] = "Raw model value"
    operations[_NUMBER_FORMAT_VISIBILITY_FORMULA_CELL] = _NUMBER_FORMAT_VISIBILITY_FORMULA
    return workbook


def _ignored_error_suppression_workbook() -> Workbook:
    """Build a stable omitted-range review surface with a downstream formula.

    The candidate later receives a raw SpreadsheetML ``ignoredError`` record.
    Every ordinary cell and formula stays fixed, so the pair observes a stored
    decision to suppress one Excel error-checking warning without asserting
    whether a spreadsheet client would display that warning.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB ignored-error suppression fixture")
    operations = workbook.active
    operations.title = _IGNORED_ERROR_SUPPRESSION_SHEET
    operations.append(["Period", "Amount", "Reported total"])
    for row, amount in enumerate((10, 20, _IGNORED_ERROR_SUPPRESSION_ADJACENT_VALUE), start=2):
        operations.cell(row, 1, f"Period {row - 1}")
        operations.cell(row, 2, amount)
    operations["A5"] = "Reported total"
    operations[_IGNORED_ERROR_SUPPRESSION_TARGET_RANGE] = _IGNORED_ERROR_SUPPRESSION_FORMULA
    operations[_IGNORED_ERROR_SUPPRESSION_DOWNSTREAM_CELL] = (
        _IGNORED_ERROR_SUPPRESSION_DOWNSTREAM_FORMULA
    )
    return workbook


def _auto_filter_criteria_workbook(filter_value: str) -> Workbook:
    """Build a report with an active stored AutoFilter criterion.

    The generated pair differs only in the criterion text. Its ``SUBTOTAL``
    formula and downstream consumer deliberately remain fixed: WCAB records
    the stored filter boundary and never calculates which rows or result an
    Excel client would display.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB AutoFilter criterion fixture")
    report = workbook.active
    report.title = "Report"
    report.append(["Region", "Amount"])
    for region, amount in (("North", 100), ("North", 200), ("South", 300), ("South", 400)):
        report.append([region, amount])
    report["D1"] = "Visible amount total"
    report["D2"] = _AUTO_FILTER_SUBTOTAL_FORMULA
    report.auto_filter.ref = _AUTO_FILTER_REF
    report.auto_filter.add_filter_column(_AUTO_FILTER_COLUMN_ID, [filter_value])

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A4"] = "Visible regional amount total"
    dashboard["B4"] = _AUTO_FILTER_DASHBOARD_FORMULA
    return workbook


def _named_sheet_view_filter_workbook() -> Workbook:
    """Build a report whose alternate Sheet View later changes in raw OOXML.

    The ordinary worksheet AutoFilter stays present but has no active criterion.
    A raw, relationship-backed Named Sheet View will retain the alternate
    criterion instead, so the pair records a saved review surface without
    changing cells, formulas, or the worksheet's active filter.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB Named Sheet View fixture")
    report = workbook.active
    report.title = _NAMED_SHEET_VIEW_REPORT_SHEET
    report.append(["Region", "Amount"])
    for region, amount in (("North", 100), ("North", 200), ("South", 300), ("South", 400)):
        report.append([region, amount])
    report["D1"] = "Visible amount total"
    report[_NAMED_SHEET_VIEW_SUBTOTAL_CELL] = _NAMED_SHEET_VIEW_SUBTOTAL_FORMULA
    report.auto_filter.ref = _NAMED_SHEET_VIEW_FILTER_REF

    dashboard = workbook.create_sheet(_NAMED_SHEET_VIEW_DASHBOARD_SHEET)
    dashboard["A4"] = "Visible regional amount total"
    dashboard[_NAMED_SHEET_VIEW_DASHBOARD_CELL] = _NAMED_SHEET_VIEW_DASHBOARD_FORMULA
    return workbook


def _inject_named_sheet_view_filter(path: Path, *, filter_value: str) -> None:
    """Attach WCAB's one stored alternate filter to a generated workbook.

    ``openpyxl`` does not author Named Sheet View parts, so this writes the
    narrow relationship-backed OOXML contract after ordinary cells and formulas
    are generated. The generated base AutoFilter supplies the documented
    binding target; no workbook client is opened or used to apply the view.
    """

    def serialize(element: ElementTree.Element) -> bytes:
        return ElementTree.tostring(element, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(members[worksheet_member])
        auto_filters = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}autoFilter")
        if len(auto_filters) != 1 or auto_filters[0].get("ref") != _NAMED_SHEET_VIEW_FILTER_REF:
            raise ValueError("Named Sheet View fixture requires one Report AutoFilter")
        auto_filter = auto_filters[0]
        if auto_filter.attrib != {"ref": _NAMED_SHEET_VIEW_FILTER_REF}:
            raise ValueError("Named Sheet View fixture has unexpected base AutoFilter metadata")
        auto_filter.set(f"{{{_OFFICE_2014_REVISION_NS}}}uid", _NAMED_SHEET_VIEW_FILTER_ID)
        members[worksheet_member] = serialize(worksheet)

        if _NAMED_SHEET_VIEW_RELATIONSHIP_MEMBER in members:
            relationships = ElementTree.fromstring(members[_NAMED_SHEET_VIEW_RELATIONSHIP_MEMBER])
        else:
            relationships = ElementTree.Element(f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships")
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        if any(
            relationship.get("Type") == _NAMED_SHEET_VIEW_RELATIONSHIP
            for relationship in relationships.findall(relationship_tag)
        ):
            raise ValueError("Named Sheet View fixture already has a Named Sheet View relationship")
        ElementTree.SubElement(
            relationships,
            relationship_tag,
            {
                "Id": _NAMED_SHEET_VIEW_RELATIONSHIP_ID,
                "Type": _NAMED_SHEET_VIEW_RELATIONSHIP,
                "Target": "../namedSheetViews/namedSheetView1.xml",
            },
        )
        members[_NAMED_SHEET_VIEW_RELATIONSHIP_MEMBER] = serialize(relationships)

        named_views = ElementTree.Element(f"{{{_NAMED_SHEET_VIEW_NS}}}namedSheetViews")
        named_view = ElementTree.SubElement(
            named_views,
            f"{{{_NAMED_SHEET_VIEW_NS}}}namedSheetView",
            {"name": _NAMED_SHEET_VIEW_NAME, "id": _NAMED_SHEET_VIEW_ID},
        )
        named_filter = ElementTree.SubElement(
            named_view,
            f"{{{_NAMED_SHEET_VIEW_NS}}}nsvFilter",
            {
                "filterId": _NAMED_SHEET_VIEW_FILTER_ID,
                "ref": _NAMED_SHEET_VIEW_FILTER_REF,
                "tableId": "0",
            },
        )
        column_filter = ElementTree.SubElement(
            named_filter,
            f"{{{_NAMED_SHEET_VIEW_NS}}}columnFilter",
            {"colId": str(_NAMED_SHEET_VIEW_FILTER_COLUMN_ID), "id": _NAMED_SHEET_VIEW_COLUMN_ID},
        )
        filter_column = ElementTree.SubElement(
            column_filter,
            f"{{{_NAMED_SHEET_VIEW_NS}}}filter",
            {"colId": str(_NAMED_SHEET_VIEW_FILTER_COLUMN_ID)},
        )
        filters = ElementTree.SubElement(
            filter_column,
            f"{{{_SPREADSHEETML_NS}}}filters",
            {"blank": "0"},
        )
        ElementTree.SubElement(filters, f"{{{_SPREADSHEETML_NS}}}filter", {"val": filter_value})
        members[_NAMED_SHEET_VIEW_MEMBER] = serialize(named_views)

        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        if any(
            override.get("PartName") == f"/{_NAMED_SHEET_VIEW_MEMBER}"
            for override in content_types.findall(override_tag)
        ):
            raise ValueError("Named Sheet View fixture already has a content-type override")
        ElementTree.SubElement(
            content_types,
            override_tag,
            {
                "PartName": f"/{_NAMED_SHEET_VIEW_MEMBER}",
                "ContentType": _NAMED_SHEET_VIEW_CONTENT_TYPE,
            },
        )
        members["[Content_Types].xml"] = serialize(content_types)

    _rewrite_xlsx_parts(path, mutate)


def _xml_map_table_workbook() -> Workbook:
    """Build a stable table whose XML-map binding is added after saving.

    The ordinary workbook contains one invoice-line table, a table-driven
    total, and a dashboard consumer.  A raw OOXML step adds a conventional XML
    Map binding because openpyxl does not author XML Maps.  The pair later
    retargets just that binding; it does not import, export, calculate, or
    otherwise materialize data through the map.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB XML Map export fixture")
    export = workbook.active
    export.title = _XML_MAP_EXPORT_SHEET
    export.append(["Item", _XML_MAP_TABLE_COLUMN_NAME])
    export.append(["Consulting", 120])
    export.append(["Software", 180])
    table = Table(displayName=_XML_MAP_TABLE_NAME, ref=_XML_MAP_TABLE_REF)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    export.add_table(table)
    export["D1"] = "Mapped table total"
    export[_XML_MAP_TOTAL_CELL] = _XML_MAP_TOTAL_FORMULA
    export["E1"] = "Export as of"
    export[_XML_MAP_SINGLE_CELL] = date(2024, 1, 1)
    export[_XML_MAP_SINGLE_CELL].number_format = "yyyy-mm-dd"

    dashboard = workbook.create_sheet(_XML_MAP_DASHBOARD_SHEET)
    dashboard["A4"] = "Mapped invoice total"
    dashboard[_XML_MAP_DASHBOARD_CELL] = _XML_MAP_DASHBOARD_FORMULA
    return workbook


def _inject_xml_map_table_binding(path: Path, *, xpath: str) -> None:
    """Attach one local XML Map package and mapped table-column declaration.

    The fixture has a stable schema, map declaration, file binding, and
    worksheet single-cell mapping.  Only the mapped table column's XPath is
    supplied by the caller.  All names and targets are synthetic and local;
    this routine records a package declaration rather than contacting a file,
    executing an import/export operation, or opening an Excel client.
    """

    if xpath not in {_XML_MAP_BASELINE_XPATH, _XML_MAP_CANDIDATE_XPATH}:
        raise ValueError(f"unsupported XML Map XPath {xpath!r}")

    def serialize(element: ElementTree.Element) -> bytes:
        return ElementTree.tostring(element, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        table_relationship = f"{_DOCUMENT_RELATIONSHIPS_NS}/table"
        table_column_tag = f"{{{_SPREADSHEETML_NS}}}tableColumn"
        table_columns_tag = f"{{{_SPREADSHEETML_NS}}}tableColumns"
        relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"

        table = ElementTree.fromstring(members[_XML_MAP_TABLE_MEMBER])
        if table.tag != f"{{{_SPREADSHEETML_NS}}}table":
            raise ValueError("XML Map fixture table has an unexpected root")
        if (
            table.get("displayName") != _XML_MAP_TABLE_NAME
            or table.get("ref") != _XML_MAP_TABLE_REF
        ):
            raise ValueError("XML Map fixture table metadata is unexpected")
        if "tableType" in table.attrib or "connectionId" in table.attrib:
            raise ValueError("XML Map fixture table already has XML mapping metadata")
        table_columns = table.findall(table_columns_tag)
        if len(table_columns) != 1:
            raise ValueError("XML Map fixture requires one tableColumns element")
        columns = table_columns[0].findall(table_column_tag)
        if len(columns) != 2:
            raise ValueError("XML Map fixture requires two table columns")
        mapped_column = columns[1]
        if (
            mapped_column.get("id") != str(_XML_MAP_TABLE_COLUMN_ID)
            or mapped_column.get("name") != _XML_MAP_TABLE_COLUMN_NAME
            or len(mapped_column) != 0
        ):
            raise ValueError("XML Map fixture has an unexpected mapped table column")
        table.set("tableType", "xml")
        table.set("connectionId", str(_XML_MAP_CONNECTION_ID))
        ElementTree.SubElement(
            mapped_column,
            f"{{{_SPREADSHEETML_NS}}}xmlColumnPr",
            {
                "mapId": str(_XML_MAP_ID),
                "xpath": xpath,
                "denormalized": "false",
                "xmlDataType": "double",
            },
        )
        members[_XML_MAP_TABLE_MEMBER] = serialize(table)

        map_info = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}MapInfo",
            {"SelectionNamespaces": f"xmlns:wcab='{_XML_MAP_NAMESPACE}'"},
        )
        schema = ElementTree.SubElement(
            map_info,
            f"{{{_SPREADSHEETML_NS}}}Schema",
            {"ID": _XML_MAP_SCHEMA_ID},
        )
        schema_definition = ElementTree.SubElement(
            schema,
            f"{{{_XML_SCHEMA_NS}}}schema",
            {
                "targetNamespace": _XML_MAP_NAMESPACE,
                "elementFormDefault": "qualified",
            },
        )
        invoice = ElementTree.SubElement(
            schema_definition,
            f"{{{_XML_SCHEMA_NS}}}element",
            {"name": _XML_MAP_ROOT_ELEMENT},
        )
        invoice_type = ElementTree.SubElement(invoice, f"{{{_XML_SCHEMA_NS}}}complexType")
        invoice_sequence = ElementTree.SubElement(invoice_type, f"{{{_XML_SCHEMA_NS}}}sequence")
        header = ElementTree.SubElement(
            invoice_sequence,
            f"{{{_XML_SCHEMA_NS}}}element",
            {"name": "Header", "minOccurs": "0"},
        )
        header_type = ElementTree.SubElement(header, f"{{{_XML_SCHEMA_NS}}}complexType")
        header_sequence = ElementTree.SubElement(header_type, f"{{{_XML_SCHEMA_NS}}}sequence")
        ElementTree.SubElement(
            header_sequence,
            f"{{{_XML_SCHEMA_NS}}}element",
            {"name": "AsOf", "type": "xs:date"},
        )
        line = ElementTree.SubElement(
            invoice_sequence,
            f"{{{_XML_SCHEMA_NS}}}element",
            {"name": "Line", "minOccurs": "0", "maxOccurs": "unbounded"},
        )
        line_type = ElementTree.SubElement(line, f"{{{_XML_SCHEMA_NS}}}complexType")
        line_sequence = ElementTree.SubElement(line_type, f"{{{_XML_SCHEMA_NS}}}sequence")
        for name, data_type in (
            ("Item", "xs:string"),
            ("NetAmount", "xs:decimal"),
            ("TaxAmount", "xs:decimal"),
        ):
            ElementTree.SubElement(
                line_sequence,
                f"{{{_XML_SCHEMA_NS}}}element",
                {"name": name, "type": data_type},
            )
        xml_map = ElementTree.SubElement(
            map_info,
            f"{{{_SPREADSHEETML_NS}}}Map",
            {
                "ID": str(_XML_MAP_ID),
                "Name": "WCAB invoice export",
                "RootElement": _XML_MAP_ROOT_ELEMENT,
                "SchemaID": _XML_MAP_SCHEMA_ID,
                "ShowImportExportValidationErrors": "false",
                "AutoFit": "true",
                "Append": "false",
                "PreserveSortAFLayout": "true",
                "PreserveFormat": "true",
            },
        )
        ElementTree.SubElement(
            xml_map,
            f"{{{_SPREADSHEETML_NS}}}DataBinding",
            {
                "DataBindingName": "WCAB invoice export binding",
                "FileBinding": "true",
                "ConnectionID": str(_XML_MAP_CONNECTION_ID),
                "FileBindingName": "wcab-invoice-export.xml",
                "DataBindingLoadMode": "1",
            },
        )
        members[_XML_MAP_MEMBER] = serialize(map_info)

        workbook_relationships = ElementTree.fromstring(members["xl/_rels/workbook.xml.rels"])
        if any(
            relationship.get("Type") == _XML_MAP_RELATIONSHIP
            for relationship in workbook_relationships.findall(relationship_tag)
        ):
            raise ValueError("XML Map fixture already has an XML Maps relationship")
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": _XML_MAP_WORKBOOK_RELATIONSHIP_ID,
                "Type": _XML_MAP_RELATIONSHIP,
                "Target": "xmlMaps.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        worksheet_relationships_member = "xl/worksheets/_rels/sheet1.xml.rels"
        worksheet_relationships = ElementTree.fromstring(members[worksheet_relationships_member])
        table_relationships = [
            relationship
            for relationship in worksheet_relationships.findall(relationship_tag)
            if relationship.get("Type") == table_relationship
        ]
        if len(table_relationships) != 1:
            raise ValueError("XML Map fixture cannot find its worksheet table relationship")
        if any(
            relationship.get("Type") == _TABLE_SINGLE_CELLS_RELATIONSHIP
            for relationship in worksheet_relationships.findall(relationship_tag)
        ):
            raise ValueError("XML Map fixture already has a single-cell relationship")

        worksheet = ElementTree.fromstring(members["xl/worksheets/sheet1.xml"])
        table_parts = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}tableParts")
        if len(table_parts) != 1 or len(table_parts[0]) != 1:
            raise ValueError("XML Map fixture requires one worksheet table part")
        table_part_relationship_id = table_parts[0][0].get(relationship_id_attribute)
        if table_part_relationship_id != table_relationships[0].get("Id"):
            raise ValueError("XML Map fixture table relationship does not bind the worksheet")

        single_cells = ElementTree.Element(f"{{{_SPREADSHEETML_NS}}}singleXmlCells")
        single_cell = ElementTree.SubElement(
            single_cells,
            f"{{{_SPREADSHEETML_NS}}}singleXmlCell",
            {
                "id": "1",
                "r": _XML_MAP_SINGLE_CELL,
                "connectionId": str(_XML_MAP_CONNECTION_ID),
            },
        )
        cell_properties = ElementTree.SubElement(
            single_cell,
            f"{{{_SPREADSHEETML_NS}}}xmlCellPr",
            {
                "id": "1",
                "uniqueName": "WCAB invoice export as of",
            },
        )
        ElementTree.SubElement(
            cell_properties,
            f"{{{_SPREADSHEETML_NS}}}xmlPr",
            {
                "mapId": str(_XML_MAP_ID),
                "xpath": _XML_MAP_SINGLE_CELL_XPATH,
                "xmlDataType": "date",
            },
        )
        members[_XML_MAP_SINGLE_CELL_MEMBER] = serialize(single_cells)
        ElementTree.SubElement(
            worksheet_relationships,
            relationship_tag,
            {
                "Id": _XML_MAP_SINGLE_CELL_RELATIONSHIP_ID,
                "Type": _TABLE_SINGLE_CELLS_RELATIONSHIP,
                "Target": "../singleCellTables/singleCellTable1.xml",
            },
        )
        members[worksheet_relationships_member] = serialize(worksheet_relationships)

        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        xml_defaults = [
            default
            for default in content_types.findall(default_tag)
            if default.get("Extension") == "xml" and default.get("ContentType") == "application/xml"
        ]
        if len(xml_defaults) != 1:
            raise ValueError("XML Map fixture requires one generic XML content type")
        if any(
            override.get("PartName") == f"/{_XML_MAP_SINGLE_CELL_MEMBER}"
            for override in content_types.findall(override_tag)
        ):
            raise ValueError("XML Map fixture already has a single-cell content-type override")
        ElementTree.SubElement(
            content_types,
            override_tag,
            {
                "PartName": f"/{_XML_MAP_SINGLE_CELL_MEMBER}",
                "ContentType": "application/vnd.ms-excel.tableSingleCells",
            },
        )
        members["[Content_Types].xml"] = serialize(content_types)

    _rewrite_xlsx_parts(path, mutate)


def _office_web_addin_auto_show_workbook() -> Workbook:
    """Build a small model with a later-added Office Web Add-in declaration.

    The ordinary input/model/dashboard values establish stable workbook context
    only. A raw OOXML step adds the document-linked task-pane declaration; it
    does not install, load, or execute an add-in.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB Office Web Add-in control fixture")
    inputs = workbook.active
    inputs.title = _OFFICE_WEB_ADDIN_INPUT_SHEET
    inputs["A1"] = "Approved units"
    inputs[_OFFICE_WEB_ADDIN_INPUT_CELL] = _OFFICE_WEB_ADDIN_INPUT_VALUE

    model = workbook.create_sheet(_OFFICE_WEB_ADDIN_MODEL_SHEET)
    model["A1"] = "Projected units"
    model[_OFFICE_WEB_ADDIN_MODEL_CELL] = _OFFICE_WEB_ADDIN_MODEL_FORMULA

    dashboard = workbook.create_sheet(_OFFICE_WEB_ADDIN_DASHBOARD_SHEET)
    dashboard["A4"] = "Projected units"
    dashboard[_OFFICE_WEB_ADDIN_DASHBOARD_CELL] = _OFFICE_WEB_ADDIN_DASHBOARD_FORMULA
    return workbook


def _inject_office_web_addin_auto_show(path: Path, *, auto_show: bool) -> None:
    """Attach one synthetic document-linked Office Web Add-in declaration.

    Only the persisted Office.AutoShowTaskpaneWithDocument property is supplied
    by the caller. The add-in identifier and FileSystem store reference are
    synthetic; the fixture has no manifest payload or network relationship.
    This routine writes a narrow relationship graph without launching a client
    or resolving a reference.
    """

    if type(auto_show) is not bool:
        raise ValueError("Office Web Add-in auto_show must be boolean")

    def serialize(element: ElementTree.Element) -> bytes:
        return ElementTree.tostring(element, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        if any(member.startswith("xl/webextensions/") for member in members):
            raise ValueError("Office Web Add-in fixture already has web-extension parts")

        workbook_relationships = ElementTree.fromstring(members["xl/_rels/workbook.xml.rels"])
        if any(
            relationship.get("Type") == _WEB_EXTENSION_TASKPANES_RELATIONSHIP
            for relationship in workbook_relationships.findall(relationship_tag)
        ):
            raise ValueError("Office Web Add-in fixture already has a task-pane relationship")
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": _OFFICE_WEB_ADDIN_WORKBOOK_RELATIONSHIP_ID,
                "Type": _WEB_EXTENSION_TASKPANES_RELATIONSHIP,
                "Target": "webextensions/taskpanes.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        for member, content_type in (
            (_OFFICE_WEB_ADDIN_TASKPANES_MEMBER, _WEB_EXTENSION_TASKPANES_CONTENT_TYPE),
            (_OFFICE_WEB_ADDIN_EXTENSION_MEMBER, _WEB_EXTENSION_CONTENT_TYPE),
        ):
            if any(
                override.get("PartName") == f"/{member}"
                for override in content_types.findall(override_tag)
            ):
                raise ValueError(f"Office Web Add-in fixture already has {member!r}")
            ElementTree.SubElement(
                content_types,
                override_tag,
                {"PartName": f"/{member}", "ContentType": content_type},
            )
        members["[Content_Types].xml"] = serialize(content_types)

        taskpanes = ElementTree.Element(f"{{{_WEB_EXTENSION_TASKPANES_NS}}}taskpanes")
        taskpane = ElementTree.SubElement(
            taskpanes,
            f"{{{_WEB_EXTENSION_TASKPANES_NS}}}taskpane",
            {
                "dockstate": "right",
                "visibility": "0",
                "width": "350",
                "row": "4",
                "locked": "1",
            },
        )
        ElementTree.SubElement(
            taskpane,
            f"{{{_WEB_EXTENSION_TASKPANES_NS}}}webextension",
            {relationship_id_attribute: _OFFICE_WEB_ADDIN_EXTENSION_RELATIONSHIP_ID},
        )
        members[_OFFICE_WEB_ADDIN_TASKPANES_MEMBER] = serialize(taskpanes)

        taskpane_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            taskpane_relationships,
            relationship_tag,
            {
                "Id": _OFFICE_WEB_ADDIN_EXTENSION_RELATIONSHIP_ID,
                "Type": _WEB_EXTENSION_RELATIONSHIP,
                "Target": "webextension1.xml",
            },
        )
        members[_OFFICE_WEB_ADDIN_TASKPANES_RELATIONSHIPS_MEMBER] = serialize(
            taskpane_relationships
        )

        extension = ElementTree.Element(
            f"{{{_WEB_EXTENSION_NS}}}webextension",
            {"id": _OFFICE_WEB_ADDIN_ID},
        )
        ElementTree.SubElement(
            extension,
            f"{{{_WEB_EXTENSION_NS}}}reference",
            {
                "id": _OFFICE_WEB_ADDIN_REFERENCE_ID,
                "version": _OFFICE_WEB_ADDIN_REFERENCE_VERSION,
                "store": _OFFICE_WEB_ADDIN_STORE,
                "storeType": _OFFICE_WEB_ADDIN_STORE_TYPE,
            },
        )
        properties = ElementTree.SubElement(
            extension,
            f"{{{_WEB_EXTENSION_NS}}}properties",
        )
        ElementTree.SubElement(
            properties,
            f"{{{_WEB_EXTENSION_NS}}}property",
            {
                "name": _OFFICE_WEB_ADDIN_AUTO_SHOW_PROPERTY,
                "value": str(auto_show).lower(),
            },
        )
        members[_OFFICE_WEB_ADDIN_EXTENSION_MEMBER] = serialize(extension)

    _rewrite_xlsx_parts(path, mutate)


def _ole_object_auto_load_workbook() -> Workbook:
    """Build a small model with a later-added embedded OLE declaration.

    The ordinary input/model/dashboard values establish stable workbook context
    only. A raw OOXML step adds a relationship-backed opaque embedded-object
    declaration; it does not deserialize the bytes or start an object server.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB embedded OLE auto-load fixture")
    inputs = workbook.active
    inputs.title = _OLE_OBJECT_INPUT_SHEET
    inputs["A1"] = "Approved units"
    inputs[_OLE_OBJECT_INPUT_CELL] = _OLE_OBJECT_INPUT_VALUE

    model = workbook.create_sheet(_OLE_OBJECT_MODEL_SHEET)
    model["A1"] = "Projected units"
    model[_OLE_OBJECT_MODEL_CELL] = _OLE_OBJECT_MODEL_FORMULA

    dashboard = workbook.create_sheet(_OLE_OBJECT_DASHBOARD_SHEET)
    dashboard["A4"] = "Projected units"
    dashboard[_OLE_OBJECT_DASHBOARD_CELL] = _OLE_OBJECT_DASHBOARD_FORMULA
    return workbook


def _inject_ole_object_auto_load(path: Path, *, auto_load: bool) -> None:
    """Attach one local embedded OLE object with an explicit load-on-open flag.

    The fixture uses the standard worksheet-to-embedded-object relationship and
    content type, but stores only fixed opaque ASCII bytes under a synthetic,
    unregistered ProgID. It never supplies an ActiveX control, a linked target,
    an object presentation, or executable content. The caller changes only the
    persisted ``oleObject/@autoLoad`` boolean.
    """

    if type(auto_load) is not bool:
        raise ValueError("OLE object auto_load must be boolean")

    def serialize(element: ElementTree.Element) -> bytes:
        return ElementTree.tostring(element, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        ole_objects_tag = f"{{{_SPREADSHEETML_NS}}}oleObjects"
        ole_object_tag = f"{{{_SPREADSHEETML_NS}}}oleObject"

        if _OLE_OBJECT_MEMBER in members:
            raise ValueError("OLE auto-load fixture already has its embedded-object part")
        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        binary_defaults = [
            default
            for default in content_types.findall(default_tag)
            if default.get("Extension") == "bin"
        ]
        if not binary_defaults:
            ElementTree.SubElement(
                content_types,
                default_tag,
                {
                    "Extension": "bin",
                    "ContentType": _OLE_OBJECT_CONTENT_TYPE,
                },
            )
        elif (
            len(binary_defaults) != 1
            or binary_defaults[0].get("ContentType") != _OLE_OBJECT_CONTENT_TYPE
        ):
            raise ValueError("OLE auto-load fixture has an incompatible binary content type")
        members["[Content_Types].xml"] = serialize(content_types)

        worksheet = ElementTree.fromstring(members[_OLE_OBJECT_WORKSHEET_MEMBER])
        if worksheet.findall(ole_objects_tag):
            raise ValueError("OLE auto-load fixture already has worksheet OLE objects")
        ole_objects = ElementTree.Element(ole_objects_tag)
        ElementTree.SubElement(
            ole_objects,
            ole_object_tag,
            {
                "progId": _OLE_OBJECT_PROG_ID,
                "dvAspect": _OLE_OBJECT_DV_ASPECT,
                "autoLoad": str(auto_load).lower(),
                "shapeId": str(_OLE_OBJECT_SHAPE_ID),
                relationship_id_attribute: _OLE_OBJECT_RELATIONSHIP_ID,
            },
        )
        trailing_tags = {
            f"{{{_SPREADSHEETML_NS}}}controls",
            f"{{{_SPREADSHEETML_NS}}}webPublishItems",
            f"{{{_SPREADSHEETML_NS}}}tableParts",
            f"{{{_SPREADSHEETML_NS}}}extLst",
        }
        insert_index = next(
            (index for index, child in enumerate(worksheet) if child.tag in trailing_tags),
            len(worksheet),
        )
        worksheet.insert(insert_index, ole_objects)
        members[_OLE_OBJECT_WORKSHEET_MEMBER] = serialize(worksheet)

        relationships = ElementTree.fromstring(
            members.get(
                _OLE_OBJECT_RELATIONSHIPS_MEMBER,
                b'<?xml version="1.0" encoding="UTF-8"?>'
                b'<Relationships xmlns="http://schemas.openxmlformats.org/'
                b'package/2006/relationships"/>',
            )
        )
        if any(
            relationship.get("Id") == _OLE_OBJECT_RELATIONSHIP_ID
            or relationship.get("Type") == _OLE_OBJECT_RELATIONSHIP
            for relationship in relationships.findall(relationship_tag)
        ):
            raise ValueError("OLE auto-load fixture already has its worksheet relationship")
        ElementTree.SubElement(
            relationships,
            relationship_tag,
            {
                "Id": _OLE_OBJECT_RELATIONSHIP_ID,
                "Type": _OLE_OBJECT_RELATIONSHIP,
                "Target": "../embeddings/wcab-review-embedded-object.bin",
            },
        )
        members[_OLE_OBJECT_RELATIONSHIPS_MEMBER] = serialize(relationships)
        members[_OLE_OBJECT_MEMBER] = _OLE_OBJECT_PAYLOAD

    _rewrite_xlsx_parts(path, mutate)


def _governance_workbook() -> Workbook:
    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB governed calculation")
    sheet = workbook.active
    sheet.title = "Controls"
    sheet.append(["Control", "Units", "Rate", "Calculated amount"])
    sheet["A10"] = "Approved estimate"
    sheet["B10"] = 12
    sheet["C10"] = 5
    sheet["D10"] = "=B10*C10"
    sheet["D11"] = "=D10+1"
    sheet.protection.sheet = True
    sheet.protection.enable()
    hidden = workbook.create_sheet("ReviewControls")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "Internal review notes"
    return workbook


def _sheet_protection_sort_permission_workbook() -> Workbook:
    """Build a protected calculation sheet with one stored sort control.

    The candidate keeps sheet protection enabled and changes only the explicit
    ``sheetProtection/@sort`` lock. The fixture records that stored permission
    boundary; it does not ask Excel to sort anything or assert a client action.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB sheet-protection sort fixture")
    controls = workbook.active
    controls.title = _SHEET_PROTECTION_SORT_SHEET
    controls.append(["Control", "Units", "Rate", "Calculated amount"])
    controls["A2"] = "Approved estimate"
    controls["B2"] = 12
    controls["C2"] = 5
    controls[_SHEET_PROTECTION_SORT_FORMULA_CELL] = _SHEET_PROTECTION_SORT_FORMULA
    controls.protection.sheet = True
    controls.protection.enable()

    dashboard = workbook.create_sheet(_SHEET_PROTECTION_SORT_DASHBOARD_SHEET)
    dashboard["A1"] = "Board output"
    dashboard[_SHEET_PROTECTION_SORT_DASHBOARD_CELL] = _SHEET_PROTECTION_SORT_DASHBOARD_FORMULA
    return workbook


def _protected_range_security_descriptor_workbook() -> Workbook:
    """Build a locked protected-range target with a small formula context.

    The raw OOXML helper writes the standard nested descriptor after openpyxl
    has created the sheet. The case records a stored declaration only: it does
    not prove an identity, verifier, authorization decision, or client action.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB protected-range descriptor fixture")
    controls = workbook.active
    controls.title = _PROTECTED_RANGE_SECURITY_DESCRIPTOR_SHEET
    controls.append(["Control", "Units", "Rate", "Calculated amount"])
    controls["A2"] = "Controlled input"
    controls[_PROTECTED_RANGE_SECURITY_DESCRIPTOR_INPUT_CELL] = (
        _PROTECTED_RANGE_SECURITY_DESCRIPTOR_INPUT_VALUE
    )
    controls["C2"] = 5
    controls[_PROTECTED_RANGE_SECURITY_DESCRIPTOR_FORMULA_CELL] = (
        _PROTECTED_RANGE_SECURITY_DESCRIPTOR_FORMULA
    )
    controls.protection.sheet = True
    controls.protection.enable()

    dashboard = workbook.create_sheet(_PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_SHEET)
    dashboard["A1"] = "Board output"
    dashboard[_PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_CELL] = (
        _PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_FORMULA
    )
    return workbook


def _add_protected_range_security_descriptor(path: Path, *, descriptor: str) -> None:
    """Attach one standard nested descriptor without invoking a workbook client.

    ``protectedRange/securityDescriptor`` is a stored ISO/IEC SpreadsheetML
    declaration. The fixture keeps its sheet protection, target cell, legacy
    verifier, range name, formulas, and every other package member fixed while
    the child text changes. It never turns that declaration into an access
    decision.
    """

    worksheet_member = _PROTECTED_RANGE_SECURITY_DESCRIPTOR_WORKSHEET_MEMBER
    protected_ranges_tag = f"{{{_SPREADSHEETML_NS}}}protectedRanges"
    protected_range_tag = f"{{{_SPREADSHEETML_NS}}}protectedRange"
    descriptor_tag = f"{{{_SPREADSHEETML_NS}}}securityDescriptor"
    sheet_protection_tag = f"{{{_SPREADSHEETML_NS}}}sheetProtection"

    def mutate(members: dict[str, bytes]) -> None:
        worksheet = ElementTree.fromstring(members[worksheet_member])
        if worksheet.find(protected_ranges_tag) is not None:
            raise ValueError("protected-range descriptor fixture already has protected ranges")
        sheet_protection = worksheet.find(sheet_protection_tag)
        if sheet_protection is None:
            raise ValueError("protected-range descriptor fixture lacks sheet protection")
        protected_ranges = ElementTree.Element(protected_ranges_tag)
        protected_range = ElementTree.SubElement(
            protected_ranges,
            protected_range_tag,
            {
                "name": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_RANGE_NAME,
                "sqref": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_RANGE_REF,
                "password": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_LEGACY_VERIFIER,
            },
        )
        ElementTree.SubElement(protected_range, descriptor_tag).text = descriptor
        worksheet.insert(list(worksheet).index(sheet_protection) + 1, protected_ranges)
        members[worksheet_member] = ElementTree.tostring(
            worksheet,
            encoding="utf-8",
            xml_declaration=True,
        )

    _rewrite_xlsx_parts(path, mutate)


def _workbook_structure_protection_workbook() -> Workbook:
    """Build a structural-lock fixture without workbook encryption.

    The pair intentionally retains a hidden review-control sheet and one small
    formula context.  Only the raw workbook-level ``lockStructure`` control
    will change, so the case records a structural-governance change without
    asserting a password, authorization, or client-side action outcome.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB workbook-structure protection fixture")
    inputs = workbook.active
    inputs.title = _WORKBOOK_STRUCTURE_PROTECTION_INPUT_SHEET
    inputs["A1"] = "Approved estimate"
    inputs["B2"] = 12
    inputs["C2"] = 5
    inputs[_WORKBOOK_STRUCTURE_PROTECTION_FORMULA_CELL] = _WORKBOOK_STRUCTURE_PROTECTION_FORMULA
    review_controls = workbook.create_sheet(_WORKBOOK_STRUCTURE_PROTECTION_HIDDEN_SHEET)
    review_controls.sheet_state = "hidden"
    review_controls["A1"] = "Internal review notes"
    workbook.security.lockStructure = True
    return workbook


def _three_d_workbook(*, include_adjustment: bool = False) -> Workbook:
    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB 3-D scope fixture")
    for name, amount in (("Jan", 10), ("Feb", 20), ("Mar", 30)):
        sheet = workbook.create_sheet(name)
        sheet["A1"] = name
        sheet["B5"] = amount
    if include_adjustment:
        adjustment = workbook.create_sheet("FebAdjustment", 2)
        adjustment["A1"] = "Inserted adjustment period"
        adjustment["B5"] = 5
    summary = workbook.create_sheet("Summary")
    summary["A1"] = "3-D total"
    summary["B5"] = "=SUM(Jan:Mar!B5)"
    workbook.remove(workbook["Sheet"])
    return workbook


def _refactor_workbook(*, rewritten: bool = False) -> Workbook:
    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB structural formula rewrite")
    model = workbook.active
    model.title = "Model"
    model["A1"] = "Input model"
    model["B5"] = 2
    if rewritten:
        model["C5"] = "Inserted descriptive column"
        model["D5"] = 3
        model["E5"] = "=B5*D5"
        output_cell = "E5"
    else:
        model["C5"] = 3
        model["D5"] = "=B5*C5"
        output_cell = "D5"
    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Output"
    summary["B2"] = f"=Model!{output_cell}"
    return workbook


def _structured_table_workbook() -> Workbook:
    """Build a table-backed model with a stable structured-reference summary."""

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB structured table scope fixture")
    ledger = workbook.active
    ledger.title = "Ledger"
    ledger.append(["Region", "Units", "Price", "Amount"])
    for row_number, (region, units, price) in enumerate(
        (("North", 10, 8), ("South", 12, 9), ("West", 7, 11)), start=2
    ):
        ledger.append([region, units, price, f"=B{row_number}*C{row_number}"])
    table = Table(displayName="SalesLedger", ref="A1:D4")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ledger.add_table(table)

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Structured table total"
    summary["B2"] = "=SUM(SalesLedger[Amount])"
    return workbook


def _table_calculated_column_formula_workbook() -> Workbook:
    """Build a Table with stable row formulas and one raw formula master.

    The stored ``calculatedColumnFormula`` is attached after openpyxl writes
    the ordinary table.  Its formula is a Table-level declaration, distinct
    from the already saved row formulas and the downstream structured-reference
    formula.  The fixture deliberately records that raw definition only; it
    does not ask a spreadsheet client to fill or calculate a column.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB Table calculated-column formula fixture")
    ledger = workbook.active
    ledger.title = _TABLE_CALCULATED_COLUMN_SHEET
    ledger.append(["Units", "Rate", _TABLE_CALCULATED_COLUMN_HEADER])
    for row_number, (units, rate) in enumerate(((12, 5), (10, 7), (8, 9)), start=2):
        ledger.append([units, rate, f"=A{row_number}*B{row_number}"])
    ledger.add_table(
        Table(displayName=_TABLE_CALCULATED_COLUMN_NAME, ref=_TABLE_CALCULATED_COLUMN_REF)
    )

    dashboard = workbook.create_sheet(_TABLE_CALCULATED_COLUMN_DASHBOARD_SHEET)
    dashboard["A1"] = "Table calculated-column total"
    dashboard[_TABLE_CALCULATED_COLUMN_DASHBOARD_CELL] = _TABLE_CALCULATED_COLUMN_DASHBOARD_FORMULA
    return workbook


def _power_pivot_data_model_workbook() -> Workbook:
    """Build stable local source Tables for a raw Data Model declaration.

    The relationship declaration and its opaque model payload are attached
    after openpyxl writes this ordinary workbook.  The local Tables make the
    relationship names concrete review context, but WCAB never asks a client
    to load them into a Data Model, execute DAX, or materialize a Pivot result.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB Power Pivot Data Model relationship fixture")

    sales = workbook.active
    sales.title = "Sales"
    sales.append([_POWER_PIVOT_DATA_FROM_COLUMN, "Amount"])
    sales.append([20240101, 100])
    sales.append([20240102, 125])
    sales.add_table(Table(displayName=_POWER_PIVOT_DATA_FROM_TABLE, ref="A1:B3"))

    calendar = workbook.create_sheet("Calendar")
    calendar.append([_POWER_PIVOT_DATA_BASELINE_TO_COLUMN, _POWER_PIVOT_DATA_CANDIDATE_TO_COLUMN])
    calendar.append([20240101, "FY24-P01"])
    calendar.append([20240102, "FY24-P01"])
    calendar.add_table(Table(displayName=_POWER_PIVOT_DATA_TO_TABLE, ref="A1:B3"))

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Data Model report outputs are intentionally not materialized."
    return workbook


def _xlm_auto_open_binding_workbook() -> Workbook:
    """Build ordinary workbook context for a raw XLM automatic-macro binding.

    The XLM macro sheet is attached after openpyxl writes this normal workbook.
    Its two fixed ``HALT()`` cells are deliberately inert fixture material, and
    neither the generator nor the validator ever opens the pair in Excel.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB XLM Auto_Open binding fixture")

    inputs = workbook.active
    inputs.title = _XLM_AUTO_OPEN_INPUT_SHEET
    inputs["A1"] = "Stable input"
    inputs[_XLM_AUTO_OPEN_INPUT_CELL] = _XLM_AUTO_OPEN_INPUT_VALUE

    model = workbook.create_sheet(_XLM_AUTO_OPEN_MODEL_SHEET)
    model["A1"] = "Stable ordinary formula"
    model[_XLM_AUTO_OPEN_MODEL_CELL] = _XLM_AUTO_OPEN_MODEL_FORMULA

    dashboard = workbook.create_sheet(_XLM_AUTO_OPEN_DASHBOARD_SHEET)
    dashboard["A1"] = "XLM dispatch is intentionally not executed."
    dashboard[_XLM_AUTO_OPEN_DASHBOARD_CELL] = _XLM_AUTO_OPEN_DASHBOARD_FORMULA
    return workbook


def _power_query_local_table_workbook() -> Workbook:
    """Build a local Excel Table consumed by a connection-only M query.

    The Data Mashup payload is attached after openpyxl writes the ordinary
    worksheet/table package. Its formula is a stored query definition, not a
    request to execute Power Query or materialize a query result.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB Power Query local-table filter fixture")
    source = workbook.active
    source.title = _POWER_QUERY_SOURCE_SHEET
    source.append([_POWER_QUERY_FILTER_COLUMN, "Amount"])
    for region, amount in (("North", 10), ("South", 20), ("North", 15), ("South", 25)):
        source.append([region, amount])
    table = Table(displayName=_POWER_QUERY_SOURCE_TABLE, ref=_POWER_QUERY_SOURCE_REF)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    source.add_table(table)

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Connection-only query output is not materialized"
    return workbook


def _scenario_manager_workbook() -> Workbook:
    """Build a small model whose alternate inputs live in Scenario Manager.

    The Scenario Manager declaration is attached after openpyxl saves the
    ordinary workbook. The visible input cells and formula path stay fixed;
    the benchmark concerns a stored alternate input value, not an applied
    scenario or a calculated result.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB Scenario Manager stored-input fixture")
    inputs = workbook.active
    inputs.title = _SCENARIO_MANAGER_SHEET
    inputs["A1"] = "Scenario-controlled inputs"
    inputs["A2"] = "Margin"
    inputs[_SCENARIO_MANAGER_CHANGING_CELL] = _SCENARIO_MANAGER_WORKSHEET_INPUT_VALUE
    inputs["A3"] = "Volume"
    inputs[_SCENARIO_MANAGER_STABLE_INPUT_CELL] = _SCENARIO_MANAGER_WORKSHEET_STABLE_INPUT_VALUE
    inputs["C2"] = "Scenario summary result"
    inputs[_SCENARIO_MANAGER_RESULT_CELL] = _SCENARIO_MANAGER_FORMULA

    dashboard = workbook.create_sheet(_SCENARIO_MANAGER_DASHBOARD_SHEET)
    dashboard["A1"] = "Board scenario output"
    dashboard["A4"] = "Scenario result"
    dashboard[_SCENARIO_MANAGER_DASHBOARD_CELL] = _SCENARIO_MANAGER_DASHBOARD_FORMULA
    return workbook


def _what_if_data_table_workbook() -> Workbook:
    """Build one column-oriented What-If Data Table without calculating it.

    The data-table master is a raw SpreadsheetML formula control rather than
    an ordinary Excel Table or a calculated value. Both possible input cells
    remain part of a stable direct model path; the fixture changes only which
    local input reference the sensitivity declaration uses.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB What-If Data Table input fixture")
    sensitivity = workbook.active
    sensitivity.title = _WHAT_IF_DATA_TABLE_SHEET
    sensitivity["A1"] = "Sensitivity controls"
    sensitivity["A2"] = "Primary rate"
    sensitivity[_WHAT_IF_DATA_TABLE_PRIMARY_INPUT_CELL] = _WHAT_IF_DATA_TABLE_PRIMARY_INPUT_VALUE
    sensitivity["A3"] = "Alternative rate"
    sensitivity[_WHAT_IF_DATA_TABLE_ALTERNATE_INPUT_CELL] = (
        _WHAT_IF_DATA_TABLE_ALTERNATE_INPUT_VALUE
    )
    sensitivity["A4"] = "Scale"
    sensitivity[_WHAT_IF_DATA_TABLE_SCALE_CELL] = _WHAT_IF_DATA_TABLE_SCALE_VALUE
    sensitivity["C2"] = "Scenario rate"
    sensitivity[_WHAT_IF_DATA_TABLE_OUTPUT_FORMULA_CELL] = _WHAT_IF_DATA_TABLE_OUTPUT_FORMULA
    for coordinate, value in zip(("C3", "C4", "C5"), _WHAT_IF_DATA_TABLE_GRID_VALUES, strict=True):
        sensitivity[coordinate] = value
    sensitivity[_WHAT_IF_DATA_TABLE_MASTER_CELL] = DataTableFormula(
        ref=_WHAT_IF_DATA_TABLE_OUTPUT_RANGE,
        ca=True,
        dt2D=False,
        dtr=False,
        r1=_WHAT_IF_DATA_TABLE_PRIMARY_INPUT_CELL,
    )

    model = workbook.create_sheet(_WHAT_IF_DATA_TABLE_MODEL_SHEET)
    model["A1"] = "Sensitivity result"
    model[_WHAT_IF_DATA_TABLE_MODEL_CELL] = _WHAT_IF_DATA_TABLE_MODEL_FORMULA

    dashboard = workbook.create_sheet(_WHAT_IF_DATA_TABLE_DASHBOARD_SHEET)
    dashboard["A1"] = "Board output"
    dashboard[_WHAT_IF_DATA_TABLE_DASHBOARD_CELL] = _WHAT_IF_DATA_TABLE_DASHBOARD_FORMULA
    return workbook


def _dynamic_reference_workbook() -> Workbook:
    """Build a small model whose candidate can introduce ``INDIRECT``.

    The text driver is deliberately editable workbook data.  The fixture does
    not calculate it; it tests whether a tool makes the resulting static
    dependency-coverage boundary visible.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB dynamic reference fixture")
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Address driver"
    inputs["E12"] = "Revenue!C8"

    revenue = workbook.create_sheet("Revenue")
    revenue["A1"] = "Revenue sources"
    revenue["B8"] = 100
    revenue["C8"] = 120

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Selected revenue"
    summary["B2"] = "=Revenue!$C$8"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Board output"
    dashboard["B4"] = "=Summary!$B$2"
    return workbook


def _dynamic_reference_driver_workbook(*, function: str, driver_value: str | int) -> Workbook:
    """Build a model with a stable dynamic formula and an editable driver.

    ``INDIRECT`` changes a reference by changing text; ``OFFSET`` changes one
    by changing displacement. In both cases the formula is deliberately stable
    across the fixture pair, so reviewers must not equate unchanged formula
    text with unchanged effective dependency selection.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title=f"WCAB {function} driver fixture")
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Dynamic reference driver"
    inputs["E12"] = driver_value

    revenue = workbook.create_sheet("Revenue")
    revenue["A1"] = "Revenue sources"
    revenue["B8"] = 100
    revenue["C8"] = 120

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Selected revenue"
    if function == "INDIRECT":
        summary["B2"] = "=INDIRECT(Inputs!$E$12)"
    elif function == "OFFSET":
        summary["B2"] = "=OFFSET(Revenue!$B$8,0,Inputs!$E$12)"
    else:  # Internal callers use the two explicitly supported fixture forms.
        raise ValueError(f"unsupported dynamic-reference fixture function {function!r}")

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Board output"
    dashboard["B4"] = "=Summary!$B$2"
    return workbook


def _external_data_refresh_workbook() -> Workbook:
    """Build a workbook whose raw connection control is added after saving.

    The cells intentionally remain ordinary static data and formulas. The case
    isolates the stored external-data refresh setting rather than pretending to
    fetch a source or calculate an imported result.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB external data refresh fixture")
    data = workbook.active
    data.title = "ImportedData"
    data["A1"] = "Saved imported revenue"
    data["B2"] = 100

    summary = workbook.create_sheet("Summary")
    summary["A1"] = "Saved imported revenue"
    summary["B2"] = "=ImportedData!$B$2"

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Board output"
    dashboard["B4"] = "=Summary!$B$2"
    return workbook


def _query_table_refresh_workbook() -> Workbook:
    """Build stable saved cells around one raw QueryTable control.

    The query table and its synthetic web-query connection are attached only
    after openpyxl writes the ordinary workbook. The stored cells are context,
    not an imported result: generation and validation never open a connection,
    refresh a query, or calculate an outcome.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB QueryTable refresh fixture")
    imported_data = workbook.active
    imported_data.title = _QUERY_TABLE_SHEET
    imported_data["A1"] = "Saved query-table revenue"
    imported_data[_QUERY_TABLE_SAVED_VALUE_CELL] = _QUERY_TABLE_SAVED_VALUE

    summary = workbook.create_sheet(_QUERY_TABLE_SUMMARY_SHEET)
    summary["A1"] = "Saved query-table revenue"
    summary[_QUERY_TABLE_SUMMARY_CELL] = _QUERY_TABLE_SUMMARY_FORMULA

    dashboard = workbook.create_sheet(_QUERY_TABLE_DASHBOARD_SHEET)
    dashboard["A1"] = "Board output"
    dashboard[_QUERY_TABLE_DASHBOARD_CELL] = _QUERY_TABLE_DASHBOARD_FORMULA
    return workbook


def _cell_hyperlink_target_workbook() -> Workbook:
    """Build stable visible cells around one external worksheet hyperlink.

    The initial URL is normalized in raw OOXML after saving, so the generated
    pair can differ only in the worksheet relationship's ``Target``. The
    ordinary cell value and its formula consumers remain intentionally fixed.
    Generation never resolves, opens, or fetches the stored URL.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB cell hyperlink target fixture")
    inputs = workbook.active
    inputs.title = _CELL_HYPERLINK_SHEET
    inputs["A1"] = "Vendor portal"
    inputs[_CELL_HYPERLINK_CELL] = _CELL_HYPERLINK_VALUE
    inputs[_CELL_HYPERLINK_CELL].hyperlink = _CELL_HYPERLINK_BASELINE_TARGET

    summary = workbook.create_sheet(_CELL_HYPERLINK_SUMMARY_SHEET)
    summary["A1"] = "Vendor portal label"
    summary[_CELL_HYPERLINK_SUMMARY_CELL] = _CELL_HYPERLINK_SUMMARY_FORMULA

    dashboard = workbook.create_sheet(_CELL_HYPERLINK_DASHBOARD_SHEET)
    dashboard["A1"] = "Board output"
    dashboard[_CELL_HYPERLINK_DASHBOARD_CELL] = _CELL_HYPERLINK_DASHBOARD_FORMULA
    return workbook


def _pivot_cache_refresh_workbook() -> Workbook:
    """Build an ordinary workbook around one raw PivotCache control.

    The raw PivotTable package is added after saving because openpyxl preserves
    PivotTable parts but does not create them. The source data, stored report
    cells, and direct dashboard consumer stay fixed across the pair: WCAB
    records only the cache's stored refresh-on-open request, never a rendered
    PivotTable result.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB PivotCache refresh fixture")
    source = workbook.active
    source.title = _PIVOT_CACHE_SOURCE_SHEET
    source.append(["Region", "Amount"])
    for region, amount in (("North", 100), ("North", 200), ("South", 300), ("South", 400)):
        source.append([region, amount])

    report = workbook.create_sheet(_PIVOT_REPORT_SHEET)
    report["A1"] = "Region"
    report["B1"] = "Pivot amount"
    report["A2"] = "North"
    report["B2"] = 300

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A4"] = "Reported pivot amount"
    dashboard["B4"] = _PIVOT_CACHE_DASHBOARD_FORMULA
    return workbook


def _chart_series_reference_workbook() -> Workbook:
    """Build a dashboard chart with two stable local value columns.

    The candidate's raw chart-series value reference is switched after saving.
    The source cells, chart anchor, title reference, category reference, and
    all other package members remain stable. WCAB records the stored chart
    binding, not a rendered chart or a calculated visual result.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB chart series reference fixture")
    source = workbook.active
    source.title = _CHART_SERIES_SOURCE_SHEET
    source.append(["Quarter", "Quarterly revenue", "Quarterly revenue"])
    source.append(["Q1", 100, 140])
    source.append(["Q2", 120, 160])
    source.append(["Q3", 140, 180])

    dashboard = workbook.create_sheet(_CHART_SERIES_DASHBOARD_SHEET)
    dashboard["A1"] = "Quarterly revenue chart"
    chart = BarChart()
    chart.title = "Quarterly revenue"
    chart.add_data(Reference(source, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    chart.set_categories(Reference(source, min_col=1, min_row=2, max_row=4))
    dashboard.add_chart(chart, _CHART_SERIES_ANCHOR)
    return workbook


def _external_workbook_link_policy_workbook() -> Workbook:
    """Build a workbook with one unchanged, synthetic external-link formula.

    The source workbook intentionally does not exist. The paired case changes
    only the raw workbook-open policy, so neither generation nor validation
    can resolve the link, execute its formula, or claim a saved result.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB external workbook link policy fixture")
    model = workbook.active
    model.title = "LinkedModel"
    model["A1"] = "Synthetic external workbook driver"
    model["B2"] = _EXTERNAL_WORKBOOK_LINK_FORMULA

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Board output"
    dashboard["B4"] = "=LinkedModel!$B$2"
    return workbook


def _external_defined_name_source_workbook() -> Workbook:
    """Build a model whose local name is a stored external reference.

    The source workbooks are synthetic and absent. The workbook contains no
    relationship-backed ``externalLink`` package: its sole external reference
    is text stored in the local defined name. This keeps the case distinct from
    the relationship-target fixture while preserving an ordinary local formula
    path for review context.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB external defined-name source fixture")
    model = workbook.active
    model.title = _EXTERNAL_DEFINED_NAME_SOURCE_MODEL_SHEET
    model["A1"] = "Scenario-adjusted rate"
    model[_EXTERNAL_DEFINED_NAME_SOURCE_MODEL_CELL] = _EXTERNAL_DEFINED_NAME_SOURCE_MODEL_FORMULA

    dashboard = workbook.create_sheet(_EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_SHEET)
    dashboard["A1"] = "Board output"
    dashboard[_EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_CELL] = (
        _EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_FORMULA
    )
    workbook.defined_names.add(
        DefinedName(
            _EXTERNAL_DEFINED_NAME_SOURCE_NAME,
            attr_text=_EXTERNAL_DEFINED_NAME_SOURCE_BASELINE_REFERS_TO,
        )
    )
    return workbook


def _named_lambda_definition_workbook() -> Workbook:
    """Build a workbook whose reusable calculation lives in one named LAMBDA.

    Excel stores workbook-scoped named LAMBDAs as defined-name text.  The
    generated candidate will change that single definition after ordinary
    cells have been saved, so this fixture records a reusable formula-program
    boundary rather than a calculated result.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB named LAMBDA definition fixture")
    inputs = workbook.active
    inputs.title = _NAMED_LAMBDA_INPUT_SHEET
    inputs["A1"] = "Scenario inputs"
    inputs["A2"] = "Rate"
    inputs[_NAMED_LAMBDA_RATE_CELL] = _NAMED_LAMBDA_RATE_VALUE
    inputs["A3"] = "Amount"
    inputs[_NAMED_LAMBDA_AMOUNT_CELL] = _NAMED_LAMBDA_AMOUNT_VALUE

    model = workbook.create_sheet(_NAMED_LAMBDA_MODEL_SHEET)
    model["A1"] = "Reusable scenario value"
    model[_NAMED_LAMBDA_MODEL_CELL] = _NAMED_LAMBDA_MODEL_FORMULA

    dashboard = workbook.create_sheet(_NAMED_LAMBDA_DASHBOARD_SHEET)
    dashboard["A1"] = "Board output"
    dashboard[_NAMED_LAMBDA_DASHBOARD_CELL] = _NAMED_LAMBDA_DASHBOARD_FORMULA
    workbook.defined_names.add(
        DefinedName(_NAMED_LAMBDA_NAME, attr_text=_NAMED_LAMBDA_BASELINE_REFERS_TO)
    )
    return workbook


def _iterative_calculation_workbook() -> Workbook:
    """Build an unchanged, intentionally circular model with explicit bounds.

    The baseline records that iteration is disabled; the candidate changes only
    that stored calculation control. The direct self-reference is deliberate
    evidence of why the control matters, not a request for WCAB to calculate
    the model or assert a converged value.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB iterative calculation fixture")
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Convergence target"
    inputs["B2"] = 10

    model = workbook.create_sheet("Model")
    model["A1"] = "Iterative circular model"
    model["B2"] = _ITERATIVE_CALCULATION_FORMULA

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Board output"
    dashboard["B4"] = "=Model!$B$2"

    workbook.calculation.iterate = False
    workbook.calculation.iterateCount = _ITERATION_COUNT
    workbook.calculation.iterateDelta = _ITERATION_DELTA
    return workbook


def _precision_as_displayed_workbook() -> Workbook:
    """Build a model whose stored precision control can change in isolation.

    The generated pair deliberately retains the input's stored 10.005 value,
    its two-decimal display format, and every formula. A desktop Excel client
    can apply its own precision-as-displayed behavior later, but fixture
    generation never opens Excel, calculates a formula, or rewrites that value.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB precision-as-displayed fixture")
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Stored input with a two-decimal display"
    inputs["B2"] = _PRECISION_AS_DISPLAYED_INPUT
    inputs["B2"].number_format = _PRECISION_AS_DISPLAYED_NUMBER_FORMAT

    model = workbook.create_sheet("Model")
    model["A1"] = "Doubled input"
    model["B2"] = _PRECISION_AS_DISPLAYED_FORMULA
    model["B2"].number_format = _PRECISION_AS_DISPLAYED_NUMBER_FORMAT

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Published output"
    dashboard["B4"] = "=Model!$B$2"
    dashboard["B4"].number_format = _PRECISION_AS_DISPLAYED_NUMBER_FORMAT

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullPrecision = True
    workbook.calculation.calcCompleted = True
    workbook.calculation.calcOnSave = True
    return workbook


def _workbook_date_system_workbook() -> Workbook:
    """Build a model whose raw date base can change without a cell edit.

    The input deliberately remains a numeric serial with a date display format.
    The pair is later mutated only in ``workbookPr``: WCAB does not calculate a
    formula, convert the serial to a date, or assert what an Excel client will
    display after it opens the package.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB workbook date-system fixture")
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Stored date serial"
    inputs["B2"] = _WORKBOOK_DATE_SYSTEM_SERIAL
    inputs["B2"].number_format = _WORKBOOK_DATE_SYSTEM_NUMBER_FORMAT

    model = workbook.create_sheet("Model")
    model["A1"] = "Date serial plus 30"
    model["B2"] = _WORKBOOK_DATE_SYSTEM_FORMULA
    model["B2"].number_format = _WORKBOOK_DATE_SYSTEM_NUMBER_FORMAT

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Published date"
    dashboard["B4"] = _WORKBOOK_DATE_SYSTEM_DASHBOARD_FORMULA
    dashboard["B4"].number_format = _WORKBOOK_DATE_SYSTEM_NUMBER_FORMAT
    return workbook


def _set_workbook_date_system(path: Path, *, date_1904: bool, date_compatibility: bool) -> None:
    """Set only the two raw ``workbookPr`` date-system controls."""

    def mutate(members: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(members["xl/workbook.xml"])
        properties = workbook.find(f"{{{_SPREADSHEETML_NS}}}workbookPr")
        if properties is None:
            raise ValueError("date-system fixture has no workbookPr element")
        properties.set("date1904", "1" if date_1904 else "0")
        properties.set("dateCompatibility", "1" if date_compatibility else "0")
        members["xl/workbook.xml"] = ElementTree.tostring(
            workbook, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _formula_cached_result_workbook() -> Workbook:
    """Build a model whose saved formula result can change in isolation.

    Formula text and the direct input remain ordinary workbook content. The
    generated pair receives a raw ``<v>`` value for ``Model!B2`` only after
    saving, since openpyxl intentionally writes formula text rather than a
    cached result. The stable manual calculation metadata is descriptive save
    state, not a request for WCAB to evaluate either formula.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB stored formula result fixture")
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "Unchanged direct input"
    inputs["B2"] = _FORMULA_CACHED_RESULT_INPUT

    model = workbook.create_sheet("Model")
    model["A1"] = "Saved formula result"
    model["B2"] = _FORMULA_CACHED_RESULT_FORMULA

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Local formula consumer"
    dashboard["B4"] = "=Model!$B$2"

    workbook.calculation.calcMode = "manual"
    workbook.calculation.calcCompleted = True
    workbook.calculation.calcOnSave = False
    return workbook


def _set_formula_cached_result(path: Path, *, result: int) -> None:
    """Set the raw numeric ``<v>`` cache for the generated ``Model!B2``.

    This narrowly mutates the serialized worksheet rather than calculating or
    opening the workbook in an Excel client. The structural guards make a
    changed generator dependency fail loudly instead of silently broadening
    the fixture's claim.
    """

    def mutate(members: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet2.xml"
        worksheet = ElementTree.fromstring(members[worksheet_member])
        cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
        formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
        value_tag = f"{{{_SPREADSHEETML_NS}}}v"
        cells = [cell for cell in worksheet.iter(cell_tag) if cell.get("r") == "B2"]
        if len(cells) != 1:
            raise ValueError("formula-cache fixture has an unexpected Model!B2 cell")
        cell = cells[0]
        formula = cell.find(formula_tag)
        if formula is None or f"={formula.text or ''}" != _FORMULA_CACHED_RESULT_FORMULA:
            raise ValueError("formula-cache fixture has an unexpected Model!B2 formula")
        if cell.get("t") not in {None, "n"}:
            raise ValueError("formula-cache fixture has a nonnumeric Model!B2 cell")
        cached_value = cell.find(value_tag)
        if cached_value is None:
            cached_value = ElementTree.SubElement(cell, value_tag)
        cached_value.text = str(result)
        members[worksheet_member] = ElementTree.tostring(
            worksheet, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _array_formula_semantics_workbook() -> Workbook:
    """Build a fixed legacy CSE array with a direct downstream consumer.

    The candidate receives raw dynamic-array metadata after saving. Its anchor
    formula and currently stored output range intentionally remain unchanged:
    the review-relevant change is the formula mode, not a claimed calculation.
    """

    workbook = Workbook()
    _configure_workbook(workbook, title="WCAB array formula semantics fixture")
    inputs = workbook.active
    inputs.title = "Inputs"
    inputs["A1"] = "a"
    inputs["A2"] = "bb"
    inputs["A3"] = "ccc"

    model = workbook.create_sheet("Model")
    model["A1"] = "Array formula anchor"
    model["B1"].value = ArrayFormula(ref="B1:B3", text="=LEN(Inputs!A1:A3)")

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Anchor consumer"
    dashboard["B2"] = "=Model!B1"
    return workbook


def _add_dynamic_array_metadata(path: Path) -> None:
    """Mark the generated ``Model!B1`` array formula as a dynamic array.

    Openpyxl serializes the legacy CSE anchor but does not write the Office
    dynamic-array cell-metadata binding. This compact, public OOXML shape
    follows the serialization documented by XlsxWriter: the formula cell gets
    ``cm=1`` and that cell-metadata record resolves to an ``XLDAPR`` record
    whose ``fDynamic`` property is true.
    """

    def mutate(members: dict[str, bytes]) -> None:
        def insert_before_closing(member: str, closing: bytes, addition: bytes) -> None:
            contents = members[member]
            if contents.count(closing) != 1:
                raise ValueError(f"array formula fixture has unexpected {member} markup")
            members[member] = contents.replace(closing, addition + closing, 1)

        anchor = b'<c r="B1"><f t="array" ref="B1:B3">'
        marked_anchor = b'<c r="B1" cm="1"><f t="array" ref="B1:B3">'
        worksheet_member = "xl/worksheets/sheet2.xml"
        if members[worksheet_member].count(anchor) != 1:
            raise ValueError("array formula fixture has an unexpected Model!B1 contract")
        members[worksheet_member] = members[worksheet_member].replace(anchor, marked_anchor, 1)

        insert_before_closing(
            "[Content_Types].xml",
            b"</Types>",
            (
                b'<Override PartName="/xl/metadata.xml" '
                b'ContentType="application/vnd.openxmlformats-officedocument.'
                b'spreadsheetml.sheetMetadata+xml"/>'
            ),
        )
        insert_before_closing(
            "xl/_rels/workbook.xml.rels",
            b"</Relationships>",
            (
                b'<Relationship Id="rIdWCABDynamicArrayMetadata" '
                b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
                b'relationships/sheetMetadata" Target="metadata.xml"/>'
            ),
        )
        members["xl/metadata.xml"] = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<metadata xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            b'xmlns:xda="' + _DYNAMIC_ARRAY_NS.encode("ascii") + b'">'
            b'<metadataTypes count="1"><metadataType name="XLDAPR" '
            b'minSupportedVersion="120000" copy="1" pasteAll="1" pasteValues="1" '
            b'merge="1" splitFirst="1" rowColShift="1" clearFormats="1" '
            b'clearComments="1" assign="1" coerce="1" cellMeta="1"/>'
            b'</metadataTypes><futureMetadata name="XLDAPR" count="1"><bk><extLst>'
            b'<ext uri="{bdbb8cdc-fa1e-496e-a857-3c3f30c029c3}">'
            b'<xda:dynamicArrayProperties fDynamic="1" fCollapsed="0"/>'
            b'</ext></extLst></bk></futureMetadata><cellMetadata count="1"><bk>'
            b'<rc t="1" v="0"/></bk></cellMetadata></metadata>'
        )

    _rewrite_xlsx_parts(path, mutate)


def _add_external_data_connection(path: Path, *, refresh_on_load: bool, url: str) -> None:
    """Add one relationship-backed, non-routable external-data connection.

    The caller supplies a reserved ``.invalid`` URL, keeping generated source
    declarations inspectable without naming a real provider. Individual cases
    choose whether their isolated difference is this source text or the stored
    connection ``refreshOnLoad`` flag.
    """

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        if not any(
            item.get("PartName") == f"/{_EXTERNAL_DATA_CONNECTION_MEMBER}" for item in content_types
        ):
            ElementTree.SubElement(
                content_types,
                override_tag,
                {
                    "PartName": f"/{_EXTERNAL_DATA_CONNECTION_MEMBER}",
                    "ContentType": _EXTERNAL_DATA_CONNECTIONS_CONTENT_TYPE,
                },
            )
        members["[Content_Types].xml"] = serialize(content_types)

        relationships = ElementTree.fromstring(members["xl/_rels/workbook.xml.rels"])
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        ElementTree.SubElement(
            relationships,
            relationship_tag,
            {
                "Id": _EXTERNAL_DATA_CONNECTION_WORKBOOK_RELATIONSHIP_ID,
                "Type": _EXTERNAL_DATA_CONNECTIONS_RELATIONSHIP,
                "Target": "connections.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = serialize(relationships)

        connections = ElementTree.Element(f"{{{_SPREADSHEETML_NS}}}connections")
        connection = ElementTree.SubElement(
            connections,
            f"{{{_SPREADSHEETML_NS}}}connection",
            {
                "id": str(_EXTERNAL_DATA_CONNECTION_ID),
                "name": _EXTERNAL_DATA_CONNECTION_NAME,
                "type": "4",
                "refreshedVersion": "1",
                "refreshOnLoad": "1" if refresh_on_load else "0",
            },
        )
        ElementTree.SubElement(
            connection,
            f"{{{_SPREADSHEETML_NS}}}webPr",
            {"url": url},
        )
        members[_EXTERNAL_DATA_CONNECTION_MEMBER] = serialize(connections)

    _rewrite_xlsx_parts(path, mutate)


def _add_package_signature_manifest(
    path: Path,
    *,
    manifest_uri: str,
    relationship_selector_source_id: str | None = None,
) -> None:
    """Add a structurally shaped OPC signature whose Manifest names one part.

    The fixture deliberately uses non-cryptographic sentinel digest/signature
    values. It exists to isolate declaration scope, not to represent a valid
    signer, digest, transform result, certificate, or trust decision.
    """

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(
        content_types: ElementTree.Element,
        member: str,
        content_type: str,
    ) -> None:
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        existing = [
            item
            for item in content_types.findall(override_tag)
            if item.get("PartName") == f"/{member}"
        ]
        if not existing:
            ElementTree.SubElement(
                content_types,
                override_tag,
                {"PartName": f"/{member}", "ContentType": content_type},
            )
            return
        if len(existing) != 1 or existing[0].get("ContentType") != content_type:
            raise ValueError(f"package-signature fixture has unexpected content type for {member}")

    if relationship_selector_source_id is not None and manifest_uri != (
        _PACKAGE_SIGNATURE_ROOT_RELATIONSHIPS_MANIFEST_URI
    ):
        raise ValueError(
            "relationship-selector signature fixture must reference root relationships"
        )

    def mutate(members: dict[str, bytes]) -> None:
        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"
        sigs_defaults = [
            item for item in content_types.findall(default_tag) if item.get("Extension") == "sigs"
        ]
        if not sigs_defaults:
            ElementTree.SubElement(
                content_types,
                default_tag,
                {
                    "Extension": "sigs",
                    "ContentType": _PACKAGE_SIGNATURE_ORIGIN_CONTENT_TYPE,
                },
            )
        elif (
            len(sigs_defaults) != 1
            or sigs_defaults[0].get("ContentType") != _PACKAGE_SIGNATURE_ORIGIN_CONTENT_TYPE
        ):
            raise ValueError("package-signature fixture has unexpected .sigs content type")
        add_override(
            content_types,
            _PACKAGE_SIGNATURE_MEMBER,
            _PACKAGE_SIGNATURE_XML_CONTENT_TYPE,
        )
        members["[Content_Types].xml"] = serialize(content_types)

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        root_relationships = ElementTree.fromstring(members["_rels/.rels"])
        if relationship_selector_source_id is not None:
            office_document_relationships = [
                relationship
                for relationship in root_relationships.findall(relationship_tag)
                if relationship.get("Type") == _PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP
            ]
            existing_ids = {
                relationship.get("Id")
                for relationship in root_relationships.findall(relationship_tag)
            }
            if len(office_document_relationships) != 1 or (
                _PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP_ID in existing_ids
                and office_document_relationships[0].get("Id")
                != _PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP_ID
            ):
                raise ValueError(
                    "package-signature selector fixture has an unexpected office document relationship"
                )
            office_document_relationships[0].set(
                "Id", _PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP_ID
            )
        if any(
            relationship.get("Id") == _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP_ID
            for relationship in root_relationships.findall(relationship_tag)
        ):
            raise ValueError("package-signature fixture already has its origin relationship")
        ElementTree.SubElement(
            root_relationships,
            relationship_tag,
            {
                "Id": _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP_ID,
                "Type": _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP,
                "Target": _PACKAGE_SIGNATURE_ORIGIN_MEMBER,
            },
        )
        members["_rels/.rels"] = serialize(root_relationships)

        origin_relationships = ElementTree.Element(f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships")
        ElementTree.SubElement(
            origin_relationships,
            relationship_tag,
            {
                "Id": _PACKAGE_SIGNATURE_XML_RELATIONSHIP_ID,
                "Type": _PACKAGE_SIGNATURE_SIGNATURE_RELATIONSHIP,
                "Target": "sig1.xml",
            },
        )
        members[_PACKAGE_SIGNATURE_ORIGIN_MEMBER] = b""
        members[_PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIPS_MEMBER] = serialize(origin_relationships)

        signature = _XML_DIGITAL_SIGNATURE_NS
        envelope = ElementTree.Element(f"{{{signature}}}Signature")
        signed_info = ElementTree.SubElement(envelope, f"{{{signature}}}SignedInfo")
        ElementTree.SubElement(
            signed_info,
            f"{{{signature}}}CanonicalizationMethod",
            {"Algorithm": "http://www.w3.org/TR/2001/REC-xml-c14n-20010315"},
        )
        ElementTree.SubElement(
            signed_info,
            f"{{{signature}}}SignatureMethod",
            {"Algorithm": "http://www.w3.org/2001/04/xmldsig-more#rsa-sha256"},
        )
        signed_info_reference = ElementTree.SubElement(
            signed_info,
            f"{{{signature}}}Reference",
            {"URI": f"#{_PACKAGE_SIGNATURE_OBJECT_ID}"},
        )
        ElementTree.SubElement(
            signed_info_reference,
            f"{{{signature}}}DigestMethod",
            {"Algorithm": "http://www.w3.org/2001/04/xmlenc#sha256"},
        )
        ElementTree.SubElement(
            signed_info_reference,
            f"{{{signature}}}DigestValue",
        ).text = "WCAB-PRIVATE-SIGNED-INFO-DIGEST"
        ElementTree.SubElement(
            envelope, f"{{{signature}}}SignatureValue"
        ).text = "WCAB-PRIVATE-SIGNATURE-VALUE"
        package_object = ElementTree.SubElement(
            envelope,
            f"{{{signature}}}Object",
            {"Id": _PACKAGE_SIGNATURE_OBJECT_ID},
        )
        manifest = ElementTree.SubElement(
            package_object,
            f"{{{signature}}}Manifest",
        )
        manifest_reference = ElementTree.SubElement(
            manifest,
            f"{{{signature}}}Reference",
            {"URI": manifest_uri},
        )
        if relationship_selector_source_id is not None:
            if relationship_selector_source_id not in {
                _PACKAGE_SIGNATURE_BASELINE_SELECTOR_SOURCE_ID,
                _PACKAGE_SIGNATURE_CANDIDATE_SELECTOR_SOURCE_ID,
            }:
                raise ValueError("package-signature selector fixture has an unexpected source ID")
            transforms = ElementTree.SubElement(
                manifest_reference,
                f"{{{signature}}}Transforms",
            )
            relationship_transform = ElementTree.SubElement(
                transforms,
                f"{{{signature}}}Transform",
                {"Algorithm": _PACKAGE_SIGNATURE_RELATIONSHIP_TRANSFORM},
            )
            ElementTree.SubElement(
                relationship_transform,
                "{http://schemas.openxmlformats.org/package/2006/digital-signature}"
                "RelationshipReference",
                {"SourceId": relationship_selector_source_id},
            )
            ElementTree.SubElement(
                transforms,
                f"{{{signature}}}Transform",
                {"Algorithm": _PACKAGE_SIGNATURE_C14N_ALGORITHM},
            )
        ElementTree.SubElement(
            manifest_reference,
            f"{{{signature}}}DigestMethod",
            {"Algorithm": "http://www.w3.org/2001/04/xmlenc#sha256"},
        )
        ElementTree.SubElement(
            manifest_reference,
            f"{{{signature}}}DigestValue",
        ).text = "WCAB-PRIVATE-MANIFEST-DIGEST"
        members[_PACKAGE_SIGNATURE_MEMBER] = serialize(envelope)

    _rewrite_xlsx_parts(path, mutate)


def _add_threaded_comment_resolution_state(path: Path, *, resolved: bool) -> None:
    """Attach one bounded modern-comment thread with a chosen resolution state.

    The generated comment is deliberately synthetic. Both packages retain the
    same comment text, person record, content types, and relationship graph;
    only the top-level ``threadedComment/@done`` token differs. This records a
    stored discussion-state control, not a notification, approval, identity,
    authorization, client action, or completed human workflow.
    """

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(
        content_types: ElementTree.Element,
        member: str,
        content_type: str,
    ) -> None:
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        existing = [
            item
            for item in content_types.findall(override_tag)
            if item.get("PartName") == f"/{member}"
        ]
        if not existing:
            ElementTree.SubElement(
                content_types,
                override_tag,
                {"PartName": f"/{member}", "ContentType": content_type},
            )
            return
        if len(existing) != 1 or existing[0].get("ContentType") != content_type:
            raise ValueError(f"threaded-comment fixture has unexpected content type for {member}")

    def mutate(members: dict[str, bytes]) -> None:
        if _THREADED_COMMENT_MEMBER in members or _THREADED_COMMENT_PERSON_MEMBER in members:
            raise ValueError("threaded-comment fixture already has its comment package parts")
        if _THREADED_COMMENT_WORKSHEET_RELATIONSHIPS_MEMBER in members:
            raise ValueError("threaded-comment fixture expected no worksheet relationships")

        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        add_override(
            content_types,
            _THREADED_COMMENT_MEMBER,
            _THREADED_COMMENT_CONTENT_TYPE,
        )
        add_override(
            content_types,
            _THREADED_COMMENT_PERSON_MEMBER,
            _THREADED_COMMENT_PERSON_CONTENT_TYPE,
        )
        members["[Content_Types].xml"] = serialize(content_types)

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        workbook_relationships = ElementTree.fromstring(members["xl/_rels/workbook.xml.rels"])
        if any(
            relationship.get("Id") == _THREADED_COMMENT_PERSON_WORKBOOK_RELATIONSHIP_ID
            for relationship in workbook_relationships.findall(relationship_tag)
        ):
            raise ValueError("threaded-comment fixture already has its person relationship")
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": _THREADED_COMMENT_PERSON_WORKBOOK_RELATIONSHIP_ID,
                "Type": _THREADED_COMMENT_PERSON_RELATIONSHIP,
                "Target": "persons/person.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        worksheet_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            worksheet_relationships,
            relationship_tag,
            {
                "Id": _THREADED_COMMENT_WORKSHEET_RELATIONSHIP_ID,
                "Type": _THREADED_COMMENT_RELATIONSHIP,
                "Target": "../threadedComments/threadedComment1.xml",
            },
        )
        members[_THREADED_COMMENT_WORKSHEET_RELATIONSHIPS_MEMBER] = serialize(
            worksheet_relationships
        )

        comments = ElementTree.Element(f"{{{_THREADED_COMMENT_NS}}}ThreadedComments")
        comment = ElementTree.SubElement(
            comments,
            f"{{{_THREADED_COMMENT_NS}}}threadedComment",
            {
                "ref": _THREADED_COMMENT_CELL,
                "dT": _THREADED_COMMENT_TIMESTAMP,
                "personId": _THREADED_COMMENT_PERSON_ID,
                "id": _THREADED_COMMENT_ROOT_ID,
                "done": "1" if resolved else "0",
            },
        )
        ElementTree.SubElement(
            comment, f"{{{_THREADED_COMMENT_NS}}}text"
        ).text = _THREADED_COMMENT_TEXT
        members[_THREADED_COMMENT_MEMBER] = serialize(comments)

        people = ElementTree.Element(f"{{{_THREADED_COMMENT_NS}}}personList")
        ElementTree.SubElement(
            people,
            f"{{{_THREADED_COMMENT_NS}}}person",
            {
                "displayName": _THREADED_COMMENT_PERSON_NAME,
                "id": _THREADED_COMMENT_PERSON_ID,
            },
        )
        members[_THREADED_COMMENT_PERSON_MEMBER] = serialize(people)

    _rewrite_xlsx_parts(path, mutate)


def _add_shared_workbook_revision_log(path: Path, *, historic_value: str) -> None:
    """Attach one bounded legacy shared-workbook revision history.

    Both packages retain the same workbook/header/log package graph, tracking
    controls, author metadata, record shape, ordinary cells, and calculation
    properties. Only a historic value inside the relationship-backed revision
    log changes. This represents stored audit-trail material, not an identity,
    a verified provenance record, a conflict-resolution result, an approval,
    or a completed review workflow.
    """

    if historic_value not in {
        _SHARED_WORKBOOK_REVISION_BASELINE_HISTORIC_VALUE,
        _SHARED_WORKBOOK_REVISION_CANDIDATE_HISTORIC_VALUE,
    }:
        raise ValueError("shared-workbook revision fixture has an unexpected historic value")

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(
        content_types: ElementTree.Element,
        member: str,
        content_type: str,
    ) -> None:
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        existing = [
            item
            for item in content_types.findall(override_tag)
            if item.get("PartName") == f"/{member}"
        ]
        if not existing:
            ElementTree.SubElement(
                content_types,
                override_tag,
                {"PartName": f"/{member}", "ContentType": content_type},
            )
            return
        if len(existing) != 1 or existing[0].get("ContentType") != content_type:
            raise ValueError(
                f"shared-workbook revision fixture has unexpected content type for {member}"
            )

    def mutate(members: dict[str, bytes]) -> None:
        revision_members = {
            _SHARED_WORKBOOK_REVISION_HEADERS_MEMBER,
            _SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER,
            _SHARED_WORKBOOK_REVISION_LOG_MEMBER,
        }
        if revision_members & members.keys():
            raise ValueError("shared-workbook revision fixture already has revision parts")

        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        add_override(
            content_types,
            _SHARED_WORKBOOK_REVISION_HEADERS_MEMBER,
            _SHARED_WORKBOOK_REVISION_HEADERS_CONTENT_TYPE,
        )
        add_override(
            content_types,
            _SHARED_WORKBOOK_REVISION_LOG_MEMBER,
            _SHARED_WORKBOOK_REVISION_LOG_CONTENT_TYPE,
        )
        members["[Content_Types].xml"] = serialize(content_types)

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        workbook_relationships = ElementTree.fromstring(members["xl/_rels/workbook.xml.rels"])
        if any(
            relationship.get("Id") == _SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIP_ID
            for relationship in workbook_relationships.findall(relationship_tag)
        ):
            raise ValueError("shared-workbook revision fixture already has its header relationship")
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": _SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIP_ID,
                "Type": _SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIP,
                "Target": "revisions/revisionHeaders.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        headers = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}headers",
            {
                "diskRevisions": "1",
                "exclusive": "0",
                "history": "true",
                "keepChangeHistory": "true",
                "protected": "true",
                "shared": "1",
                "trackRevisions": "true",
                "preserveHistory": "30",
                "revisionId": "7",
                "version": "+2",
            },
        )
        ElementTree.SubElement(
            headers,
            f"{{{_SPREADSHEETML_NS}}}header",
            {
                "guid": _SHARED_WORKBOOK_REVISION_HEADER_GUID,
                "dateTime": _SHARED_WORKBOOK_REVISION_HEADER_TIMESTAMP,
                "maxSheetId": "4",
                "userName": _SHARED_WORKBOOK_REVISION_HEADER_AUTHOR,
                f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": (
                    _SHARED_WORKBOOK_REVISION_LOG_RELATIONSHIP_ID
                ),
            },
        )
        members[_SHARED_WORKBOOK_REVISION_HEADERS_MEMBER] = serialize(headers)

        header_relationships = ElementTree.Element(f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships")
        ElementTree.SubElement(
            header_relationships,
            relationship_tag,
            {
                "Id": _SHARED_WORKBOOK_REVISION_LOG_RELATIONSHIP_ID,
                "Type": _SHARED_WORKBOOK_REVISION_LOG_RELATIONSHIP,
                "Target": "revisionLog1.xml",
            },
        )
        members[_SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER] = serialize(
            header_relationships
        )

        revisions = ElementTree.Element(f"{{{_SPREADSHEETML_NS}}}revisions")
        changed_cells = ElementTree.SubElement(
            revisions,
            f"{{{_SPREADSHEETML_NS}}}rcc",
            {"rId": "1", "sId": "1"},
        )
        old_cell = ElementTree.SubElement(
            changed_cells,
            f"{{{_SPREADSHEETML_NS}}}oc",
            {"r": "C20"},
        )
        ElementTree.SubElement(old_cell, f"{{{_SPREADSHEETML_NS}}}v").text = historic_value
        new_cell = ElementTree.SubElement(
            changed_cells,
            f"{{{_SPREADSHEETML_NS}}}nc",
            {"r": "C20"},
        )
        ElementTree.SubElement(
            new_cell, f"{{{_SPREADSHEETML_NS}}}v"
        ).text = _SHARED_WORKBOOK_REVISION_RECORDED_VALUE
        ElementTree.SubElement(
            revisions,
            f"{{{_SPREADSHEETML_NS}}}rrc",
            {
                "rId": "2",
                "sId": "1",
                "ref": "A1:A1048576",
                "action": "insertCol",
            },
        )
        ElementTree.SubElement(
            revisions,
            f"{{{_SPREADSHEETML_NS}}}rfmt",
            {"sheetId": "1", "sqref": "C20"},
        )
        members[_SHARED_WORKBOOK_REVISION_LOG_MEMBER] = serialize(revisions)

    _rewrite_xlsx_parts(path, mutate)


def _add_query_table_refresh_control(path: Path, *, refresh_on_load: bool) -> None:
    """Attach one inert, relationship-backed QueryTable refresh declaration.

    Both packages contain the same internal worksheet-to-QueryTable and
    workbook-to-connections graph. The source URL is reserved and non-routable;
    the only baseline/candidate difference is raw queryTable/@refreshOnLoad.
    No query result, credentials, table range, or external relationship is
    created.
    """

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_override(content_types: ElementTree.Element, part_name: str, content_type: str) -> None:
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        existing = [
            item
            for item in content_types.findall(override_tag)
            if item.get("PartName") == part_name
        ]
        if not existing:
            ElementTree.SubElement(
                content_types,
                override_tag,
                {"PartName": part_name, "ContentType": content_type},
            )
            return
        if len(existing) != 1 or existing[0].get("ContentType") != content_type:
            raise ValueError(f"query-table fixture has unexpected content type for {part_name}")

    def mutate(members: dict[str, bytes]) -> None:
        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        add_override(
            content_types,
            f"/{_QUERY_TABLE_CONNECTION_MEMBER}",
            _QUERY_TABLE_CONNECTIONS_CONTENT_TYPE,
        )
        add_override(
            content_types,
            f"/{_QUERY_TABLE_MEMBER}",
            _QUERY_TABLE_CONTENT_TYPE,
        )
        members["[Content_Types].xml"] = serialize(content_types)

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        workbook_relationships = ElementTree.fromstring(members["xl/_rels/workbook.xml.rels"])
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": _QUERY_TABLE_WORKBOOK_RELATIONSHIP_ID,
                "Type": _QUERY_TABLE_CONNECTIONS_RELATIONSHIP,
                "Target": "connections.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        if _QUERY_TABLE_WORKSHEET_RELATIONSHIPS_MEMBER in members:
            raise ValueError("query-table fixture expected no pre-existing worksheet relationships")
        worksheet_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            worksheet_relationships,
            relationship_tag,
            {
                "Id": _QUERY_TABLE_WORKSHEET_RELATIONSHIP_ID,
                "Type": _QUERY_TABLE_RELATIONSHIP,
                "Target": "../queryTables/queryTable1.xml",
            },
        )
        members[_QUERY_TABLE_WORKSHEET_RELATIONSHIPS_MEMBER] = serialize(worksheet_relationships)

        connections = ElementTree.Element(f"{{{_SPREADSHEETML_NS}}}connections")
        connection = ElementTree.SubElement(
            connections,
            f"{{{_SPREADSHEETML_NS}}}connection",
            {
                "id": str(_QUERY_TABLE_CONNECTION_ID),
                "name": _QUERY_TABLE_CONNECTION_NAME,
                "type": "4",
                "refreshedVersion": "1",
                "background": "0",
                "refreshOnLoad": "0",
            },
        )
        ElementTree.SubElement(
            connection,
            f"{{{_SPREADSHEETML_NS}}}webPr",
            {"url": _QUERY_TABLE_SOURCE_URL},
        )
        members[_QUERY_TABLE_CONNECTION_MEMBER] = serialize(connections)

        query_table = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}queryTable",
            {
                "name": _QUERY_TABLE_NAME,
                "connectionId": str(_QUERY_TABLE_CONNECTION_ID),
                "refreshOnLoad": "1" if refresh_on_load else "0",
                "backgroundRefresh": "1" if _QUERY_TABLE_BACKGROUND_REFRESH else "0",
                "disableRefresh": "1" if _QUERY_TABLE_REFRESH_DISABLED else "0",
                "removeDataOnSave": "1" if _QUERY_TABLE_REMOVE_DATA_ON_SAVE else "0",
                "fillFormulas": "1" if _QUERY_TABLE_FILL_FORMULAS else "0",
                "disableEdit": "1" if _QUERY_TABLE_CONNECTION_EDIT_DISABLED else "0",
                "growShrinkType": _QUERY_TABLE_GROWTH_BEHAVIOR,
            },
        )
        members[_QUERY_TABLE_MEMBER] = serialize(query_table)

    _rewrite_xlsx_parts(path, mutate)


def _set_cell_hyperlink_target(path: Path, *, target: str) -> None:
    """Normalize one worksheet hyperlink binding and set its stored target.

    This deliberately changes only the local OOXML relationship. Both fixture
    packages use the same explicit relationship ID and worksheet hyperlink
    declaration; their sole intended distinction is ``Relationship/@Target``.
    """

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        try:
            worksheet = ElementTree.fromstring(members[_CELL_HYPERLINK_WORKSHEET_MEMBER])
            relationships = ElementTree.fromstring(
                members[_CELL_HYPERLINK_WORKSHEET_RELATIONSHIPS_MEMBER]
            )
        except (KeyError, ElementTree.ParseError) as error:
            raise ValueError(
                "cell-hyperlink fixture is missing its worksheet relationship graph"
            ) from error

        hyperlink_sets = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}hyperlinks")
        hyperlink_tag = f"{{{_SPREADSHEETML_NS}}}hyperlink"
        relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        if (
            len(hyperlink_sets) != 1
            or hyperlink_sets[0].attrib
            or len(hyperlink_sets[0]) != 1
            or hyperlink_sets[0][0].tag != hyperlink_tag
            or set(hyperlink_sets[0][0].attrib) != {"ref", relationship_id_attribute}
            or hyperlink_sets[0][0].get("ref") != _CELL_HYPERLINK_CELL
        ):
            raise ValueError(
                "cell-hyperlink fixture has an unexpected worksheet hyperlink declaration"
            )
        hyperlink = hyperlink_sets[0][0]
        original_relationship_id = hyperlink.get(relationship_id_attribute)
        if not isinstance(original_relationship_id, str):
            raise ValueError("cell-hyperlink fixture has no worksheet hyperlink relationship ID")

        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        hyperlink_relationships = [
            relationship
            for relationship in relationships.findall(relationship_tag)
            if relationship.get("Id") == original_relationship_id
            and relationship.get("Type") == _CELL_HYPERLINK_RELATIONSHIP
        ]
        if (
            len(relationships) != 1
            or len(hyperlink_relationships) != 1
            or set(hyperlink_relationships[0].attrib) != {"Id", "Type", "Target", "TargetMode"}
            or hyperlink_relationships[0].get("TargetMode") != "External"
        ):
            raise ValueError("cell-hyperlink fixture has an unexpected hyperlink relationship")

        hyperlink.set(relationship_id_attribute, _CELL_HYPERLINK_RELATIONSHIP_ID)
        hyperlink_relationships[0].set("Id", _CELL_HYPERLINK_RELATIONSHIP_ID)
        hyperlink_relationships[0].set("Target", target)
        members[_CELL_HYPERLINK_WORKSHEET_MEMBER] = serialize(worksheet)
        members[_CELL_HYPERLINK_WORKSHEET_RELATIONSHIPS_MEMBER] = serialize(relationships)

    _rewrite_xlsx_parts(path, mutate)


def _add_pivot_cache_refresh_control(path: Path, *, refresh_on_load: bool) -> None:
    """Attach one small PivotTable/PivotCache graph with a raw open-time flag.

    The fixture has a local worksheet source and no external provider. It is a
    compact, relationship-backed package shape that openpyxl can read, but this
    generator never asks Excel or openpyxl to refresh, render, or calculate the
    PivotTable. The sole baseline/candidate distinction is
    ``pivotCacheDefinition/@refreshOnLoad``.
    """

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(members["xl/workbook.xml"])
        sheets = workbook.find(f"{{{_SPREADSHEETML_NS}}}sheets")
        sheet_tag = f"{{{_SPREADSHEETML_NS}}}sheet"
        relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        report_sheets = (
            [
                sheet
                for sheet in sheets.findall(sheet_tag)
                if sheet.get("name") == _PIVOT_REPORT_SHEET
            ]
            if sheets is not None
            else []
        )
        if len(report_sheets) != 1:
            raise ValueError("PivotCache fixture has an unexpected Report worksheet")
        report_relationship_id = report_sheets[0].get(relationship_id_attribute)
        if not isinstance(report_relationship_id, str):
            raise ValueError("PivotCache fixture Report worksheet has no relationship")

        workbook_relationships = ElementTree.fromstring(members["xl/_rels/workbook.xml.rels"])
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        report_relationships = [
            relationship
            for relationship in workbook_relationships.findall(relationship_tag)
            if relationship.get("Id") == report_relationship_id
            and relationship.get("Type") == f"{_DOCUMENT_RELATIONSHIPS_NS}/worksheet"
        ]
        if len(report_relationships) != 1:
            raise ValueError("PivotCache fixture cannot resolve Report worksheet relationship")
        report_target = report_relationships[0].get("Target")
        if (
            not isinstance(report_target, str)
            or not report_target
            or report_target.startswith("../")
        ):
            raise ValueError("PivotCache fixture has an unsafe Report worksheet target")
        if report_target.startswith("/"):
            if not report_target.startswith("/xl/"):
                raise ValueError("PivotCache fixture has a non-workbook Report worksheet target")
            report_member = report_target.lstrip("/")
        else:
            report_member = f"xl/{report_target.lstrip('./')}"
        if not report_member.startswith("xl/worksheets/") or ".." in report_member.split("/"):
            raise ValueError("PivotCache fixture has an unsafe Report worksheet member")
        if report_member not in members:
            raise ValueError("PivotCache fixture Report worksheet member is absent")

        pivot_caches_tag = f"{{{_SPREADSHEETML_NS}}}pivotCaches"
        if workbook.find(pivot_caches_tag) is not None:
            raise ValueError("PivotCache fixture unexpectedly already has pivot caches")
        pivot_caches = ElementTree.SubElement(workbook, pivot_caches_tag)
        ElementTree.SubElement(
            pivot_caches,
            f"{{{_SPREADSHEETML_NS}}}pivotCache",
            {
                "cacheId": str(_PIVOT_CACHE_ID),
                relationship_id_attribute: "rIdWCABPivotCache",
            },
        )
        members["xl/workbook.xml"] = serialize(workbook)

        if any(
            relationship.get("Id") == "rIdWCABPivotCache"
            for relationship in workbook_relationships.findall(relationship_tag)
        ):
            raise ValueError("PivotCache fixture relationship ID is already in use")
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdWCABPivotCache",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/pivotCacheDefinition",
                "Target": "pivotCache/pivotCacheDefinition1.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        report = ElementTree.fromstring(members[report_member])
        pivot_table_parts_tag = f"{{{_SPREADSHEETML_NS}}}pivotTableParts"
        if report.find(pivot_table_parts_tag) is not None:
            raise ValueError("PivotCache fixture Report worksheet already has pivot tables")
        pivot_table_parts = ElementTree.SubElement(report, pivot_table_parts_tag, {"count": "1"})
        ElementTree.SubElement(
            pivot_table_parts,
            f"{{{_SPREADSHEETML_NS}}}pivotTablePart",
            {relationship_id_attribute: "rIdWCABPivotTable"},
        )
        members[report_member] = serialize(report)

        report_directory, report_filename = report_member.rsplit("/", maxsplit=1)
        report_relationship_member = f"{report_directory}/_rels/{report_filename}.rels"
        if report_relationship_member in members:
            report_relationships_root = ElementTree.fromstring(members[report_relationship_member])
        else:
            report_relationships_root = ElementTree.Element(
                f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
            )
        if any(
            relationship.get("Id") == "rIdWCABPivotTable"
            for relationship in report_relationships_root.findall(relationship_tag)
        ):
            raise ValueError("PivotCache fixture Report relationship ID is already in use")
        ElementTree.SubElement(
            report_relationships_root,
            relationship_tag,
            {
                "Id": "rIdWCABPivotTable",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/pivotTable",
                "Target": "../pivotTables/pivotTable1.xml",
            },
        )
        members[report_relationship_member] = serialize(report_relationships_root)

        pivot_table_member = "xl/pivotTables/pivotTable1.xml"
        pivot_table = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}pivotTableDefinition",
            {
                "name": "WCAB Pivot Report",
                "cacheId": str(_PIVOT_CACHE_ID),
                "dataCaption": "Amount",
            },
        )
        ElementTree.SubElement(
            pivot_table,
            f"{{{_SPREADSHEETML_NS}}}location",
            {
                "ref": _PIVOT_REPORT_REF,
                "firstHeaderRow": "1",
                "firstDataRow": "2",
                "firstDataCol": "1",
            },
        )
        pivot_fields = ElementTree.SubElement(
            pivot_table, f"{{{_SPREADSHEETML_NS}}}pivotFields", {"count": "2"}
        )
        ElementTree.SubElement(
            pivot_fields,
            f"{{{_SPREADSHEETML_NS}}}pivotField",
            {"axis": "axisRow", "showAll": "0"},
        )
        ElementTree.SubElement(
            pivot_fields,
            f"{{{_SPREADSHEETML_NS}}}pivotField",
            {"dataField": "1", "showAll": "0"},
        )
        row_fields = ElementTree.SubElement(
            pivot_table, f"{{{_SPREADSHEETML_NS}}}rowFields", {"count": "1"}
        )
        ElementTree.SubElement(row_fields, f"{{{_SPREADSHEETML_NS}}}field", {"x": "0"})
        data_fields = ElementTree.SubElement(
            pivot_table, f"{{{_SPREADSHEETML_NS}}}dataFields", {"count": "1"}
        )
        ElementTree.SubElement(
            data_fields,
            f"{{{_SPREADSHEETML_NS}}}dataField",
            {"fld": "1", "subtotal": "sum"},
        )
        members[pivot_table_member] = serialize(pivot_table)

        pivot_relationships = ElementTree.Element(f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships")
        ElementTree.SubElement(
            pivot_relationships,
            relationship_tag,
            {
                "Id": "rIdWCABPivotCache",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/pivotCacheDefinition",
                "Target": "../pivotCache/pivotCacheDefinition1.xml",
            },
        )
        members["xl/pivotTables/_rels/pivotTable1.xml.rels"] = serialize(pivot_relationships)

        cache_definition = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}pivotCacheDefinition",
            {
                "recordCount": "4",
                "refreshOnLoad": "1" if refresh_on_load else "0",
                "saveData": "1",
                relationship_id_attribute: "rIdWCABPivotRecords",
            },
        )
        cache_source = ElementTree.SubElement(
            cache_definition, f"{{{_SPREADSHEETML_NS}}}cacheSource", {"type": "worksheet"}
        )
        ElementTree.SubElement(
            cache_source,
            f"{{{_SPREADSHEETML_NS}}}worksheetSource",
            {"ref": _PIVOT_CACHE_SOURCE_REF, "sheet": _PIVOT_CACHE_SOURCE_SHEET},
        )
        cache_fields = ElementTree.SubElement(
            cache_definition, f"{{{_SPREADSHEETML_NS}}}cacheFields", {"count": "2"}
        )
        for name, values in (
            ("Region", (("s", "North"), ("s", "South"))),
            (
                "Amount",
                (("n", "100"), ("n", "200"), ("n", "300"), ("n", "400")),
            ),
        ):
            cache_field = ElementTree.SubElement(
                cache_fields,
                f"{{{_SPREADSHEETML_NS}}}cacheField",
                {"name": name, "numFmtId": "0"},
            )
            shared_items = ElementTree.SubElement(
                cache_field,
                f"{{{_SPREADSHEETML_NS}}}sharedItems",
                {"count": str(len(values))},
            )
            for element_name, value in values:
                ElementTree.SubElement(
                    shared_items, f"{{{_SPREADSHEETML_NS}}}{element_name}", {"v": value}
                )
        members["xl/pivotCache/pivotCacheDefinition1.xml"] = serialize(cache_definition)

        cache_definition_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            cache_definition_relationships,
            relationship_tag,
            {
                "Id": "rIdWCABPivotRecords",
                "Type": f"{_DOCUMENT_RELATIONSHIPS_NS}/pivotCacheRecords",
                "Target": "pivotCacheRecords1.xml",
            },
        )
        members["xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels"] = serialize(
            cache_definition_relationships
        )

        cache_records = ElementTree.Element(
            f"{{{_SPREADSHEETML_NS}}}pivotCacheRecords", {"count": "4"}
        )
        for region_index, amount_index in ((0, 0), (0, 1), (1, 2), (1, 3)):
            record = ElementTree.SubElement(cache_records, f"{{{_SPREADSHEETML_NS}}}r")
            ElementTree.SubElement(record, f"{{{_SPREADSHEETML_NS}}}x", {"v": str(region_index)})
            ElementTree.SubElement(record, f"{{{_SPREADSHEETML_NS}}}x", {"v": str(amount_index)})
        members["xl/pivotCache/pivotCacheRecords1.xml"] = serialize(cache_records)

        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        for part_name, content_type in (
            (
                "/xl/pivotTables/pivotTable1.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml",
            ),
            (
                "/xl/pivotCache/pivotCacheDefinition1.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml."
                "pivotCacheDefinition+xml",
            ),
            (
                "/xl/pivotCache/pivotCacheRecords1.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml",
            ),
        ):
            if any(
                item.get("PartName") == part_name for item in content_types.findall(override_tag)
            ):
                raise ValueError(f"PivotCache fixture already has content type {part_name!r}")
            ElementTree.SubElement(
                content_types,
                override_tag,
                {"PartName": part_name, "ContentType": content_type},
            )
        members["[Content_Types].xml"] = serialize(content_types)

    _rewrite_xlsx_parts(path, mutate)


def _power_query_nested_zip(parts: dict[str, bytes]) -> bytes:
    """Build a deterministic compact package embedded in a Data Mashup."""

    payload = io.BytesIO()
    with ZipFile(payload, "w", compression=ZIP_DEFLATED, compresslevel=9) as archive:
        for name in sorted(parts):
            info = ZipInfo(filename=name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o100644 << 16
            archive.writestr(info, parts[name])
    return payload.getvalue()


def _power_query_m_filter_formula(filter_value: str) -> str:
    """Return WCAB's compact connection-only local-table M program."""

    if filter_value not in {
        _POWER_QUERY_BASELINE_FILTER_VALUE,
        _POWER_QUERY_CANDIDATE_FILTER_VALUE,
    }:
        raise ValueError(f"unsupported Power Query filter value {filter_value!r}")
    return (
        f"section {_POWER_QUERY_SECTION};\n\n"
        f"shared {_POWER_QUERY_NAME} = let\n"
        f'    Source = Excel.CurrentWorkbook(){{[Name="{_POWER_QUERY_SOURCE_TABLE}"]}}[Content],\n'
        f'    FilteredRows = Table.SelectRows(Source, each [{_POWER_QUERY_FILTER_COLUMN}] = "{filter_value}")\n'
        "in\n"
        "    FilteredRows;\n"
    )


def _power_query_local_table_data_mashup(filter_value: str) -> bytes:
    """Return a deterministic compact Data Mashup with one stored M program.

    The package has no external source, load target, embedded content, or
    user-bound permission payload. It describes a connection-only query over
    the local generated Excel Table and is not evaluated while building.
    """

    package_parts = _power_query_nested_zip(
        {
            "[Content_Types].xml": b"<Types/>",
            "Config/Package.xml": b"<Package>WCAB connection-only local table query</Package>",
            f"Formulas/{_POWER_QUERY_SECTION}.m": _power_query_m_filter_formula(
                filter_value
            ).encode("utf-8"),
        }
    )
    metadata_xml = (
        '<?xml version="1.0" encoding="utf-8"?>\n'
        f'<LocalPackageMetadataFile xmlns="{_DATA_MASHUP_NS}">\n'
        "  <Items>\n"
        "    <Item>\n"
        "      <ItemLocation>\n"
        "        <ItemType>Formula</ItemType>\n"
        f"        <ItemPath>{_POWER_QUERY_SECTION}/{_POWER_QUERY_NAME}</ItemPath>\n"
        "      </ItemLocation>\n"
        "      <StableEntries>\n"
        '        <Entry Type="FillEnabled" Value="l0" />\n'
        "      </StableEntries>\n"
        "    </Item>\n"
        "  </Items>\n"
        "</LocalPackageMetadataFile>\n"
    ).encode()
    metadata = struct.pack("<II", 0, len(metadata_xml)) + metadata_xml + struct.pack("<I", 0)
    permissions = (
        b'<?xml version="1.0" encoding="utf-8"?>\n'
        b"<PermissionList>\n"
        b"  <CanEvaluateFuturePackages>false</CanEvaluateFuturePackages>\n"
        b"  <FirewallEnabled>true</FirewallEnabled>\n"
        b"</PermissionList>\n"
    )
    stream = struct.pack("<I", 0)
    for field in (package_parts, permissions, metadata, b""):
        stream += struct.pack("<I", len(field)) + field
    root = ElementTree.Element(f"{{{_DATA_MASHUP_NS}}}DataMashup")
    root.text = base64.b64encode(stream).decode("ascii")
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)


def _add_power_query_local_table_filter(path: Path, *, filter_value: str) -> None:
    """Attach one generated local-table Data Mashup query to an XLSX package."""

    # Validate before editing the package so callers cannot create an ambiguous
    # fixture with a value outside its explicit two-value truth contract.
    _power_query_m_filter_formula(filter_value)

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        relationships = ElementTree.fromstring(members["_rels/.rels"])
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        custom_xml_relationship_type = f"{_DOCUMENT_RELATIONSHIPS_NS}/customXml"
        if any(
            relationship.get("Id") == _POWER_QUERY_ROOT_RELATIONSHIP_ID
            or relationship.get("Type") == custom_xml_relationship_type
            for relationship in relationships.findall(relationship_tag)
        ):
            raise ValueError("Power Query fixture unexpectedly already has custom XML")
        ElementTree.SubElement(
            relationships,
            relationship_tag,
            {
                "Id": _POWER_QUERY_ROOT_RELATIONSHIP_ID,
                "Type": custom_xml_relationship_type,
                "Target": _POWER_QUERY_CUSTOM_XML_MEMBER,
            },
        )
        members["_rels/.rels"] = serialize(relationships)

        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"
        xml_defaults = [
            item
            for item in content_types.findall(default_tag)
            if item.get("Extension") == "xml" and item.get("ContentType") == "application/xml"
        ]
        if len(xml_defaults) != 1:
            raise ValueError("Power Query fixture lacks the default XML content type")
        members[_POWER_QUERY_CUSTOM_XML_MEMBER] = _power_query_local_table_data_mashup(filter_value)

    _rewrite_xlsx_parts(path, mutate)


def _add_scenario_manager_stored_input(path: Path, *, stored_value: str) -> None:
    """Attach WCAB's one raw Scenario Manager declaration to a worksheet.

    Scenario Manager is represented directly inside the worksheet part rather
    than as ordinary cells. This helper creates one selected, locked scenario
    with two stored input records; callers vary only the first record's value.
    It does not ask a spreadsheet client to show a scenario or calculate the
    resulting workbook.
    """

    if stored_value not in {
        _SCENARIO_MANAGER_BASELINE_STORED_VALUE,
        _SCENARIO_MANAGER_CANDIDATE_STORED_VALUE,
    }:
        raise ValueError(f"unsupported Scenario Manager stored value {stored_value!r}")

    def mutate(members: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(members[worksheet_member])
        scenarios_tag = f"{{{_SPREADSHEETML_NS}}}scenarios"
        if worksheet.findall(scenarios_tag):
            raise ValueError("Scenario Manager fixture unexpectedly already has scenarios")
        sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
        sheet_data_indexes = [
            index for index, child in enumerate(worksheet) if child.tag == sheet_data_tag
        ]
        if len(sheet_data_indexes) != 1:
            raise ValueError("Scenario Manager fixture lacks one sheetData element")

        scenarios = ElementTree.Element(
            scenarios_tag,
            {"current": "0", "show": "0", "sqref": _SCENARIO_MANAGER_RESULT_CELL},
        )
        scenario = ElementTree.SubElement(
            scenarios,
            f"{{{_SPREADSHEETML_NS}}}scenario",
            {
                "name": _SCENARIO_MANAGER_NAME,
                "locked": "1",
                "count": "2",
                "user": _SCENARIO_MANAGER_USER,
                "comment": _SCENARIO_MANAGER_COMMENT,
            },
        )
        ElementTree.SubElement(
            scenario,
            f"{{{_SPREADSHEETML_NS}}}inputCells",
            {
                "r": _SCENARIO_MANAGER_CHANGING_CELL,
                "val": stored_value,
                "numFmtId": str(_SCENARIO_MANAGER_INPUT_NUMBER_FORMAT_ID),
            },
        )
        ElementTree.SubElement(
            scenario,
            f"{{{_SPREADSHEETML_NS}}}inputCells",
            {
                "r": _SCENARIO_MANAGER_STABLE_INPUT_CELL,
                "val": _SCENARIO_MANAGER_STABLE_STORED_VALUE,
            },
        )
        worksheet.insert(sheet_data_indexes[0] + 1, scenarios)
        members[worksheet_member] = ElementTree.tostring(
            worksheet, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _set_what_if_data_table_input_reference(path: Path, *, input_cell: str) -> None:
    """Set the one generated Data Table master's local input reference.

    ``f/@r1`` is raw SpreadsheetML metadata on the Data Table master. The
    helper edits that one attribute after openpyxl serializes the ordinary
    workbook, so the pair does not manufacture or recalculate a table result.
    """

    if input_cell not in {
        _WHAT_IF_DATA_TABLE_PRIMARY_INPUT_CELL,
        _WHAT_IF_DATA_TABLE_ALTERNATE_INPUT_CELL,
    }:
        raise ValueError(f"unsupported What-If Data Table input cell {input_cell!r}")

    def mutate(members: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(members[worksheet_member])
        cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
        formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
        masters = [
            cell
            for cell in worksheet.iter(cell_tag)
            if cell.get("r") == _WHAT_IF_DATA_TABLE_MASTER_CELL
        ]
        if len(masters) != 1:
            raise ValueError("What-If Data Table fixture lacks one master cell")
        formulas = masters[0].findall(formula_tag)
        if (
            len(formulas) != 1
            or formulas[0].get("t") != "dataTable"
            or formulas[0].get("ref") != _WHAT_IF_DATA_TABLE_OUTPUT_RANGE
            or formulas[0].get("r1") != _WHAT_IF_DATA_TABLE_PRIMARY_INPUT_CELL
        ):
            raise ValueError("What-If Data Table fixture has an unexpected master declaration")
        formulas[0].set("r1", input_cell)
        members[worksheet_member] = ElementTree.tostring(
            worksheet, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _set_data_validation_list_source(path: Path, *, source_formula: str) -> None:
    """Set WCAB's one list-validation source declaration without saving cells.

    Excel stores a list validation's source as the ``formula1`` child of the
    target's ``dataValidation`` record. Editing that small raw declaration
    after openpyxl saves the baseline shape makes the candidate's package
    boundary explicit: only the Inputs worksheet part is allowed to differ.
    """

    if source_formula not in {
        _DATA_VALIDATION_LIST_BASELINE_SOURCE_FORMULA,
        _DATA_VALIDATION_LIST_CANDIDATE_SOURCE_FORMULA,
    }:
        raise ValueError(f"unsupported list-validation source {source_formula!r}")

    def mutate(members: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(members[worksheet_member])
        data_validations_tag = f"{{{_SPREADSHEETML_NS}}}dataValidations"
        data_validation_tag = f"{{{_SPREADSHEETML_NS}}}dataValidation"
        formula1_tag = f"{{{_SPREADSHEETML_NS}}}formula1"
        containers = worksheet.findall(data_validations_tag)
        if len(containers) != 1:
            raise ValueError("list-validation fixture lacks one dataValidations container")
        validations = list(containers[0])
        if len(validations) != 1 or validations[0].tag != data_validation_tag:
            raise ValueError("list-validation fixture lacks one target validation")
        validation = validations[0]
        formulas = validation.findall(formula1_tag)
        if (
            validation.get("sqref") != _DATA_VALIDATION_LIST_TARGET_RANGE
            or validation.get("type") != "list"
            or len(validation) != 1
            or len(formulas) != 1
            or formulas[0].attrib
            or len(formulas[0])
            or formulas[0].text != _DATA_VALIDATION_LIST_BASELINE_SOURCE_FORMULA
        ):
            raise ValueError("list-validation fixture has an unexpected source declaration")
        formulas[0].text = source_formula
        members[worksheet_member] = ElementTree.tostring(
            worksheet, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _set_conditional_formatting_threshold(path: Path, *, formula: str) -> None:
    """Set WCAB's one conditional-formatting threshold without saving cells.

    The stored ``formula`` child of a ``cellIs`` conditional-formatting rule
    defines the exception threshold. Editing it after openpyxl writes the
    baseline shape keeps the candidate's boundary narrow: only its operations
    worksheet part may differ, and no conditional formatting is evaluated.
    """

    if formula not in {
        _CONDITIONAL_FORMATTING_THRESHOLD_BASELINE_FORMULA,
        _CONDITIONAL_FORMATTING_THRESHOLD_CANDIDATE_FORMULA,
    }:
        raise ValueError(f"unsupported conditional-formatting threshold {formula!r}")

    def mutate(members: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(members[worksheet_member])
        conditional_formatting_tag = f"{{{_SPREADSHEETML_NS}}}conditionalFormatting"
        rule_tag = f"{{{_SPREADSHEETML_NS}}}cfRule"
        formula_tag = f"{{{_SPREADSHEETML_NS}}}formula"
        controls = worksheet.findall(conditional_formatting_tag)
        if (
            len(controls) != 1
            or controls[0].get("sqref") != _CONDITIONAL_FORMATTING_THRESHOLD_RANGE
        ):
            raise ValueError("conditional-formatting fixture lacks one target control")
        rules = list(controls[0])
        if len(rules) != 1 or rules[0].tag != rule_tag:
            raise ValueError("conditional-formatting fixture lacks one rule")
        rule = rules[0]
        formulas = rule.findall(formula_tag)
        if (
            tuple(sorted(rule.attrib.items()))
            != (
                ("dxfId", "0"),
                ("operator", "greaterThan"),
                ("priority", "1"),
                ("type", "cellIs"),
            )
            or len(rule) != 1
            or len(formulas) != 1
            or formulas[0].attrib
            or len(formulas[0])
            or formulas[0].text != _CONDITIONAL_FORMATTING_THRESHOLD_BASELINE_FORMULA
        ):
            raise ValueError("conditional-formatting fixture has an unexpected threshold rule")
        formulas[0].text = formula
        members[worksheet_member] = ElementTree.tostring(
            worksheet, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _set_number_format_visibility_format(path: Path, *, format_code: str) -> None:
    """Set WCAB's one custom format code without saving cells or styles.

    Custom number formats are declared in styles.xml. Replacing only the
    selected formatCode after openpyxl saves the baseline shape keeps the
    candidate boundary narrow: the target cell retains the same style index and
    only the referenced custom display declaration changes.
    """

    if format_code != _NUMBER_FORMAT_VISIBILITY_CANDIDATE_FORMAT:
        raise ValueError(f"unsupported number format {format_code!r}")

    def mutate(members: dict[str, bytes]) -> None:
        styles_member = "xl/styles.xml"
        styles = ElementTree.fromstring(members[styles_member])
        num_fmts_tag = f"{{{_SPREADSHEETML_NS}}}numFmts"
        num_fmt_tag = f"{{{_SPREADSHEETML_NS}}}numFmt"
        containers = styles.findall(num_fmts_tag)
        if len(containers) != 1 or tuple(sorted(containers[0].attrib.items())) != (("count", "1"),):
            raise ValueError("number-format fixture lacks one custom-number-format container")
        number_formats = containers[0].findall(num_fmt_tag)
        if len(number_formats) != 1 or tuple(sorted(number_formats[0].attrib.items())) != (
            ("formatCode", _NUMBER_FORMAT_VISIBILITY_BASELINE_FORMAT),
            ("numFmtId", str(_NUMBER_FORMAT_VISIBILITY_CUSTOM_ID)),
        ):
            raise ValueError("number-format fixture has an unexpected custom format")
        number_formats[0].set("formatCode", format_code)
        members[styles_member] = ElementTree.tostring(
            styles, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _add_ignored_error_formula_range_suppression(path: Path) -> None:
    """Add WCAB's one stored formula-range warning suppression.

    ``ignoredErrors/ignoredError`` records a per-range decision to suppress a
    category of Excel error checking. The raw mutation occurs after openpyxl
    creates stable cells and formulas, making the sole candidate boundary the
    generated Operations worksheet declaration rather than a rendered warning.
    """

    def mutate(members: dict[str, bytes]) -> None:
        worksheet_member = "xl/worksheets/sheet1.xml"
        worksheet = ElementTree.fromstring(members[worksheet_member])
        ignored_errors_tag = f"{{{_SPREADSHEETML_NS}}}ignoredErrors"
        ignored_error_tag = f"{{{_SPREADSHEETML_NS}}}ignoredError"
        sheet_data_tag = f"{{{_SPREADSHEETML_NS}}}sheetData"
        cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
        formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
        value_tag = f"{{{_SPREADSHEETML_NS}}}v"
        if worksheet.findall(ignored_errors_tag):
            raise ValueError("ignored-error fixture unexpectedly already has controls")
        target_cells = [
            cell
            for cell in worksheet.iter(cell_tag)
            if cell.get("r") == _IGNORED_ERROR_SUPPRESSION_TARGET_RANGE
        ]
        if len(target_cells) != 1:
            raise ValueError("ignored-error fixture lacks one target formula cell")
        target_children = list(target_cells[0])
        formulas = target_cells[0].findall(formula_tag)
        if (
            len(formulas) != 1
            or len(target_children) not in {1, 2}
            or target_children[0].tag != formula_tag
            or (len(target_children) == 2 and target_children[1].tag != value_tag)
            or formulas[0].attrib
            or len(formulas[0])
            or formulas[0].text != _IGNORED_ERROR_SUPPRESSION_FORMULA.removeprefix("=")
        ):
            raise ValueError("ignored-error fixture has an unexpected target formula")
        sheet_data_indexes = [
            index for index, child in enumerate(worksheet) if child.tag == sheet_data_tag
        ]
        if len(sheet_data_indexes) != 1:
            raise ValueError("ignored-error fixture lacks one sheetData element")
        controls = ElementTree.Element(ignored_errors_tag)
        ElementTree.SubElement(
            controls,
            ignored_error_tag,
            {
                "sqref": _IGNORED_ERROR_SUPPRESSION_TARGET_RANGE,
                _IGNORED_ERROR_SUPPRESSION_FLAG: "1",
            },
        )
        worksheet.insert(sheet_data_indexes[0] + 1, controls)
        members[worksheet_member] = ElementTree.tostring(
            worksheet, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _add_pivot_slicer_selection(path: Path, *, selected_item_index: int) -> None:
    """Attach one local PivotTable slicer with one explicitly selected cache item.

    ``slicerCacheDefinition/data/tabular/items/i/@s`` is a stored Slicer-cache
    state.  This package-only fixture deliberately records that declaration
    after openpyxl writes ordinary worksheet cells; it neither creates a Slicer
    drawing nor asks a spreadsheet client to apply, refresh, calculate, or
    render the filter.
    """

    if type(selected_item_index) is not int or selected_item_index not in range(
        _PIVOT_SLICER_ITEM_COUNT
    ):
        raise ValueError(f"unsupported PivotTable slicer selection {selected_item_index!r}")

    def serialize(root: ElementTree.Element) -> bytes:
        return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(members["xl/workbook.xml"])
        relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        sheets = workbook.find(f"{{{_SPREADSHEETML_NS}}}sheets")
        report_sheets = (
            [
                sheet
                for sheet in sheets.findall(f"{{{_SPREADSHEETML_NS}}}sheet")
                if sheet.get("name") == _PIVOT_REPORT_SHEET
            ]
            if sheets is not None
            else []
        )
        if len(report_sheets) != 1:
            raise ValueError("PivotTable slicer fixture has an unexpected Report worksheet")
        report_sheet_id = report_sheets[0].get("sheetId")
        try:
            if (
                not isinstance(report_sheet_id, str)
                or int(report_sheet_id) != _PIVOT_SLICER_PIVOT_TAB_ID
            ):
                raise ValueError
        except ValueError as error:
            raise ValueError(
                "PivotTable slicer fixture Report worksheet has no sheet id"
            ) from error

        pivot_caches = workbook.findall(f"{{{_SPREADSHEETML_NS}}}pivotCaches")
        if len(pivot_caches) != 1 or len(pivot_caches[0]) != 1:
            raise ValueError("PivotTable slicer fixture must contain exactly one PivotCache")
        pivot_cache = pivot_caches[0][0]
        if pivot_cache.tag != f"{{{_SPREADSHEETML_NS}}}pivotCache":
            raise ValueError("PivotTable slicer fixture has an unexpected PivotCache declaration")
        if workbook.find(f"{{{_SPREADSHEETML_NS}}}extLst") is not None:
            raise ValueError(
                "PivotTable slicer fixture unexpectedly already has workbook extensions"
            )

        workbook_extensions = ElementTree.SubElement(workbook, f"{{{_SPREADSHEETML_NS}}}extLst")
        workbook_extension = ElementTree.SubElement(
            workbook_extensions,
            f"{{{_SPREADSHEETML_NS}}}ext",
            {"uri": _PIVOT_SLICER_CACHE_EXTENSION_URI},
        )
        slicer_caches = ElementTree.SubElement(
            workbook_extension, f"{{{_OFFICE_2010_SPREADSHEET_NS}}}slicerCaches"
        )
        ElementTree.SubElement(
            slicer_caches,
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}slicerCache",
            {relationship_id_attribute: "rIdWCABSlicerCache"},
        )
        members["xl/workbook.xml"] = serialize(workbook)

        workbook_relationships = ElementTree.fromstring(members["xl/_rels/workbook.xml.rels"])
        if any(
            relationship.get("Id") == "rIdWCABSlicerCache"
            for relationship in workbook_relationships.findall(relationship_tag)
        ):
            raise ValueError("PivotTable slicer fixture relationship ID is already in use")
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": "rIdWCABSlicerCache",
                "Type": _PIVOT_SLICER_CACHE_RELATIONSHIP,
                "Target": "slicerCaches/slicerCache1.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        cache_definition_member = "xl/pivotCache/pivotCacheDefinition1.xml"
        cache_definition = ElementTree.fromstring(members[cache_definition_member])
        if cache_definition.tag != f"{{{_SPREADSHEETML_NS}}}pivotCacheDefinition":
            raise ValueError("PivotTable slicer fixture has an unexpected cache definition")
        if cache_definition.find(f"{{{_SPREADSHEETML_NS}}}extLst") is not None:
            raise ValueError("PivotTable slicer fixture unexpectedly already has cache extensions")
        cache_extensions = ElementTree.SubElement(
            cache_definition, f"{{{_SPREADSHEETML_NS}}}extLst"
        )
        cache_extension = ElementTree.SubElement(
            cache_extensions,
            f"{{{_SPREADSHEETML_NS}}}ext",
            {"uri": _PIVOT_SLICER_PIVOT_CACHE_EXTENSION_URI},
        )
        ElementTree.SubElement(
            cache_extension,
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}pivotCacheDefinition",
            {"pivotCacheId": str(_PIVOT_CACHE_ID)},
        )
        members[cache_definition_member] = serialize(cache_definition)

        slicer_member = "xl/slicerCaches/slicerCache1.xml"
        slicer_cache = ElementTree.Element(
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}slicerCacheDefinition",
            {"name": _PIVOT_SLICER_NAME, "sourceName": _PIVOT_SLICER_SOURCE_NAME},
        )
        pivot_tables = ElementTree.SubElement(
            slicer_cache,
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}pivotTables",
            {"count": "1"},
        )
        ElementTree.SubElement(
            pivot_tables,
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}pivotTable",
            {"tabId": report_sheet_id, "name": _PIVOT_SLICER_PIVOT_TABLE_NAME},
        )
        data = ElementTree.SubElement(slicer_cache, f"{{{_OFFICE_2010_SPREADSHEET_NS}}}data")
        tabular = ElementTree.SubElement(
            data,
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}tabular",
            {"pivotCacheId": str(_PIVOT_CACHE_ID)},
        )
        items = ElementTree.SubElement(
            tabular,
            f"{{{_OFFICE_2010_SPREADSHEET_NS}}}items",
            {"count": str(_PIVOT_SLICER_ITEM_COUNT)},
        )
        for item_index in range(_PIVOT_SLICER_ITEM_COUNT):
            ElementTree.SubElement(
                items,
                f"{{{_OFFICE_2010_SPREADSHEET_NS}}}i",
                {"x": str(item_index), "s": "1" if item_index == selected_item_index else "0"},
            )
        members[slicer_member] = serialize(slicer_cache)

        content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        if any(
            item.get("PartName") == f"/{slicer_member}"
            for item in content_types.findall(override_tag)
        ):
            raise ValueError("PivotTable slicer fixture already has a slicer content type")
        ElementTree.SubElement(
            content_types,
            override_tag,
            {"PartName": f"/{slicer_member}", "ContentType": _PIVOT_SLICER_CACHE_CONTENT_TYPE},
        )
        members["[Content_Types].xml"] = serialize(content_types)

    _rewrite_xlsx_parts(path, mutate)


def _set_pivot_data_field_subtotal(path: Path, *, subtotal: str) -> None:
    """Set the sole generated PivotTable value field's aggregation control.

    ``dataField/@subtotal`` stores the aggregate function for a PivotTable
    value field.  This is deliberately a raw OOXML change after openpyxl has
    written the ordinary workbook, so the fixture records a PivotTable
    declaration rather than a recalculated or rendered output.
    """

    if subtotal not in {
        _PIVOT_DATA_FIELD_BASELINE_SUBTOTAL,
        _PIVOT_DATA_FIELD_CANDIDATE_SUBTOTAL,
    }:
        raise ValueError(f"unsupported PivotTable data-field subtotal {subtotal!r}")

    def mutate(members: dict[str, bytes]) -> None:
        pivot_members = sorted(
            name
            for name in members
            if name.startswith("xl/pivotTables/")
            and name.endswith(".xml")
            and "/_rels/" not in name
        )
        if pivot_members != ["xl/pivotTables/pivotTable1.xml"]:
            raise ValueError("pivot aggregation fixture must contain exactly one PivotTable part")
        pivot_table = ElementTree.fromstring(members[pivot_members[0]])
        data_fields_tag = f"{{{_SPREADSHEETML_NS}}}dataFields"
        data_field_tag = f"{{{_SPREADSHEETML_NS}}}dataField"
        data_fields = pivot_table.findall(data_fields_tag)
        if (
            len(data_fields) != 1
            or data_fields[0].get("count") != "1"
            or len(data_fields[0]) != 1
            or data_fields[0][0].tag != data_field_tag
            or data_fields[0][0].get("fld") != str(_PIVOT_DATA_FIELD_SOURCE_INDEX)
        ):
            raise ValueError("pivot aggregation fixture has an unexpected data-field declaration")
        data_fields[0][0].set("subtotal", subtotal)
        members[pivot_members[0]] = ElementTree.tostring(
            pivot_table, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _set_chart_series_value_reference(path: Path, *, value_reference: str) -> None:
    """Set the sole generated chart's stored numeric-series reference.

    The mutation is deliberately limited to the one ``c:ser/c:val/c:numRef/c:f``
    text node. Rewriting it after openpyxl saves the package keeps this fixture
    about a DrawingML chart binding rather than a worksheet edit or a rendered
    chart cache.
    """

    if value_reference not in {
        _CHART_SERIES_BASELINE_VALUE_REFERENCE,
        _CHART_SERIES_CANDIDATE_VALUE_REFERENCE,
    }:
        raise ValueError(f"unsupported chart-series value reference {value_reference!r}")

    def mutate(members: dict[str, bytes]) -> None:
        chart_members = sorted(
            name
            for name in members
            if name.startswith("xl/charts/") and name.endswith(".xml") and "/_rels/" not in name
        )
        if len(chart_members) != 1:
            raise ValueError("chart-series fixture must contain exactly one chart part")
        chart_member = chart_members[0]
        chart = ElementTree.fromstring(members[chart_member])
        series = chart.findall(f".//{{{_DRAWINGML_CHART_NS}}}ser")
        if len(series) != 1:
            raise ValueError("chart-series fixture must contain exactly one series")
        formulas = series[0].findall(
            f"{{{_DRAWINGML_CHART_NS}}}val/"
            f"{{{_DRAWINGML_CHART_NS}}}numRef/"
            f"{{{_DRAWINGML_CHART_NS}}}f"
        )
        if len(formulas) != 1:
            raise ValueError("chart-series fixture must contain exactly one numeric value formula")
        formulas[0].text = value_reference
        members[chart_member] = ElementTree.tostring(chart, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_parts(path, mutate)


def _set_table_calculated_column_formula(path: Path, *, formula: str) -> None:
    """Store WCAB's sole Table calculated-column formula declaration.

    ``calculatedColumnFormula`` belongs to the table definition rather than a
    worksheet cell.  This narrowly writes one formula-master text node after
    the workbook is saved and rejects any unexpected Table shape, keeping the
    resulting pair about this one raw structural boundary.
    """

    if formula not in {
        _TABLE_CALCULATED_COLUMN_BASELINE_FORMULA,
        _TABLE_CALCULATED_COLUMN_CANDIDATE_FORMULA,
    }:
        raise ValueError(f"unsupported Table calculated-column formula {formula!r}")

    def mutate(members: dict[str, bytes]) -> None:
        if _TABLE_CALCULATED_COLUMN_MEMBER not in members:
            raise ValueError("calculated-column fixture has no Table definition")
        table = ElementTree.fromstring(members[_TABLE_CALCULATED_COLUMN_MEMBER])
        table_tag = f"{{{_SPREADSHEETML_NS}}}table"
        auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
        columns_tag = f"{{{_SPREADSHEETML_NS}}}tableColumns"
        column_tag = f"{{{_SPREADSHEETML_NS}}}tableColumn"
        formula_tag = f"{{{_SPREADSHEETML_NS}}}calculatedColumnFormula"
        expected_table_attributes = {
            "id": "1",
            "name": _TABLE_CALCULATED_COLUMN_NAME,
            "displayName": _TABLE_CALCULATED_COLUMN_NAME,
            "ref": _TABLE_CALCULATED_COLUMN_REF,
            "headerRowCount": "1",
        }
        expected_columns = (
            {"id": "1", "name": "Units"},
            {"id": "2", "name": "Rate"},
            {
                "id": str(_TABLE_CALCULATED_COLUMN_ID),
                "name": _TABLE_CALCULATED_COLUMN_HEADER,
            },
        )
        auto_filters = table.findall(auto_filter_tag)
        containers = table.findall(columns_tag)
        if (
            table.tag != table_tag
            or table.attrib != expected_table_attributes
            or len(table) != 2
            or len(auto_filters) != 1
            or auto_filters[0].attrib != {"ref": _TABLE_CALCULATED_COLUMN_REF}
            or len(auto_filters[0])
            or len(containers) != 1
            or containers[0].attrib != {"count": "3"}
        ):
            raise ValueError("calculated-column fixture has an unexpected Table shape")
        columns = list(containers[0])
        if (
            len(columns) != 3
            or any(column.tag != column_tag for column in columns)
            or tuple(column.attrib for column in columns) != expected_columns
            or any(len(column) for column in columns)
        ):
            raise ValueError("calculated-column fixture has unexpected Table columns")
        calculated_formula = ElementTree.SubElement(columns[-1], formula_tag)
        calculated_formula.text = formula
        members[_TABLE_CALCULATED_COLUMN_MEMBER] = ElementTree.tostring(
            table, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _attach_power_pivot_data_model(path: Path, *, to_column: str) -> None:
    """Attach WCAB's one relationship-backed, opaque Data Model package.

    Excel's embedded Data Model stream is an Analysis Services payload.  This
    fixture deliberately leaves it as fixed inert bytes and writes only the
    compact workbook-level x15 declaration needed to express one relationship.
    No Data Model is opened, deserialized, refreshed, or calculated here.
    """

    if to_column not in {
        _POWER_PIVOT_DATA_BASELINE_TO_COLUMN,
        _POWER_PIVOT_DATA_CANDIDATE_TO_COLUMN,
    }:
        raise ValueError(f"unsupported Power Pivot Data Model target column {to_column!r}")

    def serialize(element: ElementTree.Element) -> bytes:
        return ElementTree.tostring(element, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        try:
            workbook = ElementTree.fromstring(members[_POWER_PIVOT_DATA_WORKBOOK_MEMBER])
            workbook_relationships = ElementTree.fromstring(
                members[_POWER_PIVOT_DATA_WORKBOOK_RELATIONSHIPS_MEMBER]
            )
            content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        except (KeyError, ElementTree.ParseError) as error:
            raise ValueError("Power Pivot Data Model fixture has no base OOXML package") from error

        ext_list_tag = f"{{{_SPREADSHEETML_NS}}}extLst"
        extension_tag = f"{{{_SPREADSHEETML_NS}}}ext"
        data_model_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}dataModel"
        model_tables_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}modelTables"
        model_table_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}modelTable"
        model_relationships_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}modelRelationships"
        model_relationship_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}modelRelationship"
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"

        if workbook.findall(ext_list_tag):
            raise ValueError("Power Pivot Data Model fixture has unexpected workbook extensions")
        extensions = ElementTree.SubElement(workbook, ext_list_tag)
        extension = ElementTree.SubElement(
            extensions,
            extension_tag,
            {"uri": _POWER_PIVOT_DATA_EXTENSION_URI},
        )
        data_model = ElementTree.SubElement(
            extension,
            data_model_tag,
            {"minVersionLoad": _POWER_PIVOT_DATA_MIN_VERSION_LOAD},
        )
        model_tables = ElementTree.SubElement(data_model, model_tables_tag)
        for table in _POWER_PIVOT_DATA_MODEL_TABLES:
            ElementTree.SubElement(model_tables, model_table_tag, table)
        model_relationships = ElementTree.SubElement(data_model, model_relationships_tag)
        ElementTree.SubElement(
            model_relationships,
            model_relationship_tag,
            {
                "fromTable": _POWER_PIVOT_DATA_FROM_TABLE,
                "fromColumn": _POWER_PIVOT_DATA_FROM_COLUMN,
                "toTable": _POWER_PIVOT_DATA_TO_TABLE,
                "toColumn": to_column,
            },
        )

        if (
            any(
                relationship.get("Id") == _POWER_PIVOT_DATA_WORKBOOK_RELATIONSHIP_ID
                or relationship.get("Type") == _POWER_PIVOT_DATA_RELATIONSHIP
                for relationship in workbook_relationships.findall(relationship_tag)
            )
            or any(
                content_type.get("Extension") == "data"
                for content_type in content_types.findall(default_tag)
            )
            or _POWER_PIVOT_DATA_MEMBER in members
        ):
            raise ValueError("Power Pivot Data Model fixture already has a Data Model package")
        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": _POWER_PIVOT_DATA_WORKBOOK_RELATIONSHIP_ID,
                "Type": _POWER_PIVOT_DATA_RELATIONSHIP,
                "Target": "model/item.data",
            },
        )
        ElementTree.SubElement(
            content_types,
            default_tag,
            {"Extension": "data", "ContentType": _POWER_PIVOT_DATA_CONTENT_TYPE},
        )

        members[_POWER_PIVOT_DATA_WORKBOOK_MEMBER] = serialize(workbook)
        members[_POWER_PIVOT_DATA_WORKBOOK_RELATIONSHIPS_MEMBER] = serialize(workbook_relationships)
        members["[Content_Types].xml"] = serialize(content_types)
        members[_POWER_PIVOT_DATA_MEMBER] = _POWER_PIVOT_DATA_PAYLOAD

    _rewrite_xlsx_parts(path, mutate)


def _attach_xlm_auto_open_binding(path: Path, *, target: str) -> None:
    """Attach one fixed XLM macro sheet and a workbook-scoped Auto_Open name.

    The macro sheet is a tiny raw OOXML payload containing only two identical
    ``HALT()`` cells. The candidate changes the special workbook defined-name
    target only; WCAB never interprets a macro formula, opens Excel, or permits
    a macro to run while building the pair.
    """

    if target not in {_XLM_AUTO_OPEN_BASELINE_TARGET, _XLM_AUTO_OPEN_CANDIDATE_TARGET}:
        raise ValueError(f"unsupported XLM Auto_Open target {target!r}")

    def serialize(element: ElementTree.Element) -> bytes:
        return ElementTree.tostring(element, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        try:
            workbook = ElementTree.fromstring(members[_XLM_AUTO_OPEN_WORKBOOK_MEMBER])
            workbook_relationships = ElementTree.fromstring(
                members[_XLM_AUTO_OPEN_WORKBOOK_RELATIONSHIPS_MEMBER]
            )
            content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        except (KeyError, ElementTree.ParseError) as error:
            raise ValueError("XLM Auto_Open fixture has no base OOXML package") from error

        sheets_tag = f"{{{_SPREADSHEETML_NS}}}sheets"
        sheet_tag = f"{{{_SPREADSHEETML_NS}}}sheet"
        defined_names_tag = f"{{{_SPREADSHEETML_NS}}}definedNames"
        defined_name_tag = f"{{{_SPREADSHEETML_NS}}}definedName"
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        sheets = workbook.findall(sheets_tag)
        defined_names = workbook.findall(defined_names_tag)
        if (
            len(sheets) != 1
            or len(defined_names) != 1
            or defined_names[0].attrib
            or len(defined_names[0])
            or any(sheet.get("name") == _XLM_AUTO_OPEN_MACRO_SHEET_NAME for sheet in sheets[0])
            or any(sheet.tag != sheet_tag for sheet in sheets[0])
            or workbook_relationships.tag != f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
            or content_types.tag != f"{{{_CONTENT_TYPES_NS}}}Types"
        ):
            raise ValueError("XLM Auto_Open fixture has an unexpected base workbook shape")
        sheet_ids = [sheet.get("sheetId") for sheet in sheets[0]]
        if sheet_ids != ["1", "2", "3"]:
            raise ValueError("XLM Auto_Open fixture has unexpected ordinary sheet IDs")
        if any(
            relationship.get("Id") == _XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP_ID
            or relationship.get("Type") == _XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP
            for relationship in workbook_relationships.findall(relationship_tag)
        ):
            raise ValueError("XLM Auto_Open fixture already has a macro-sheet binding")
        if _XLM_AUTO_OPEN_MACRO_SHEET_MEMBER in members:
            raise ValueError("XLM Auto_Open fixture already has a macro sheet")

        ElementTree.SubElement(
            sheets[0],
            sheet_tag,
            {
                "name": _XLM_AUTO_OPEN_MACRO_SHEET_NAME,
                "sheetId": _XLM_AUTO_OPEN_MACRO_SHEET_ID,
                "state": _XLM_AUTO_OPEN_MACRO_SHEET_STATE,
                f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": _XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP_ID,
            },
        )
        ElementTree.SubElement(
            defined_names[0],
            defined_name_tag,
            {"name": _XLM_AUTO_OPEN_DEFINED_NAME},
        ).text = target

        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": _XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP_ID,
                "Type": _XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP,
                "Target": "macrosheets/sheet1.xml",
            },
        )

        workbook_overrides = [
            override
            for override in content_types.findall(override_tag)
            if override.get("PartName") == "/xl/workbook.xml"
        ]
        if len(workbook_overrides) != 1 or any(
            override.get("PartName") == f"/{_XLM_AUTO_OPEN_MACRO_SHEET_MEMBER}"
            for override in content_types.findall(override_tag)
        ):
            raise ValueError("XLM Auto_Open fixture has unexpected content types")
        workbook_overrides[0].set("ContentType", _XLM_AUTO_OPEN_WORKBOOK_CONTENT_TYPE)
        ElementTree.SubElement(
            content_types,
            override_tag,
            {
                "PartName": f"/{_XLM_AUTO_OPEN_MACRO_SHEET_MEMBER}",
                "ContentType": _XLM_AUTO_OPEN_MACRO_SHEET_CONTENT_TYPE,
            },
        )

        members[_XLM_AUTO_OPEN_WORKBOOK_MEMBER] = serialize(workbook)
        members[_XLM_AUTO_OPEN_WORKBOOK_RELATIONSHIPS_MEMBER] = serialize(workbook_relationships)
        members["[Content_Types].xml"] = serialize(content_types)
        members[_XLM_AUTO_OPEN_MACRO_SHEET_MEMBER] = _XLM_AUTO_OPEN_MACRO_SHEET_PAYLOAD

    _rewrite_xlsx_parts(path, mutate)


def _set_external_workbook_link_update_policy(path: Path, *, update_links: str) -> None:
    """Set the stored workbook-open policy for external-workbook links.

    ``workbookPr/@updateLinks`` is distinct from the relationship-backed
    connection control used by the earlier external-data case. The allowed
    values here deliberately make the generated transition unambiguous:
    never update linked workbooks at open versus always update them at open.
    """

    if update_links not in {"never", "always"}:
        raise ValueError(f"unsupported external-workbook update policy {update_links!r}")

    def mutate(members: dict[str, bytes]) -> None:
        workbook = ElementTree.fromstring(members["xl/workbook.xml"])
        properties_tag = f"{{{_SPREADSHEETML_NS}}}workbookPr"
        properties = workbook.find(properties_tag)
        if properties is None:
            properties = ElementTree.Element(properties_tag)
            workbook.insert(0, properties)
        properties.set("updateLinks", update_links)
        members["xl/workbook.xml"] = ElementTree.tostring(
            workbook, encoding="utf-8", xml_declaration=True
        )

    _rewrite_xlsx_parts(path, mutate)


def _add_external_workbook_link_source(path: Path, *, target: str) -> None:
    """Attach one compact external-workbook package with a stored source target.

    The generated source is represented only by the standard ``externalLink``
    relationship graph. It deliberately has no local source workbook, cached
    source cells, defined names, DDE/OLE material, credentials, or executable
    content. This routine never resolves, opens, fetches, or refreshes the
    external target.
    """

    if not isinstance(target, str) or not target:
        raise ValueError("external-workbook source target must be a non-empty string")

    def serialize(element: ElementTree.Element) -> bytes:
        return ElementTree.tostring(element, encoding="utf-8", xml_declaration=True)

    def mutate(members: dict[str, bytes]) -> None:
        try:
            workbook = ElementTree.fromstring(members["xl/workbook.xml"])
            workbook_relationships = ElementTree.fromstring(members["xl/_rels/workbook.xml.rels"])
            content_types = ElementTree.fromstring(members["[Content_Types].xml"])
        except (KeyError, ElementTree.ParseError) as error:
            raise ValueError("external-workbook-link fixture has no base OOXML package") from error

        external_references_tag = f"{{{_SPREADSHEETML_NS}}}externalReferences"
        external_reference_tag = f"{{{_SPREADSHEETML_NS}}}externalReference"
        relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
        relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
        if workbook.findall(external_references_tag):
            raise ValueError("external-workbook-link fixture already has external references")
        if any(
            relationship.get("Id") == _EXTERNAL_WORKBOOK_LINK_WORKBOOK_RELATIONSHIP_ID
            or relationship.get("Type") == _EXTERNAL_WORKBOOK_LINK_RELATIONSHIP
            for relationship in workbook_relationships.findall(relationship_tag)
        ):
            raise ValueError(
                "external-workbook-link fixture already has an external-link relationship"
            )
        if any(
            override.get("PartName") == f"/{_EXTERNAL_WORKBOOK_LINK_MEMBER}"
            for override in content_types.findall(override_tag)
        ):
            raise ValueError("external-workbook-link fixture already has an external-link part")
        if (
            _EXTERNAL_WORKBOOK_LINK_MEMBER in members
            or _EXTERNAL_WORKBOOK_LINK_RELATIONSHIPS_MEMBER in members
        ):
            raise ValueError("external-workbook-link fixture already has external-link payload")

        external_references = ElementTree.Element(external_references_tag)
        ElementTree.SubElement(
            external_references,
            external_reference_tag,
            {
                relationship_id_attribute: _EXTERNAL_WORKBOOK_LINK_WORKBOOK_RELATIONSHIP_ID,
            },
        )
        calculation_properties = workbook.find(f"{{{_SPREADSHEETML_NS}}}calcPr")
        workbook.insert(
            (
                list(workbook).index(calculation_properties)
                if calculation_properties is not None
                else len(workbook)
            ),
            external_references,
        )
        members["xl/workbook.xml"] = serialize(workbook)

        ElementTree.SubElement(
            workbook_relationships,
            relationship_tag,
            {
                "Id": _EXTERNAL_WORKBOOK_LINK_WORKBOOK_RELATIONSHIP_ID,
                "Type": _EXTERNAL_WORKBOOK_LINK_RELATIONSHIP,
                "Target": "externalLinks/externalLink1.xml",
            },
        )
        members["xl/_rels/workbook.xml.rels"] = serialize(workbook_relationships)

        ElementTree.SubElement(
            content_types,
            override_tag,
            {
                "PartName": f"/{_EXTERNAL_WORKBOOK_LINK_MEMBER}",
                "ContentType": _EXTERNAL_WORKBOOK_LINK_CONTENT_TYPE,
            },
        )
        members["[Content_Types].xml"] = serialize(content_types)

        external_link = ElementTree.Element(f"{{{_SPREADSHEETML_NS}}}externalLink")
        external_book = ElementTree.SubElement(
            external_link,
            f"{{{_SPREADSHEETML_NS}}}externalBook",
            {
                relationship_id_attribute: _EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP_ID,
            },
        )
        sheet_names = ElementTree.SubElement(
            external_book,
            f"{{{_SPREADSHEETML_NS}}}sheetNames",
        )
        ElementTree.SubElement(
            sheet_names,
            f"{{{_SPREADSHEETML_NS}}}sheetName",
            {"val": _EXTERNAL_WORKBOOK_LINK_EXTERNAL_SHEET},
        )
        members[_EXTERNAL_WORKBOOK_LINK_MEMBER] = serialize(external_link)

        external_link_relationships = ElementTree.Element(
            f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        )
        ElementTree.SubElement(
            external_link_relationships,
            relationship_tag,
            {
                "Id": _EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP_ID,
                "Type": _EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP,
                "Target": target,
                "TargetMode": "External",
            },
        )
        members[_EXTERNAL_WORKBOOK_LINK_RELATIONSHIPS_MEMBER] = serialize(
            external_link_relationships
        )

    _rewrite_xlsx_parts(path, mutate)


def _portfolio_workbook(*, driver_value: int) -> tuple[Workbook, Workbook]:
    drivers = Workbook()
    _configure_workbook(drivers, title="WCAB portfolio drivers")
    inputs = drivers.active
    inputs.title = "Inputs"
    inputs["A1"] = "Approved driver"
    inputs["B2"] = driver_value

    model = Workbook()
    _configure_workbook(model, title="WCAB portfolio model")
    summary = model.active
    summary.title = "Summary"
    summary["A1"] = "Externally supplied driver"
    summary["B2"] = "='[drivers.xlsx]Inputs'!$B$2"
    summary["A3"] = "Derived output"
    summary["B3"] = "=B2*10"
    return drivers, model


def _build_finance_formula_to_value(root: Path) -> None:
    truth = _truth(
        case_id="finance.formula_to_value",
        title="Revenue formula is replaced with a hard-coded value",
        family="finance",
        review_expectation="block",
        facts=[{"kind": "formula_to_value", "sheet": "Revenue", "cell": "C8"}],
        must_reach=[
            {
                "source": {"sheet": "Revenue", "cell": "C8"},
                "targets": [
                    {"sheet": "Revenue", "cell": "C10"},
                    {"sheet": "Summary", "cell": "C5"},
                    {"sheet": "Summary", "cell": "C6"},
                    {"sheet": "Dashboard", "cell": "B4"},
                ],
            }
        ],
        coverage=["Static dependency lower bound; numerical recalculation is out of scope."],
    )
    _write_pair(
        root / "finance" / "formula_to_value",
        _finance_workbook,
        lambda wb: setattr(wb["Revenue"]["C8"], "value", 999),
        truth,
    )


def _build_finance_wrong_period_reference(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook["Revenue"]["C8"] = "=B5*C6"

    truth = _truth(
        case_id="finance.wrong_period_reference",
        title="Revenue formula uses a prior-period units reference",
        family="finance",
        review_expectation="block",
        facts=[{"kind": "formula_changed", "sheet": "Revenue", "cell": "C8"}],
        must_reach=[
            {
                "source": {"sheet": "Revenue", "cell": "C8"},
                "targets": [
                    {"sheet": "Revenue", "cell": "C10"},
                    {"sheet": "Summary", "cell": "C5"},
                    {"sheet": "Dashboard", "cell": "B4"},
                ],
            }
        ],
        coverage=[
            "The benchmark asserts the reference drift, not the business correctness of every formula."
        ],
    )
    _write_pair(root / "finance" / "wrong_period_reference", _finance_workbook, mutate, truth)


def _build_finance_input_change(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook["Assumptions"]["C2"] = 140

    truth = _truth(
        case_id="finance.input_value_change",
        title="Approved input changes and propagates into outputs",
        family="finance",
        review_expectation="review",
        facts=[{"kind": "value_changed", "sheet": "Assumptions", "cell": "C2"}],
        must_reach=[
            {
                "source": {"sheet": "Assumptions", "cell": "C2"},
                "targets": [
                    {"sheet": "Revenue", "cell": "C5"},
                    {"sheet": "Revenue", "cell": "C8"},
                    {"sheet": "Revenue", "cell": "C10"},
                    {"sheet": "Summary", "cell": "C6"},
                    {"sheet": "Dashboard", "cell": "B4"},
                ],
            }
        ],
        coverage=[
            "An input change is not inherently an error; it remains material review evidence."
        ],
    )
    _write_pair(root / "finance" / "input_value_change", _finance_workbook, mutate, truth)


def _build_finance_external_formula(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook["Summary"]["C8"] = "='[MarketData.xlsx]Rates'!$B$2"

    truth = _truth(
        case_id="finance.external_formula_reference",
        title="An internal driver is replaced by an external workbook reference",
        family="finance",
        review_expectation="block",
        facts=[
            {"kind": "formula_changed", "sheet": "Summary", "cell": "C8"},
            {"kind": "external_formula_added", "sheet": "Summary", "cell": "C8"},
        ],
        coverage=[
            "The external workbook is intentionally absent and must not be opened by a benchmark tool."
        ],
    )
    _write_pair(root / "finance" / "external_formula_reference", _finance_workbook, mutate, truth)


def _build_finance_defined_name(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook.defined_names["DiscountRate"] = DefinedName(
            "DiscountRate", attr_text="Assumptions!$E$11", comment="WCAB synthetic driver"
        )

    truth = _truth(
        case_id="finance.defined_name_redirect",
        title="A named rate redirects to a different assumption",
        family="finance",
        review_expectation="block",
        facts=[{"kind": "defined_name_changed", "name": "DiscountRate"}],
        coverage=[
            "The assertion compares the defined-name destination; name resolution beyond this fixture is out of scope."
        ],
    )
    _write_pair(root / "finance" / "defined_name_redirect", _finance_workbook, mutate, truth)


def _build_operations_formula_interruption(root: Path) -> None:
    truth = _truth(
        case_id="operations.copied_formula_interruption",
        title="One copied margin formula is replaced by a manual number",
        family="operations",
        review_expectation="block",
        facts=[{"kind": "formula_to_value", "sheet": "Operations", "cell": "F7"}],
        coverage=[
            "Immediate formula peers provide the copied-block context; no formula execution is required."
        ],
    )
    _write_pair(
        root / "operations" / "copied_formula_interruption",
        _operations_workbook,
        lambda wb: setattr(wb["Operations"]["F7"], "value", 42),
        truth,
    )


def _build_operations_sumifs_shape(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook["Operations"]["H6"] = "=SUMIFS($F$6:$F$9,$A$6:$A$8,A6)"

    truth = _truth(
        case_id="operations.sumifs_range_shape",
        title="A conditional aggregate has mismatched static ranges",
        family="operations",
        review_expectation="block",
        facts=[{"kind": "formula_changed", "sheet": "Operations", "cell": "H6"}],
        coverage=[
            "The range-shape defect is deliberately static and does not require evaluating SUMIFS."
        ],
    )
    _write_pair(root / "operations" / "sumifs_range_shape", _operations_workbook, mutate, truth)


def _build_operations_data_validation(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook["Operations"].data_validations.dataValidation.clear()

    truth = _truth(
        case_id="operations.data_validation_removed",
        title="Input validation is removed from an operational tracker",
        family="operations",
        review_expectation="block",
        facts=[
            {
                "kind": "data_validation_count_changed",
                "sheet": "Operations",
                "baseline_count": 1,
                "candidate_count": 0,
            }
        ],
        coverage=[
            "The contract observes validation presence, not whether a future user enters an invalid value."
        ],
    )
    _write_pair(
        root / "operations" / "data_validation_removed", _operations_workbook, mutate, truth
    )


def _build_operations_data_validation_list_source(root: Path) -> None:
    """Build a stored list-source retarget without altering input/result cells."""

    directory = root / "operations" / "data_validation_list_source_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_data_validation_list_source_workbook(), baseline)
    _save_workbook(_data_validation_list_source_workbook(), candidate)
    _set_data_validation_list_source(
        candidate, source_formula=_DATA_VALIDATION_LIST_CANDIDATE_SOURCE_FORMULA
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="operations.data_validation_list_source_changed",
            title="A list validation switches its permitted-status source",
            family="operations",
            review_expectation="block",
            facts=[
                {
                    "kind": "data_validation_list_source_changed",
                    "validation_sheet": _DATA_VALIDATION_LIST_SHEET,
                    "target_range": _DATA_VALIDATION_LIST_TARGET_RANGE,
                    "validation_type": "list",
                    "baseline_source_formula": _DATA_VALIDATION_LIST_BASELINE_SOURCE_FORMULA,
                    "candidate_source_formula": _DATA_VALIDATION_LIST_CANDIDATE_SOURCE_FORMULA,
                    "allow_blank": False,
                    "dropdown_hidden": False,
                    "show_error_message": True,
                    "error_style": _DATA_VALIDATION_LIST_ERROR_STYLE,
                    "error_title": _DATA_VALIDATION_LIST_ERROR_TITLE,
                    "error": _DATA_VALIDATION_LIST_ERROR,
                    "show_input_message": False,
                    "prompt_title": _DATA_VALIDATION_LIST_PROMPT_TITLE,
                    "prompt": _DATA_VALIDATION_LIST_PROMPT,
                    "source_sheet": _DATA_VALIDATION_LIST_SOURCE_SHEET,
                    "baseline_source_range": _DATA_VALIDATION_LIST_BASELINE_SOURCE_RANGE,
                    "candidate_source_range": _DATA_VALIDATION_LIST_CANDIDATE_SOURCE_RANGE,
                    "baseline_source_values": list(_DATA_VALIDATION_LIST_BASELINE_SOURCE_VALUES),
                    "candidate_source_values": list(_DATA_VALIDATION_LIST_CANDIDATE_SOURCE_VALUES),
                    "input_cell": _DATA_VALIDATION_LIST_TARGET_RANGE,
                    "input_value": _DATA_VALIDATION_LIST_INPUT_VALUE,
                    "model_sheet": _DATA_VALIDATION_LIST_MODEL_SHEET,
                    "model_cell": _DATA_VALIDATION_LIST_MODEL_CELL,
                    "model_formula": _DATA_VALIDATION_LIST_MODEL_FORMULA,
                    "dashboard_sheet": _DATA_VALIDATION_LIST_DASHBOARD_SHEET,
                    "dashboard_cell": _DATA_VALIDATION_LIST_DASHBOARD_CELL,
                    "dashboard_formula": _DATA_VALIDATION_LIST_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {
                        "sheet": _DATA_VALIDATION_LIST_SHEET,
                        "cell": _DATA_VALIDATION_LIST_TARGET_RANGE,
                    },
                    "targets": [
                        {
                            "sheet": _DATA_VALIDATION_LIST_MODEL_SHEET,
                            "cell": _DATA_VALIDATION_LIST_MODEL_CELL,
                        },
                        {
                            "sheet": _DATA_VALIDATION_LIST_DASHBOARD_SHEET,
                            "cell": _DATA_VALIDATION_LIST_DASHBOARD_CELL,
                        },
                    ],
                }
            ],
            coverage=[
                "The pair changes only the raw formula1 source declaration in the Inputs worksheet's one list data-validation record. It does not edit the target range, rule metadata, source-list values, current input, ordinary formulas, or calculation properties.",
                "The observed contract is a stored validation source. WCAB does not evaluate the list formula, determine whether any future entry is valid, accept or reject an entry, calculate the workbook, or claim Excel-client behavior.",
                "The direct static Inputs-to-model-to-dashboard path is a lower bound only if a user later enters a value. It is not proof that a validation rule permits that value or of a resulting calculation.",
            ],
        ),
    )


def _build_operations_conditional_formatting(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        worksheet = workbook["Operations"]
        for rule_range in list(worksheet.conditional_formatting):
            del worksheet.conditional_formatting[str(rule_range.sqref)]

    truth = _truth(
        case_id="operations.conditional_formatting_removed",
        title="A material exception highlight is removed",
        family="operations",
        review_expectation="review",
        facts=[
            {
                "kind": "conditional_formatting_count_changed",
                "sheet": "Operations",
                "baseline_count": 1,
                "candidate_count": 0,
            }
        ],
        coverage=[
            "The benchmark observes the control removal but does not judge the intended visual design."
        ],
    )
    _write_pair(
        root / "operations" / "conditional_formatting_removed", _operations_workbook, mutate, truth
    )


def _build_operations_conditional_formatting_threshold(root: Path) -> None:
    """Build a stored exception-threshold transition without changing cells."""

    directory = root / "operations" / "conditional_formatting_threshold_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_conditional_formatting_threshold_workbook(), baseline)
    _save_workbook(_conditional_formatting_threshold_workbook(), candidate)
    _set_conditional_formatting_threshold(
        candidate, formula=_CONDITIONAL_FORMATTING_THRESHOLD_CANDIDATE_FORMULA
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="operations.conditional_formatting_threshold_changed",
            title="A conditional-formatting exception threshold is lowered",
            family="operations",
            review_expectation="review",
            facts=[
                {
                    "kind": "conditional_formatting_threshold_changed",
                    "sheet": _CONDITIONAL_FORMATTING_THRESHOLD_SHEET,
                    "target_range": _CONDITIONAL_FORMATTING_THRESHOLD_RANGE,
                    "priority": 1,
                    "rule_type": "cellIs",
                    "operator": "greaterThan",
                    "baseline_formula": _CONDITIONAL_FORMATTING_THRESHOLD_BASELINE_FORMULA,
                    "candidate_formula": _CONDITIONAL_FORMATTING_THRESHOLD_CANDIDATE_FORMULA,
                    "metric_values": list(_CONDITIONAL_FORMATTING_THRESHOLD_VALUES),
                    "fill_rgb": _CONDITIONAL_FORMATTING_THRESHOLD_FILL_RGB,
                }
            ],
            coverage=[
                "The pair changes only the raw formula child in one cellIs conditional-formatting rule. It does not edit the target range, priority, operator, differential fill, worksheet values, calculation properties, or any other package member.",
                "The observed contract is a stored visual exception threshold. WCAB does not evaluate the rule, determine which cells a client formats, calculate the workbook, or claim Excel-client behavior.",
            ],
        ),
    )


def _build_operations_number_format_visibility(root: Path) -> None:
    """Build a custom-format transition without a cell or formula edit."""

    directory = root / "operations" / "number_format_value_hidden"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_number_format_visibility_workbook(), baseline)
    _save_workbook(_number_format_visibility_workbook(), candidate)
    _set_number_format_visibility_format(
        candidate, format_code=_NUMBER_FORMAT_VISIBILITY_CANDIDATE_FORMAT
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="operations.number_format_value_hidden",
            title="A reported margin receives a hide-value number format",
            family="operations",
            review_expectation="review",
            facts=[
                {
                    "kind": "cell_number_format_changed",
                    "sheet": _NUMBER_FORMAT_VISIBILITY_SHEET,
                    "cell": _NUMBER_FORMAT_VISIBILITY_CELL,
                    "value": _NUMBER_FORMAT_VISIBILITY_VALUE,
                    "custom_number_format_id": _NUMBER_FORMAT_VISIBILITY_CUSTOM_ID,
                    "baseline_format": _NUMBER_FORMAT_VISIBILITY_BASELINE_FORMAT,
                    "candidate_format": _NUMBER_FORMAT_VISIBILITY_CANDIDATE_FORMAT,
                    "formula_cell": _NUMBER_FORMAT_VISIBILITY_FORMULA_CELL,
                    "formula": _NUMBER_FORMAT_VISIBILITY_FORMULA,
                }
            ],
            coverage=[
                "The pair changes only one custom numFmt formatCode declaration in styles.xml. It does not edit the target cell's style index or stored numeric text, the neighboring formula, calculation properties, or any other package member.",
                "The observed contract is stored display metadata. WCAB does not render a number format, resolve locale or column-width behavior, decide what a client displays, calculate the workbook, or claim Excel-client behavior.",
            ],
        ),
    )


def _build_operations_ignored_error_suppression(root: Path) -> None:
    """Build a stored error-warning suppression without an ordinary cell edit."""

    directory = root / "operations" / "ignored_error_formula_range_suppressed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_ignored_error_suppression_workbook(), baseline)
    _save_workbook(_ignored_error_suppression_workbook(), candidate)
    _add_ignored_error_formula_range_suppression(candidate)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="operations.ignored_error_formula_range_suppressed",
            title="An omitted-range error-checking warning is marked ignored",
            family="operations",
            review_expectation="review",
            facts=[
                {
                    "kind": "ignored_error_rule_added",
                    "sheet": _IGNORED_ERROR_SUPPRESSION_SHEET,
                    "target_range": _IGNORED_ERROR_SUPPRESSION_TARGET_RANGE,
                    "warning_flag": _IGNORED_ERROR_SUPPRESSION_FLAG,
                    "formula": _IGNORED_ERROR_SUPPRESSION_FORMULA,
                    "adjacent_populated_cell": _IGNORED_ERROR_SUPPRESSION_ADJACENT_CELL,
                    "adjacent_populated_value": _IGNORED_ERROR_SUPPRESSION_ADJACENT_VALUE,
                    "downstream_formula_cell": _IGNORED_ERROR_SUPPRESSION_DOWNSTREAM_CELL,
                    "downstream_formula": _IGNORED_ERROR_SUPPRESSION_DOWNSTREAM_FORMULA,
                }
            ],
            coverage=[
                "The pair adds one standard ignoredErrors/ignoredError record with formulaRange=1 for Operations!B5. It does not edit the worksheet's ordinary cells or formulas, calculation properties, or any other package member.",
                "The observed contract is a stored request to suppress one class of Excel error checking. WCAB does not determine whether a client would otherwise show a warning, evaluate the formula, decide whether a formula omission is justified, render an indicator, change application-level error-checking options, calculate the workbook, or claim Excel-client behavior.",
            ],
        ),
    )


def _build_operations_auto_filter_criteria(root: Path) -> None:
    """Build an active-filter transition with stable cells and formulas."""

    directory = root / "operations" / "auto_filter_criteria_changed"
    directory.mkdir(parents=True, exist_ok=True)
    _save_workbook(
        _auto_filter_criteria_workbook(_AUTO_FILTER_BASELINE_VALUE), directory / "baseline.xlsx"
    )
    _save_workbook(
        _auto_filter_criteria_workbook(_AUTO_FILTER_CANDIDATE_VALUE), directory / "candidate.xlsx"
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="operations.auto_filter_criteria_changed",
            title="An active AutoFilter criterion changes a subtotal control",
            family="operations",
            review_expectation="block",
            facts=[
                {
                    "kind": "auto_filter_criteria_changed",
                    "sheet": "Report",
                    "filter_ref": _AUTO_FILTER_REF,
                    "filter_column_id": _AUTO_FILTER_COLUMN_ID,
                    "baseline_filter_value": _AUTO_FILTER_BASELINE_VALUE,
                    "candidate_filter_value": _AUTO_FILTER_CANDIDATE_VALUE,
                    "subtotal_cell": "D2",
                    "subtotal_formula": _AUTO_FILTER_SUBTOTAL_FORMULA,
                    "dashboard_sheet": "Dashboard",
                    "dashboard_cell": "B4",
                    "dashboard_formula": _AUTO_FILTER_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": "Report", "cell": "D2"},
                    "targets": [{"sheet": "Dashboard", "cell": "B4"}],
                }
            ],
            coverage=[
                "The pair changes only one raw worksheet AutoFilter criterion. It does not edit a cell value, formula text, style, calculation property, row state, or dependency edge.",
                "Excel filters control which rows are shown, and SUBTOTAL excludes rows omitted by a filter. WCAB does not apply the filter, calculate a subtotal, infer a visible row set, or claim what an Excel client will display or print.",
                "The validator reads raw OOXML filter declarations, formula text, and package-member differences. It treats the criterion transition as stored-control review evidence, not proof of a calculated outcome.",
            ],
        ),
    )


def _build_operations_named_sheet_view_filter_criterion(root: Path) -> None:
    """Build a saved alternate filter change without an active-filter edit."""

    directory = root / "operations" / "named_sheet_view_filter_criterion_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_named_sheet_view_filter_workbook(), baseline)
    _save_workbook(_named_sheet_view_filter_workbook(), candidate)
    _inject_named_sheet_view_filter(baseline, filter_value=_NAMED_SHEET_VIEW_BASELINE_VALUE)
    _inject_named_sheet_view_filter(candidate, filter_value=_NAMED_SHEET_VIEW_CANDIDATE_VALUE)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="operations.named_sheet_view_filter_criterion_changed",
            title="A saved Sheet View changes its alternate regional filter",
            family="operations",
            review_expectation="review",
            facts=[
                {
                    "kind": "named_sheet_view_filter_criterion_changed",
                    "sheet": _NAMED_SHEET_VIEW_REPORT_SHEET,
                    "view_member": _NAMED_SHEET_VIEW_MEMBER,
                    "base_filter_ref": _NAMED_SHEET_VIEW_FILTER_REF,
                    "filter_column_id": _NAMED_SHEET_VIEW_FILTER_COLUMN_ID,
                    "baseline_filter_value": _NAMED_SHEET_VIEW_BASELINE_VALUE,
                    "candidate_filter_value": _NAMED_SHEET_VIEW_CANDIDATE_VALUE,
                    "subtotal_cell": _NAMED_SHEET_VIEW_SUBTOTAL_CELL,
                    "subtotal_formula": _NAMED_SHEET_VIEW_SUBTOTAL_FORMULA,
                    "dashboard_sheet": _NAMED_SHEET_VIEW_DASHBOARD_SHEET,
                    "dashboard_cell": _NAMED_SHEET_VIEW_DASHBOARD_CELL,
                    "dashboard_formula": _NAMED_SHEET_VIEW_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {
                        "sheet": _NAMED_SHEET_VIEW_REPORT_SHEET,
                        "cell": _NAMED_SHEET_VIEW_SUBTOTAL_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _NAMED_SHEET_VIEW_DASHBOARD_SHEET,
                            "cell": _NAMED_SHEET_VIEW_DASHBOARD_CELL,
                        }
                    ],
                }
            ],
            coverage=[
                "The pair changes only one raw Named Sheet View list criterion. It does not edit an ordinary cell, formula text, style, active AutoFilter criterion, row state, or dependency edge.",
                "A Named Sheet View is an alternate stored sort/filter declaration. WCAB does not activate or render a view, apply a filter, calculate a subtotal, infer a visible row set, or claim what an Excel client will display or print.",
                "The validator follows the Report worksheet relationship to the generated Named Sheet View part, reconciles its filter ID to the stable worksheet AutoFilter, and treats the criterion transition as stored-control review evidence rather than proof of a calculated outcome.",
            ],
        ),
    )


def _build_operations_xml_map_table_xpath(root: Path) -> None:
    """Build a mapped-table retargeting pair with stable local model context."""

    directory = root / "operations" / "xml_map_table_xpath_retargeted"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_xml_map_table_workbook(), baseline)
    _save_workbook(_xml_map_table_workbook(), candidate)
    _inject_xml_map_table_binding(baseline, xpath=_XML_MAP_BASELINE_XPATH)
    _inject_xml_map_table_binding(candidate, xpath=_XML_MAP_CANDIDATE_XPATH)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="operations.xml_map_table_xpath_retargeted",
            title="An XML-mapped invoice table field is retargeted for export",
            family="operations",
            review_expectation="review",
            facts=[
                {
                    "kind": "xml_map_table_column_xpath_retargeted",
                    "sheet": _XML_MAP_EXPORT_SHEET,
                    "table_member": _XML_MAP_TABLE_MEMBER,
                    "table_name": _XML_MAP_TABLE_NAME,
                    "table_ref": _XML_MAP_TABLE_REF,
                    "mapped_column_id": _XML_MAP_TABLE_COLUMN_ID,
                    "mapped_column_name": _XML_MAP_TABLE_COLUMN_NAME,
                    "map_member": _XML_MAP_MEMBER,
                    "map_id": _XML_MAP_ID,
                    "schema_id": _XML_MAP_SCHEMA_ID,
                    "connection_id": _XML_MAP_CONNECTION_ID,
                    "baseline_xpath": _XML_MAP_BASELINE_XPATH,
                    "candidate_xpath": _XML_MAP_CANDIDATE_XPATH,
                    "single_cell_member": _XML_MAP_SINGLE_CELL_MEMBER,
                    "single_cell": _XML_MAP_SINGLE_CELL,
                    "single_cell_xpath": _XML_MAP_SINGLE_CELL_XPATH,
                    "total_cell": _XML_MAP_TOTAL_CELL,
                    "total_formula": _XML_MAP_TOTAL_FORMULA,
                    "dashboard_sheet": _XML_MAP_DASHBOARD_SHEET,
                    "dashboard_cell": _XML_MAP_DASHBOARD_CELL,
                    "dashboard_formula": _XML_MAP_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": _XML_MAP_EXPORT_SHEET, "cell": _XML_MAP_TOTAL_CELL},
                    "targets": [
                        {
                            "sheet": _XML_MAP_DASHBOARD_SHEET,
                            "cell": _XML_MAP_DASHBOARD_CELL,
                        }
                    ],
                }
            ],
            coverage=[
                "The pair changes only one raw xmlColumnPr XPath in xl/tables/table1.xml. It does not edit table cells, ordinary formulas, map/schema declarations, the stable file binding, the sheet-level single-cell mapping, calculation properties, or any other package member.",
                "The observed contract is a stored XML Map table-field binding used by Excel import/export workflows. WCAB does not access a file, validate a schema, import or export XML, materialize mapped data, calculate formulas, or claim an Excel-client result.",
                "The direct Export!D2-to-Dashboard!B4 dependency is a stable lower-bound context for the stored table data. It is not evidence that an import or export uses either XML path or that future mapped values would produce a particular total.",
            ],
        ),
    )


def _build_governance_office_web_addin_auto_show(root: Path) -> None:
    """Build an embedded add-in auto-show request without a cell edit."""

    directory = root / "governance" / "office_web_addin_auto_show_enabled"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_office_web_addin_auto_show_workbook(), baseline)
    _save_workbook(_office_web_addin_auto_show_workbook(), candidate)
    _inject_office_web_addin_auto_show(baseline, auto_show=_OFFICE_WEB_ADDIN_BASELINE_AUTO_SHOW)
    _inject_office_web_addin_auto_show(candidate, auto_show=_OFFICE_WEB_ADDIN_CANDIDATE_AUTO_SHOW)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.office_web_addin_auto_show_enabled",
            title="An embedded Office Web Add-in requests task-pane auto-show",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "office_web_addin_auto_show_enabled",
                    "taskpane_member": _OFFICE_WEB_ADDIN_TASKPANES_MEMBER,
                    "web_extension_member": _OFFICE_WEB_ADDIN_EXTENSION_MEMBER,
                    "addin_id": _OFFICE_WEB_ADDIN_ID,
                    "reference_id": _OFFICE_WEB_ADDIN_REFERENCE_ID,
                    "reference_version": _OFFICE_WEB_ADDIN_REFERENCE_VERSION,
                    "store": _OFFICE_WEB_ADDIN_STORE,
                    "store_type": _OFFICE_WEB_ADDIN_STORE_TYPE,
                    "baseline_auto_show": _OFFICE_WEB_ADDIN_BASELINE_AUTO_SHOW,
                    "candidate_auto_show": _OFFICE_WEB_ADDIN_CANDIDATE_AUTO_SHOW,
                    "input_sheet": _OFFICE_WEB_ADDIN_INPUT_SHEET,
                    "input_cell": _OFFICE_WEB_ADDIN_INPUT_CELL,
                    "input_value": _OFFICE_WEB_ADDIN_INPUT_VALUE,
                    "model_sheet": _OFFICE_WEB_ADDIN_MODEL_SHEET,
                    "model_cell": _OFFICE_WEB_ADDIN_MODEL_CELL,
                    "model_formula": _OFFICE_WEB_ADDIN_MODEL_FORMULA,
                    "dashboard_sheet": _OFFICE_WEB_ADDIN_DASHBOARD_SHEET,
                    "dashboard_cell": _OFFICE_WEB_ADDIN_DASHBOARD_CELL,
                    "dashboard_formula": _OFFICE_WEB_ADDIN_DASHBOARD_FORMULA,
                }
            ],
            coverage=[
                "The pair changes only one Office.AutoShowTaskpaneWithDocument property value in xl/webextensions/webextension1.xml. It does not edit ordinary cells, formulas, calculation properties, task-pane layout, add-in identity, or any other package member.",
                "The synthetic add-in reference uses a local FileSystem store name and has no manifest payload or external relationship. WCAB does not install, load, execute, or fetch an add-in or manifest, and it does not claim that any task pane opens.",
                "The stable Inputs!B2-to-Model!B2-to-Dashboard!B4 formula path is workbook context only. It is not evidence that an add-in reads, writes, calculates, or displays any of those cells.",
            ],
        ),
    )


def _build_governance_ole_object_auto_load(root: Path) -> None:
    """Build one opaque embedded-object auto-load request without a cell edit."""

    directory = root / "governance" / "ole_object_auto_load_enabled"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_ole_object_auto_load_workbook(), baseline)
    _save_workbook(_ole_object_auto_load_workbook(), candidate)
    _inject_ole_object_auto_load(baseline, auto_load=_OLE_OBJECT_BASELINE_AUTO_LOAD)
    _inject_ole_object_auto_load(candidate, auto_load=_OLE_OBJECT_CANDIDATE_AUTO_LOAD)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.ole_object_auto_load_enabled",
            title="An embedded OLE object requests automatic loading on workbook open",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "ole_object_auto_load_enabled",
                    "sheet": _OLE_OBJECT_INPUT_SHEET,
                    "worksheet_member": _OLE_OBJECT_WORKSHEET_MEMBER,
                    "worksheet_relationships_member": _OLE_OBJECT_RELATIONSHIPS_MEMBER,
                    "relationship_id": _OLE_OBJECT_RELATIONSHIP_ID,
                    "relationship_type": _OLE_OBJECT_RELATIONSHIP,
                    "target": "../embeddings/wcab-review-embedded-object.bin",
                    "embedded_object_member": _OLE_OBJECT_MEMBER,
                    "content_type": _OLE_OBJECT_CONTENT_TYPE,
                    "prog_id": _OLE_OBJECT_PROG_ID,
                    "dv_aspect": _OLE_OBJECT_DV_ASPECT,
                    "shape_id": _OLE_OBJECT_SHAPE_ID,
                    "baseline_auto_load": _OLE_OBJECT_BASELINE_AUTO_LOAD,
                    "candidate_auto_load": _OLE_OBJECT_CANDIDATE_AUTO_LOAD,
                    "input_sheet": _OLE_OBJECT_INPUT_SHEET,
                    "input_cell": _OLE_OBJECT_INPUT_CELL,
                    "input_value": _OLE_OBJECT_INPUT_VALUE,
                    "model_sheet": _OLE_OBJECT_MODEL_SHEET,
                    "model_cell": _OLE_OBJECT_MODEL_CELL,
                    "model_formula": _OLE_OBJECT_MODEL_FORMULA,
                    "dashboard_sheet": _OLE_OBJECT_DASHBOARD_SHEET,
                    "dashboard_cell": _OLE_OBJECT_DASHBOARD_CELL,
                    "dashboard_formula": _OLE_OBJECT_DASHBOARD_FORMULA,
                }
            ],
            coverage=[
                "The pair changes only raw xl/worksheets/sheet1.xml oleObject/@autoLoad from false to true. It does not edit ordinary cells, formulas, calculation properties, content types, relationships, or the embedded-object bytes.",
                "The object has one fixed internal worksheet relationship and an opaque synthetic ASCII payload under a synthetic unregistered ProgID. There is no linked target, ActiveX control, object presentation, macro, or external relationship.",
                "WCAB does not deserialize, open, render, execute, register, or invoke an object server, and it does not claim that an object loads successfully. The stable Inputs!B2-to-Model!B2-to-Dashboard!B4 formula path is workbook context only.",
            ],
        ),
    )


def _build_governance_visibility(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook["ReviewControls"].sheet_state = "visible"

    truth = _truth(
        case_id="governance.hidden_sheet_revealed",
        title="A hidden review-control sheet is made visible",
        family="governance",
        review_expectation="review",
        facts=[
            {
                "kind": "sheet_visibility_changed",
                "sheet": "ReviewControls",
                "baseline_state": "hidden",
                "candidate_state": "visible",
            }
        ],
        coverage=[
            "Visibility is an observable workbook control; cell content is not part of this assertion."
        ],
    )
    _write_pair(root / "governance" / "hidden_sheet_revealed", _governance_workbook, mutate, truth)


def _build_governance_protection(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook["Controls"]["D10"].protection = Protection(locked=False)

    truth = _truth(
        case_id="governance.formula_cell_unlocked",
        title="A protected calculated cell is explicitly unlocked",
        family="governance",
        review_expectation="block",
        facts=[{"kind": "formula_cell_unlocked", "sheet": "Controls", "cell": "D10"}],
        coverage=[
            "This is a direct cell-protection assertion; style inheritance and allowed edit ranges are out of scope."
        ],
    )
    _write_pair(root / "governance" / "formula_cell_unlocked", _governance_workbook, mutate, truth)


def _build_governance_sheet_protection_sort_permission(root: Path) -> None:
    """Build a pair that enables sorting while retaining sheet protection."""

    def mutate(workbook: Workbook) -> None:
        workbook[_SHEET_PROTECTION_SORT_SHEET].protection.sort = False

    truth = _truth(
        case_id="governance.sheet_protection_sort_permission_enabled",
        title="A protected sheet permits sorting without a cell edit",
        family="governance",
        review_expectation="block",
        facts=[
            {
                "kind": "sheet_protection_sort_permission_enabled",
                "sheet": _SHEET_PROTECTION_SORT_SHEET,
                "worksheet_member": "xl/worksheets/sheet1.xml",
                "baseline_sort_locked": True,
                "candidate_sort_locked": False,
                "formula_cell": _SHEET_PROTECTION_SORT_FORMULA_CELL,
                "formula": _SHEET_PROTECTION_SORT_FORMULA,
                "dashboard_sheet": _SHEET_PROTECTION_SORT_DASHBOARD_SHEET,
                "dashboard_cell": _SHEET_PROTECTION_SORT_DASHBOARD_CELL,
                "dashboard_formula": _SHEET_PROTECTION_SORT_DASHBOARD_FORMULA,
            }
        ],
        must_reach=[
            {
                "source": {
                    "sheet": _SHEET_PROTECTION_SORT_SHEET,
                    "cell": _SHEET_PROTECTION_SORT_FORMULA_CELL,
                },
                "targets": [
                    {
                        "sheet": _SHEET_PROTECTION_SORT_DASHBOARD_SHEET,
                        "cell": _SHEET_PROTECTION_SORT_DASHBOARD_CELL,
                    }
                ],
            }
        ],
        coverage=[
            'The pair changes only xl/worksheets/sheet1.xml sheetProtection/@sort from "1" (locked) to "0" while sheet protection remains enabled. It does not edit cells, formulas, calculation properties, styles, or workbook-level protection.',
            "The fact is the stored sort-permission transition on one protected worksheet. WCAB does not test a password, encryption, authorization, editable ranges, whether a client permits a particular sort, or what a sort would change.",
            "Controls!D2 and its direct Dashboard!B4 consumer remain fixed. That local formula path is review context only, not an assertion that sorting changes either stored value or calculation result.",
        ],
    )
    _write_pair(
        root / "governance" / "sheet_protection_sort_permission_enabled",
        _sheet_protection_sort_permission_workbook,
        mutate,
        truth,
    )


def _build_governance_protected_range_security_descriptor(root: Path) -> None:
    """Build one standards-form protected-range descriptor transition."""

    directory = root / "governance" / "protected_range_security_descriptor_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_protected_range_security_descriptor_workbook(), baseline)
    _save_workbook(_protected_range_security_descriptor_workbook(), candidate)
    _add_protected_range_security_descriptor(
        baseline,
        descriptor=_PROTECTED_RANGE_SECURITY_DESCRIPTOR_BASELINE,
    )
    _add_protected_range_security_descriptor(
        candidate,
        descriptor=_PROTECTED_RANGE_SECURITY_DESCRIPTOR_CANDIDATE,
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.protected_range_security_descriptor_changed",
            title="A protected range changes its stored security descriptor",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "protected_range_security_descriptor_changed",
                    "sheet": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_SHEET,
                    "worksheet_member": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_WORKSHEET_MEMBER,
                    "protected_range_ref": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_RANGE_REF,
                    "protected_range_count": 1,
                    "security_descriptor_count": 1,
                    "has_legacy_verifier": True,
                    "input_cell": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_INPUT_CELL,
                    "input_value": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_INPUT_VALUE,
                    "formula_cell": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_FORMULA_CELL,
                    "formula": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_FORMULA,
                    "dashboard_sheet": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_SHEET,
                    "dashboard_cell": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_CELL,
                    "dashboard_formula": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {
                        "sheet": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_SHEET,
                        "cell": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_INPUT_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_SHEET,
                            "cell": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_FORMULA_CELL,
                        },
                        {
                            "sheet": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_SHEET,
                            "cell": _PROTECTED_RANGE_SECURITY_DESCRIPTOR_DASHBOARD_CELL,
                        },
                    ],
                }
            ],
            coverage=[
                "The pair changes only xl/worksheets/sheet1.xml: one standard protectedRange/securityDescriptor child text differs. The worksheet protection, locked target, legacy verifier, range name, range reference, cells, formulas, calculation properties, and every other package member stay fixed.",
                "The fact is a stored descriptor change on one protected range. WCAB does not test a password or verifier, encryption, identity, authentication, authorization, editable-range enforcement, a spreadsheet client, or any resulting value.",
                "The raw validator checks the exact compact generated XML privately. An adapter must instead require FormulaFence's equal safe one-range profile, security-descriptor-material signal, and FF022 evidence; it must not require a descriptor, range name, verifier, or identity value.",
            ],
        ),
    )


def _build_governance_workbook_structure_protection(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook.security.lockStructure = False

    truth = _truth(
        case_id="governance.workbook_structure_lock_removed",
        title="A workbook structure lock is disabled",
        family="governance",
        review_expectation="block",
        facts=[
            {
                "kind": "workbook_structure_lock_removed",
                "baseline_lock_structure": True,
                "candidate_lock_structure": False,
                "hidden_sheet": _WORKBOOK_STRUCTURE_PROTECTION_HIDDEN_SHEET,
                "hidden_sheet_state": "hidden",
                "formula_sheet": _WORKBOOK_STRUCTURE_PROTECTION_INPUT_SHEET,
                "formula_cell": _WORKBOOK_STRUCTURE_PROTECTION_FORMULA_CELL,
                "formula": _WORKBOOK_STRUCTURE_PROTECTION_FORMULA,
            }
        ],
        coverage=[
            "This is a stored workbook-structure control assertion, not a password, encryption, authorization, or client-action test."
        ],
    )
    _write_pair(
        root / "governance" / "workbook_structure_lock_removed",
        _workbook_structure_protection_workbook,
        mutate,
        truth,
    )


def _build_governance_manual_calculation(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook.calculation.calcMode = "manual"
        workbook.calculation.calcCompleted = False

    truth = _truth(
        case_id="governance.manual_calculation_incomplete",
        title="A formula workbook records incomplete manual calculation",
        family="governance",
        review_expectation="block",
        facts=[{"kind": "manual_calculation_incomplete"}],
        coverage=[
            "Calculation metadata is evidence of save state, not proof that any cached result is wrong."
        ],
    )
    _write_pair(
        root / "governance" / "manual_calculation_incomplete", _governance_workbook, mutate, truth
    )


def _build_governance_iterative_calculation(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook.calculation.iterate = True

    truth = _truth(
        case_id="governance.iterative_calculation_enabled",
        title="An unchanged direct circular formula enables iterative calculation",
        family="governance",
        review_expectation="block",
        facts=[
            {
                "kind": "iterative_calculation_enabled",
                "sheet": "Model",
                "cell": "B2",
                "formula": _ITERATIVE_CALCULATION_FORMULA,
                "baseline_iterate": False,
                "candidate_iterate": True,
                "iteration_count": _ITERATION_COUNT,
                "iteration_delta": _ITERATION_DELTA,
            }
        ],
        must_reach=[
            {
                "source": {"sheet": "Model", "cell": "B2"},
                "targets": [{"sheet": "Dashboard", "cell": "B4"}],
            }
        ],
        coverage=[
            "The direct self-reference is intentional and remains unchanged. The fixture records calculation controls; it does not calculate the circular formula or claim a converged, cached, or business-correct result.",
            "The validator checks the explicit iterate flag plus the shared iteration count and delta. It does not predict the number of recalculations, convergence, numerical precision, or Excel-client behavior.",
            "The pair differs only in workbook calculation metadata; no worksheet value, formula text, or dependency edge is edited.",
        ],
    )
    _write_pair(
        root / "governance" / "iterative_calculation_enabled",
        _iterative_calculation_workbook,
        mutate,
        truth,
    )


def _build_governance_precision_as_displayed(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook.calculation.fullPrecision = False

    truth = _truth(
        case_id="governance.precision_as_displayed_enabled",
        title="Precision-as-displayed calculation is enabled without a cell edit",
        family="governance",
        review_expectation="block",
        facts=[
            {
                "kind": "precision_as_displayed_enabled",
                "input_sheet": "Inputs",
                "input_cell": "B2",
                "input_value": _PRECISION_AS_DISPLAYED_INPUT,
                "number_format": _PRECISION_AS_DISPLAYED_NUMBER_FORMAT,
                "formula_sheet": "Model",
                "formula_cell": "B2",
                "formula": _PRECISION_AS_DISPLAYED_FORMULA,
                "baseline_full_precision": True,
                "candidate_full_precision": False,
            }
        ],
        must_reach=[
            {
                "source": {"sheet": "Inputs", "cell": "B2"},
                "targets": [
                    {"sheet": "Model", "cell": "B2"},
                    {"sheet": "Dashboard", "cell": "B4"},
                ],
            }
        ],
        coverage=[
            "The pair records the raw calcPr fullPrecision switch that corresponds to Excel's precision-as-displayed setting. WCAB does not emulate Excel's calculation engine or claim that a client applied the control.",
            "Both generated packages retain the exact stored 10.005 input and its two-decimal number format. The validator does not claim that opening, calculating, or saving either workbook rounds a value or produces a particular formula result.",
            "The pair differs only in workbook calculation metadata; no worksheet value, formula text, number format, or dependency edge is edited.",
        ],
    )
    _write_pair(
        root / "governance" / "precision_as_displayed_enabled",
        _precision_as_displayed_workbook,
        mutate,
        truth,
    )


def _build_governance_workbook_date_system(root: Path) -> None:
    """Build a raw date-system-control transition with stable cell content."""

    directory = root / "governance" / "workbook_date_system_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_workbook_date_system_workbook(), baseline)
    _save_workbook(_workbook_date_system_workbook(), candidate)
    _set_workbook_date_system(baseline, date_1904=False, date_compatibility=True)
    _set_workbook_date_system(candidate, date_1904=True, date_compatibility=True)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.workbook_date_system_changed",
            title="Workbook serial-date controls change without a cell edit",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "workbook_date_system_changed",
                    "baseline_date_1904": False,
                    "candidate_date_1904": True,
                    "date_compatibility": True,
                    "serial_sheet": "Inputs",
                    "serial_cell": "B2",
                    "serial_value": _WORKBOOK_DATE_SYSTEM_SERIAL,
                    "number_format": _WORKBOOK_DATE_SYSTEM_NUMBER_FORMAT,
                    "formula_sheet": "Model",
                    "formula_cell": "B2",
                    "formula": _WORKBOOK_DATE_SYSTEM_FORMULA,
                    "dashboard_sheet": "Dashboard",
                    "dashboard_cell": "B4",
                    "dashboard_formula": _WORKBOOK_DATE_SYSTEM_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": "Inputs", "cell": "B2"},
                    "targets": [
                        {"sheet": "Model", "cell": "B2"},
                        {"sheet": "Dashboard", "cell": "B4"},
                    ],
                },
                {
                    "source": {"sheet": "Model", "cell": "B2"},
                    "targets": [{"sheet": "Dashboard", "cell": "B4"}],
                },
            ],
            coverage=[
                "The pair changes only raw workbookPr date1904 from false to true while dateCompatibility remains explicitly true. It does not edit a worksheet cell, formula, style, calculation property, or dependency edge.",
                "The stored numeric serial and date number format remain unchanged. WCAB does not calculate a formula, convert the serial into a date, infer a visible value, or claim what an Excel client will do on opening or saving either workbook.",
                "The validator reads raw OOXML controls, numeric cell text, formula text, number format, and package-member differences. It treats the date-system transition as a stored-control review boundary, not proof of a displayed outcome.",
            ],
        ),
    )


def _build_governance_formula_cached_result(root: Path) -> None:
    """Build a pair with exactly one unexplained saved formula-result change."""

    directory = root / "governance" / "formula_cached_result_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_formula_cached_result_workbook(), baseline)
    _save_workbook(_formula_cached_result_workbook(), candidate)
    _set_formula_cached_result(baseline, result=_FORMULA_CACHED_RESULT_BASELINE)
    _set_formula_cached_result(candidate, result=_FORMULA_CACHED_RESULT_CANDIDATE)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.formula_cached_result_changed",
            title="A saved formula result changes without a formula or input edit",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "formula_cached_result_changed",
                    "sheet": "Model",
                    "cell": "B2",
                    "formula": _FORMULA_CACHED_RESULT_FORMULA,
                    "input_sheet": "Inputs",
                    "input_cell": "B2",
                    "input_value": _FORMULA_CACHED_RESULT_INPUT,
                    "result_type": "numeric",
                    "baseline_cached_result": _FORMULA_CACHED_RESULT_BASELINE,
                    "candidate_cached_result": _FORMULA_CACHED_RESULT_CANDIDATE,
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": "Inputs", "cell": "B2"},
                    "targets": [
                        {"sheet": "Model", "cell": "B2"},
                        {"sheet": "Dashboard", "cell": "B4"},
                    ],
                },
                {
                    "source": {"sheet": "Model", "cell": "B2"},
                    "targets": [{"sheet": "Dashboard", "cell": "B4"}],
                },
            ],
            coverage=[
                "The pair changes exactly one raw OOXML CellValue (<v>) saved beside an unchanged formula; it does not calculate either formula or change a visible input, formula, dependency edge, or calculation control.",
                "A saved formula result is evidence of the last stored calculation, not proof that either result is current, stale, tampered, mathematically correct, or what an Excel client will display after opening or recalculating.",
                "The validator compares the raw formula cache, formula text, input, and package members. It does not execute a formula, infer volatile or external inputs, or claim a downstream displayed result changed.",
            ],
        ),
    )


def _build_governance_static_cycle(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook["Controls"]["D10"] = "=D11+1"
        workbook["Controls"]["D11"] = "=D10+1"

    truth = _truth(
        case_id="governance.static_cycle_introduced",
        title="Two direct formulas form a static cycle",
        family="governance",
        review_expectation="block",
        facts=[
            {
                "kind": "static_cycle_introduced",
                "cells": [
                    {"sheet": "Controls", "cell": "D10"},
                    {"sheet": "Controls", "cell": "D11"},
                ],
            }
        ],
        coverage=[
            "The cycle is direct and static; iterative calculation settings and dynamic references are out of scope."
        ],
    )
    _write_pair(
        root / "governance" / "static_cycle_introduced", _governance_workbook, mutate, truth
    )


def _build_governance_external_data_refresh(root: Path) -> None:
    """Build a pair that differs only in a raw connection refresh control."""

    directory = root / "governance" / "external_data_refresh_on_open"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_external_data_refresh_workbook(), baseline)
    _save_workbook(_external_data_refresh_workbook(), candidate)
    _add_external_data_connection(
        baseline,
        refresh_on_load=False,
        url=_EXTERNAL_DATA_CONNECTION_REFRESH_URL,
    )
    _add_external_data_connection(
        candidate,
        refresh_on_load=True,
        url=_EXTERNAL_DATA_CONNECTION_REFRESH_URL,
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.external_data_refresh_on_open",
            title="An external-data connection starts refreshing when the workbook opens",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "external_data_connection_refresh_on_load_changed",
                    "connection_id": _EXTERNAL_DATA_CONNECTION_ID,
                    "baseline_refresh_on_load": False,
                    "candidate_refresh_on_load": True,
                }
            ],
            coverage=[
                "The connection endpoint is synthetic and non-routable; the benchmark never opens it or tests credentials, trust, or returned data.",
                "The observable contract is the stored refresh-on-open control, not a claim that downstream saved values will recalculate.",
            ],
        ),
    )


def _build_governance_external_data_connection_source(root: Path) -> None:
    """Build a pair whose only stored difference is a web-query endpoint."""

    directory = root / "governance" / "external_data_connection_source_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_external_data_refresh_workbook(), baseline)
    _save_workbook(_external_data_refresh_workbook(), candidate)
    _add_external_data_connection(
        baseline,
        refresh_on_load=False,
        url=_EXTERNAL_DATA_CONNECTION_SOURCE_BASELINE_URL,
    )
    _add_external_data_connection(
        candidate,
        refresh_on_load=False,
        url=_EXTERNAL_DATA_CONNECTION_SOURCE_CANDIDATE_URL,
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.external_data_connection_source_changed",
            title="An external-data web query points to a different stored source",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "external_data_connection_web_query_url_changed",
                    "connection_id": _EXTERNAL_DATA_CONNECTION_ID,
                    "connection_member": _EXTERNAL_DATA_CONNECTION_MEMBER,
                    "workbook_relationships_member": "xl/_rels/workbook.xml.rels",
                    "relationship_id": _EXTERNAL_DATA_CONNECTION_WORKBOOK_RELATIONSHIP_ID,
                    "relationship_type": _EXTERNAL_DATA_CONNECTIONS_RELATIONSHIP,
                    "connection_content_type": _EXTERNAL_DATA_CONNECTIONS_CONTENT_TYPE,
                    "baseline_url": _EXTERNAL_DATA_CONNECTION_SOURCE_BASELINE_URL,
                    "candidate_url": _EXTERNAL_DATA_CONNECTION_SOURCE_CANDIDATE_URL,
                    "refresh_on_load": False,
                    "saved_value_sheet": "ImportedData",
                    "saved_value_cell": "B2",
                    "saved_value": 100,
                    "summary_sheet": "Summary",
                    "summary_cell": "B2",
                    "summary_formula": "=ImportedData!$B$2",
                    "dashboard_sheet": "Dashboard",
                    "dashboard_cell": "B4",
                    "dashboard_formula": "=Summary!$B$2",
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": "ImportedData", "cell": "B2"},
                    "targets": [
                        {"sheet": "Summary", "cell": "B2"},
                        {"sheet": "Dashboard", "cell": "B4"},
                    ],
                },
                {
                    "source": {"sheet": "Summary", "cell": "B2"},
                    "targets": [{"sheet": "Dashboard", "cell": "B4"}],
                },
            ],
            coverage=[
                "The pair changes only raw xl/connections.xml webPr/@url from one reserved example.invalid endpoint to another. It does not edit ordinary cells, formulas, calculation properties, connection identity, connection refresh controls, content types, or relationships.",
                "The URLs are synthetic and non-routable. The benchmark records only their stored text and does not open a connection, fetch a URL, refresh a query, materialize rows, calculate a workbook, inspect credentials or trust, or claim that a client returns a value.",
                "The stable saved cells and direct formula path are context only; a stored source declaration does not prove source reachability, returned data, or a recalculated display.",
            ],
        ),
    )


def _build_governance_package_signature_manifest_retarget(root: Path) -> None:
    """Build a pair whose only delta is one OPC Manifest target declaration."""

    directory = root / "governance" / "package_signature_manifest_retargeted"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_governance_workbook(), baseline)
    _save_workbook(_governance_workbook(), candidate)
    _add_package_signature_manifest(
        baseline,
        manifest_uri=_PACKAGE_SIGNATURE_BASELINE_MANIFEST_URI,
    )
    _add_package_signature_manifest(
        candidate,
        manifest_uri=_PACKAGE_SIGNATURE_CANDIDATE_MANIFEST_URI,
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.package_signature_manifest_retargeted",
            title="An OPC package signature declares a different package part",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "package_signature_manifest_direct_part_retargeted",
                    "root_relationships_member": "_rels/.rels",
                    "origin_member": _PACKAGE_SIGNATURE_ORIGIN_MEMBER,
                    "origin_relationships_member": _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIPS_MEMBER,
                    "origin_relationship_id": _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP_ID,
                    "origin_relationship_type": _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP,
                    "signature_member": _PACKAGE_SIGNATURE_MEMBER,
                    "signature_relationship_id": _PACKAGE_SIGNATURE_XML_RELATIONSHIP_ID,
                    "signature_relationship_type": _PACKAGE_SIGNATURE_SIGNATURE_RELATIONSHIP,
                    "signature_content_type": _PACKAGE_SIGNATURE_XML_CONTENT_TYPE,
                    "signed_info_reference_uri": f"#{_PACKAGE_SIGNATURE_OBJECT_ID}",
                    "baseline_manifest_uri": _PACKAGE_SIGNATURE_BASELINE_MANIFEST_URI,
                    "candidate_manifest_uri": _PACKAGE_SIGNATURE_CANDIDATE_MANIFEST_URI,
                    "baseline_direct_part_class": "workbook",
                    "candidate_direct_part_class": "worksheet",
                    "stable_sheet": "Controls",
                    "stable_value_cell": "B10",
                    "stable_value": 12,
                    "stable_formula_cell": "D10",
                    "stable_formula": "=B10*C10",
                }
            ],
            coverage=[
                "The pair changes only _xmlsignatures/sig1.xml: its OPC Object/Manifest direct part URI moves from xl/workbook.xml to xl/worksheets/sheet1.xml. The origin graph, XML signature shape, content type, ordinary cells, formulas, and every other package member stay fixed.",
                "The XMLDSIG digest and signature values are deliberately synthetic. The benchmark records only declared package scope; it does not verify cryptography, XML transform processing, digest values, certificate identity/trust, or a package consumer's trust decision.",
                "The raw manifest URIs are public WCAB truth used by its local validator. An adapter must instead require FormulaFence's redacted aggregate before/after coverage profile and FF050 evidence, never a URI or selector.",
            ],
        ),
    )


def _build_governance_package_signature_relationship_selector_retarget(
    root: Path,
) -> None:
    """Build a same-count OPC relationship-selector scope retarget."""

    directory = root / "governance" / "package_signature_relationship_selector_retargeted"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_governance_workbook(), baseline)
    _save_workbook(_governance_workbook(), candidate)
    _add_package_signature_manifest(
        baseline,
        manifest_uri=_PACKAGE_SIGNATURE_ROOT_RELATIONSHIPS_MANIFEST_URI,
        relationship_selector_source_id=_PACKAGE_SIGNATURE_BASELINE_SELECTOR_SOURCE_ID,
    )
    _add_package_signature_manifest(
        candidate,
        manifest_uri=_PACKAGE_SIGNATURE_ROOT_RELATIONSHIPS_MANIFEST_URI,
        relationship_selector_source_id=_PACKAGE_SIGNATURE_CANDIDATE_SELECTOR_SOURCE_ID,
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.package_signature_relationship_selector_retargeted",
            title="An OPC package signature selects a different root relationship",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "package_signature_manifest_relationship_selector_retargeted",
                    "root_relationships_member": "_rels/.rels",
                    "origin_member": _PACKAGE_SIGNATURE_ORIGIN_MEMBER,
                    "origin_relationships_member": _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIPS_MEMBER,
                    "origin_relationship_id": _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP_ID,
                    "origin_relationship_type": _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP,
                    "signature_member": _PACKAGE_SIGNATURE_MEMBER,
                    "signature_relationship_id": _PACKAGE_SIGNATURE_XML_RELATIONSHIP_ID,
                    "signature_relationship_type": _PACKAGE_SIGNATURE_SIGNATURE_RELATIONSHIP,
                    "signature_content_type": _PACKAGE_SIGNATURE_XML_CONTENT_TYPE,
                    "signed_info_reference_uri": f"#{_PACKAGE_SIGNATURE_OBJECT_ID}",
                    "manifest_uri": _PACKAGE_SIGNATURE_ROOT_RELATIONSHIPS_MANIFEST_URI,
                    "relationship_transform_algorithm": _PACKAGE_SIGNATURE_RELATIONSHIP_TRANSFORM,
                    "canonicalization_algorithm": _PACKAGE_SIGNATURE_C14N_ALGORITHM,
                    "office_document_relationship_id": _PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP_ID,
                    "office_document_relationship_type": _PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP,
                    "office_document_relationship_target": "xl/workbook.xml",
                    "baseline_selector_source_id": _PACKAGE_SIGNATURE_BASELINE_SELECTOR_SOURCE_ID,
                    "candidate_selector_source_id": _PACKAGE_SIGNATURE_CANDIDATE_SELECTOR_SOURCE_ID,
                    "baseline_selected_relationship_type": _PACKAGE_SIGNATURE_OFFICE_DOCUMENT_RELATIONSHIP,
                    "baseline_selected_relationship_target": "xl/workbook.xml",
                    "candidate_selected_relationship_type": _PACKAGE_SIGNATURE_ORIGIN_RELATIONSHIP,
                    "candidate_selected_relationship_target": _PACKAGE_SIGNATURE_ORIGIN_MEMBER,
                    "stable_sheet": "Controls",
                    "stable_value_cell": "B10",
                    "stable_value": 12,
                    "stable_formula_cell": "D10",
                    "stable_formula": "=B10*C10",
                }
            ],
            coverage=[
                "The pair changes only _xmlsignatures/sig1.xml: one Object/Manifest Relationships Transform selector moves from the root officeDocument relationship to the root digital-signature-origin relationship. The manifest URI, required following C14N transform, root/origin graph, XML signature shape, ordinary cells, formulas, and every other package member stay fixed.",
                "A relationship selector identifies a Relationship entry, not proof that its target part was signed. The XMLDSIG digest and signature values are deliberately synthetic; the validator does not execute a transform, verify cryptography, assess digest values, certificates, identity/trust, or a package consumer decision.",
                "The raw selector IDs are public WCAB truth used only by its local validator. An adapter must instead require FormulaFence's redacted equal-count aggregate plus a Manifest-coverage-change signal and FF050 evidence; it must never require a URI, selector, digest, certificate, or trust assertion.",
            ],
        ),
    )


def _build_governance_threaded_comment_resolution_state(root: Path) -> None:
    """Build one modern-comment thread whose stored resolution state changes."""

    directory = root / "governance" / "threaded_comment_resolution_state_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_governance_workbook(), baseline)
    _save_workbook(_governance_workbook(), candidate)
    _add_threaded_comment_resolution_state(baseline, resolved=False)
    _add_threaded_comment_resolution_state(candidate, resolved=True)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.threaded_comment_resolution_state_changed",
            title="A threaded comment thread is marked resolved",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "threaded_comment_resolution_state_changed",
                    "threaded_comment_sheet": _THREADED_COMMENT_SHEET,
                    "worksheet_member": _THREADED_COMMENT_WORKSHEET_MEMBER,
                    "worksheet_relationships_member": (
                        _THREADED_COMMENT_WORKSHEET_RELATIONSHIPS_MEMBER
                    ),
                    "workbook_relationships_member": "xl/_rels/workbook.xml.rels",
                    "threaded_comment_member": _THREADED_COMMENT_MEMBER,
                    "threaded_comment_content_type": _THREADED_COMMENT_CONTENT_TYPE,
                    "threaded_comment_relationship_type": _THREADED_COMMENT_RELATIONSHIP,
                    "person_member": _THREADED_COMMENT_PERSON_MEMBER,
                    "person_content_type": _THREADED_COMMENT_PERSON_CONTENT_TYPE,
                    "person_relationship_type": _THREADED_COMMENT_PERSON_RELATIONSHIP,
                    "worksheet_threaded_comment_sheet_count": 1,
                    "threaded_comment_part_count": 1,
                    "comment_thread_count": 1,
                    "comment_count": 1,
                    "reply_count": 0,
                    "baseline_resolved_comment_count": 0,
                    "candidate_resolved_comment_count": 1,
                    "comment_with_text_count": 1,
                    "mention_count": 0,
                    "mentioned_person_count": 0,
                    "person_part_count": 1,
                    "person_count": 1,
                    "orphan_person_count": 0,
                    "binding_relationship_count": 2,
                    "external_relationship_count": 0,
                    "unrecognized_threaded_comment_count": 0,
                    "stable_sheet": "Controls",
                    "stable_value_cell": "B10",
                    "stable_value": 12,
                    "stable_formula_cell": "D10",
                    "stable_formula": "=B10*C10",
                }
            ],
            coverage=[
                "The pair changes only xl/threadedComments/threadedComment1.xml: one top-level threadedComment/@done token moves from 0 to 1. Its synthetic text, timestamps, IDs, cell binding, person record, worksheet/workbook relationships, content types, ordinary cells, formulas, calculation properties, and every other package member stay fixed.",
                "The pair records one stored thread-resolution state only. It does not reveal or assess discussion content, prove that a person reviewed or approved anything, resolve a notification, authenticate an author, enforce authorization, open a client, or claim a completed workflow.",
                "The raw XML validates the exact synthetic package shape locally. An adapter must instead require FormulaFence's redacted one-thread profile, the resolved-comment count transition, threaded-comment definition-material signal, and FF045 evidence; it must never require text, a comment-cell reference, timestamp, relationship ID, comment ID, person ID, or identity data.",
            ],
        ),
    )


def _build_governance_shared_workbook_revision_log(root: Path) -> None:
    """Build one private revision-log mutation outside ordinary worksheet cells."""

    directory = root / "governance" / "shared_workbook_revision_log_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_governance_workbook(), baseline)
    _save_workbook(_governance_workbook(), candidate)
    _add_shared_workbook_revision_log(
        baseline,
        historic_value=_SHARED_WORKBOOK_REVISION_BASELINE_HISTORIC_VALUE,
    )
    _add_shared_workbook_revision_log(
        candidate,
        historic_value=_SHARED_WORKBOOK_REVISION_CANDIDATE_HISTORIC_VALUE,
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.shared_workbook_revision_log_changed",
            title="A legacy shared-workbook revision log changes",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "shared_workbook_revision_log_changed",
                    "workbook_relationships_member": "xl/_rels/workbook.xml.rels",
                    "revision_headers_member": _SHARED_WORKBOOK_REVISION_HEADERS_MEMBER,
                    "revision_headers_relationships_member": (
                        _SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIPS_MEMBER
                    ),
                    "revision_log_member": _SHARED_WORKBOOK_REVISION_LOG_MEMBER,
                    "revision_headers_content_type": (
                        _SHARED_WORKBOOK_REVISION_HEADERS_CONTENT_TYPE
                    ),
                    "revision_log_content_type": _SHARED_WORKBOOK_REVISION_LOG_CONTENT_TYPE,
                    "revision_headers_relationship_type": (
                        _SHARED_WORKBOOK_REVISION_HEADERS_RELATIONSHIP
                    ),
                    "revision_log_relationship_type": (_SHARED_WORKBOOK_REVISION_LOG_RELATIONSHIP),
                    "revision_header_part_count": 1,
                    "revision_header_count": 1,
                    "revision_log_part_count": 1,
                    "revision_log_entry_count": 3,
                    "shared_workbook_enabled_count": 1,
                    "track_revisions_enabled_count": 1,
                    "revision_history_enabled_count": 1,
                    "keep_change_history_enabled_count": 1,
                    "revision_history_protected_count": 1,
                    "unrecognized_shared_workbook_revision_count": 0,
                    "stable_sheet": "Controls",
                    "stable_value_cell": "B10",
                    "stable_value": 12,
                    "stable_formula_cell": "D10",
                    "stable_formula": "=B10*C10",
                }
            ],
            coverage=[
                "The pair changes only xl/revisions/revisionLog1.xml: one synthetic historic value differs inside one stored revision record. Its revision-header/log relationship graph, content types, tracking and retention controls, record shape, ordinary cells, formulas, calculation properties, and every other package member stay fixed.",
                "The pair records a stored legacy revision-history difference only. It does not reveal or validate historic cell values, locations, author identity, timestamps, GUIDs, relationship IDs, provenance, conflict resolution, review/approval, authentication, authorization, or a completed workflow.",
                "The raw XML validates the exact synthetic package shape locally. An adapter must instead require FormulaFence's redacted one-header/one-log equal-count profile, revision-log-material signal, and FF062 evidence; it must never require a historic value, cell reference, author, timestamp, GUID, or relationship ID.",
            ],
        ),
    )


def _build_governance_query_table_refresh(root: Path) -> None:
    """Build a pair whose QueryTable, not connection, requests refresh on open."""

    directory = root / "governance" / "query_table_refresh_on_open"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_query_table_refresh_workbook(), baseline)
    _save_workbook(_query_table_refresh_workbook(), candidate)
    _add_query_table_refresh_control(
        baseline, refresh_on_load=_QUERY_TABLE_BASELINE_REFRESH_ON_LOAD
    )
    _add_query_table_refresh_control(
        candidate, refresh_on_load=_QUERY_TABLE_CANDIDATE_REFRESH_ON_LOAD
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.query_table_refresh_on_open",
            title="A QueryTable starts refreshing when the workbook opens",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "query_table_refresh_on_load_changed",
                    "sheet": _QUERY_TABLE_SHEET,
                    "connection_id": _QUERY_TABLE_CONNECTION_ID,
                    "connection_member": _QUERY_TABLE_CONNECTION_MEMBER,
                    "connection_url": _QUERY_TABLE_SOURCE_URL,
                    "query_table_member": _QUERY_TABLE_MEMBER,
                    "worksheet_member": _QUERY_TABLE_WORKSHEET_MEMBER,
                    "worksheet_relationships_member": _QUERY_TABLE_WORKSHEET_RELATIONSHIPS_MEMBER,
                    "relationship_id": _QUERY_TABLE_WORKSHEET_RELATIONSHIP_ID,
                    "relationship_type": _QUERY_TABLE_RELATIONSHIP,
                    "baseline_refresh_on_load": _QUERY_TABLE_BASELINE_REFRESH_ON_LOAD,
                    "candidate_refresh_on_load": _QUERY_TABLE_CANDIDATE_REFRESH_ON_LOAD,
                    "background_refresh": _QUERY_TABLE_BACKGROUND_REFRESH,
                    "refresh_disabled": _QUERY_TABLE_REFRESH_DISABLED,
                    "remove_data_on_save": _QUERY_TABLE_REMOVE_DATA_ON_SAVE,
                    "fill_formulas": _QUERY_TABLE_FILL_FORMULAS,
                    "connection_edit_disabled": _QUERY_TABLE_CONNECTION_EDIT_DISABLED,
                    "growth_behavior": _QUERY_TABLE_GROWTH_BEHAVIOR,
                    "saved_value_cell": _QUERY_TABLE_SAVED_VALUE_CELL,
                    "saved_value": _QUERY_TABLE_SAVED_VALUE,
                    "summary_sheet": _QUERY_TABLE_SUMMARY_SHEET,
                    "summary_cell": _QUERY_TABLE_SUMMARY_CELL,
                    "summary_formula": _QUERY_TABLE_SUMMARY_FORMULA,
                    "dashboard_sheet": _QUERY_TABLE_DASHBOARD_SHEET,
                    "dashboard_cell": _QUERY_TABLE_DASHBOARD_CELL,
                    "dashboard_formula": _QUERY_TABLE_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {
                        "sheet": _QUERY_TABLE_SHEET,
                        "cell": _QUERY_TABLE_SAVED_VALUE_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _QUERY_TABLE_SUMMARY_SHEET,
                            "cell": _QUERY_TABLE_SUMMARY_CELL,
                        },
                        {
                            "sheet": _QUERY_TABLE_DASHBOARD_SHEET,
                            "cell": _QUERY_TABLE_DASHBOARD_CELL,
                        },
                    ],
                },
                {
                    "source": {
                        "sheet": _QUERY_TABLE_SUMMARY_SHEET,
                        "cell": _QUERY_TABLE_SUMMARY_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _QUERY_TABLE_DASHBOARD_SHEET,
                            "cell": _QUERY_TABLE_DASHBOARD_CELL,
                        }
                    ],
                },
            ],
            coverage=[
                "The pair changes only raw xl/queryTables/queryTable1.xml queryTable/@refreshOnLoad from false to true. It does not edit ordinary cells, formulas, calculation properties, the workbook connection, content types, or relationships.",
                "The QueryTable retains one direct internal worksheet relationship and a fixed web-query connection to a reserved example.invalid URL. There is no credential, query result, table range, or external OOXML relationship.",
                "WCAB does not open a connection, fetch a URL, refresh a query, materialize rows, calculate a workbook, or claim that a client refreshes successfully. The stable saved cells and direct formula path are context only.",
            ],
        ),
    )


def _build_governance_cell_hyperlink_target(root: Path) -> None:
    """Build a pair whose visible hyperlink text stays fixed while its target moves."""

    directory = root / "governance" / "cell_hyperlink_target_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_cell_hyperlink_target_workbook(), baseline)
    _save_workbook(_cell_hyperlink_target_workbook(), candidate)
    _set_cell_hyperlink_target(baseline, target=_CELL_HYPERLINK_BASELINE_TARGET)
    _set_cell_hyperlink_target(candidate, target=_CELL_HYPERLINK_CANDIDATE_TARGET)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.cell_hyperlink_target_changed",
            title="A worksheet cell hyperlink points to a different target",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "cell_hyperlink_target_changed",
                    "sheet": _CELL_HYPERLINK_SHEET,
                    "cell": _CELL_HYPERLINK_CELL,
                    "cell_value": _CELL_HYPERLINK_VALUE,
                    "worksheet_member": _CELL_HYPERLINK_WORKSHEET_MEMBER,
                    "worksheet_relationships_member": _CELL_HYPERLINK_WORKSHEET_RELATIONSHIPS_MEMBER,
                    "relationship_id": _CELL_HYPERLINK_RELATIONSHIP_ID,
                    "relationship_type": _CELL_HYPERLINK_RELATIONSHIP,
                    "target_mode": "External",
                    "baseline_target": _CELL_HYPERLINK_BASELINE_TARGET,
                    "candidate_target": _CELL_HYPERLINK_CANDIDATE_TARGET,
                    "summary_sheet": _CELL_HYPERLINK_SUMMARY_SHEET,
                    "summary_cell": _CELL_HYPERLINK_SUMMARY_CELL,
                    "summary_formula": _CELL_HYPERLINK_SUMMARY_FORMULA,
                    "dashboard_sheet": _CELL_HYPERLINK_DASHBOARD_SHEET,
                    "dashboard_cell": _CELL_HYPERLINK_DASHBOARD_CELL,
                    "dashboard_formula": _CELL_HYPERLINK_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": _CELL_HYPERLINK_SHEET, "cell": _CELL_HYPERLINK_CELL},
                    "targets": [
                        {
                            "sheet": _CELL_HYPERLINK_SUMMARY_SHEET,
                            "cell": _CELL_HYPERLINK_SUMMARY_CELL,
                        },
                        {
                            "sheet": _CELL_HYPERLINK_DASHBOARD_SHEET,
                            "cell": _CELL_HYPERLINK_DASHBOARD_CELL,
                        },
                    ],
                },
                {
                    "source": {
                        "sheet": _CELL_HYPERLINK_SUMMARY_SHEET,
                        "cell": _CELL_HYPERLINK_SUMMARY_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _CELL_HYPERLINK_DASHBOARD_SHEET,
                            "cell": _CELL_HYPERLINK_DASHBOARD_CELL,
                        }
                    ],
                },
            ],
            coverage=[
                "The pair changes only xl/worksheets/_rels/sheet1.xml.rels Relationship/@Target. Ordinary cells, formulas, calculation properties, worksheet XML, and the relationship declaration remain byte-identical.",
                "Each package contains one standard worksheet cell hyperlink declaration bound to one external OOXML hyperlink relationship. The reserved example.invalid targets have no location, display, tooltip, or HYPERLINK formula.",
                "WCAB does not resolve, open, fetch, visit, execute, calculate, or otherwise interact with either target, and does not claim that any client follows it.",
            ],
        ),
    )


def _build_governance_pivot_cache_refresh(root: Path) -> None:
    """Build a local PivotCache whose open-time refresh request changes."""

    directory = root / "governance" / "pivot_cache_refresh_on_open"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_pivot_cache_refresh_workbook(), baseline)
    _save_workbook(_pivot_cache_refresh_workbook(), candidate)
    _add_pivot_cache_refresh_control(baseline, refresh_on_load=False)
    _add_pivot_cache_refresh_control(candidate, refresh_on_load=True)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.pivot_cache_refresh_on_open",
            title="A PivotTable cache starts refreshing when the workbook opens",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "pivot_cache_refresh_on_load_changed",
                    "cache_id": _PIVOT_CACHE_ID,
                    "source_type": "worksheet",
                    "source_sheet": _PIVOT_CACHE_SOURCE_SHEET,
                    "source_ref": _PIVOT_CACHE_SOURCE_REF,
                    "pivot_sheet": _PIVOT_REPORT_SHEET,
                    "pivot_ref": _PIVOT_REPORT_REF,
                    "pivot_output_cell": "B2",
                    "dashboard_sheet": "Dashboard",
                    "dashboard_cell": "B4",
                    "dashboard_formula": _PIVOT_CACHE_DASHBOARD_FORMULA,
                    "baseline_refresh_on_load": False,
                    "candidate_refresh_on_load": True,
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": _PIVOT_REPORT_SHEET, "cell": "B2"},
                    "targets": [{"sheet": "Dashboard", "cell": "B4"}],
                }
            ],
            coverage=[
                "The pair changes only raw pivotCacheDefinition/@refreshOnLoad. It does not edit source cells, stored PivotTable display cells, formula text, or calculation properties.",
                "PivotTable cache refresh-on-open is stored control evidence. WCAB does not open the workbook in Excel, refresh a cache, calculate, or render the PivotTable, or claim a changed report result.",
                "The validator reads relationship-backed raw OOXML, checks the cache/PivotTable binding and direct local dashboard edge, and treats that edge as a lower bound only if a client refresh changes the PivotTable display.",
            ],
        ),
    )


def _build_structural_pivot_data_field_aggregation(root: Path) -> None:
    """Build a PivotTable whose stored value aggregation switches in isolation."""

    directory = root / "structural" / "pivot_data_field_aggregation_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_pivot_cache_refresh_workbook(), baseline)
    _save_workbook(_pivot_cache_refresh_workbook(), candidate)
    _add_pivot_cache_refresh_control(baseline, refresh_on_load=False)
    _add_pivot_cache_refresh_control(candidate, refresh_on_load=False)
    _set_pivot_data_field_subtotal(baseline, subtotal=_PIVOT_DATA_FIELD_BASELINE_SUBTOTAL)
    _set_pivot_data_field_subtotal(candidate, subtotal=_PIVOT_DATA_FIELD_CANDIDATE_SUBTOTAL)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.pivot_data_field_aggregation_changed",
            title="A PivotTable value field switches from Sum to Average",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "pivot_data_field_aggregation_changed",
                    "cache_id": _PIVOT_CACHE_ID,
                    "source_type": "worksheet",
                    "source_sheet": _PIVOT_CACHE_SOURCE_SHEET,
                    "source_ref": _PIVOT_CACHE_SOURCE_REF,
                    "pivot_sheet": _PIVOT_REPORT_SHEET,
                    "pivot_ref": _PIVOT_REPORT_REF,
                    "pivot_output_cell": "B2",
                    "dashboard_sheet": "Dashboard",
                    "dashboard_cell": "B4",
                    "dashboard_formula": _PIVOT_CACHE_DASHBOARD_FORMULA,
                    "data_field_source_index": _PIVOT_DATA_FIELD_SOURCE_INDEX,
                    "baseline_subtotal": _PIVOT_DATA_FIELD_BASELINE_SUBTOTAL,
                    "candidate_subtotal": _PIVOT_DATA_FIELD_CANDIDATE_SUBTOTAL,
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": _PIVOT_REPORT_SHEET, "cell": "B2"},
                    "targets": [{"sheet": "Dashboard", "cell": "B4"}],
                }
            ],
            coverage=[
                "The pair changes only raw pivotTableDefinition/dataFields/dataField/@subtotal from sum to average. It does not edit source cells, cache records, stored PivotTable display cells, formula text, refresh controls, or calculation properties.",
                "The observable contract is the stored aggregation declaration. WCAB does not calculate, refresh, or render the PivotTable, infer a changed displayed value, or claim client behavior.",
                "The validator follows the local cache/PivotTable relationship graph, verifies the stable source and direct dashboard edge, and treats that edge as a lower bound only if a client refresh changes the PivotTable display.",
            ],
        ),
    )


def _build_structural_pivot_slicer_selection(root: Path) -> None:
    """Build a PivotTable Slicer whose selected cache item switches in isolation."""

    directory = root / "structural" / "pivot_slicer_selection_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_pivot_cache_refresh_workbook(), baseline)
    _save_workbook(_pivot_cache_refresh_workbook(), candidate)
    _add_pivot_cache_refresh_control(baseline, refresh_on_load=False)
    _add_pivot_cache_refresh_control(candidate, refresh_on_load=False)
    _add_pivot_slicer_selection(baseline, selected_item_index=_PIVOT_SLICER_BASELINE_SELECTED_INDEX)
    _add_pivot_slicer_selection(
        candidate, selected_item_index=_PIVOT_SLICER_CANDIDATE_SELECTED_INDEX
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.pivot_slicer_selection_changed",
            title="A PivotTable slicer switches from North to South",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "pivot_slicer_selection_changed",
                    "cache_id": _PIVOT_CACHE_ID,
                    "source_type": "worksheet",
                    "source_sheet": _PIVOT_CACHE_SOURCE_SHEET,
                    "source_ref": _PIVOT_CACHE_SOURCE_REF,
                    "pivot_sheet": _PIVOT_REPORT_SHEET,
                    "pivot_ref": _PIVOT_REPORT_REF,
                    "pivot_output_cell": "B2",
                    "dashboard_sheet": "Dashboard",
                    "dashboard_cell": "B4",
                    "dashboard_formula": _PIVOT_CACHE_DASHBOARD_FORMULA,
                    "slicer_name": _PIVOT_SLICER_NAME,
                    "slicer_source_name": _PIVOT_SLICER_SOURCE_NAME,
                    "slicer_pivot_table_name": _PIVOT_SLICER_PIVOT_TABLE_NAME,
                    "slicer_pivot_tab_id": _PIVOT_SLICER_PIVOT_TAB_ID,
                    "item_count": _PIVOT_SLICER_ITEM_COUNT,
                    "baseline_selected_item_index": _PIVOT_SLICER_BASELINE_SELECTED_INDEX,
                    "candidate_selected_item_index": _PIVOT_SLICER_CANDIDATE_SELECTED_INDEX,
                    "baseline_selected_value": "North",
                    "candidate_selected_value": "South",
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": _PIVOT_REPORT_SHEET, "cell": "B2"},
                    "targets": [{"sheet": "Dashboard", "cell": "B4"}],
                }
            ],
            coverage=[
                "The pair changes only raw slicerCacheDefinition/data/tabular/items/i/@s selection state. It does not edit source cells, cache records, stored PivotTable display cells, formula text, refresh controls, or calculation properties.",
                "The observable contract is the stored Slicer-cache selection declaration. This fixture has no Slicer drawing or view geometry; WCAB does not apply the filter, refresh, calculate, render, infer a changed display value, or claim client behavior.",
                "The validator follows the local workbook-to-Slicer-cache-to-PivotCache/PivotTable relationship graph, verifies the stable source and direct dashboard edge, and treats that edge as a lower bound only if a client applies the Slicer filter.",
            ],
        ),
    )


def _build_structural_power_query_m_filter(root: Path) -> None:
    """Build a connection-only local-table Power Query M filter transition."""

    directory = root / "structural" / "power_query_m_filter_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_power_query_local_table_workbook(), baseline)
    _save_workbook(_power_query_local_table_workbook(), candidate)
    _add_power_query_local_table_filter(baseline, filter_value=_POWER_QUERY_BASELINE_FILTER_VALUE)
    _add_power_query_local_table_filter(candidate, filter_value=_POWER_QUERY_CANDIDATE_FILTER_VALUE)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.power_query_m_filter_changed",
            title="A Power Query local-table filter switches from North to South",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "power_query_m_filter_changed",
                    "data_mashup_part": _POWER_QUERY_CUSTOM_XML_MEMBER,
                    "source_sheet": _POWER_QUERY_SOURCE_SHEET,
                    "source_table": _POWER_QUERY_SOURCE_TABLE,
                    "source_ref": _POWER_QUERY_SOURCE_REF,
                    "query_section": _POWER_QUERY_SECTION,
                    "query_name": _POWER_QUERY_NAME,
                    "filter_column": _POWER_QUERY_FILTER_COLUMN,
                    "baseline_filter_value": _POWER_QUERY_BASELINE_FILTER_VALUE,
                    "candidate_filter_value": _POWER_QUERY_CANDIDATE_FILTER_VALUE,
                    "fill_enabled": False,
                    "firewall_enabled": True,
                    "future_packages_allowed": False,
                }
            ],
            coverage=[
                "The pair changes only the stored M filter literal in one Data Mashup custom-XML part. It does not edit the local source table, worksheet cells, table definition, workbook calculation properties, connection settings, or a stored query result.",
                "The query is connection-only (FillEnabled=false) and reads the generated workbook's local SourceData table. WCAB does not execute M, apply the filter, refresh a query, materialize output, calculate formulas, or infer returned rows or client behavior.",
                "The validator follows the package-root customXml relationship, parses only this compact generated Data Mashup envelope, and verifies its one local Table source plus stable metadata and permission controls. It is not a general M parser or a proof of query execution semantics.",
            ],
        ),
    )


def _build_structural_scenario_manager_stored_input(root: Path) -> None:
    """Build a Scenario Manager alternate-input transition without cell churn."""

    directory = root / "structural" / "scenario_manager_stored_input_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_scenario_manager_workbook(), baseline)
    _save_workbook(_scenario_manager_workbook(), candidate)
    _add_scenario_manager_stored_input(
        baseline, stored_value=_SCENARIO_MANAGER_BASELINE_STORED_VALUE
    )
    _add_scenario_manager_stored_input(
        candidate, stored_value=_SCENARIO_MANAGER_CANDIDATE_STORED_VALUE
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.scenario_manager_stored_input_changed",
            title="A Scenario Manager downside input changes without a worksheet-cell edit",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "scenario_manager_stored_input_value_changed",
                    "scenario_sheet": _SCENARIO_MANAGER_SHEET,
                    "scenario_name": _SCENARIO_MANAGER_NAME,
                    "changing_cell": _SCENARIO_MANAGER_CHANGING_CELL,
                    "stable_input_cell": _SCENARIO_MANAGER_STABLE_INPUT_CELL,
                    "baseline_stored_value": _SCENARIO_MANAGER_BASELINE_STORED_VALUE,
                    "candidate_stored_value": _SCENARIO_MANAGER_CANDIDATE_STORED_VALUE,
                    "stable_stored_value": _SCENARIO_MANAGER_STABLE_STORED_VALUE,
                    "input_number_format_id": _SCENARIO_MANAGER_INPUT_NUMBER_FORMAT_ID,
                    "summary_ref": _SCENARIO_MANAGER_RESULT_CELL,
                    "worksheet_input_value": _SCENARIO_MANAGER_WORKSHEET_INPUT_VALUE,
                    "worksheet_stable_input_value": _SCENARIO_MANAGER_WORKSHEET_STABLE_INPUT_VALUE,
                    "result_cell": _SCENARIO_MANAGER_RESULT_CELL,
                    "result_formula": _SCENARIO_MANAGER_FORMULA,
                    "dashboard_sheet": _SCENARIO_MANAGER_DASHBOARD_SHEET,
                    "dashboard_cell": _SCENARIO_MANAGER_DASHBOARD_CELL,
                    "dashboard_formula": _SCENARIO_MANAGER_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {
                        "sheet": _SCENARIO_MANAGER_SHEET,
                        "cell": _SCENARIO_MANAGER_CHANGING_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _SCENARIO_MANAGER_SHEET,
                            "cell": _SCENARIO_MANAGER_RESULT_CELL,
                        },
                        {
                            "sheet": _SCENARIO_MANAGER_DASHBOARD_SHEET,
                            "cell": _SCENARIO_MANAGER_DASHBOARD_CELL,
                        },
                    ],
                }
            ],
            coverage=[
                "The pair changes only one raw scenarios/scenario/inputCells/@val record in the Inputs worksheet part. It does not edit visible worksheet cells, formula text, calculation properties, or the Scenario Manager selection, protection, comment, user, summary-reference, or number-format metadata.",
                "The observed contract is one stored alternate Scenario Manager input value. WCAB does not show or apply a scenario, calculate the model, predict a result, generate a Scenario Summary, or claim a client will use the stored input.",
                "The stable local formula path from the declared changing cell to the dashboard is a lower bound only if a client applies the scenario. The validator reads raw worksheet OOXML and formula text without executing either action.",
            ],
        ),
    )


def _build_structural_what_if_data_table_input(root: Path) -> None:
    """Build a Data Table input-reference switch without cell/result churn."""

    directory = root / "structural" / "what_if_data_table_input_reference_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_what_if_data_table_workbook(), baseline)
    _save_workbook(_what_if_data_table_workbook(), candidate)
    _set_what_if_data_table_input_reference(
        candidate, input_cell=_WHAT_IF_DATA_TABLE_ALTERNATE_INPUT_CELL
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.what_if_data_table_input_reference_changed",
            title="A What-If Data Table switches its one-variable input cell",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "what_if_data_table_input_reference_changed",
                    "table_sheet": _WHAT_IF_DATA_TABLE_SHEET,
                    "master_cell": _WHAT_IF_DATA_TABLE_MASTER_CELL,
                    "output_range": _WHAT_IF_DATA_TABLE_OUTPUT_RANGE,
                    "baseline_input_cell": _WHAT_IF_DATA_TABLE_PRIMARY_INPUT_CELL,
                    "candidate_input_cell": _WHAT_IF_DATA_TABLE_ALTERNATE_INPUT_CELL,
                    "orientation": "column",
                    "recalculation_requested": True,
                    "input_value_range": _WHAT_IF_DATA_TABLE_INPUT_VALUE_RANGE,
                    "input_values": list(_WHAT_IF_DATA_TABLE_GRID_VALUES),
                    "primary_input_value": _WHAT_IF_DATA_TABLE_PRIMARY_INPUT_VALUE,
                    "alternate_input_value": _WHAT_IF_DATA_TABLE_ALTERNATE_INPUT_VALUE,
                    "scale_cell": _WHAT_IF_DATA_TABLE_SCALE_CELL,
                    "scale_value": _WHAT_IF_DATA_TABLE_SCALE_VALUE,
                    "output_formula_cell": _WHAT_IF_DATA_TABLE_OUTPUT_FORMULA_CELL,
                    "output_formula": _WHAT_IF_DATA_TABLE_OUTPUT_FORMULA,
                    "model_sheet": _WHAT_IF_DATA_TABLE_MODEL_SHEET,
                    "model_cell": _WHAT_IF_DATA_TABLE_MODEL_CELL,
                    "model_formula": _WHAT_IF_DATA_TABLE_MODEL_FORMULA,
                    "dashboard_sheet": _WHAT_IF_DATA_TABLE_DASHBOARD_SHEET,
                    "dashboard_cell": _WHAT_IF_DATA_TABLE_DASHBOARD_CELL,
                    "dashboard_formula": _WHAT_IF_DATA_TABLE_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {
                        "sheet": _WHAT_IF_DATA_TABLE_SHEET,
                        "cell": _WHAT_IF_DATA_TABLE_PRIMARY_INPUT_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _WHAT_IF_DATA_TABLE_MODEL_SHEET,
                            "cell": _WHAT_IF_DATA_TABLE_MODEL_CELL,
                        },
                        {
                            "sheet": _WHAT_IF_DATA_TABLE_DASHBOARD_SHEET,
                            "cell": _WHAT_IF_DATA_TABLE_DASHBOARD_CELL,
                        },
                    ],
                },
                {
                    "source": {
                        "sheet": _WHAT_IF_DATA_TABLE_SHEET,
                        "cell": _WHAT_IF_DATA_TABLE_ALTERNATE_INPUT_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _WHAT_IF_DATA_TABLE_MODEL_SHEET,
                            "cell": _WHAT_IF_DATA_TABLE_MODEL_CELL,
                        },
                        {
                            "sheet": _WHAT_IF_DATA_TABLE_DASHBOARD_SHEET,
                            "cell": _WHAT_IF_DATA_TABLE_DASHBOARD_CELL,
                        },
                    ],
                },
            ],
            coverage=[
                "The pair changes only the one-variable Data Table master's raw f/@r1 input-cell reference in the Sensitivity worksheet. It does not edit the table output range, orientation, recalculation request, visible cells, ordinary formula text, calculation properties, or saved table results.",
                "The table is a column-oriented one-variable What-If Data Table. WCAB records its stored declaration only; it does not substitute input values, recalculate the workbook or table, infer output values, resolve a circular dependency, or claim Excel-client behavior.",
                "Both possible input cells retain ordinary static local paths to the model and dashboard. Those are lower-bound formula edges, not a proof of which temporary value Excel substitutes or of any Data Table result.",
            ],
        ),
    )


def _build_governance_external_workbook_link_update_policy(root: Path) -> None:
    """Build a pair differing only in the global external-link open policy."""

    directory = root / "governance" / "external_workbook_link_update_on_open"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_external_workbook_link_policy_workbook(), baseline)
    _save_workbook(_external_workbook_link_policy_workbook(), candidate)
    _set_external_workbook_link_update_policy(baseline, update_links="never")
    _set_external_workbook_link_update_policy(candidate, update_links="always")
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.external_workbook_link_update_on_open",
            title="An external-workbook link switches from never to always updating on open",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "external_workbook_link_update_policy_changed",
                    "sheet": "LinkedModel",
                    "cell": "B2",
                    "formula": _EXTERNAL_WORKBOOK_LINK_FORMULA,
                    "baseline_update_links": "never",
                    "candidate_update_links": "always",
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": "LinkedModel", "cell": "B2"},
                    "targets": [{"sheet": "Dashboard", "cell": "B4"}],
                }
            ],
            coverage=[
                "The external workbook name is synthetic and absent; the benchmark never opens, resolves, refreshes, or authenticates to it.",
                "The observed contract is only workbookPr/@updateLinks. It does not establish source reachability, trust, changed source data, a returned value, or recalculation success.",
                "The external-link formula and its local downstream reference remain unchanged; the validator compares stored XML and formula text without executing either formula.",
            ],
        ),
    )


def _build_governance_external_workbook_link_source(root: Path) -> None:
    """Build a pair whose external-workbook source target moves invisibly."""

    directory = root / "governance" / "external_workbook_link_source_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_external_workbook_link_policy_workbook(), baseline)
    _save_workbook(_external_workbook_link_policy_workbook(), candidate)
    _add_external_workbook_link_source(baseline, target=_EXTERNAL_WORKBOOK_LINK_BASELINE_TARGET)
    _add_external_workbook_link_source(candidate, target=_EXTERNAL_WORKBOOK_LINK_CANDIDATE_TARGET)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.external_workbook_link_source_changed",
            title="An external-workbook link points to a different source",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "external_workbook_link_source_changed",
                    "sheet": _EXTERNAL_WORKBOOK_LINK_HOST_SHEET,
                    "cell": _EXTERNAL_WORKBOOK_LINK_HOST_CELL,
                    "formula": _EXTERNAL_WORKBOOK_LINK_FORMULA,
                    "workbook_member": "xl/workbook.xml",
                    "workbook_relationships_member": "xl/_rels/workbook.xml.rels",
                    "workbook_relationship_id": _EXTERNAL_WORKBOOK_LINK_WORKBOOK_RELATIONSHIP_ID,
                    "workbook_relationship_type": _EXTERNAL_WORKBOOK_LINK_RELATIONSHIP,
                    "external_link_member": _EXTERNAL_WORKBOOK_LINK_MEMBER,
                    "external_link_relationships_member": _EXTERNAL_WORKBOOK_LINK_RELATIONSHIPS_MEMBER,
                    "external_link_content_type": _EXTERNAL_WORKBOOK_LINK_CONTENT_TYPE,
                    "external_link_relationship_id": _EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP_ID,
                    "external_link_relationship_type": _EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP,
                    "external_sheet": _EXTERNAL_WORKBOOK_LINK_EXTERNAL_SHEET,
                    "target_mode": "External",
                    "baseline_target": _EXTERNAL_WORKBOOK_LINK_BASELINE_TARGET,
                    "candidate_target": _EXTERNAL_WORKBOOK_LINK_CANDIDATE_TARGET,
                    "dashboard_sheet": _EXTERNAL_WORKBOOK_LINK_DASHBOARD_SHEET,
                    "dashboard_cell": _EXTERNAL_WORKBOOK_LINK_DASHBOARD_CELL,
                    "dashboard_formula": _EXTERNAL_WORKBOOK_LINK_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {
                        "sheet": _EXTERNAL_WORKBOOK_LINK_HOST_SHEET,
                        "cell": _EXTERNAL_WORKBOOK_LINK_HOST_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _EXTERNAL_WORKBOOK_LINK_DASHBOARD_SHEET,
                            "cell": _EXTERNAL_WORKBOOK_LINK_DASHBOARD_CELL,
                        }
                    ],
                }
            ],
            coverage=[
                "The pair changes only xl/externalLinks/_rels/externalLink1.xml.rels Relationship/@Target. It does not edit formula text, stored cells, calculation properties, workbook XML, workbook relationships, the externalLink declaration, or content types.",
                "Each package has one workbook externalReferences binding to one externalLink/externalBook declaration with one externalLinkPath relationship. The external workbook name and worksheet declaration remain fixed; only the reserved example.invalid source target changes.",
                "WCAB reads local OOXML only. It does not resolve, open, fetch, authenticate to, trust, refresh, calculate, or otherwise interact with either source, and does not claim that a client updates the link or returns a value.",
            ],
        ),
    )


def _build_governance_external_defined_name_source(root: Path) -> None:
    """Build a pair whose local defined name moves between absent sources."""

    def mutate(workbook: Workbook) -> None:
        workbook.defined_names[_EXTERNAL_DEFINED_NAME_SOURCE_NAME] = DefinedName(
            _EXTERNAL_DEFINED_NAME_SOURCE_NAME,
            attr_text=_EXTERNAL_DEFINED_NAME_SOURCE_CANDIDATE_REFERS_TO,
        )

    truth = _truth(
        case_id="governance.external_defined_name_source_changed",
        title="A defined name points to a different external workbook source",
        family="governance",
        review_expectation="block",
        facts=[
            {
                "kind": "external_defined_name_source_changed",
                "name": _EXTERNAL_DEFINED_NAME_SOURCE_NAME,
                "workbook_member": "xl/workbook.xml",
                "baseline_refers_to": _EXTERNAL_DEFINED_NAME_SOURCE_BASELINE_REFERS_TO,
                "candidate_refers_to": _EXTERNAL_DEFINED_NAME_SOURCE_CANDIDATE_REFERS_TO,
                "formula_sheet": _EXTERNAL_DEFINED_NAME_SOURCE_MODEL_SHEET,
                "formula_cell": _EXTERNAL_DEFINED_NAME_SOURCE_MODEL_CELL,
                "formula": _EXTERNAL_DEFINED_NAME_SOURCE_MODEL_FORMULA,
                "dashboard_sheet": _EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_SHEET,
                "dashboard_cell": _EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_CELL,
                "dashboard_formula": _EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_FORMULA,
            }
        ],
        must_reach=[
            {
                "source": {
                    "sheet": _EXTERNAL_DEFINED_NAME_SOURCE_MODEL_SHEET,
                    "cell": _EXTERNAL_DEFINED_NAME_SOURCE_MODEL_CELL,
                },
                "targets": [
                    {
                        "sheet": _EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_SHEET,
                        "cell": _EXTERNAL_DEFINED_NAME_SOURCE_DASHBOARD_CELL,
                    }
                ],
            }
        ],
        coverage=[
            "The pair changes only the one local xl/workbook.xml definedName text for ScenarioRate. The formula cells, calculation properties, sheet declarations, workbook relationships, and all other package members remain unchanged.",
            "Both stored expressions are qualified references to synthetic absent workbooks. This compact package has no externalLink part or workbook externalReferences declaration; WCAB records the local defined-name text only.",
            "WCAB does not resolve, open, fetch, authenticate to, trust, refresh, calculate, or otherwise interact with either source, and does not claim that a client updates the name or returns a value. The stable Model!B2-to-Dashboard!B4 path is local review context only.",
        ],
    )
    _write_pair(
        root / "governance" / "external_defined_name_source_changed",
        _external_defined_name_source_workbook,
        mutate,
        truth,
    )


def _build_governance_xlm_auto_open_binding(root: Path) -> None:
    """Build a macro-enabled pair whose Auto_Open dispatch target changes."""

    directory = root / "governance" / "xlm_auto_open_binding_retargeted"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsm"
    candidate = directory / "candidate.xlsm"
    _save_workbook(_xlm_auto_open_binding_workbook(), baseline)
    _save_workbook(_xlm_auto_open_binding_workbook(), candidate)
    _attach_xlm_auto_open_binding(baseline, target=_XLM_AUTO_OPEN_BASELINE_TARGET)
    _attach_xlm_auto_open_binding(candidate, target=_XLM_AUTO_OPEN_CANDIDATE_TARGET)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="governance.xlm_auto_open_binding_retargeted",
            title="A hidden XLM Auto_Open binding retargets a macro-sheet cell",
            family="governance",
            review_expectation="block",
            facts=[
                {
                    "kind": "xlm_auto_open_binding_retargeted",
                    "workbook_member": _XLM_AUTO_OPEN_WORKBOOK_MEMBER,
                    "workbook_relationships_member": (_XLM_AUTO_OPEN_WORKBOOK_RELATIONSHIPS_MEMBER),
                    "macro_sheet_member": _XLM_AUTO_OPEN_MACRO_SHEET_MEMBER,
                    "macro_sheet_relationship_id": (_XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP_ID),
                    "macro_sheet_relationship_type": _XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP,
                    "macro_sheet_relationship_target": "macrosheets/sheet1.xml",
                    "macro_sheet_content_type": _XLM_AUTO_OPEN_MACRO_SHEET_CONTENT_TYPE,
                    "workbook_content_type": _XLM_AUTO_OPEN_WORKBOOK_CONTENT_TYPE,
                    "macro_sheet_name": _XLM_AUTO_OPEN_MACRO_SHEET_NAME,
                    "macro_sheet_sheet_id": _XLM_AUTO_OPEN_MACRO_SHEET_ID,
                    "macro_sheet_state": _XLM_AUTO_OPEN_MACRO_SHEET_STATE,
                    "automatic_macro_name": _XLM_AUTO_OPEN_DEFINED_NAME,
                    "automatic_macro_event": _XLM_AUTO_OPEN_EVENT,
                    "baseline_target": _XLM_AUTO_OPEN_BASELINE_TARGET,
                    "candidate_target": _XLM_AUTO_OPEN_CANDIDATE_TARGET,
                    "macro_sheet_formula": _XLM_AUTO_OPEN_MACRO_FORMULA,
                    "macro_sheet_formula_cells": ["A1", "A2"],
                    "macro_sheet_sha256": sha256(_XLM_AUTO_OPEN_MACRO_SHEET_PAYLOAD).hexdigest(),
                    "macro_sheet_size": len(_XLM_AUTO_OPEN_MACRO_SHEET_PAYLOAD),
                    "input_sheet": _XLM_AUTO_OPEN_INPUT_SHEET,
                    "input_cell": _XLM_AUTO_OPEN_INPUT_CELL,
                    "input_value": _XLM_AUTO_OPEN_INPUT_VALUE,
                    "model_sheet": _XLM_AUTO_OPEN_MODEL_SHEET,
                    "model_cell": _XLM_AUTO_OPEN_MODEL_CELL,
                    "model_formula": _XLM_AUTO_OPEN_MODEL_FORMULA,
                    "dashboard_sheet": _XLM_AUTO_OPEN_DASHBOARD_SHEET,
                    "dashboard_cell": _XLM_AUTO_OPEN_DASHBOARD_CELL,
                    "dashboard_formula": _XLM_AUTO_OPEN_DASHBOARD_FORMULA,
                }
            ],
            must_reach=[
                {
                    "source": {
                        "sheet": _XLM_AUTO_OPEN_INPUT_SHEET,
                        "cell": _XLM_AUTO_OPEN_INPUT_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _XLM_AUTO_OPEN_MODEL_SHEET,
                            "cell": _XLM_AUTO_OPEN_MODEL_CELL,
                        },
                        {
                            "sheet": _XLM_AUTO_OPEN_DASHBOARD_SHEET,
                            "cell": _XLM_AUTO_OPEN_DASHBOARD_CELL,
                        },
                    ],
                },
                {
                    "source": {
                        "sheet": _XLM_AUTO_OPEN_MODEL_SHEET,
                        "cell": _XLM_AUTO_OPEN_MODEL_CELL,
                    },
                    "targets": [
                        {
                            "sheet": _XLM_AUTO_OPEN_DASHBOARD_SHEET,
                            "cell": _XLM_AUTO_OPEN_DASHBOARD_CELL,
                        }
                    ],
                },
            ],
            coverage=[
                "The macro-enabled pair changes only the one workbook-scoped _xlnm.Auto_Open definedName text in xl/workbook.xml, from Macro Automation!$A$1 to Macro Automation!$A$2. The raw XLM macro-sheet part, macro-sheet relationship, macro-enabled workbook and macro-sheet content types, ordinary cells/formulas, calculation properties, and every other package member remain unchanged.",
                "Both direct targets name cells on one very-hidden xlMacrosheet relationship. That fixed part contains only two identical HALT() formula cells and no VBA project, related payload, or external target. The ordinary Inputs!B2-to-Model!B2-to-Dashboard!B4 path is stable local review context; it is not a macro dependency claim.",
                "WCAB does not open Excel, enable or execute XLM code, parse or emulate macro instructions, resolve a dynamic name, inspect macro-security/trust settings, infer a dispatch result, calculate a workbook, or claim client behavior. It records only the stored workbook dispatch declaration and bounded raw package shape.",
            ],
        ),
    )


def _build_structural_named_lambda_definition(root: Path) -> None:
    """Build a pair whose reusable named-function body changes in isolation."""

    def mutate(workbook: Workbook) -> None:
        workbook.defined_names[_NAMED_LAMBDA_NAME] = DefinedName(
            _NAMED_LAMBDA_NAME,
            attr_text=_NAMED_LAMBDA_CANDIDATE_REFERS_TO,
        )

    truth = _truth(
        case_id="structural.named_lambda_definition_changed",
        title="A reusable named LAMBDA changes its calculation body",
        family="structural",
        review_expectation="block",
        facts=[
            {
                "kind": "named_lambda_definition_changed",
                "name": _NAMED_LAMBDA_NAME,
                "workbook_member": "xl/workbook.xml",
                "parameters": ["rate", "amount"],
                "baseline_refers_to": _NAMED_LAMBDA_BASELINE_REFERS_TO,
                "candidate_refers_to": _NAMED_LAMBDA_CANDIDATE_REFERS_TO,
                "input_sheet": _NAMED_LAMBDA_INPUT_SHEET,
                "rate_cell": _NAMED_LAMBDA_RATE_CELL,
                "rate_value": _NAMED_LAMBDA_RATE_VALUE,
                "amount_cell": _NAMED_LAMBDA_AMOUNT_CELL,
                "amount_value": _NAMED_LAMBDA_AMOUNT_VALUE,
                "formula_sheet": _NAMED_LAMBDA_MODEL_SHEET,
                "formula_cell": _NAMED_LAMBDA_MODEL_CELL,
                "formula": _NAMED_LAMBDA_MODEL_FORMULA,
                "dashboard_sheet": _NAMED_LAMBDA_DASHBOARD_SHEET,
                "dashboard_cell": _NAMED_LAMBDA_DASHBOARD_CELL,
                "dashboard_formula": _NAMED_LAMBDA_DASHBOARD_FORMULA,
            }
        ],
        must_reach=[
            {
                "source": {
                    "sheet": _NAMED_LAMBDA_INPUT_SHEET,
                    "cell": _NAMED_LAMBDA_RATE_CELL,
                },
                "targets": [
                    {
                        "sheet": _NAMED_LAMBDA_MODEL_SHEET,
                        "cell": _NAMED_LAMBDA_MODEL_CELL,
                    },
                    {
                        "sheet": _NAMED_LAMBDA_DASHBOARD_SHEET,
                        "cell": _NAMED_LAMBDA_DASHBOARD_CELL,
                    },
                ],
            },
            {
                "source": {
                    "sheet": _NAMED_LAMBDA_INPUT_SHEET,
                    "cell": _NAMED_LAMBDA_AMOUNT_CELL,
                },
                "targets": [
                    {
                        "sheet": _NAMED_LAMBDA_MODEL_SHEET,
                        "cell": _NAMED_LAMBDA_MODEL_CELL,
                    },
                    {
                        "sheet": _NAMED_LAMBDA_DASHBOARD_SHEET,
                        "cell": _NAMED_LAMBDA_DASHBOARD_CELL,
                    },
                ],
            },
            {
                "source": {
                    "sheet": _NAMED_LAMBDA_MODEL_SHEET,
                    "cell": _NAMED_LAMBDA_MODEL_CELL,
                },
                "targets": [
                    {
                        "sheet": _NAMED_LAMBDA_DASHBOARD_SHEET,
                        "cell": _NAMED_LAMBDA_DASHBOARD_CELL,
                    }
                ],
            },
        ],
        coverage=[
            "The pair changes only one xl/workbook.xml definedName text from =LAMBDA(rate,amount,rate*amount) to =LAMBDA(rate,amount,rate*(amount+10)). Inputs, ordinary formulas, calculation properties, workbook relationships, sheet declarations, and every other package member remain unchanged.",
            "The fixture has one workbook-scoped defined name and no externalLink package or workbook externalReferences declaration. Model!B2 calls that name, but WCAB records the stored formula definition rather than evaluating the LAMBDA body or resolving all named-formula dependencies.",
            "WCAB does not calculate a result, execute code, infer Excel-version support, inspect Name Manager behavior, or claim that a client recalculates, spills, or persists a value. The stable Model!B2-to-Dashboard!B4 path is local review context only.",
        ],
    )
    _write_pair(
        root / "structural" / "named_lambda_definition_changed",
        _named_lambda_definition_workbook,
        mutate,
        truth,
    )


def _build_structural_power_pivot_data_model_relationship(root: Path) -> None:
    """Build a pair whose stored Data Model relationship retargets one key."""

    directory = root / "structural" / "power_pivot_data_model_relationship_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_power_pivot_data_model_workbook(), baseline)
    _save_workbook(_power_pivot_data_model_workbook(), candidate)
    _attach_power_pivot_data_model(
        baseline,
        to_column=_POWER_PIVOT_DATA_BASELINE_TO_COLUMN,
    )
    _attach_power_pivot_data_model(
        candidate,
        to_column=_POWER_PIVOT_DATA_CANDIDATE_TO_COLUMN,
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.power_pivot_data_model_relationship_changed",
            title="A Power Pivot Data Model relationship retargets its key column",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "power_pivot_data_model_relationship_changed",
                    "workbook_member": _POWER_PIVOT_DATA_WORKBOOK_MEMBER,
                    "workbook_relationships_member": (
                        _POWER_PIVOT_DATA_WORKBOOK_RELATIONSHIPS_MEMBER
                    ),
                    "data_model_member": _POWER_PIVOT_DATA_MEMBER,
                    "workbook_relationship_id": _POWER_PIVOT_DATA_WORKBOOK_RELATIONSHIP_ID,
                    "workbook_relationship_type": _POWER_PIVOT_DATA_RELATIONSHIP,
                    "workbook_relationship_target": "model/item.data",
                    "data_model_content_type": _POWER_PIVOT_DATA_CONTENT_TYPE,
                    "extension_uri": _POWER_PIVOT_DATA_EXTENSION_URI,
                    "min_version_load": _POWER_PIVOT_DATA_MIN_VERSION_LOAD,
                    "model_tables": [table["name"] for table in _POWER_PIVOT_DATA_MODEL_TABLES],
                    "from_table": _POWER_PIVOT_DATA_FROM_TABLE,
                    "from_column": _POWER_PIVOT_DATA_FROM_COLUMN,
                    "to_table": _POWER_PIVOT_DATA_TO_TABLE,
                    "baseline_to_column": _POWER_PIVOT_DATA_BASELINE_TO_COLUMN,
                    "candidate_to_column": _POWER_PIVOT_DATA_CANDIDATE_TO_COLUMN,
                    "data_model_payload_sha256": sha256(_POWER_PIVOT_DATA_PAYLOAD).hexdigest(),
                    "data_model_payload_size": len(_POWER_PIVOT_DATA_PAYLOAD),
                }
            ],
            coverage=[
                "The pair changes only the x15:modelRelationship/@toColumn value in xl/workbook.xml, from CalendarModel.DateKey to CalendarModel.FiscalDateKey. The workbook-to-model relationship, content type, local Tables, opaque xl/model/item.data bytes, calculation properties, and every other package member remain unchanged.",
                "The stored model relationship binds the declared foreign-key SalesModel.CalendarKey to a named primary key on CalendarModel. It is relationship-definition evidence only: WCAB does not treat either local Excel Table as loaded Data Model content or claim that the fixed opaque payload is a complete executable model.",
                "WCAB does not deserialize the Analysis Services payload, evaluate DAX, refresh a model, calculate or render a PivotTable or chart, infer model-to-cell impact, or fetch an external target. The raw declaration is deliberately outside the ordinary A1 dependency graph, so this case has no formula reachability assertion.",
            ],
        ),
    )


def _build_structural_chart_series_reference(root: Path) -> None:
    """Build a dashboard chart whose local value-series binding changes."""

    directory = root / "structural" / "chart_series_reference_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_chart_series_reference_workbook(), baseline)
    _save_workbook(_chart_series_reference_workbook(), candidate)
    _set_chart_series_value_reference(
        baseline, value_reference=_CHART_SERIES_BASELINE_VALUE_REFERENCE
    )
    _set_chart_series_value_reference(
        candidate, value_reference=_CHART_SERIES_CANDIDATE_VALUE_REFERENCE
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.chart_series_reference_changed",
            title="A dashboard chart switches to a different local value series",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "chart_series_value_reference_changed",
                    "chart_sheet": _CHART_SERIES_DASHBOARD_SHEET,
                    "chart_anchor": _CHART_SERIES_ANCHOR,
                    "source_sheet": _CHART_SERIES_SOURCE_SHEET,
                    "series_title_ref": _CHART_SERIES_TITLE_REFERENCE,
                    "category_ref": _CHART_SERIES_CATEGORY_REFERENCE,
                    "baseline_value_ref": _CHART_SERIES_BASELINE_VALUE_REFERENCE,
                    "candidate_value_ref": _CHART_SERIES_CANDIDATE_VALUE_REFERENCE,
                }
            ],
            coverage=[
                "The pair changes only the raw DrawingML c:ser/c:val/c:numRef/c:f value-reference text in one chart part. It does not edit worksheet cells, chart anchor, category/title references, or calculation properties.",
                "A chart-series source reference is stored report-binding evidence. WCAB does not open Excel, calculate values, refresh chart data, render a chart, infer a visual difference, or claim any client will use the stored reference.",
                "The validator follows the local worksheet-to-drawing-to-chart relationship chain and compares raw chart XML after removing the declared value reference. It does not resolve arbitrary chart formulas or add chart references to the formula dependency graph.",
            ],
        ),
    )


def _build_structural_array_formula_mode(root: Path) -> None:
    """Build a legacy-CSE to dynamic-array transition without formula text churn."""

    directory = root / "structural" / "array_formula_mode_changed"
    directory.mkdir(parents=True, exist_ok=True)
    _save_workbook(_array_formula_semantics_workbook(), directory / "baseline.xlsx")
    candidate = directory / "candidate.xlsx"
    _save_workbook(_array_formula_semantics_workbook(), candidate)
    _add_dynamic_array_metadata(candidate)
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.array_formula_mode_changed",
            title="An unchanged array formula switches from fixed CSE to dynamic spill semantics",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "array_formula_mode_changed",
                    "sheet": "Model",
                    "cell": "B1",
                    "formula": "=LEN(Inputs!A1:A3)",
                    "baseline_mode": "legacy_cse",
                    "candidate_mode": "dynamic",
                    "baseline_output_range": "B1:B3",
                    "candidate_output_range": "B1:B3",
                }
            ],
            must_reach=[
                {
                    "source": {"sheet": "Model", "cell": "B1"},
                    "targets": [{"sheet": "Dashboard", "cell": "B2"}],
                }
            ],
            coverage=[
                "The legacy CSE range and the candidate's currently stored output range coincide; a dynamic array can resize or be blocked when Excel recalculates.",
                "The validator checks stored formula mode and metadata only. It does not calculate a result, predict a future spill extent, identify blockers, or assert client-version compatibility.",
            ],
        ),
    )


def _build_structural_three_d_scope(root: Path) -> None:
    directory = root / "structural" / "three_d_scope_expansion"
    directory.mkdir(parents=True, exist_ok=True)
    _save_workbook(_three_d_workbook(), directory / "baseline.xlsx")
    _save_workbook(_three_d_workbook(include_adjustment=True), directory / "candidate.xlsx")
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.three_d_scope_expansion",
            title="A new period silently enters an unchanged 3-D formula span",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "three_d_scope_changed",
                    "formula_sheet": "Summary",
                    "formula_cell": "B5",
                    "inserted_sheet": "FebAdjustment",
                    "after_sheet": "Jan",
                    "before_sheet": "Feb",
                }
            ],
            coverage=[
                "The formula text is intentionally unchanged; the observable risk is tab-order scope."
            ],
        ),
    )


def _build_structural_formula_rewrite(root: Path) -> None:
    directory = root / "structural" / "formula_rewrite_after_column_insert"
    directory.mkdir(parents=True, exist_ok=True)
    _save_workbook(_refactor_workbook(), directory / "baseline.xlsx")
    _save_workbook(_refactor_workbook(rewritten=True), directory / "candidate.xlsx")
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.formula_rewrite_after_column_insert",
            title="A column insertion rewrites formulas while retaining the declared inputs",
            family="structural",
            review_expectation="review",
            facts=[
                {
                    "kind": "structural_formula_rewrite",
                    "baseline": {"sheet": "Model", "cell": "D5", "formula": "=B5*C5"},
                    "candidate": {"sheet": "Model", "cell": "E5", "formula": "=B5*D5"},
                }
            ],
            coverage=[
                "The fixture declares equivalent logical input positions but does not prove general Excel semantic equivalence."
            ],
        ),
    )


def _build_structural_table_scope(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        ledger = workbook["Ledger"]
        ledger.append(["East", 15, 10, "=B5*C5"])
        ledger.tables["SalesLedger"].ref = "A1:D5"

    truth = _truth(
        case_id="structural.structured_table_scope_expansion",
        title="An Excel Table grows while a dependent structured reference stays unchanged",
        family="structural",
        review_expectation="block",
        facts=[
            {
                "kind": "structured_table_scope_changed",
                "table_sheet": "Ledger",
                "table": "SalesLedger",
                "baseline_ref": "A1:D4",
                "candidate_ref": "A1:D5",
                "formula_sheet": "Summary",
                "formula_cell": "B2",
            }
        ],
        coverage=[
            "The summary formula text intentionally remains unchanged; the observable risk is the stored Excel Table range.",
            "The contract does not calculate the structured reference, apply filters, or infer table-total semantics.",
        ],
    )
    _write_pair(
        root / "structural" / "structured_table_scope_expansion",
        _structured_table_workbook,
        mutate,
        truth,
    )


def _build_structural_table_calculated_column_formula(root: Path) -> None:
    """Build a pair whose Table-level calculated-column master changes."""

    directory = root / "structural" / "table_calculated_column_formula_changed"
    directory.mkdir(parents=True, exist_ok=True)
    baseline = directory / "baseline.xlsx"
    candidate = directory / "candidate.xlsx"
    _save_workbook(_table_calculated_column_formula_workbook(), baseline)
    _save_workbook(_table_calculated_column_formula_workbook(), candidate)
    _set_table_calculated_column_formula(
        baseline,
        formula=_TABLE_CALCULATED_COLUMN_BASELINE_FORMULA,
    )
    _set_table_calculated_column_formula(
        candidate,
        formula=_TABLE_CALCULATED_COLUMN_CANDIDATE_FORMULA,
    )
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="structural.table_calculated_column_formula_changed",
            title="An Excel Table calculated-column master changes without a cell edit",
            family="structural",
            review_expectation="block",
            facts=[
                {
                    "kind": "table_calculated_column_formula_changed",
                    "table_sheet": _TABLE_CALCULATED_COLUMN_SHEET,
                    "table_member": _TABLE_CALCULATED_COLUMN_MEMBER,
                    "table": _TABLE_CALCULATED_COLUMN_NAME,
                    "table_ref": _TABLE_CALCULATED_COLUMN_REF,
                    "calculated_column_id": _TABLE_CALCULATED_COLUMN_ID,
                    "calculated_column_name": _TABLE_CALCULATED_COLUMN_HEADER,
                    "baseline_formula": _TABLE_CALCULATED_COLUMN_BASELINE_FORMULA,
                    "candidate_formula": _TABLE_CALCULATED_COLUMN_CANDIDATE_FORMULA,
                    "stable_formula_cells": list(_TABLE_CALCULATED_COLUMN_STABLE_FORMULA_CELLS),
                    "dashboard_sheet": _TABLE_CALCULATED_COLUMN_DASHBOARD_SHEET,
                    "dashboard_cell": _TABLE_CALCULATED_COLUMN_DASHBOARD_CELL,
                    "dashboard_formula": _TABLE_CALCULATED_COLUMN_DASHBOARD_FORMULA,
                }
            ],
            coverage=[
                "The pair changes only xl/tables/table1.xml tableColumn/calculatedColumnFormula text from A2*B2 to A2*(B2+1). The Table name, range, headers, direct worksheet-to-Table relationship, ordinary Ledger!C2:C4 formulas, dashboard formula, calculation properties, and every other package member remain unchanged.",
                "The raw calculatedColumnFormula is a Table-level formula master. WCAB records that stored declaration and its local Table binding only; it does not fill formulas down a column, reconcile the stored cell formulas to the master, evaluate a structured reference, or infer a total.",
                "WCAB does not open Excel, calculate a workbook, apply a calculated-column update, add a row, render a Table, or claim client behavior. The stable Dashboard!B4 structured-reference formula is review context only.",
            ],
        ),
    )


def _build_structural_dynamic_reference(root: Path) -> None:
    def mutate(workbook: Workbook) -> None:
        workbook["Summary"]["B2"] = "=INDIRECT(Inputs!$E$12)"

    truth = _truth(
        case_id="structural.dynamic_reference_introduced",
        title="A direct output reference becomes an INDIRECT-driven reference",
        family="structural",
        review_expectation="block",
        facts=[
            {
                "kind": "dynamic_formula_reference_added",
                "sheet": "Summary",
                "cell": "B2",
                "functions": ["INDIRECT"],
            }
        ],
        must_reach=[
            {
                "source": {"sheet": "Summary", "cell": "B2"},
                "targets": [{"sheet": "Dashboard", "cell": "B4"}],
            }
        ],
        coverage=[
            "INDIRECT returns a reference from workbook text; this fixture does not evaluate that text or claim complete target resolution.",
            "The scoreable coverage expectation requires an explicit static-dependency boundary disclosure, not a universal formula-evaluation claim.",
        ],
        coverage_expectations=[
            {
                "kind": "dynamic_reference_static_coverage",
                "sheet": "Summary",
                "cell": "B2",
                "functions": ["INDIRECT"],
            }
        ],
    )
    _write_pair(
        root / "structural" / "dynamic_reference_introduced",
        _dynamic_reference_workbook,
        mutate,
        truth,
    )


def _build_structural_dynamic_reference_drivers(root: Path) -> None:
    def write_case(
        *,
        case_name: str,
        case_id: str,
        function: str,
        baseline_driver: str | int,
        candidate_driver: str | int,
        title: str,
        coverage: list[str],
    ) -> None:
        truth = _truth(
            case_id=case_id,
            title=title,
            family="structural",
            review_expectation="block",
            facts=[{"kind": "value_changed", "sheet": "Inputs", "cell": "E12"}],
            must_reach=[
                {
                    "source": {"sheet": "Inputs", "cell": "E12"},
                    "targets": [
                        {"sheet": "Summary", "cell": "B2"},
                        {"sheet": "Dashboard", "cell": "B4"},
                    ],
                }
            ],
            coverage=coverage,
            coverage_expectations=[
                {
                    "kind": "dynamic_reference_driver_changed",
                    "driver": {"sheet": "Inputs", "cell": "E12"},
                    "formula": {"sheet": "Summary", "cell": "B2"},
                    "functions": [function],
                }
            ],
        )
        _write_pair(
            root / "structural" / case_name,
            lambda: _dynamic_reference_driver_workbook(
                function=function, driver_value=baseline_driver
            ),
            lambda workbook: setattr(workbook["Inputs"]["E12"], "value", candidate_driver),
            truth,
        )

    write_case(
        case_name="indirect_reference_driver_changed",
        case_id="structural.indirect_reference_driver_changed",
        function="INDIRECT",
        baseline_driver="Revenue!B8",
        candidate_driver="Revenue!C8",
        title="An unchanged INDIRECT formula receives a different address driver",
        coverage=[
            "INDIRECT returns a reference from workbook text; this fixture does not evaluate or calculate the selected address.",
            "The unchanged formula still directly reads the changed driver, but its effective target selection is outside a complete static dependency graph.",
        ],
    )
    write_case(
        case_name="offset_reference_driver_changed",
        case_id="structural.offset_reference_driver_changed",
        function="OFFSET",
        baseline_driver=0,
        candidate_driver=1,
        title="An unchanged OFFSET formula receives a different column displacement",
        coverage=[
            "OFFSET returns a reference displaced from another reference; this fixture does not calculate the selected cell.",
            "The unchanged formula still directly reads the changed driver, but its effective target selection is outside a complete static dependency graph.",
        ],
    )


def _build_portfolio_external_driver(root: Path) -> None:
    directory = root / "portfolio" / "external_driver_value_change"
    baseline = directory / "baseline"
    candidate = directory / "candidate"
    baseline.mkdir(parents=True, exist_ok=True)
    candidate.mkdir(parents=True, exist_ok=True)
    base_drivers, base_model = _portfolio_workbook(driver_value=10)
    candidate_drivers, candidate_model = _portfolio_workbook(driver_value=15)
    _save_workbook(base_drivers, baseline / "drivers.xlsx")
    _save_workbook(base_model, baseline / "model.xlsx")
    _save_workbook(candidate_drivers, candidate / "drivers.xlsx")
    _save_workbook(candidate_model, candidate / "model.xlsx")
    _write_json(
        directory / "truth.json",
        _truth(
            case_id="portfolio.external_driver_value_change",
            title="An upstream workbook driver changes while a downstream workbook consumes it",
            family="portfolio",
            review_expectation="review",
            topology="portfolio",
            facts=[
                {
                    "kind": "portfolio_value_changed",
                    "workbook": "drivers.xlsx",
                    "sheet": "Inputs",
                    "cell": "B2",
                },
                {
                    "kind": "portfolio_external_reference",
                    "workbook": "model.xlsx",
                    "sheet": "Summary",
                    "cell": "B2",
                    "target_workbook": "drivers.xlsx",
                },
            ],
            coverage=[
                "The target workbook is present only as a local sibling; no network or filesystem path resolution is permitted."
            ],
        ),
    )


_BUILDERS: tuple[Callable[[Path], None], ...] = (
    _build_finance_formula_to_value,
    _build_finance_wrong_period_reference,
    _build_finance_input_change,
    _build_finance_external_formula,
    _build_finance_defined_name,
    _build_operations_formula_interruption,
    _build_operations_sumifs_shape,
    _build_operations_data_validation,
    _build_operations_data_validation_list_source,
    _build_operations_conditional_formatting,
    _build_operations_conditional_formatting_threshold,
    _build_operations_number_format_visibility,
    _build_operations_ignored_error_suppression,
    _build_operations_auto_filter_criteria,
    _build_operations_named_sheet_view_filter_criterion,
    _build_operations_xml_map_table_xpath,
    _build_governance_office_web_addin_auto_show,
    _build_governance_ole_object_auto_load,
    _build_governance_visibility,
    _build_governance_protection,
    _build_governance_sheet_protection_sort_permission,
    _build_governance_protected_range_security_descriptor,
    _build_governance_workbook_structure_protection,
    _build_governance_manual_calculation,
    _build_governance_iterative_calculation,
    _build_governance_precision_as_displayed,
    _build_governance_workbook_date_system,
    _build_governance_formula_cached_result,
    _build_governance_static_cycle,
    _build_governance_external_data_refresh,
    _build_governance_external_data_connection_source,
    _build_governance_package_signature_manifest_retarget,
    _build_governance_package_signature_relationship_selector_retarget,
    _build_governance_threaded_comment_resolution_state,
    _build_governance_shared_workbook_revision_log,
    _build_governance_query_table_refresh,
    _build_governance_cell_hyperlink_target,
    _build_governance_pivot_cache_refresh,
    _build_governance_external_workbook_link_update_policy,
    _build_governance_external_workbook_link_source,
    _build_governance_external_defined_name_source,
    _build_governance_xlm_auto_open_binding,
    _build_structural_named_lambda_definition,
    _build_structural_power_pivot_data_model_relationship,
    _build_structural_pivot_data_field_aggregation,
    _build_structural_pivot_slicer_selection,
    _build_structural_power_query_m_filter,
    _build_structural_scenario_manager_stored_input,
    _build_structural_what_if_data_table_input,
    _build_structural_chart_series_reference,
    _build_structural_array_formula_mode,
    _build_structural_three_d_scope,
    _build_structural_formula_rewrite,
    _build_structural_table_scope,
    _build_structural_table_calculated_column_formula,
    _build_structural_dynamic_reference,
    _build_structural_dynamic_reference_drivers,
    _build_portfolio_external_driver,
)

CASE_IDS = (
    "finance.formula_to_value",
    "finance.wrong_period_reference",
    "finance.input_value_change",
    "finance.external_formula_reference",
    "finance.defined_name_redirect",
    "operations.copied_formula_interruption",
    "operations.sumifs_range_shape",
    "operations.data_validation_removed",
    "operations.data_validation_list_source_changed",
    "operations.conditional_formatting_removed",
    "operations.conditional_formatting_threshold_changed",
    "operations.number_format_value_hidden",
    "operations.ignored_error_formula_range_suppressed",
    "operations.auto_filter_criteria_changed",
    "operations.named_sheet_view_filter_criterion_changed",
    "operations.xml_map_table_xpath_retargeted",
    "governance.office_web_addin_auto_show_enabled",
    "governance.ole_object_auto_load_enabled",
    "governance.hidden_sheet_revealed",
    "governance.formula_cell_unlocked",
    "governance.sheet_protection_sort_permission_enabled",
    "governance.protected_range_security_descriptor_changed",
    "governance.workbook_structure_lock_removed",
    "governance.manual_calculation_incomplete",
    "governance.iterative_calculation_enabled",
    "governance.precision_as_displayed_enabled",
    "governance.workbook_date_system_changed",
    "governance.formula_cached_result_changed",
    "governance.static_cycle_introduced",
    "governance.external_data_refresh_on_open",
    "governance.external_data_connection_source_changed",
    "governance.package_signature_manifest_retargeted",
    "governance.package_signature_relationship_selector_retargeted",
    "governance.threaded_comment_resolution_state_changed",
    "governance.shared_workbook_revision_log_changed",
    "governance.query_table_refresh_on_open",
    "governance.cell_hyperlink_target_changed",
    "governance.pivot_cache_refresh_on_open",
    "governance.external_workbook_link_update_on_open",
    "governance.external_workbook_link_source_changed",
    "governance.external_defined_name_source_changed",
    "governance.xlm_auto_open_binding_retargeted",
    "structural.named_lambda_definition_changed",
    "structural.power_pivot_data_model_relationship_changed",
    "structural.pivot_data_field_aggregation_changed",
    "structural.pivot_slicer_selection_changed",
    "structural.power_query_m_filter_changed",
    "structural.scenario_manager_stored_input_changed",
    "structural.what_if_data_table_input_reference_changed",
    "structural.chart_series_reference_changed",
    "structural.array_formula_mode_changed",
    "structural.three_d_scope_expansion",
    "structural.formula_rewrite_after_column_insert",
    "structural.structured_table_scope_expansion",
    "structural.table_calculated_column_formula_changed",
    "structural.dynamic_reference_introduced",
    "structural.indirect_reference_driver_changed",
    "structural.offset_reference_driver_changed",
    "portfolio.external_driver_value_change",
)


def build_all(root: str | Path) -> list[Path]:
    """Build all WCAB fixtures below *root* and return their truth manifests."""

    fixture_root = Path(root).resolve()
    fixture_root.mkdir(parents=True, exist_ok=True)
    for builder in _BUILDERS:
        builder(fixture_root)
    manifests = sorted(fixture_root.rglob("truth.json"))
    if len(manifests) != len(CASE_IDS):
        raise RuntimeError(f"expected {len(CASE_IDS)} manifests, found {len(manifests)}")
    write_manifest(fixture_root, expected_ids=CASE_IDS)
    return manifests
