"""Validation for WCAB's observable fixture truth contract."""

from __future__ import annotations

import base64
import binascii
import io
import json
import posixpath
import re
import struct
from collections import deque
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from .build import CASE_IDS, FIXTURE_SCHEMA_VERSION

_CELL_REFERENCE = re.compile(
    r"(?:(?:'(?P<quoted>(?:[^']|'')+)'|(?P<plain>[A-Za-z_][A-Za-z0-9_. ]*))!)?"
    r"(?P<cell>\$?[A-Z]{1,3}\$?[1-9][0-9]*)"
)
_DYNAMIC_REFERENCE_FUNCTION = re.compile(r"(?i)\b(?P<function>INDIRECT|OFFSET)\s*\(")
_SPREADSHEETML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_PACKAGE_RELATIONSHIPS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_DOCUMENT_RELATIONSHIPS_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
_OFFICE_2010_SPREADSHEET_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
_OFFICE_2013_SPREADSHEET_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
_OFFICE_2014_REVISION_NS = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
_NAMED_SHEET_VIEW_NS = "http://schemas.microsoft.com/office/spreadsheetml/2019/namedsheetviews"
_NAMED_SHEET_VIEW_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2019/04/relationships/namedSheetView"
)
_XML_SCHEMA_NS = "http://www.w3.org/2001/XMLSchema"
_XML_MAP_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/xmlMaps"
_TABLE_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/table"
_TABLE_SINGLE_CELLS_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/tableSingleCells"
_WEB_EXTENSION_TASKPANES_NS = "http://schemas.microsoft.com/office/webextensions/taskpanes/2010/11"
_WEB_EXTENSION_NS = "http://schemas.microsoft.com/office/webextensions/webextension/2010/11"
_WEB_EXTENSION_TASKPANES_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2011/relationships/webextensiontaskpanes"
)
_WEB_EXTENSION_RELATIONSHIP = "http://schemas.microsoft.com/office/2011/relationships/webextension"
_OLE_OBJECT_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/oleObject"
_OLE_OBJECT_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.oleObject"
_OLE_OBJECT_PAYLOAD = (
    b"WCAB opaque synthetic embedded-object fixture bytes; never deserialized or opened."
)
_QUERY_TABLE_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/queryTable"
_QUERY_TABLE_CONNECTIONS_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/connections"
_QUERY_TABLE_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.queryTable+xml"
)
_QUERY_TABLE_CONNECTIONS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml"
)
_EXTERNAL_DATA_CONNECTIONS_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/connections"
_EXTERNAL_DATA_CONNECTIONS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml"
)
_POWER_PIVOT_DATA_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/powerPivotData"
_POWER_PIVOT_DATA_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.model+data"
_POWER_PIVOT_DATA_PAYLOAD = (
    b"WCAB opaque synthetic Power Pivot Data Model payload v1; never deserialized or opened."
)
_XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP = (
    "http://schemas.microsoft.com/office/2006/relationships/xlMacrosheet"
)
_XLM_AUTO_OPEN_MACRO_SHEET_CONTENT_TYPE = "application/vnd.ms-excel.macrosheet+xml"
_XLM_AUTO_OPEN_WORKBOOK_CONTENT_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
_XLM_AUTO_OPEN_MACRO_SHEET_NS = "http://schemas.microsoft.com/office/excel/2006/main"
_XLM_AUTO_OPEN_MACRO_SHEET_PAYLOAD = (
    b'<?xml version="1.0" encoding="UTF-8"?>'
    b'<macrosheet xmlns="http://schemas.microsoft.com/office/excel/2006/main">'
    b'<sheetData><row r="1"><c r="A1"><f>HALT()</f></c></row>'
    b'<row r="2"><c r="A2"><f>HALT()</f></c></row></sheetData></macrosheet>'
)
_CELL_HYPERLINK_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/hyperlink"
_EXTERNAL_WORKBOOK_LINK_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/externalLink"
_EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/externalLinkPath"
_EXTERNAL_WORKBOOK_LINK_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"
)
_SLICER_CACHE_RELATIONSHIP = "http://schemas.microsoft.com/office/2007/relationships/slicerCache"
_DATA_MASHUP_NS = "http://schemas.microsoft.com/DataMashup"
_POWER_QUERY_CUSTOM_XML_RELATIONSHIP = f"{_DOCUMENT_RELATIONSHIPS_NS}/customXml"
_POWER_QUERY_STREAM_MAX_ENCODED_BYTES = 128 * 1024
_POWER_QUERY_PACKAGE_MAX_UNCOMPRESSED_BYTES = 128 * 1024
_DYNAMIC_ARRAY_NS = "http://schemas.microsoft.com/office/spreadsheetml/2017/dynamicarray"
_DRAWINGML_CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
_DRAWINGML_SPREADSHEET_DRAWING_NS = (
    "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
)
_POWER_QUERY_M_FILTER = re.compile(
    r"\Asection (?P<section>[A-Za-z_][A-Za-z0-9_]*);\n\n"
    r"shared (?P<query>[A-Za-z_][A-Za-z0-9_]*) = let\n"
    r'    Source = Excel\.CurrentWorkbook\(\)\{\[Name="(?P<source_table>[A-Za-z_][A-Za-z0-9_]*)"\]\}\[Content\],\n'
    r'    FilteredRows = Table\.SelectRows\(Source, each \[(?P<filter_column>[A-Za-z_][A-Za-z0-9_]*)\] = "(?P<filter_value>[A-Za-z_][A-Za-z0-9_]*)"\)\n'
    r"in\n"
    r"    FilteredRows;\n\Z"
)
_PAIR_WORKBOOK_EXTENSIONS = ("xlsx", "xlsm")


class FixtureValidationError(ValueError):
    """One or more fixture assertions did not match the generated workbooks."""


def _load_truth(path: Path) -> dict[str, Any]:
    try:
        result = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise FixtureValidationError(f"{path}: cannot load truth manifest: {error}") from error
    if not isinstance(result, dict):
        raise FixtureValidationError(f"{path}: truth manifest must be a JSON object")
    return result


def _load_workbook(path: Path) -> Workbook:
    try:
        return load_workbook(path, data_only=False, keep_links=True)
    except Exception as error:  # Openpyxl has several parse-specific exception types.
        raise FixtureValidationError(f"{path}: cannot load workbook: {error}") from error


def _cell_kind(cell: Any) -> str:
    if cell.value is None:
        return "blank"
    if cell.data_type == "f":
        return "formula"
    return "value"


def _dynamic_reference_functions(formula: str) -> tuple[str, ...]:
    """Return dynamic-reference functions in first-appearance order.

    This is deliberately a narrow fixture validator, not a general formula
    parser.  It only establishes the direct, generated-workbook invariant that
    a declared dynamic-reference boundary is present in the candidate.
    """

    return tuple(
        dict.fromkeys(
            match.group("function").upper()
            for match in _DYNAMIC_REFERENCE_FUNCTION.finditer(formula)
        )
    )


def _defined_name_text(workbook: Workbook, name: str) -> str | None:
    definition = workbook.defined_names.get(name)
    return None if definition is None else definition.attr_text


def _raw_external_data_connection_state(
    path: Path,
    connection_id: int,
) -> dict[str, Any] | None:
    """Read WCAB's compact relationship-backed web connection locally.

    This accepts only the generated one-connection package shape. It retains a
    stored URL as text but never resolves, opens, fetches, authenticates to,
    trusts, refreshes, or otherwise interacts with it.
    """

    try:
        with ZipFile(path) as archive:
            relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            connection_relationships = [
                relationship
                for relationship in relationships.findall(relationship_tag)
                if relationship.get("Type") == _EXTERNAL_DATA_CONNECTIONS_RELATIONSHIP
                and relationship.get("Target") == "connections.xml"
                and relationship.get("TargetMode") is None
            ]
            if len(connection_relationships) != 1:
                return None
            connections = ElementTree.fromstring(archive.read("xl/connections.xml"))
            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None

    connection_tag = f"{{{_SPREADSHEETML_NS}}}connection"
    web_properties_tag = f"{{{_SPREADSHEETML_NS}}}webPr"
    if (
        connections.tag != f"{{{_SPREADSHEETML_NS}}}connections"
        or connections.attrib
        or len(connections) != 1
        or any(connection.tag != connection_tag for connection in connections)
    ):
        return None
    matches = [
        connection
        for connection in connections.findall(connection_tag)
        if connection.get("id") == str(connection_id)
    ]
    if len(matches) != 1:
        return None
    connection = matches[0]
    web_properties = connection.findall(web_properties_tag)
    if len(connection) != 1 or len(web_properties) != 1 or len(web_properties[0]):
        return None
    url = web_properties[0].get("url")
    if not isinstance(url, str) or not url:
        return None
    override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
    connection_overrides = [
        override
        for override in content_types.findall(override_tag)
        if override.get("PartName") == "/xl/connections.xml"
    ]
    if len(connection_overrides) != 1 or set(connection_overrides[0].attrib) != {
        "PartName",
        "ContentType",
    }:
        return None
    return {
        "workbook_relationships_member": "xl/_rels/workbook.xml.rels",
        "connection_member": "xl/connections.xml",
        "relationship_attributes": tuple(sorted(connection_relationships[0].attrib.items())),
        "connection_attributes": tuple(sorted(connection.attrib.items())),
        "web_properties_attributes": tuple(sorted(web_properties[0].attrib.items())),
        "connection_content_type_attributes": tuple(sorted(connection_overrides[0].attrib.items())),
        "url": url,
    }


def _external_data_connection_refresh_on_load(path: Path, connection_id: int) -> bool | None:
    """Read the generated connection's explicit refresh-on-open flag."""

    state = _raw_external_data_connection_state(path, connection_id)
    if state is None:
        return None
    attributes = dict(state["connection_attributes"])
    return {"0": False, "1": True}.get(attributes.get("refreshOnLoad"))


def _external_data_connection_without_web_query_url(path: Path, connection_id: int) -> bytes | None:
    """Return one generated Connections part with only its web URL erased."""

    if _raw_external_data_connection_state(path, connection_id) is None:
        return None
    try:
        with ZipFile(path) as archive:
            connections = ElementTree.fromstring(archive.read("xl/connections.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    connection_tag = f"{{{_SPREADSHEETML_NS}}}connection"
    web_properties_tag = f"{{{_SPREADSHEETML_NS}}}webPr"
    matches = [
        connection
        for connection in connections.findall(connection_tag)
        if connection.get("id") == str(connection_id)
    ]
    if len(matches) != 1:
        return None
    web_properties = matches[0].findall(web_properties_tag)
    if len(web_properties) != 1 or "url" not in web_properties[0].attrib:
        return None
    web_properties[0].attrib.pop("url")
    return ElementTree.tostring(connections, encoding="utf-8", xml_declaration=True)


def _raw_query_table_refresh_state(path: Path, sheet_name: str) -> dict[str, Any] | None:
    """Read WCAB's compact QueryTable and connection relationship graph.

    This accepts only the generated one-QueryTable package shape. It follows
    local OOXML parts without opening a connection, fetching its stored URL,
    refreshing a query, materializing rows, or calculating a workbook.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
            worksheet_directory, worksheet_filename = worksheet_member.rsplit("/", maxsplit=1)
            worksheet_relationships_member = (
                f"{worksheet_directory}/_rels/{worksheet_filename}.rels"
            )
            worksheet_relationships = ElementTree.fromstring(
                archive.read(worksheet_relationships_member)
            )
            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            query_relationships = [
                relationship
                for relationship in worksheet_relationships.findall(relationship_tag)
                if relationship.get("Type") == _QUERY_TABLE_RELATIONSHIP
                and relationship.get("TargetMode") is None
            ]
            if len(query_relationships) != 1 or len(worksheet_relationships) != 1:
                return None
            query_relationship_id = query_relationships[0].get("Id")
            if not isinstance(query_relationship_id, str):
                return None
            query_table_member = _relationship_part_member(
                worksheet_relationships,
                query_relationship_id,
                worksheet_member,
                relationship_type=_QUERY_TABLE_RELATIONSHIP,
            )
            if query_table_member is None:
                return None
            query_table = ElementTree.fromstring(archive.read(query_table_member))

            workbook_relationships = ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )
            connection_relationships = [
                relationship
                for relationship in workbook_relationships.findall(relationship_tag)
                if relationship.get("Type") == _QUERY_TABLE_CONNECTIONS_RELATIONSHIP
                and relationship.get("TargetMode") is None
            ]
            if len(connection_relationships) != 1:
                return None
            connection_relationship_id = connection_relationships[0].get("Id")
            if not isinstance(connection_relationship_id, str):
                return None
            connection_member = _relationship_part_member(
                workbook_relationships,
                connection_relationship_id,
                "xl/workbook.xml",
                relationship_type=_QUERY_TABLE_CONNECTIONS_RELATIONSHIP,
            )
            if connection_member is None:
                return None
            connections = ElementTree.fromstring(archive.read(connection_member))
            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError, ValueError):
        return None

    if (
        worksheet.tag != f"{{{_SPREADSHEETML_NS}}}worksheet"
        or query_table.tag != f"{{{_SPREADSHEETML_NS}}}queryTable"
        or len(query_table)
        or connections.tag != f"{{{_SPREADSHEETML_NS}}}connections"
        or connections.attrib
        or len(connections) != 1
    ):
        return None
    connection = connections[0]
    connection_tag = f"{{{_SPREADSHEETML_NS}}}connection"
    web_properties_tag = f"{{{_SPREADSHEETML_NS}}}webPr"
    web_properties = connection.findall(web_properties_tag)
    if (
        connection.tag != connection_tag
        or len(connection) != 1
        or len(web_properties) != 1
        or len(web_properties[0])
    ):
        return None
    refresh_on_load = _ooxml_boolean(query_table.get("refreshOnLoad"))
    if refresh_on_load is None:
        return None
    override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
    query_table_overrides = [
        override
        for override in content_types.findall(override_tag)
        if override.get("PartName") == f"/{query_table_member}"
    ]
    connection_overrides = [
        override
        for override in content_types.findall(override_tag)
        if override.get("PartName") == f"/{connection_member}"
    ]
    if len(query_table_overrides) != 1 or len(connection_overrides) != 1:
        return None
    return {
        "worksheet_member": worksheet_member,
        "worksheet_relationships_member": worksheet_relationships_member,
        "worksheet_relationship_attributes": tuple(sorted(query_relationships[0].attrib.items())),
        "query_table_member": query_table_member,
        "query_table_attributes": tuple(sorted(query_table.attrib.items())),
        "connection_relationship_attributes": tuple(
            sorted(connection_relationships[0].attrib.items())
        ),
        "connection_member": connection_member,
        "connection_attributes": tuple(sorted(connection.attrib.items())),
        "web_properties_attributes": tuple(sorted(web_properties[0].attrib.items())),
        "query_table_content_type_attributes": tuple(
            sorted(query_table_overrides[0].attrib.items())
        ),
        "connection_content_type_attributes": tuple(sorted(connection_overrides[0].attrib.items())),
        "refresh_on_load": refresh_on_load,
    }


def _query_table_without_refresh_on_load(path: Path, sheet_name: str) -> bytes | None:
    """Return WCAB's query-table part with only its refresh-on-load flag erased."""

    state = _raw_query_table_refresh_state(path, sheet_name)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            query_table = ElementTree.fromstring(archive.read(state["query_table_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    if "refreshOnLoad" not in query_table.attrib:
        return None
    query_table.attrib.pop("refreshOnLoad")
    return ElementTree.tostring(query_table, encoding="utf-8", xml_declaration=True)


def _raw_cell_hyperlink_target_state(
    path: Path, sheet_name: str, coordinate: str
) -> dict[str, Any] | None:
    """Read WCAB's one relationship-backed external cell hyperlink.

    This accepts only the compact generated package shape: one worksheet
    ``hyperlink`` declaration with a relationship ID and one external hyperlink
    relationship. It reads neither the target nor any external content.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
            worksheet_directory, worksheet_filename = worksheet_member.rsplit("/", maxsplit=1)
            worksheet_relationships_member = (
                f"{worksheet_directory}/_rels/{worksheet_filename}.rels"
            )
            relationships = ElementTree.fromstring(archive.read(worksheet_relationships_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError, ValueError):
        return None

    hyperlink_sets = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}hyperlinks")
    hyperlink_tag = f"{{{_SPREADSHEETML_NS}}}hyperlink"
    relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
    if (
        worksheet.tag != f"{{{_SPREADSHEETML_NS}}}worksheet"
        or len(hyperlink_sets) != 1
        or hyperlink_sets[0].attrib
        or len(hyperlink_sets[0]) != 1
        or hyperlink_sets[0][0].tag != hyperlink_tag
        or set(hyperlink_sets[0][0].attrib) != {"ref", relationship_id_attribute}
        or hyperlink_sets[0][0].get("ref") != coordinate
    ):
        return None
    hyperlink = hyperlink_sets[0][0]
    relationship_id = hyperlink.get(relationship_id_attribute)
    if not isinstance(relationship_id, str):
        return None

    relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
    if (
        relationships.tag != f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        or relationships.attrib
        or len(relationships) != 1
        or relationships[0].tag != relationship_tag
        or len(relationships[0])
        or set(relationships[0].attrib) != {"Id", "Type", "Target", "TargetMode"}
        or relationships[0].get("Id") != relationship_id
        or relationships[0].get("Type") != _CELL_HYPERLINK_RELATIONSHIP
        or relationships[0].get("TargetMode") != "External"
    ):
        return None
    target = relationships[0].get("Target")
    if not isinstance(target, str) or not target:
        return None
    return {
        "worksheet_member": worksheet_member,
        "worksheet_relationships_member": worksheet_relationships_member,
        "hyperlink_attributes": tuple(sorted(hyperlink.attrib.items())),
        "relationship_id": relationship_id,
        "relationship_attributes": tuple(sorted(relationships[0].attrib.items())),
        "target_mode": "External",
        "target": target,
    }


def _cell_hyperlink_relationship_without_target(
    path: Path, sheet_name: str, coordinate: str
) -> bytes | None:
    """Return WCAB's hyperlink relationship part with only ``Target`` erased."""

    state = _raw_cell_hyperlink_target_state(path, sheet_name, coordinate)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            relationships = ElementTree.fromstring(
                archive.read(state["worksheet_relationships_member"])
            )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
    matches = [
        relationship
        for relationship in relationships.findall(relationship_tag)
        if relationship.get("Id") == state["relationship_id"]
        and relationship.get("Type") == _CELL_HYPERLINK_RELATIONSHIP
    ]
    if len(matches) != 1 or "Target" not in matches[0].attrib:
        return None
    matches[0].attrib.pop("Target")
    return ElementTree.tostring(relationships, encoding="utf-8", xml_declaration=True)


def _raw_external_workbook_link_source_state(path: Path) -> dict[str, Any] | None:
    """Read WCAB's one compact external-workbook link package locally.

    The validator follows only the two in-package relationships required to
    reach the ``externalLink`` XML and its relationships part. The final path
    is retained as stored text only: it is never resolved, opened, fetched,
    authenticated, trusted, refreshed, or otherwise interacted with.
    """

    workbook_member = "xl/workbook.xml"
    workbook_relationships_member = "xl/_rels/workbook.xml.rels"
    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read(workbook_member))
            workbook_relationships = ElementTree.fromstring(
                archive.read(workbook_relationships_member)
            )
            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))

            external_references_tag = f"{{{_SPREADSHEETML_NS}}}externalReferences"
            external_reference_tag = f"{{{_SPREADSHEETML_NS}}}externalReference"
            relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
            reference_sets = workbook.findall(external_references_tag)
            if (
                workbook.tag != f"{{{_SPREADSHEETML_NS}}}workbook"
                or len(reference_sets) != 1
                or reference_sets[0].attrib
                or len(reference_sets[0]) != 1
                or reference_sets[0][0].tag != external_reference_tag
                or set(reference_sets[0][0].attrib) != {relationship_id_attribute}
            ):
                return None
            workbook_relationship_id = reference_sets[0][0].get(relationship_id_attribute)
            if not isinstance(workbook_relationship_id, str):
                return None

            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            workbook_link_relationships = [
                relationship
                for relationship in workbook_relationships.findall(relationship_tag)
                if relationship.get("Type") == _EXTERNAL_WORKBOOK_LINK_RELATIONSHIP
            ]
            if (
                len(workbook_link_relationships) != 1
                or workbook_link_relationships[0].get("Id") != workbook_relationship_id
                or set(workbook_link_relationships[0].attrib) != {"Id", "Type", "Target"}
            ):
                return None
            external_link_member = _relationship_part_member(
                workbook_relationships,
                workbook_relationship_id,
                workbook_member,
                relationship_type=_EXTERNAL_WORKBOOK_LINK_RELATIONSHIP,
            )
            if external_link_member is None:
                return None
            external_link = ElementTree.fromstring(archive.read(external_link_member))
            external_link_directory, external_link_filename = external_link_member.rsplit(
                "/", maxsplit=1
            )
            external_link_relationships_member = (
                f"{external_link_directory}/_rels/{external_link_filename}.rels"
            )
            external_link_relationships = ElementTree.fromstring(
                archive.read(external_link_relationships_member)
            )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError, ValueError):
        return None

    external_link_tag = f"{{{_SPREADSHEETML_NS}}}externalLink"
    external_book_tag = f"{{{_SPREADSHEETML_NS}}}externalBook"
    sheet_names_tag = f"{{{_SPREADSHEETML_NS}}}sheetNames"
    sheet_name_tag = f"{{{_SPREADSHEETML_NS}}}sheetName"
    if (
        external_link.tag != external_link_tag
        or external_link.attrib
        or len(external_link) != 1
        or external_link[0].tag != external_book_tag
        or set(external_link[0].attrib) != {relationship_id_attribute}
        or len(external_link[0]) != 1
        or external_link[0][0].tag != sheet_names_tag
        or external_link[0][0].attrib
        or len(external_link[0][0]) != 1
        or external_link[0][0][0].tag != sheet_name_tag
        or set(external_link[0][0][0].attrib) != {"val"}
    ):
        return None
    external_link_relationship_id = external_link[0].get(relationship_id_attribute)
    external_sheet = external_link[0][0][0].get("val")
    if (
        not isinstance(external_link_relationship_id, str)
        or not isinstance(external_sheet, str)
        or not external_sheet
    ):
        return None

    if (
        external_link_relationships.tag != f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        or external_link_relationships.attrib
        or len(external_link_relationships) != 1
        or external_link_relationships[0].tag != relationship_tag
        or len(external_link_relationships[0])
        or set(external_link_relationships[0].attrib) != {"Id", "Type", "Target", "TargetMode"}
        or external_link_relationships[0].get("Id") != external_link_relationship_id
        or external_link_relationships[0].get("Type") != _EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP
        or external_link_relationships[0].get("TargetMode") != "External"
    ):
        return None
    target = external_link_relationships[0].get("Target")
    if not isinstance(target, str) or not target:
        return None

    override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
    content_type_overrides = [
        override
        for override in content_types.findall(override_tag)
        if override.get("PartName") == f"/{external_link_member}"
    ]
    if len(content_type_overrides) != 1 or set(content_type_overrides[0].attrib) != {
        "PartName",
        "ContentType",
    }:
        return None
    return {
        "workbook_member": workbook_member,
        "workbook_relationships_member": workbook_relationships_member,
        "workbook_relationship_id": workbook_relationship_id,
        "workbook_relationship_attributes": tuple(
            sorted(workbook_link_relationships[0].attrib.items())
        ),
        "external_link_member": external_link_member,
        "external_link_relationships_member": external_link_relationships_member,
        "external_book_attributes": tuple(sorted(external_link[0].attrib.items())),
        "external_sheet": external_sheet,
        "external_link_relationship_id": external_link_relationship_id,
        "external_link_relationship_attributes": tuple(
            sorted(external_link_relationships[0].attrib.items())
        ),
        "content_type_attributes": tuple(sorted(content_type_overrides[0].attrib.items())),
        "target": target,
    }


def _external_workbook_link_relationship_without_target(path: Path) -> bytes | None:
    """Return WCAB's external-link relationship part with only ``Target`` erased."""

    state = _raw_external_workbook_link_source_state(path)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            relationships = ElementTree.fromstring(
                archive.read(state["external_link_relationships_member"])
            )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
    matches = [
        relationship
        for relationship in relationships.findall(relationship_tag)
        if relationship.get("Id") == state["external_link_relationship_id"]
        and relationship.get("Type") == _EXTERNAL_WORKBOOK_LINK_PATH_RELATIONSHIP
    ]
    if len(matches) != 1 or "Target" not in matches[0].attrib:
        return None
    matches[0].attrib.pop("Target")
    return ElementTree.tostring(relationships, encoding="utf-8", xml_declaration=True)


def _workbook_properties(path: Path) -> dict[str, str] | None:
    """Read the stored ``workbookPr`` attributes without loading link targets."""

    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    properties = workbook.find(f"{{{_SPREADSHEETML_NS}}}workbookPr")
    return None if properties is None else dict(properties.attrib)


def _calculation_properties(path: Path) -> dict[str, str] | None:
    """Read stored ``calcPr`` attributes without evaluating workbook formulas."""

    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    properties = workbook.find(f"{{{_SPREADSHEETML_NS}}}calcPr")
    return None if properties is None else dict(properties.attrib)


def _ooxml_boolean(value: str | None) -> bool | None:
    """Parse the strict boolean values used by WCAB's raw OOXML contracts."""

    return {"0": False, "1": True, "false": False, "true": True}.get(value)


def _calculation_iteration_state(path: Path) -> tuple[bool, int, Decimal] | None:
    """Read WCAB's explicit stored iteration switch and shared bounds."""

    properties = _calculation_properties(path)
    if properties is None:
        return None
    iterate = _ooxml_boolean(properties.get("iterate"))
    try:
        count = int(properties["iterateCount"])
        delta = Decimal(properties["iterateDelta"])
    except (KeyError, InvalidOperation, ValueError):
        return None
    return None if iterate is None else (iterate, count, delta)


def _calculation_full_precision(path: Path) -> bool | None:
    """Read the explicit stored full-precision calculation control."""

    properties = _calculation_properties(path)
    return None if properties is None else _ooxml_boolean(properties.get("fullPrecision"))


def _relationship_target(
    relationships: ElementTree.Element,
    relationship_id: str,
    *,
    relationship_type: str | None = None,
) -> str | None:
    """Return one generated-package relationship target without following it."""

    relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
    matches = [
        relationship
        for relationship in relationships.findall(relationship_tag)
        if relationship.get("Id") == relationship_id
        and (relationship_type is None or relationship.get("Type") == relationship_type)
    ]
    if len(matches) != 1:
        return None
    target = matches[0].get("Target")
    if not isinstance(target, str) or not target or target.startswith("../"):
        return None
    return target.lstrip("/")


def _relationship_part_member(
    relationships: ElementTree.Element,
    relationship_id: str,
    source_member: str,
    *,
    relationship_type: str,
) -> str | None:
    """Resolve one local relationship to a safe package-member name.

    PivotTable relationships conventionally use ``../`` targets, unlike the
    workbook-to-worksheet links handled by :func:`_relationship_target`. This
    helper follows only an in-package relative target and rejects external or
    escaping relationships before opening a part.
    """

    if not source_member.startswith("xl/"):
        return None
    relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
    matches = [
        relationship
        for relationship in relationships.findall(relationship_tag)
        if relationship.get("Id") == relationship_id
        and relationship.get("Type") == relationship_type
        and relationship.get("TargetMode") != "External"
    ]
    if len(matches) != 1:
        return None
    target = matches[0].get("Target")
    if not isinstance(target, str) or not target or target.startswith("//"):
        return None
    if target.startswith("/"):
        member = target.lstrip("/")
    else:
        member = posixpath.normpath(f"{posixpath.dirname(source_member)}/{target}")
    if (
        not member.startswith("xl/")
        or member.startswith("../")
        or any(segment in {"", ".", ".."} for segment in member.split("/"))
    ):
        return None
    return member


def _worksheet_member_for_sheet(archive: ZipFile, sheet_name: str) -> str | None:
    """Resolve one generated worksheet name to its local OOXML part.

    This follows only the workbook-local relationship needed by WCAB's raw
    fixture contracts. It rejects traversal targets and never opens a linked
    workbook or evaluates a formula.
    """

    try:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        sheets = workbook.find(f"{{{_SPREADSHEETML_NS}}}sheets")
        if sheets is None:
            return None
        sheet_tag = f"{{{_SPREADSHEETML_NS}}}sheet"
        matches = [sheet for sheet in sheets.findall(sheet_tag) if sheet.get("name") == sheet_name]
        if len(matches) != 1:
            return None
        relationship_id = matches[0].get(f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id")
        if not isinstance(relationship_id, str):
            return None
        relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target = _relationship_target(
            relationships,
            relationship_id,
            relationship_type=f"{_DOCUMENT_RELATIONSHIPS_NS}/worksheet",
        )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    if target is None:
        return None
    return target if target.startswith("xl/") else f"xl/{target}"


def _raw_pivot_cache_refresh_state(path: Path, pivot_sheet: str) -> dict[str, Any] | None:
    """Read WCAB's compact relationship-backed PivotCache/PivotTable graph.

    This is intentionally a narrow reader for the generated fixture. It
    establishes the stored open-time control and the local cache-to-PivotTable
    binding without opening a client, refreshing data, or interpreting a
    PivotTable display value.
    """

    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            workbook_relationships = ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )
            pivot_caches_tag = f"{{{_SPREADSHEETML_NS}}}pivotCaches"
            pivot_cache_tag = f"{{{_SPREADSHEETML_NS}}}pivotCache"
            pivot_cache_sets = workbook.findall(pivot_caches_tag)
            if len(pivot_cache_sets) != 1:
                return None
            pivot_caches = pivot_cache_sets[0]
            if len(pivot_caches) != 1 or pivot_caches[0].tag != pivot_cache_tag:
                return None
            pivot_cache = pivot_caches[0]
            try:
                cache_id = int(pivot_cache.get("cacheId", ""))
            except ValueError:
                return None
            cache_relationship_id = pivot_cache.get(f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id")
            if not isinstance(cache_relationship_id, str):
                return None
            definition_member = _relationship_part_member(
                workbook_relationships,
                cache_relationship_id,
                "xl/workbook.xml",
                relationship_type=f"{_DOCUMENT_RELATIONSHIPS_NS}/pivotCacheDefinition",
            )
            if definition_member is None:
                return None
            cache_definition = ElementTree.fromstring(archive.read(definition_member))
            cache_records_relationship_id = cache_definition.get(
                f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
            )
            if not isinstance(cache_records_relationship_id, str):
                return None
            definition_directory, definition_filename = definition_member.rsplit("/", maxsplit=1)
            cache_definition_relationships = ElementTree.fromstring(
                archive.read(f"{definition_directory}/_rels/{definition_filename}.rels")
            )
            cache_records_member = _relationship_part_member(
                cache_definition_relationships,
                cache_records_relationship_id,
                definition_member,
                relationship_type=f"{_DOCUMENT_RELATIONSHIPS_NS}/pivotCacheRecords",
            )
            if cache_records_member is None:
                return None
            cache_records = ElementTree.fromstring(archive.read(cache_records_member))

            report_member = _worksheet_member_for_sheet(archive, pivot_sheet)
            if report_member is None:
                return None
            report = ElementTree.fromstring(archive.read(report_member))
            report_directory, report_filename = report_member.rsplit("/", maxsplit=1)
            report_relationships = ElementTree.fromstring(
                archive.read(f"{report_directory}/_rels/{report_filename}.rels")
            )
            pivot_table_parts_tag = f"{{{_SPREADSHEETML_NS}}}pivotTableParts"
            pivot_table_part_tag = f"{{{_SPREADSHEETML_NS}}}pivotTablePart"
            pivot_table_parts = report.findall(pivot_table_parts_tag)
            if len(pivot_table_parts) != 1 or len(pivot_table_parts[0]) != 1:
                return None
            pivot_table_part = pivot_table_parts[0][0]
            if pivot_table_part.tag != pivot_table_part_tag:
                return None
            pivot_table_relationship_id = pivot_table_part.get(
                f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
            )
            if not isinstance(pivot_table_relationship_id, str):
                return None
            pivot_table_member = _relationship_part_member(
                report_relationships,
                pivot_table_relationship_id,
                report_member,
                relationship_type=f"{_DOCUMENT_RELATIONSHIPS_NS}/pivotTable",
            )
            if pivot_table_member is None:
                return None
            pivot_table = ElementTree.fromstring(archive.read(pivot_table_member))
            pivot_directory, pivot_filename = pivot_table_member.rsplit("/", maxsplit=1)
            pivot_relationships = ElementTree.fromstring(
                archive.read(f"{pivot_directory}/_rels/{pivot_filename}.rels")
            )
            bound_definition_member = _relationship_part_member(
                pivot_relationships,
                "rIdWCABPivotCache",
                pivot_table_member,
                relationship_type=f"{_DOCUMENT_RELATIONSHIPS_NS}/pivotCacheDefinition",
            )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None

    cache_definition_tag = f"{{{_SPREADSHEETML_NS}}}pivotCacheDefinition"
    cache_source_tag = f"{{{_SPREADSHEETML_NS}}}cacheSource"
    worksheet_source_tag = f"{{{_SPREADSHEETML_NS}}}worksheetSource"
    cache_records_tag = f"{{{_SPREADSHEETML_NS}}}pivotCacheRecords"
    record_tag = f"{{{_SPREADSHEETML_NS}}}r"
    shared_item_index_tag = f"{{{_SPREADSHEETML_NS}}}x"
    pivot_table_tag = f"{{{_SPREADSHEETML_NS}}}pivotTableDefinition"
    location_tag = f"{{{_SPREADSHEETML_NS}}}location"
    if (
        cache_definition.tag != cache_definition_tag
        or cache_records.tag != cache_records_tag
        or pivot_table.tag != pivot_table_tag
        or bound_definition_member != definition_member
    ):
        return None
    cache_sources = cache_definition.findall(cache_source_tag)
    if len(cache_sources) != 1 or len(cache_sources[0]) != 1:
        return None
    cache_source = cache_sources[0]
    worksheet_source = cache_source[0]
    if worksheet_source.tag != worksheet_source_tag:
        return None
    locations = pivot_table.findall(location_tag)
    if len(locations) != 1:
        return None
    records = cache_records.findall(record_tag)
    try:
        declared_record_count = int(cache_records.get("count", ""))
        record_indexes = tuple(
            tuple(int(value.get("v", "")) for value in record) for record in records
        )
    except (TypeError, ValueError):
        return None
    if declared_record_count != len(records) or any(
        len(record) != 2 or any(value.tag != shared_item_index_tag for value in record)
        for record in records
    ):
        return None
    return {
        "cache_id": cache_id,
        "definition_member": definition_member,
        "definition_attributes": tuple(sorted(cache_definition.attrib.items())),
        "cache_source_attributes": tuple(sorted(cache_source.attrib.items())),
        "worksheet_source_attributes": tuple(sorted(worksheet_source.attrib.items())),
        "cache_records_member": cache_records_member,
        "cache_records_attributes": tuple(sorted(cache_records.attrib.items())),
        "cache_record_indexes": record_indexes,
        "report_member": report_member,
        "pivot_table_member": pivot_table_member,
        "pivot_table_attributes": tuple(sorted(pivot_table.attrib.items())),
        "location_attributes": tuple(sorted(locations[0].attrib.items())),
    }


def _pivot_cache_definition_without_refresh_on_load(path: Path, pivot_sheet: str) -> bytes | None:
    """Return WCAB's cache definition with only its open-time flag removed."""

    state = _raw_pivot_cache_refresh_state(path, pivot_sheet)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            definition = ElementTree.fromstring(archive.read(state["definition_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    definition.attrib.pop("refreshOnLoad", None)
    return ElementTree.tostring(definition, encoding="utf-8", xml_declaration=True)


def _raw_pivot_data_field_aggregation_state(path: Path, pivot_sheet: str) -> dict[str, Any] | None:
    """Read the sole generated PivotTable value-field aggregation declaration.

    The shared PivotCache reader establishes the relationship-backed graph.
    This narrower layer then reads exactly one ``dataFields/dataField`` entry;
    it never recalculates, refreshes, or renders the PivotTable.
    """

    state = _raw_pivot_cache_refresh_state(path, pivot_sheet)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            pivot_table = ElementTree.fromstring(archive.read(state["pivot_table_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    data_fields_tag = f"{{{_SPREADSHEETML_NS}}}dataFields"
    data_field_tag = f"{{{_SPREADSHEETML_NS}}}dataField"
    data_fields = pivot_table.findall(data_fields_tag)
    if (
        len(data_fields) != 1
        or data_fields[0].get("count") != "1"
        or len(data_fields[0]) != 1
        or data_fields[0][0].tag != data_field_tag
    ):
        return None
    data_field = data_fields[0][0]
    try:
        source_index = int(data_field.get("fld", ""))
    except ValueError:
        return None
    subtotal = data_field.get("subtotal")
    if not isinstance(subtotal, str):
        return None
    return {
        **state,
        "data_field_source_index": source_index,
        "data_field_attributes": tuple(sorted(data_field.attrib.items())),
    }


def _pivot_table_without_data_field_subtotal(path: Path, pivot_sheet: str) -> bytes | None:
    """Return the generated PivotTable definition without its one subtotal attr."""

    state = _raw_pivot_data_field_aggregation_state(path, pivot_sheet)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            pivot_table = ElementTree.fromstring(archive.read(state["pivot_table_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    data_fields = pivot_table.findall(f"{{{_SPREADSHEETML_NS}}}dataFields")
    if len(data_fields) != 1 or len(data_fields[0]) != 1:
        return None
    data_fields[0][0].attrib.pop("subtotal", None)
    return ElementTree.tostring(pivot_table, encoding="utf-8", xml_declaration=True)


def _raw_pivot_slicer_selection_state(path: Path, pivot_sheet: str) -> dict[str, Any] | None:
    """Read WCAB's compact local PivotTable Slicer-cache selection graph.

    The reader establishes only a stored relationship-backed declaration. It
    does not create a Slicer view, apply a filter, refresh a cache, calculate a
    PivotTable, or infer any displayed report result.
    """

    state = _raw_pivot_cache_refresh_state(path, pivot_sheet)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            workbook_relationships = ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )
            sheets = workbook.find(f"{{{_SPREADSHEETML_NS}}}sheets")
            report_sheets = (
                [
                    sheet
                    for sheet in sheets.findall(f"{{{_SPREADSHEETML_NS}}}sheet")
                    if sheet.get("name") == pivot_sheet
                ]
                if sheets is not None
                else []
            )
            if len(report_sheets) != 1:
                return None
            report_sheet_id = report_sheets[0].get("sheetId")
            try:
                if not isinstance(report_sheet_id, str) or int(report_sheet_id) <= 0:
                    return None
            except ValueError:
                return None

            workbook_extension_lists = workbook.findall(f"{{{_SPREADSHEETML_NS}}}extLst")
            if len(workbook_extension_lists) != 1:
                return None
            workbook_extensions = workbook_extension_lists[0]
            slicer_extensions = [
                extension
                for extension in workbook_extensions.findall(f"{{{_SPREADSHEETML_NS}}}ext")
                if extension.get("uri") == "{BBE1A952-AA13-448E-AADC-164F8A28A991}"
            ]
            if len(slicer_extensions) != 1:
                return None
            slicer_cache_sets = slicer_extensions[0].findall(
                f"{{{_OFFICE_2010_SPREADSHEET_NS}}}slicerCaches"
            )
            if len(slicer_cache_sets) != 1 or len(slicer_cache_sets[0]) != 1:
                return None
            slicer_declaration = slicer_cache_sets[0][0]
            if slicer_declaration.tag != f"{{{_OFFICE_2010_SPREADSHEET_NS}}}slicerCache":
                return None
            slicer_relationship_id = slicer_declaration.get(f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id")
            if not isinstance(slicer_relationship_id, str):
                return None
            slicer_member = _relationship_part_member(
                workbook_relationships,
                slicer_relationship_id,
                "xl/workbook.xml",
                relationship_type=_SLICER_CACHE_RELATIONSHIP,
            )
            if slicer_member is None:
                return None
            slicer_cache = ElementTree.fromstring(archive.read(slicer_member))

            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
            slicer_content_types = [
                override
                for override in content_types.findall(f"{{{_CONTENT_TYPES_NS}}}Override")
                if override.get("PartName") == f"/{slicer_member}"
            ]
            if len(slicer_content_types) != 1:
                return None

            cache_definition = ElementTree.fromstring(archive.read(state["definition_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None

    if slicer_cache.tag != f"{{{_OFFICE_2010_SPREADSHEET_NS}}}slicerCacheDefinition":
        return None
    pivot_table_sets = slicer_cache.findall(f"{{{_OFFICE_2010_SPREADSHEET_NS}}}pivotTables")
    data_sets = slicer_cache.findall(f"{{{_OFFICE_2010_SPREADSHEET_NS}}}data")
    if (
        len(pivot_table_sets) != 1
        or pivot_table_sets[0].get("count") != "1"
        or len(pivot_table_sets[0]) != 1
        or pivot_table_sets[0][0].tag != f"{{{_OFFICE_2010_SPREADSHEET_NS}}}pivotTable"
        or len(data_sets) != 1
        or len(data_sets[0]) != 1
        or data_sets[0][0].tag != f"{{{_OFFICE_2010_SPREADSHEET_NS}}}tabular"
    ):
        return None
    slicer_pivot_table = pivot_table_sets[0][0]
    tabular = data_sets[0][0]
    item_sets = tabular.findall(f"{{{_OFFICE_2010_SPREADSHEET_NS}}}items")
    if len(item_sets) != 1:
        return None
    items = item_sets[0]
    slicer_items = list(items)
    if any(item.tag != f"{{{_OFFICE_2010_SPREADSHEET_NS}}}i" or len(item) for item in slicer_items):
        return None
    try:
        item_count = int(items.get("count", ""))
        item_indexes = tuple(int(item.get("x", "")) for item in slicer_items)
    except ValueError:
        return None
    selected_indexes = tuple(
        item_index
        for item_index, item in zip(item_indexes, slicer_items, strict=True)
        if _ooxml_boolean(item.get("s")) is True
    )
    if (
        item_count != len(slicer_items)
        or item_count <= 0
        or item_indexes != tuple(range(item_count))
        or len(selected_indexes) != 1
    ):
        return None

    cache_extensions = cache_definition.findall(f"{{{_SPREADSHEETML_NS}}}extLst")
    if len(cache_extensions) != 1:
        return None
    pivot_cache_extensions = [
        extension
        for extension in cache_extensions[0].findall(f"{{{_SPREADSHEETML_NS}}}ext")
        if extension.get("uri") == "{725AE2AE-9491-48BE-B2B4-4EB974FC3084}"
    ]
    if len(pivot_cache_extensions) != 1 or len(pivot_cache_extensions[0]) != 1:
        return None
    pivot_cache_extension = pivot_cache_extensions[0][0]
    if pivot_cache_extension.tag != f"{{{_OFFICE_2010_SPREADSHEET_NS}}}pivotCacheDefinition":
        return None

    cache_field_sets = cache_definition.findall(f"{{{_SPREADSHEETML_NS}}}cacheFields")
    if len(cache_field_sets) != 1:
        return None
    cache_fields = list(cache_field_sets[0])
    source_name = slicer_cache.get("sourceName")
    source_fields = [
        (index, field)
        for index, field in enumerate(cache_fields)
        if field.tag == f"{{{_SPREADSHEETML_NS}}}cacheField" and field.get("name") == source_name
    ]
    if len(source_fields) != 1:
        return None
    source_field_index, source_field = source_fields[0]
    shared_item_sets = source_field.findall(f"{{{_SPREADSHEETML_NS}}}sharedItems")
    if len(shared_item_sets) != 1:
        return None
    shared_items = list(shared_item_sets[0])
    shared_item_values = tuple(item.get("v") for item in shared_items)
    if (
        shared_item_sets[0].get("count") != str(len(shared_items))
        or any(item.tag != f"{{{_SPREADSHEETML_NS}}}s" for item in shared_items)
        or any(not isinstance(value, str) for value in shared_item_values)
        or item_count != len(shared_item_values)
    ):
        return None
    selected_item_index = selected_indexes[0]
    return {
        **state,
        "pivot_sheet_id": int(report_sheet_id),
        "slicer_member": slicer_member,
        "slicer_declaration_attributes": tuple(sorted(slicer_declaration.attrib.items())),
        "slicer_extension_attributes": tuple(sorted(slicer_extensions[0].attrib.items())),
        "slicer_content_type_attributes": tuple(sorted(slicer_content_types[0].attrib.items())),
        "slicer_attributes": tuple(sorted(slicer_cache.attrib.items())),
        "slicer_pivot_table_attributes": tuple(sorted(slicer_pivot_table.attrib.items())),
        "slicer_tabular_attributes": tuple(sorted(tabular.attrib.items())),
        "slicer_item_attributes": tuple(
            tuple(sorted(item.attrib.items())) for item in slicer_items
        ),
        "slicer_source_field_index": source_field_index,
        "slicer_source_shared_item_values": shared_item_values,
        "slicer_selected_item_index": selected_item_index,
        "slicer_selected_value": shared_item_values[selected_item_index],
        "pivot_cache_slicer_extension_attributes": tuple(
            sorted(pivot_cache_extension.attrib.items())
        ),
    }


def _slicer_cache_without_selection_state(path: Path, pivot_sheet: str) -> bytes | None:
    """Return WCAB's one Slicer cache with only selected-state attrs erased."""

    state = _raw_pivot_slicer_selection_state(path, pivot_sheet)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            slicer_cache = ElementTree.fromstring(archive.read(state["slicer_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    item_sets = slicer_cache.findall(
        f"{{{_OFFICE_2010_SPREADSHEET_NS}}}data/"
        f"{{{_OFFICE_2010_SPREADSHEET_NS}}}tabular/"
        f"{{{_OFFICE_2010_SPREADSHEET_NS}}}items"
    )
    if len(item_sets) != 1 or len(item_sets[0]) != 2:
        return None
    for item in item_sets[0]:
        item.attrib.pop("s", None)
    return ElementTree.tostring(slicer_cache, encoding="utf-8", xml_declaration=True)


def _power_query_length_prefixed_fields(payload: bytes) -> tuple[int, tuple[bytes, ...]] | None:
    """Read the four compact fields from WCAB's generated Data Mashup stream."""

    if len(payload) < 4:
        return None
    version = struct.unpack_from("<I", payload)[0]
    cursor = 4
    fields: list[bytes] = []
    for _ in range(4):
        if cursor + 4 > len(payload):
            return None
        field_size = struct.unpack_from("<I", payload, cursor)[0]
        cursor += 4
        if field_size > len(payload) - cursor:
            return None
        fields.append(payload[cursor : cursor + field_size])
        cursor += field_size
    return (version, tuple(fields)) if cursor == len(payload) else None


def _raw_power_query_m_filter_state(path: Path) -> dict[str, Any] | None:
    """Read WCAB's compact, connection-only local-table Data Mashup contract.

    This reader accepts only the small generated envelope, with bounded nested
    ZIP members and a deliberately narrow M formula pattern. It never executes
    M, refreshes a query, materializes output, or evaluates a workbook formula.
    """

    try:
        with ZipFile(path) as archive:
            relationships = ElementTree.fromstring(archive.read("_rels/.rels"))
            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            custom_xml_relationships = [
                relationship
                for relationship in relationships.findall(relationship_tag)
                if relationship.get("Id") == "rIdWCABPowerQuery"
                and relationship.get("Type") == _POWER_QUERY_CUSTOM_XML_RELATIONSHIP
                and relationship.get("Target") == "customXml/item1.xml"
                and relationship.get("TargetMode") != "External"
            ]
            if len(custom_xml_relationships) != 1:
                return None

            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
            xml_defaults = [
                item
                for item in content_types.findall(f"{{{_CONTENT_TYPES_NS}}}Default")
                if item.get("Extension") == "xml" and item.get("ContentType") == "application/xml"
            ]
            if len(xml_defaults) != 1:
                return None

            mashup_root = ElementTree.fromstring(archive.read("customXml/item1.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    if mashup_root.tag != f"{{{_DATA_MASHUP_NS}}}DataMashup" or mashup_root.attrib:
        return None
    encoded_stream = mashup_root.text
    if (
        not isinstance(encoded_stream, str)
        or not encoded_stream
        or len(encoded_stream.encode("ascii", "ignore")) > _POWER_QUERY_STREAM_MAX_ENCODED_BYTES
    ):
        return None
    try:
        decoded_stream = base64.b64decode(encoded_stream, validate=True)
    except (binascii.Error, ValueError):
        return None
    stream_fields = _power_query_length_prefixed_fields(decoded_stream)
    if stream_fields is None:
        return None
    stream_version, (package_payload, permissions_payload, metadata_payload, permission_binding) = (
        stream_fields
    )
    try:
        with ZipFile(io.BytesIO(package_payload)) as package:
            package_entries = package.infolist()
            expected_package_members = (
                "Config/Package.xml",
                "Formulas/Section1.m",
                "[Content_Types].xml",
            )
            if (
                tuple(sorted(entry.filename for entry in package_entries))
                != expected_package_members
                or any(
                    entry.file_size > _POWER_QUERY_PACKAGE_MAX_UNCOMPRESSED_BYTES
                    for entry in package_entries
                )
                or sum(entry.file_size for entry in package_entries)
                > _POWER_QUERY_PACKAGE_MAX_UNCOMPRESSED_BYTES
                or package.testzip() is not None
            ):
                return None
            package_content_types = package.read("[Content_Types].xml")
            package_configuration = package.read("Config/Package.xml")
            formula = package.read("Formulas/Section1.m").decode("utf-8")
    except (BadZipFile, KeyError, OSError, RuntimeError, UnicodeDecodeError):
        return None
    formula_match = _POWER_QUERY_M_FILTER.fullmatch(formula)
    if formula_match is None:
        return None
    if len(metadata_payload) < 12:
        return None
    metadata_version, metadata_xml_size = struct.unpack_from("<II", metadata_payload)
    metadata_xml_end = 8 + metadata_xml_size
    if metadata_xml_end + 4 > len(metadata_payload):
        return None
    metadata_content_size = struct.unpack_from("<I", metadata_payload, metadata_xml_end)[0]
    metadata_content_start = metadata_xml_end + 4
    if metadata_content_size != 0 or metadata_content_start != len(metadata_payload):
        return None
    try:
        metadata_root = ElementTree.fromstring(metadata_payload[8:metadata_xml_end])
        permissions_root = ElementTree.fromstring(permissions_payload)
    except ElementTree.ParseError:
        return None
    if (
        metadata_root.tag != f"{{{_DATA_MASHUP_NS}}}LocalPackageMetadataFile"
        or metadata_root.attrib
    ):
        return None
    item_sets = metadata_root.findall(f"{{{_DATA_MASHUP_NS}}}Items")
    if len(item_sets) != 1 or len(item_sets[0]) != 1:
        return None
    metadata_item = item_sets[0][0]
    if metadata_item.tag != f"{{{_DATA_MASHUP_NS}}}Item" or metadata_item.attrib:
        return None
    metadata_item_parts = list(metadata_item)
    if [part.tag for part in metadata_item_parts] != [
        f"{{{_DATA_MASHUP_NS}}}ItemLocation",
        f"{{{_DATA_MASHUP_NS}}}StableEntries",
    ]:
        return None
    item_location, stable_entries = metadata_item_parts
    if len(item_location) != 2 or len(stable_entries) != 1:
        return None
    item_type, item_path = list(item_location)
    stable_entry = stable_entries[0]
    if (
        item_type.tag != f"{{{_DATA_MASHUP_NS}}}ItemType"
        or item_type.text != "Formula"
        or item_path.tag != f"{{{_DATA_MASHUP_NS}}}ItemPath"
        or not isinstance(item_path.text, str)
        or stable_entry.tag != f"{{{_DATA_MASHUP_NS}}}Entry"
        or tuple(sorted(stable_entry.attrib.items())) != (("Type", "FillEnabled"), ("Value", "l0"))
        or len(stable_entry)
    ):
        return None
    if permissions_root.tag != "PermissionList" or permissions_root.attrib:
        return None
    permission_children = list(permissions_root)
    if any(child.attrib or len(child) for child in permission_children):
        return None
    permission_values = {child.tag: child.text for child in permission_children}
    if permission_values != {
        "CanEvaluateFuturePackages": "false",
        "FirewallEnabled": "true",
    }:
        return None
    return {
        "data_mashup_member": "customXml/item1.xml",
        "root_relationship_attributes": tuple(sorted(custom_xml_relationships[0].attrib.items())),
        "stream_version": stream_version,
        "package_members": tuple(sorted(entry.filename for entry in package_entries)),
        "package_content_types": package_content_types,
        "package_configuration": package_configuration,
        "formula": formula,
        "query_section": formula_match.group("section"),
        "query_name": formula_match.group("query"),
        "source_table": formula_match.group("source_table"),
        "filter_column": formula_match.group("filter_column"),
        "filter_value": formula_match.group("filter_value"),
        "metadata_version": metadata_version,
        "metadata_item_path": item_path.text,
        "fill_enabled": False,
        "firewall_enabled": True,
        "future_packages_allowed": False,
        "permission_binding": permission_binding,
    }


def _raw_scenario_manager_stored_input_state(path: Path, sheet_name: str) -> dict[str, Any] | None:
    """Read WCAB's compact raw Scenario Manager declaration.

    Scenario Manager records alternate input values inside a worksheet's
    ``scenarios`` element. This narrow reader accepts only WCAB's generated
    one-scenario/two-input shape and neither shows a scenario nor calculates a
    workbook.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    if worksheet.tag != f"{{{_SPREADSHEETML_NS}}}worksheet":
        return None
    scenarios_tag = f"{{{_SPREADSHEETML_NS}}}scenarios"
    scenario_tag = f"{{{_SPREADSHEETML_NS}}}scenario"
    input_cell_tag = f"{{{_SPREADSHEETML_NS}}}inputCells"
    scenario_sets = worksheet.findall(scenarios_tag)
    if len(scenario_sets) != 1 or len(scenario_sets[0]) != 1:
        return None
    scenarios = scenario_sets[0]
    scenario = scenarios[0]
    if scenario.tag != scenario_tag:
        return None
    input_cells = list(scenario)
    if len(input_cells) != 2 or any(
        cell.tag != input_cell_tag or len(cell) for cell in input_cells
    ):
        return None
    input_state: list[tuple[str, str, tuple[tuple[str, str], ...]]] = []
    for input_cell in input_cells:
        reference = input_cell.get("r")
        value = input_cell.get("val")
        if not isinstance(reference, str) or not isinstance(value, str):
            return None
        input_state.append((reference, value, tuple(sorted(input_cell.attrib.items()))))
    if len({reference for reference, _, _ in input_state}) != len(input_state):
        return None
    return {
        "worksheet_member": worksheet_member,
        "scenarios_attributes": tuple(sorted(scenarios.attrib.items())),
        "scenario_attributes": tuple(sorted(scenario.attrib.items())),
        "input_cells": tuple(input_state),
    }


def _scenario_manager_worksheet_without_stored_input_value(
    path: Path, sheet_name: str, changing_cell: str
) -> bytes | None:
    """Return WCAB's Scenario Manager worksheet with one stored value erased."""

    if _raw_scenario_manager_stored_input_state(path, sheet_name) is None:
        return None
    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    input_cell_tag = f"{{{_SPREADSHEETML_NS}}}inputCells"
    matches = [
        input_cell
        for input_cell in worksheet.iter(input_cell_tag)
        if input_cell.get("r") == changing_cell
    ]
    if len(matches) != 1 or "val" not in matches[0].attrib:
        return None
    matches[0].attrib.pop("val")
    return ElementTree.tostring(worksheet, encoding="utf-8", xml_declaration=True)


def _raw_data_validation_list_source_state(
    path: Path, sheet_name: str, target_range: str
) -> dict[str, Any] | None:
    """Read WCAB's one stored list-validation source declaration.

    A validation list's source is a ``formula1`` child on a worksheet-local
    ``dataValidation`` record. This narrow reader accepts only the generated
    one-container/one-rule shape and never evaluates the source expression or
    decides whether a future entry is valid.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    if worksheet.tag != f"{{{_SPREADSHEETML_NS}}}worksheet":
        return None
    data_validations_tag = f"{{{_SPREADSHEETML_NS}}}dataValidations"
    data_validation_tag = f"{{{_SPREADSHEETML_NS}}}dataValidation"
    formula1_tag = f"{{{_SPREADSHEETML_NS}}}formula1"
    containers = worksheet.findall(data_validations_tag)
    if len(containers) != 1:
        return None
    validations = list(containers[0])
    if len(validations) != 1 or validations[0].tag != data_validation_tag:
        return None
    validation = validations[0]
    formulas = validation.findall(formula1_tag)
    if validation.get("sqref") != target_range or len(validation) != 1 or len(formulas) != 1:
        return None
    formula = formulas[0]
    return {
        "worksheet_member": worksheet_member,
        "container_attributes": tuple(sorted(containers[0].attrib.items())),
        "validation_attributes": tuple(sorted(validation.attrib.items())),
        "formula1_attributes": tuple(sorted(formula.attrib.items())),
        "formula1_text": formula.text,
        "formula1_child_count": len(formula),
    }


def _data_validation_worksheet_without_source_formula(
    path: Path, sheet_name: str, target_range: str
) -> bytes | None:
    """Return WCAB's validation worksheet with its one ``formula1`` erased."""

    if _raw_data_validation_list_source_state(path, sheet_name, target_range) is None:
        return None
    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    data_validation_tag = f"{{{_SPREADSHEETML_NS}}}dataValidation"
    formula1_tag = f"{{{_SPREADSHEETML_NS}}}formula1"
    validations = [
        validation
        for validation in worksheet.iter(data_validation_tag)
        if validation.get("sqref") == target_range
    ]
    if len(validations) != 1:
        return None
    formulas = validations[0].findall(formula1_tag)
    if len(formulas) != 1:
        return None
    formulas[0].text = None
    return ElementTree.tostring(worksheet, encoding="utf-8", xml_declaration=True)


def _raw_conditional_formatting_threshold_state(
    path: Path, sheet_name: str, target_range: str
) -> dict[str, Any] | None:
    """Read WCAB's one stored ``cellIs`` exception-threshold declaration.

    This narrow reader only accepts the compact rule shape generated for the
    benchmark. It observes a stored threshold formula without evaluating the
    rule or determining which cells a spreadsheet client would format.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    if worksheet.tag != f"{{{_SPREADSHEETML_NS}}}worksheet":
        return None
    conditional_formatting_tag = f"{{{_SPREADSHEETML_NS}}}conditionalFormatting"
    rule_tag = f"{{{_SPREADSHEETML_NS}}}cfRule"
    formula_tag = f"{{{_SPREADSHEETML_NS}}}formula"
    controls = worksheet.findall(conditional_formatting_tag)
    if len(controls) != 1 or controls[0].get("sqref") != target_range:
        return None
    rules = list(controls[0])
    if len(rules) != 1 or rules[0].tag != rule_tag:
        return None
    rule = rules[0]
    formulas = rule.findall(formula_tag)
    if len(rule) != 1 or len(formulas) != 1:
        return None
    formula = formulas[0]
    return {
        "worksheet_member": worksheet_member,
        "control_attributes": tuple(sorted(controls[0].attrib.items())),
        "rule_attributes": tuple(sorted(rule.attrib.items())),
        "formula_attributes": tuple(sorted(formula.attrib.items())),
        "formula_text": formula.text,
        "formula_child_count": len(formula),
    }


def _conditional_formatting_worksheet_without_threshold_formula(
    path: Path, sheet_name: str, target_range: str
) -> bytes | None:
    """Return WCAB's conditional-formatting worksheet with its formula erased."""

    if _raw_conditional_formatting_threshold_state(path, sheet_name, target_range) is None:
        return None
    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    rule_tag = f"{{{_SPREADSHEETML_NS}}}cfRule"
    formula_tag = f"{{{_SPREADSHEETML_NS}}}formula"
    rules = [rule for rule in worksheet.iter(rule_tag) if rule.get("type") == "cellIs"]
    if len(rules) != 1:
        return None
    formulas = rules[0].findall(formula_tag)
    if len(formulas) != 1:
        return None
    formulas[0].text = None
    return ElementTree.tostring(worksheet, encoding="utf-8", xml_declaration=True)


def _conditional_formatting_differential_fill_state(
    path: Path, dxf_id: int
) -> tuple[str, str] | None:
    """Read one generated differential fill control without rendering it."""

    if dxf_id < 0:
        return None
    try:
        with ZipFile(path) as archive:
            styles = ElementTree.fromstring(archive.read("xl/styles.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    dxfs = styles.find(f"{{{_SPREADSHEETML_NS}}}dxfs")
    if dxfs is None:
        return None
    entries = list(dxfs)
    if dxf_id >= len(entries) or entries[dxf_id].tag != f"{{{_SPREADSHEETML_NS}}}dxf":
        return None
    fill = entries[dxf_id].find(f"{{{_SPREADSHEETML_NS}}}fill")
    pattern = fill.find(f"{{{_SPREADSHEETML_NS}}}patternFill") if fill is not None else None
    colour = pattern.find(f"{{{_SPREADSHEETML_NS}}}fgColor") if pattern is not None else None
    pattern_type = pattern.get("patternType") if pattern is not None else None
    colour_rgb = colour.get("rgb") if colour is not None else None
    if not isinstance(pattern_type, str) or not isinstance(colour_rgb, str):
        return None
    return (pattern_type, colour_rgb)


def _raw_what_if_data_table_state(
    path: Path, sheet_name: str, master_cell: str
) -> dict[str, Any] | None:
    """Read WCAB's one raw What-If Data Table master declaration.

    A Data Table is encoded as a ``dataTable`` formula on its top-left master
    cell. This narrow reader establishes only the small generated declaration;
    it does not substitute inputs, calculate the table, or interpret a saved
    table output.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    if worksheet.tag != f"{{{_SPREADSHEETML_NS}}}worksheet":
        return None
    cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
    formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
    value_tag = f"{{{_SPREADSHEETML_NS}}}v"
    data_table_masters = [
        (cell, formula)
        for cell in worksheet.iter(cell_tag)
        for formula in cell.findall(formula_tag)
        if formula.get("t") == "dataTable"
    ]
    if len(data_table_masters) != 1:
        return None
    cell, formula = data_table_masters[0]
    if cell.get("r") != master_cell or len(cell.findall(formula_tag)) != 1:
        return None
    values = cell.findall(value_tag)
    if len(values) != 1:
        return None
    return {
        "worksheet_member": worksheet_member,
        "master_cell_attributes": tuple(sorted(cell.attrib.items())),
        "formula_attributes": tuple(sorted(formula.attrib.items())),
        "formula_text": formula.text,
        "formula_child_count": len(formula),
        "master_value_text": values[0].text,
    }


def _what_if_data_table_worksheet_without_input_reference(
    path: Path, sheet_name: str, master_cell: str
) -> bytes | None:
    """Return WCAB's Data Table worksheet with its one ``r1`` value erased."""

    if _raw_what_if_data_table_state(path, sheet_name, master_cell) is None:
        return None
    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
    formula_tag = f"{{{_SPREADSHEETML_NS}}}f"
    masters = [cell for cell in worksheet.iter(cell_tag) if cell.get("r") == master_cell]
    if len(masters) != 1:
        return None
    formulas = [
        formula for formula in masters[0].findall(formula_tag) if formula.get("t") == "dataTable"
    ]
    if len(formulas) != 1 or "r1" not in formulas[0].attrib:
        return None
    formulas[0].attrib.pop("r1")
    return ElementTree.tostring(worksheet, encoding="utf-8", xml_declaration=True)


def _chart_anchor_cell(anchor: ElementTree.Element) -> str | None:
    """Return one zero-offset worksheet-drawing anchor as an A1 coordinate."""

    from_tag = f"{{{_DRAWINGML_SPREADSHEET_DRAWING_NS}}}from"
    anchor_from = anchor.find(from_tag)
    if anchor_from is None:
        return None
    values: dict[str, int] = {}
    for name in ("col", "colOff", "row", "rowOff"):
        element = anchor_from.find(f"{{{_DRAWINGML_SPREADSHEET_DRAWING_NS}}}{name}")
        try:
            values[name] = int(element.text) if element is not None and element.text else -1
        except ValueError:
            return None
    if values["col"] < 0 or values["row"] < 0 or values["colOff"] != 0 or values["rowOff"] != 0:
        return None
    column = values["col"] + 1
    label = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        label = chr(ord("A") + remainder) + label
    return f"{label}{values['row'] + 1}"


def _raw_chart_series_value_reference_state(path: Path, chart_sheet: str) -> dict[str, Any] | None:
    """Read WCAB's one worksheet-drawing chart and its stored references.

    This follows the generated worksheet -> drawing -> chart relationship chain
    only. It does not evaluate a chart formula, calculate source cells, or
    render a chart.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, chart_sheet)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
            drawing_references = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}drawing")
            if len(drawing_references) != 1:
                return None
            drawing_relationship_id = drawing_references[0].get(
                f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
            )
            if not isinstance(drawing_relationship_id, str):
                return None
            worksheet_directory, worksheet_filename = worksheet_member.rsplit("/", maxsplit=1)
            worksheet_relationships = ElementTree.fromstring(
                archive.read(f"{worksheet_directory}/_rels/{worksheet_filename}.rels")
            )
            drawing_member = _relationship_part_member(
                worksheet_relationships,
                drawing_relationship_id,
                worksheet_member,
                relationship_type=f"{_DOCUMENT_RELATIONSHIPS_NS}/drawing",
            )
            if drawing_member is None:
                return None
            drawing = ElementTree.fromstring(archive.read(drawing_member))
            anchors = drawing.findall(f"{{{_DRAWINGML_SPREADSHEET_DRAWING_NS}}}oneCellAnchor")
            if len(anchors) != 1:
                return None
            anchor = _chart_anchor_cell(anchors[0])
            if anchor is None:
                return None
            chart_references = list(drawing.iter(f"{{{_DRAWINGML_CHART_NS}}}chart"))
            if len(chart_references) != 1:
                return None
            chart_relationship_id = chart_references[0].get(f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id")
            if not isinstance(chart_relationship_id, str):
                return None
            drawing_directory, drawing_filename = drawing_member.rsplit("/", maxsplit=1)
            drawing_relationships = ElementTree.fromstring(
                archive.read(f"{drawing_directory}/_rels/{drawing_filename}.rels")
            )
            chart_member = _relationship_part_member(
                drawing_relationships,
                chart_relationship_id,
                drawing_member,
                relationship_type=f"{_DOCUMENT_RELATIONSHIPS_NS}/chart",
            )
            if chart_member is None:
                return None
            chart = ElementTree.fromstring(archive.read(chart_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None

    if chart.tag != f"{{{_DRAWINGML_CHART_NS}}}chartSpace":
        return None
    series = chart.findall(f".//{{{_DRAWINGML_CHART_NS}}}ser")
    if len(series) != 1:
        return None
    title_references = series[0].findall(
        f"{{{_DRAWINGML_CHART_NS}}}tx/{{{_DRAWINGML_CHART_NS}}}strRef/{{{_DRAWINGML_CHART_NS}}}f"
    )
    category_references = series[0].findall(
        f"{{{_DRAWINGML_CHART_NS}}}cat/{{{_DRAWINGML_CHART_NS}}}numRef/{{{_DRAWINGML_CHART_NS}}}f"
    )
    value_references = series[0].findall(
        f"{{{_DRAWINGML_CHART_NS}}}val/{{{_DRAWINGML_CHART_NS}}}numRef/{{{_DRAWINGML_CHART_NS}}}f"
    )
    if (
        len(title_references) != 1
        or len(category_references) != 1
        or len(value_references) != 1
        or not isinstance(title_references[0].text, str)
        or not isinstance(category_references[0].text, str)
        or not isinstance(value_references[0].text, str)
    ):
        return None
    return {
        "worksheet_member": worksheet_member,
        "drawing_member": drawing_member,
        "chart_member": chart_member,
        "anchor": anchor,
        "series_title_reference": title_references[0].text,
        "category_reference": category_references[0].text,
        "value_reference": value_references[0].text,
    }


def _chart_without_value_reference(path: Path, chart_member: str) -> bytes | None:
    """Return one generated chart XML tree with its sole value reference erased."""

    try:
        with ZipFile(path) as archive:
            chart = ElementTree.fromstring(archive.read(chart_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    series = chart.findall(f".//{{{_DRAWINGML_CHART_NS}}}ser")
    if len(series) != 1:
        return None
    value_references = series[0].findall(
        f"{{{_DRAWINGML_CHART_NS}}}val/{{{_DRAWINGML_CHART_NS}}}numRef/{{{_DRAWINGML_CHART_NS}}}f"
    )
    if len(value_references) != 1:
        return None
    value_references[0].text = None
    return ElementTree.tostring(chart, encoding="utf-8", xml_declaration=True)


def _formula_cached_result_state(
    path: Path, sheet_name: str, coordinate: str
) -> tuple[str, str, str, str, bytes] | None:
    """Read one formula expression and its raw saved ``<v>`` result.

    The tuple is ``(worksheet_member, formula, result_type, result_text,
    worksheet_without_result)``. It is intentionally a narrow raw-OOXML
    reader: the final item lets the validator prove that the fixture changed
    only the selected result text, without treating that text as an evaluated
    or business-correct outcome.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
            cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
            cells = [cell for cell in worksheet.iter(cell_tag) if cell.get("r") == coordinate]
            if len(cells) != 1:
                return None
            cell = cells[0]
            formula = cell.find(f"{{{_SPREADSHEETML_NS}}}f")
            cached_result = cell.find(f"{{{_SPREADSHEETML_NS}}}v")
            if formula is None or cached_result is None or cached_result.text is None:
                return None
            formula_text = f"={formula.text or ''}"
            cell_type = cell.get("t")
            result_type = "numeric" if cell_type in {None, "n"} else cell_type
            result_text = cached_result.text
            cached_result.text = None
            worksheet_without_result = ElementTree.tostring(
                worksheet, encoding="utf-8", xml_declaration=True
            )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    return (
        worksheet_member,
        formula_text,
        result_type,
        result_text,
        worksheet_without_result,
    )


def _raw_cell_state(
    path: Path, sheet_name: str, coordinate: str
) -> tuple[str, str | None, str | None, str | None, str | None] | None:
    """Read one generated cell's raw type, style, formula, and value text.

    This is intentionally narrower than a worksheet reader. It lets date-system
    fixtures prove that a stored serial and formula text did not change before a
    client applies any epoch-based conversion.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
    matches = [cell for cell in worksheet.iter(cell_tag) if cell.get("r") == coordinate]
    if len(matches) != 1:
        return None
    cell = matches[0]
    formula = cell.find(f"{{{_SPREADSHEETML_NS}}}f")
    value = cell.find(f"{{{_SPREADSHEETML_NS}}}v")
    return (
        worksheet_member,
        cell.get("t"),
        cell.get("s"),
        f"={formula.text or ''}" if formula is not None else None,
        value.text if value is not None else None,
    )


def _raw_auto_filter_state(
    path: Path, sheet_name: str
) -> (
    tuple[
        str,
        tuple[tuple[str, str], ...],
        int,
        tuple[tuple[str, str], ...],
        tuple[tuple[str, str], ...],
        tuple[str, ...],
    ]
    | None
):
    """Read WCAB's one explicit worksheet AutoFilter criterion.

    The narrow reader proves a raw stored control transition without applying a
    filter, changing row visibility, or calculating a ``SUBTOTAL`` result.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
    filter_column_tag = f"{{{_SPREADSHEETML_NS}}}filterColumn"
    filters_tag = f"{{{_SPREADSHEETML_NS}}}filters"
    filter_tag = f"{{{_SPREADSHEETML_NS}}}filter"
    auto_filters = worksheet.findall(auto_filter_tag)
    if len(auto_filters) != 1:
        return None
    auto_filter = auto_filters[0]
    filter_columns = auto_filter.findall(filter_column_tag)
    if len(filter_columns) != 1:
        return None
    filter_column = filter_columns[0]
    if len(auto_filter) != 1 or auto_filter[0] is not filter_column:
        return None
    try:
        column_id = int(filter_column.get("colId", ""))
    except ValueError:
        return None
    filters = filter_column.findall(filters_tag)
    if len(filters) != 1:
        return None
    if len(filter_column) != 1 or filter_column[0] is not filters[0]:
        return None
    filter_elements = filters[0].findall(filter_tag)
    if len(filter_elements) != 1 or len(filters[0]) != 1 or filters[0][0] is not filter_elements[0]:
        return None
    values: list[str] = []
    for filter_element in filter_elements:
        value = filter_element.get("val")
        if not isinstance(value, str):
            return None
        values.append(value)
    if len(values) != 1:
        return None
    return (
        worksheet_member,
        tuple(sorted(auto_filter.attrib.items())),
        column_id,
        tuple(sorted(filter_column.attrib.items())),
        tuple(sorted(filters[0].attrib.items())),
        tuple(values),
    )


def _worksheet_without_auto_filter_criteria(path: Path, sheet_name: str) -> bytes | None:
    """Return one raw worksheet with its sole AutoFilter criterion removed."""

    if _raw_auto_filter_state(path, sheet_name) is None:
        return None
    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
    filter_column_tag = f"{{{_SPREADSHEETML_NS}}}filterColumn"
    auto_filters = worksheet.findall(auto_filter_tag)
    if len(auto_filters) != 1:
        return None
    auto_filter = auto_filters[0]
    filter_columns = auto_filter.findall(filter_column_tag)
    if len(filter_columns) != 1:
        return None
    auto_filter.remove(filter_columns[0])
    return ElementTree.tostring(worksheet, encoding="utf-8", xml_declaration=True)


def _raw_named_sheet_view_filter_state(path: Path, sheet_name: str) -> dict[str, Any] | None:
    """Read WCAB's compact relationship-backed Named Sheet View declaration.

    The reader follows only the generated worksheet relationship and its local
    part. It proves stored filter metadata, not whether Excel activates a view,
    applies a filter, calculates a subtotal, or renders a report.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
            auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
            auto_filters = worksheet.findall(auto_filter_tag)
            if len(auto_filters) != 1 or len(auto_filters[0]) != 0:
                return None
            auto_filter = auto_filters[0]

            worksheet_directory, worksheet_filename = worksheet_member.rsplit("/", maxsplit=1)
            relationships_member = f"{worksheet_directory}/_rels/{worksheet_filename}.rels"
            relationships = ElementTree.fromstring(archive.read(relationships_member))
            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            named_view_relationships = [
                relationship
                for relationship in relationships.findall(relationship_tag)
                if relationship.get("Type") == _NAMED_SHEET_VIEW_RELATIONSHIP
            ]
            if len(named_view_relationships) != 1:
                return None
            relationship = named_view_relationships[0]
            relationship_id = relationship.get("Id")
            if not isinstance(relationship_id, str):
                return None
            named_view_member = _relationship_part_member(
                relationships,
                relationship_id,
                worksheet_member,
                relationship_type=_NAMED_SHEET_VIEW_RELATIONSHIP,
            )
            if named_view_member is None:
                return None

            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
            overrides = [
                override
                for override in content_types.findall(f"{{{_CONTENT_TYPES_NS}}}Override")
                if override.get("PartName") == f"/{named_view_member}"
            ]
            if len(overrides) != 1:
                return None

            named_views = ElementTree.fromstring(archive.read(named_view_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError, ValueError):
        return None

    named_views_tag = f"{{{_NAMED_SHEET_VIEW_NS}}}namedSheetViews"
    named_view_tag = f"{{{_NAMED_SHEET_VIEW_NS}}}namedSheetView"
    named_filter_tag = f"{{{_NAMED_SHEET_VIEW_NS}}}nsvFilter"
    column_filter_tag = f"{{{_NAMED_SHEET_VIEW_NS}}}columnFilter"
    filter_column_tag = f"{{{_NAMED_SHEET_VIEW_NS}}}filter"
    filters_tag = f"{{{_SPREADSHEETML_NS}}}filters"
    filter_tag = f"{{{_SPREADSHEETML_NS}}}filter"
    if named_views.tag != named_views_tag or named_views.attrib or len(named_views) != 1:
        return None
    named_view = named_views[0]
    if named_view.tag != named_view_tag or len(named_view) != 1:
        return None
    named_filter = named_view[0]
    if named_filter.tag != named_filter_tag or len(named_filter) != 1:
        return None
    column_filter = named_filter[0]
    if column_filter.tag != column_filter_tag or len(column_filter) != 1:
        return None
    filter_column = column_filter[0]
    if filter_column.tag != filter_column_tag or len(filter_column) != 1:
        return None
    filters = filter_column[0]
    if filters.tag != filters_tag or len(filters) != 1:
        return None
    criterion = filters[0]
    if criterion.tag != filter_tag or len(criterion) != 0:
        return None
    criterion_value = criterion.get("val")
    if not isinstance(criterion_value, str):
        return None
    return {
        "worksheet_member": worksheet_member,
        "base_auto_filter_attributes": tuple(sorted(auto_filter.attrib.items())),
        "relationships_member": relationships_member,
        "relationship_attributes": tuple(sorted(relationship.attrib.items())),
        "view_member": named_view_member,
        "content_type_attributes": tuple(sorted(overrides[0].attrib.items())),
        "named_view_attributes": tuple(sorted(named_view.attrib.items())),
        "named_filter_attributes": tuple(sorted(named_filter.attrib.items())),
        "column_filter_attributes": tuple(sorted(column_filter.attrib.items())),
        "filter_column_attributes": tuple(sorted(filter_column.attrib.items())),
        "filters_attributes": tuple(sorted(filters.attrib.items())),
        "criterion_attributes": tuple(sorted(criterion.attrib.items())),
        "criterion_value": criterion_value,
    }


def _named_sheet_view_without_filter_criterion(path: Path, sheet_name: str) -> bytes | None:
    """Return WCAB's Named Sheet View part with only its list value removed."""

    state = _raw_named_sheet_view_filter_state(path, sheet_name)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            named_views = ElementTree.fromstring(archive.read(state["view_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    try:
        criterion = named_views[0][0][0][0][0]
    except IndexError:
        return None
    if criterion.tag != f"{{{_SPREADSHEETML_NS}}}filter" or "val" not in criterion.attrib:
        return None
    criterion.attrib.pop("val")
    return ElementTree.tostring(named_views, encoding="utf-8", xml_declaration=True)


def _raw_xml_map_table_binding_state(path: Path, sheet_name: str) -> dict[str, Any] | None:
    """Read WCAB's compact XML Map table and single-cell binding graph.

    This reader follows only local package relationships written by the
    generated fixture. It establishes a stored schema/map/table binding, not
    whether Excel can validate the schema, open a bound file, import or export
    XML, materialize data, calculate formulas, or render a result.
    """

    try:
        with ZipFile(path) as archive:
            workbook_relationships = ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )
            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            map_relationships = [
                relationship
                for relationship in workbook_relationships.findall(relationship_tag)
                if relationship.get("Type") == _XML_MAP_RELATIONSHIP
            ]
            if len(map_relationships) != 1:
                return None
            map_relationship_id = map_relationships[0].get("Id")
            if not isinstance(map_relationship_id, str):
                return None
            map_member = _relationship_part_member(
                workbook_relationships,
                map_relationship_id,
                "xl/workbook.xml",
                relationship_type=_XML_MAP_RELATIONSHIP,
            )
            if map_member is None:
                return None
            map_info = ElementTree.fromstring(archive.read(map_member))

            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
            table_parts = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}tableParts")
            if len(table_parts) != 1 or len(table_parts[0]) != 1:
                return None
            table_part = table_parts[0][0]
            if table_part.tag != f"{{{_SPREADSHEETML_NS}}}tablePart":
                return None
            table_relationship_id = table_part.get(f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id")
            if not isinstance(table_relationship_id, str):
                return None

            worksheet_directory, worksheet_filename = worksheet_member.rsplit("/", maxsplit=1)
            worksheet_relationships_member = (
                f"{worksheet_directory}/_rels/{worksheet_filename}.rels"
            )
            worksheet_relationships = ElementTree.fromstring(
                archive.read(worksheet_relationships_member)
            )
            table_member = _relationship_part_member(
                worksheet_relationships,
                table_relationship_id,
                worksheet_member,
                relationship_type=f"{_DOCUMENT_RELATIONSHIPS_NS}/table",
            )
            if table_member is None:
                return None
            table_relationships = [
                relationship
                for relationship in worksheet_relationships.findall(relationship_tag)
                if relationship.get("Id") == table_relationship_id
                and relationship.get("Type") == f"{_DOCUMENT_RELATIONSHIPS_NS}/table"
            ]
            if len(table_relationships) != 1:
                return None
            table = ElementTree.fromstring(archive.read(table_member))

            single_cell_relationships = [
                relationship
                for relationship in worksheet_relationships.findall(relationship_tag)
                if relationship.get("Type") == _TABLE_SINGLE_CELLS_RELATIONSHIP
            ]
            if len(single_cell_relationships) != 1:
                return None
            single_cell_relationship_id = single_cell_relationships[0].get("Id")
            if not isinstance(single_cell_relationship_id, str):
                return None
            single_cell_member = _relationship_part_member(
                worksheet_relationships,
                single_cell_relationship_id,
                worksheet_member,
                relationship_type=_TABLE_SINGLE_CELLS_RELATIONSHIP,
            )
            if single_cell_member is None:
                return None
            single_cells = ElementTree.fromstring(archive.read(single_cell_member))

            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError, ValueError):
        return None

    spreadsheet = _SPREADSHEETML_NS
    if map_info.tag != f"{{{spreadsheet}}}MapInfo":
        return None
    schemas = map_info.findall(f"{{{spreadsheet}}}Schema")
    maps = map_info.findall(f"{{{spreadsheet}}}Map")
    if len(schemas) != 1 or len(maps) != 1 or len(map_info) != 2:
        return None
    schema = schemas[0]
    schema_definitions = schema.findall(f"{{{_XML_SCHEMA_NS}}}schema")
    if len(schema_definitions) != 1 or len(schema) != 1:
        return None
    schema_definition = schema_definitions[0]
    root_elements = schema_definition.findall(f"{{{_XML_SCHEMA_NS}}}element")
    if len(root_elements) != 1 or len(schema_definition) != 1:
        return None
    schema_element_attributes = tuple(
        tuple(sorted(element.attrib.items()))
        for element in schema_definition.iter(f"{{{_XML_SCHEMA_NS}}}element")
    )

    xml_map = maps[0]
    data_bindings = xml_map.findall(f"{{{spreadsheet}}}DataBinding")
    if len(data_bindings) != 1 or len(xml_map) != 1:
        return None

    if table.tag != f"{{{spreadsheet}}}table":
        return None
    table_columns = table.findall(f"{{{spreadsheet}}}tableColumns")
    if len(table_columns) != 1:
        return None
    columns = table_columns[0].findall(f"{{{spreadsheet}}}tableColumn")
    if len(columns) != 2:
        return None
    mapped_columns = [
        (column, bindings[0])
        for column in columns
        if len(bindings := column.findall(f"{{{spreadsheet}}}xmlColumnPr")) == 1
    ]
    if len(mapped_columns) != 1:
        return None
    mapped_column, xml_column = mapped_columns[0]
    if len(xml_column) != 0:
        return None

    if single_cells.tag != f"{{{spreadsheet}}}singleXmlCells" or single_cells.attrib:
        return None
    single_cells_children = single_cells.findall(f"{{{spreadsheet}}}singleXmlCell")
    if len(single_cells_children) != 1 or len(single_cells) != 1:
        return None
    single_cell = single_cells_children[0]
    cell_properties = single_cell.findall(f"{{{spreadsheet}}}xmlCellPr")
    if len(cell_properties) != 1 or len(single_cell) != 1:
        return None
    cell_property = cell_properties[0]
    cell_bindings = cell_property.findall(f"{{{spreadsheet}}}xmlPr")
    if len(cell_bindings) != 1 or len(cell_property) != 1 or len(cell_bindings[0]) != 0:
        return None

    default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"
    override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
    xml_defaults = [
        default
        for default in content_types.findall(default_tag)
        if default.get("Extension") == "xml" and default.get("ContentType") == "application/xml"
    ]
    single_cell_content_types = [
        override
        for override in content_types.findall(override_tag)
        if override.get("PartName") == f"/{single_cell_member}"
    ]
    if len(xml_defaults) != 1 or len(single_cell_content_types) != 1:
        return None

    return {
        "worksheet_member": worksheet_member,
        "worksheet_relationships_member": worksheet_relationships_member,
        "worksheet_table_part_attributes": tuple(sorted(table_part.attrib.items())),
        "table_relationship_attributes": tuple(sorted(table_relationships[0].attrib.items())),
        "table_member": table_member,
        "table_attributes": tuple(sorted(table.attrib.items())),
        "table_column_attributes": tuple(sorted(mapped_column.attrib.items())),
        "xml_column_attributes": tuple(sorted(xml_column.attrib.items())),
        "map_relationship_attributes": tuple(sorted(map_relationships[0].attrib.items())),
        "map_member": map_member,
        "map_info_attributes": tuple(sorted(map_info.attrib.items())),
        "schema_attributes": tuple(sorted(schema.attrib.items())),
        "schema_definition_attributes": tuple(sorted(schema_definition.attrib.items())),
        "schema_root_element_attributes": tuple(sorted(root_elements[0].attrib.items())),
        "schema_element_attributes": schema_element_attributes,
        "map_attributes": tuple(sorted(xml_map.attrib.items())),
        "data_binding_attributes": tuple(sorted(data_bindings[0].attrib.items())),
        "single_cell_relationship_attributes": tuple(
            sorted(single_cell_relationships[0].attrib.items())
        ),
        "single_cell_member": single_cell_member,
        "single_cell_attributes": tuple(sorted(single_cell.attrib.items())),
        "single_cell_property_attributes": tuple(sorted(cell_property.attrib.items())),
        "single_cell_binding_attributes": tuple(sorted(cell_bindings[0].attrib.items())),
        "xml_default_attributes": tuple(sorted(xml_defaults[0].attrib.items())),
        "single_cell_content_type_attributes": tuple(
            sorted(single_cell_content_types[0].attrib.items())
        ),
    }


def _xml_map_table_without_xpath(path: Path, sheet_name: str) -> bytes | None:
    """Return the generated mapped Table part with only its XPath erased."""

    state = _raw_xml_map_table_binding_state(path, sheet_name)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            table = ElementTree.fromstring(archive.read(state["table_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    bindings = table.findall(f".//{{{_SPREADSHEETML_NS}}}xmlColumnPr")
    if len(bindings) != 1 or "xpath" not in bindings[0].attrib:
        return None
    bindings[0].attrib.pop("xpath")
    return ElementTree.tostring(table, encoding="utf-8", xml_declaration=True)


def _raw_office_web_addin_auto_show_state(path: Path) -> dict[str, Any] | None:
    """Read WCAB's compact document-linked Office Web Add-in declaration.

    This follows only the generated workbook-to-taskpane-to-webextension local
    relationship graph. It proves a stored auto-show request without loading a
    manifest, following a store reference, installing an add-in, or opening an
    Office client.
    """

    try:
        with ZipFile(path) as archive:
            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            workbook_relationships = ElementTree.fromstring(
                archive.read("xl/_rels/workbook.xml.rels")
            )
            taskpane_relationships = [
                relationship
                for relationship in workbook_relationships.findall(relationship_tag)
                if relationship.get("Type") == _WEB_EXTENSION_TASKPANES_RELATIONSHIP
            ]
            if len(taskpane_relationships) != 1:
                return None
            taskpane_relationship_id = taskpane_relationships[0].get("Id")
            if not isinstance(taskpane_relationship_id, str):
                return None
            taskpane_member = _relationship_part_member(
                workbook_relationships,
                taskpane_relationship_id,
                "xl/workbook.xml",
                relationship_type=_WEB_EXTENSION_TASKPANES_RELATIONSHIP,
            )
            if taskpane_member is None:
                return None
            taskpanes = ElementTree.fromstring(archive.read(taskpane_member))

            taskpane_directory, taskpane_filename = taskpane_member.rsplit("/", maxsplit=1)
            taskpane_relationships_member = f"{taskpane_directory}/_rels/{taskpane_filename}.rels"
            taskpane_relationships_root = ElementTree.fromstring(
                archive.read(taskpane_relationships_member)
            )
            extension_relationships = [
                relationship
                for relationship in taskpane_relationships_root.findall(relationship_tag)
                if relationship.get("Type") == _WEB_EXTENSION_RELATIONSHIP
            ]
            if len(extension_relationships) != 1 or len(taskpane_relationships_root) != 1:
                return None
            extension_relationship_id = extension_relationships[0].get("Id")
            if not isinstance(extension_relationship_id, str):
                return None
            extension_member = _relationship_part_member(
                taskpane_relationships_root,
                extension_relationship_id,
                taskpane_member,
                relationship_type=_WEB_EXTENSION_RELATIONSHIP,
            )
            if extension_member is None:
                return None
            extension = ElementTree.fromstring(archive.read(extension_member))

            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError, ValueError):
        return None

    if (
        taskpanes.tag != f"{{{_WEB_EXTENSION_TASKPANES_NS}}}taskpanes"
        or taskpanes.attrib
        or len(taskpanes) != 1
    ):
        return None
    taskpane = taskpanes[0]
    if taskpane.tag != f"{{{_WEB_EXTENSION_TASKPANES_NS}}}taskpane" or len(taskpane) != 1:
        return None
    taskpane_reference = taskpane[0]
    relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
    if (
        taskpane_reference.tag != f"{{{_WEB_EXTENSION_TASKPANES_NS}}}webextension"
        or taskpane_reference.get(relationship_id_attribute) != extension_relationship_id
        or len(taskpane_reference)
    ):
        return None

    if extension.tag != f"{{{_WEB_EXTENSION_NS}}}webextension" or len(extension) != 2:
        return None
    reference_tag = f"{{{_WEB_EXTENSION_NS}}}reference"
    properties_tag = f"{{{_WEB_EXTENSION_NS}}}properties"
    property_tag = f"{{{_WEB_EXTENSION_NS}}}property"
    references = extension.findall(reference_tag)
    properties = extension.findall(properties_tag)
    if len(references) != 1 or len(properties) != 1 or len(properties[0]) != 1:
        return None
    property_element = properties[0][0]
    if property_element.tag != property_tag or len(property_element):
        return None
    auto_show = _ooxml_boolean(property_element.get("value"))
    if auto_show is None:
        return None

    override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
    taskpane_overrides = [
        override
        for override in content_types.findall(override_tag)
        if override.get("PartName") == f"/{taskpane_member}"
    ]
    extension_overrides = [
        override
        for override in content_types.findall(override_tag)
        if override.get("PartName") == f"/{extension_member}"
    ]
    if len(taskpane_overrides) != 1 or len(extension_overrides) != 1:
        return None
    return {
        "workbook_relationship_attributes": tuple(sorted(taskpane_relationships[0].attrib.items())),
        "taskpane_member": taskpane_member,
        "taskpane_content_type_attributes": tuple(sorted(taskpane_overrides[0].attrib.items())),
        "taskpane_relationships_member": taskpane_relationships_member,
        "taskpane_to_extension_relationship_attributes": tuple(
            sorted(extension_relationships[0].attrib.items())
        ),
        "taskpane_attributes": tuple(sorted(taskpane.attrib.items())),
        "taskpane_reference_attributes": tuple(sorted(taskpane_reference.attrib.items())),
        "extension_member": extension_member,
        "extension_attributes": tuple(sorted(extension.attrib.items())),
        "extension_content_type_attributes": tuple(sorted(extension_overrides[0].attrib.items())),
        "reference_attributes": tuple(sorted(references[0].attrib.items())),
        "properties_attributes": tuple(sorted(properties[0].attrib.items())),
        "property_attributes": tuple(sorted(property_element.attrib.items())),
        "auto_show": auto_show,
    }


def _office_web_addin_without_auto_show(path: Path) -> bytes | None:
    """Return the generated WebExtension part with only its auto-show value erased."""

    state = _raw_office_web_addin_auto_show_state(path)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            extension = ElementTree.fromstring(archive.read(state["extension_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    property_tag = f"{{{_WEB_EXTENSION_NS}}}property"
    properties = [
        property_element
        for property_element in extension.iter(property_tag)
        if property_element.get("name") == "Office.AutoShowTaskpaneWithDocument"
    ]
    if len(properties) != 1 or "value" not in properties[0].attrib:
        return None
    properties[0].attrib.pop("value")
    return ElementTree.tostring(extension, encoding="utf-8", xml_declaration=True)


def _raw_ole_object_auto_load_state(path: Path, sheet_name: str) -> dict[str, Any] | None:
    """Read WCAB's one relationship-backed embedded-object declaration.

    This follows only the generated internal worksheet relationship and checks
    the fixed opaque bytes as bytes. It does not deserialize an OLE container,
    render a presentation, invoke a ProgID, register an object server, or open
    an Office client.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
            ole_objects_tag = f"{{{_SPREADSHEETML_NS}}}oleObjects"
            ole_object_tag = f"{{{_SPREADSHEETML_NS}}}oleObject"
            controls_tag = f"{{{_SPREADSHEETML_NS}}}controls"
            ole_objects = worksheet.findall(ole_objects_tag)
            if (
                len(ole_objects) != 1
                or ole_objects[0].attrib
                or len(ole_objects[0]) != 1
                or worksheet.findall(controls_tag)
            ):
                return None
            ole_object = ole_objects[0][0]
            relationship_id = ole_object.get(f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id")
            if (
                ole_object.tag != ole_object_tag
                or len(ole_object)
                or not isinstance(relationship_id, str)
            ):
                return None

            worksheet_directory, worksheet_filename = worksheet_member.rsplit("/", maxsplit=1)
            worksheet_relationships_member = (
                f"{worksheet_directory}/_rels/{worksheet_filename}.rels"
            )
            relationships = ElementTree.fromstring(archive.read(worksheet_relationships_member))
            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            matching_relationships = [
                relationship
                for relationship in relationships.findall(relationship_tag)
                if relationship.get("Id") == relationship_id
                and relationship.get("Type") == _OLE_OBJECT_RELATIONSHIP
                and relationship.get("TargetMode") is None
            ]
            if len(matching_relationships) != 1 or len(relationships) != 1:
                return None
            embedded_object_member = _relationship_part_member(
                relationships,
                relationship_id,
                worksheet_member,
                relationship_type=_OLE_OBJECT_RELATIONSHIP,
            )
            if embedded_object_member is None:
                return None
            payload = archive.read(embedded_object_member)
            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError, ValueError):
        return None

    auto_load = _ooxml_boolean(ole_object.get("autoLoad"))
    default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"
    override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
    binary_defaults = [
        default
        for default in content_types.findall(default_tag)
        if default.get("Extension") == "bin"
    ]
    if (
        auto_load is None
        or payload != _OLE_OBJECT_PAYLOAD
        or len(binary_defaults) != 1
        or binary_defaults[0].get("ContentType") != _OLE_OBJECT_CONTENT_TYPE
        or any(
            override.get("PartName") == f"/{embedded_object_member}"
            for override in content_types.findall(override_tag)
        )
    ):
        return None
    return {
        "worksheet_member": worksheet_member,
        "worksheet_relationships_member": worksheet_relationships_member,
        "ole_objects_attributes": tuple(sorted(ole_objects[0].attrib.items())),
        "ole_object_attributes": tuple(sorted(ole_object.attrib.items())),
        "relationship_attributes": tuple(sorted(matching_relationships[0].attrib.items())),
        "embedded_object_member": embedded_object_member,
        "content_type_attributes": tuple(sorted(binary_defaults[0].attrib.items())),
        "payload_size": len(payload),
        "auto_load": auto_load,
    }


def _ole_object_without_auto_load(path: Path, sheet_name: str) -> bytes | None:
    """Return the generated worksheet with only the OLE auto-load flag erased."""

    state = _raw_ole_object_auto_load_state(path, sheet_name)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            worksheet = ElementTree.fromstring(archive.read(state["worksheet_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    ole_objects = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}oleObjects")
    if len(ole_objects) != 1 or len(ole_objects[0]) != 1:
        return None
    ole_object = ole_objects[0][0]
    if "autoLoad" not in ole_object.attrib:
        return None
    ole_object.attrib.pop("autoLoad")
    return ElementTree.tostring(worksheet, encoding="utf-8", xml_declaration=True)


def _raw_ignored_error_formula_range_state(path: Path, sheet_name: str) -> dict[str, Any] | None:
    """Read WCAB's compact worksheet-local ignored-error declaration.

    The generated pair uses at most one standard SpreadsheetML ``ignoredErrors``
    container and one ``ignoredError`` child. This reader establishes only that
    stored metadata; it neither evaluates the formula nor decides whether an
    Excel client would show, suppress, or justify an error indicator.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    if worksheet.tag != f"{{{_SPREADSHEETML_NS}}}worksheet":
        return None
    ignored_errors_tag = f"{{{_SPREADSHEETML_NS}}}ignoredErrors"
    ignored_error_tag = f"{{{_SPREADSHEETML_NS}}}ignoredError"
    containers = worksheet.findall(ignored_errors_tag)
    if len(containers) == 0:
        return {
            "worksheet_member": worksheet_member,
            "present": False,
            "container_attributes": None,
            "container_child_count": 0,
            "container_text": None,
            "rule_attributes": None,
            "rule_child_count": 0,
            "rule_text": None,
        }
    if len(containers) != 1:
        return None
    container = containers[0]
    rules = list(container)
    if len(rules) != 1 or rules[0].tag != ignored_error_tag:
        return None
    rule = rules[0]
    return {
        "worksheet_member": worksheet_member,
        "present": True,
        "container_attributes": tuple(sorted(container.attrib.items())),
        "container_child_count": len(container),
        "container_text": container.text,
        "rule_attributes": tuple(sorted(rule.attrib.items())),
        "rule_child_count": len(rule),
        "rule_text": rule.text,
    }


def _worksheet_without_ignored_error_controls(path: Path, sheet_name: str) -> bytes | None:
    """Return the generated worksheet with its one ignored-errors node removed."""

    if _raw_ignored_error_formula_range_state(path, sheet_name) is None:
        return None
    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    controls = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}ignoredErrors")
    if len(controls) > 1:
        return None
    if controls:
        worksheet.remove(controls[0])
    return ElementTree.tostring(worksheet, encoding="utf-8", xml_declaration=True)


def _raw_cell_number_format(path: Path, sheet_name: str, coordinate: str) -> str | None:
    """Resolve the generated cell's custom number format from raw styles XML."""

    state = _raw_cell_state(path, sheet_name, coordinate)
    if state is None:
        return None
    try:
        style_index = int(state[2] or "0")
        with ZipFile(path) as archive:
            styles = ElementTree.fromstring(archive.read("xl/styles.xml"))
    except (BadZipFile, KeyError, OSError, ValueError, ElementTree.ParseError):
        return None
    cell_xfs = styles.findall(f"./{{{_SPREADSHEETML_NS}}}cellXfs/{{{_SPREADSHEETML_NS}}}xf")
    if not 0 <= style_index < len(cell_xfs):
        return None
    try:
        number_format_id = int(cell_xfs[style_index].get("numFmtId", ""))
    except ValueError:
        return None
    custom_formats: dict[int, str] = {}
    for number_format in styles.findall(
        f"./{{{_SPREADSHEETML_NS}}}numFmts/{{{_SPREADSHEETML_NS}}}numFmt"
    ):
        try:
            custom_format_id = int(number_format.get("numFmtId", ""))
        except ValueError:
            continue
        format_code = number_format.get("formatCode")
        if format_code is not None:
            custom_formats[custom_format_id] = format_code
    return custom_formats.get(number_format_id)


def _raw_custom_number_format_state(
    path: Path, sheet_name: str, coordinate: str
) -> dict[str, Any] | None:
    """Read one generated direct-cell custom number-format declaration.

    This narrow reader follows the target cell's existing style index to one
    custom numFmt record. It observes stored display metadata only; it never
    renders a format or derives a displayed value.
    """

    cell_state = _raw_cell_state(path, sheet_name, coordinate)
    if cell_state is None:
        return None
    try:
        style_index = int(cell_state[2] or "0")
        with ZipFile(path) as archive:
            styles = ElementTree.fromstring(archive.read("xl/styles.xml"))
    except (BadZipFile, KeyError, OSError, ValueError, ElementTree.ParseError):
        return None
    cell_xfs = styles.findall(f"./{{{_SPREADSHEETML_NS}}}cellXfs/{{{_SPREADSHEETML_NS}}}xf")
    if not 0 <= style_index < len(cell_xfs):
        return None
    try:
        custom_number_format_id = int(cell_xfs[style_index].get("numFmtId", ""))
    except ValueError:
        return None
    num_fmts_tag = f"{{{_SPREADSHEETML_NS}}}numFmts"
    num_fmt_tag = f"{{{_SPREADSHEETML_NS}}}numFmt"
    containers = styles.findall(num_fmts_tag)
    if len(containers) != 1:
        return None
    matches = [
        number_format
        for number_format in containers[0].findall(num_fmt_tag)
        if number_format.get("numFmtId") == str(custom_number_format_id)
    ]
    if len(matches) != 1:
        return None
    number_format = matches[0]
    format_code = number_format.get("formatCode")
    if not isinstance(format_code, str):
        return None
    return {
        "worksheet_member": cell_state[0],
        "styles_member": "xl/styles.xml",
        "cell_style_index": style_index,
        "custom_number_format_id": custom_number_format_id,
        "cell_xf_attributes": tuple(sorted(cell_xfs[style_index].attrib.items())),
        "num_fmts_attributes": tuple(sorted(containers[0].attrib.items())),
        "num_fmt_attributes": tuple(sorted(number_format.attrib.items())),
        "format_code": format_code,
    }


def _styles_without_custom_number_format_code(
    path: Path, custom_number_format_id: int
) -> bytes | None:
    """Return generated styles XML with one declared custom format code erased."""

    if custom_number_format_id < 0:
        return None
    try:
        with ZipFile(path) as archive:
            styles = ElementTree.fromstring(archive.read("xl/styles.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    num_fmts_tag = f"{{{_SPREADSHEETML_NS}}}numFmts"
    num_fmt_tag = f"{{{_SPREADSHEETML_NS}}}numFmt"
    containers = styles.findall(num_fmts_tag)
    if len(containers) != 1:
        return None
    matches = [
        number_format
        for number_format in containers[0].findall(num_fmt_tag)
        if number_format.get("numFmtId") == str(custom_number_format_id)
    ]
    if len(matches) != 1 or "formatCode" not in matches[0].attrib:
        return None
    matches[0].attrib.pop("formatCode")
    return ElementTree.tostring(styles, encoding="utf-8", xml_declaration=True)


def _workbook_date_system_state(path: Path) -> tuple[bool, bool] | None:
    """Read WCAB's explicit raw workbook serial-date controls."""

    properties = _workbook_properties(path)
    if properties is None:
        return None
    date_1904 = _ooxml_boolean(properties.get("date1904"))
    date_compatibility = _ooxml_boolean(properties.get("dateCompatibility"))
    return (
        (date_1904, date_compatibility)
        if date_1904 is not None and date_compatibility is not None
        else None
    )


def _workbook_without_date_system_controls(path: Path) -> bytes | None:
    """Return raw workbook XML with only WCAB's two date controls removed."""

    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    properties = workbook.find(f"{{{_SPREADSHEETML_NS}}}workbookPr")
    if properties is None:
        return None
    properties.attrib.pop("date1904", None)
    properties.attrib.pop("dateCompatibility", None)
    return ElementTree.tostring(workbook, encoding="utf-8", xml_declaration=True)


def _workbook_structure_protection_state(
    path: Path,
) -> tuple[bool, tuple[tuple[str, str], ...]] | None:
    """Read WCAB's explicit workbook-structure lock and its compact attributes."""

    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    protection = workbook.find(f"{{{_SPREADSHEETML_NS}}}workbookProtection")
    if protection is None:
        return None
    lock_structure = _ooxml_boolean(protection.get("lockStructure"))
    return (
        (lock_structure, tuple(sorted(protection.attrib.items())))
        if lock_structure is not None
        else None
    )


def _workbook_without_structure_lock(path: Path) -> bytes | None:
    """Return raw workbook XML with only WCAB's ``lockStructure`` control removed."""

    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    protection = workbook.find(f"{{{_SPREADSHEETML_NS}}}workbookProtection")
    if protection is None or "lockStructure" not in protection.attrib:
        return None
    protection.attrib.pop("lockStructure")
    return ElementTree.tostring(workbook, encoding="utf-8", xml_declaration=True)


def _raw_external_defined_name_source_state(path: Path) -> dict[str, Any] | None:
    """Read WCAB's one local defined-name external-reference declaration.

    This is intentionally a narrow reader for the generated pair. It reads the
    stored OOXML name text and confirms that the package has no relationship-
    backed external-link declarations; it never follows the external text.
    """

    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            external_link_members = tuple(
                sorted(name for name in archive.namelist() if name.startswith("xl/externalLinks/"))
            )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    defined_names_tag = f"{{{_SPREADSHEETML_NS}}}definedNames"
    defined_name_tag = f"{{{_SPREADSHEETML_NS}}}definedName"
    external_references_tag = f"{{{_SPREADSHEETML_NS}}}externalReferences"
    containers = workbook.findall(defined_names_tag)
    if len(containers) != 1 or containers[0].attrib or len(containers[0]) != 1:
        return None
    definition = containers[0][0]
    if definition.tag != defined_name_tag or set(definition.attrib) != {"name"}:
        return None
    name = definition.get("name")
    refers_to = definition.text
    if not isinstance(name, str) or not name or not isinstance(refers_to, str) or not refers_to:
        return None
    return {
        "workbook_member": "xl/workbook.xml",
        "name": name,
        "attributes": tuple(sorted(definition.attrib.items())),
        "refers_to": refers_to,
        "external_references_count": len(workbook.findall(external_references_tag)),
        "external_link_members": external_link_members,
    }


def _workbook_without_external_defined_name_source(path: Path) -> bytes | None:
    """Return raw workbook XML with WCAB's sole defined-name text erased."""

    if _raw_external_defined_name_source_state(path) is None:
        return None
    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    defined_names = workbook.find(f"{{{_SPREADSHEETML_NS}}}definedNames")
    if defined_names is None or len(defined_names) != 1:
        return None
    definition = defined_names[0]
    if definition.tag != f"{{{_SPREADSHEETML_NS}}}definedName" or definition.text is None:
        return None
    definition.text = None
    return ElementTree.tostring(workbook, encoding="utf-8", xml_declaration=True)


def _raw_named_lambda_definition_state(path: Path) -> dict[str, Any] | None:
    """Read WCAB's one workbook-scoped named LAMBDA declaration.

    This is a deliberately narrow reader for the generated fixture. It
    observes the stored defined-name text and confirms that the compact
    package has no relationship-backed external-workbook declarations; it
    neither evaluates the LAMBDA nor expands its dependencies.
    """

    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            external_link_members = tuple(
                sorted(name for name in archive.namelist() if name.startswith("xl/externalLinks/"))
            )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    defined_names_tag = f"{{{_SPREADSHEETML_NS}}}definedNames"
    defined_name_tag = f"{{{_SPREADSHEETML_NS}}}definedName"
    external_references_tag = f"{{{_SPREADSHEETML_NS}}}externalReferences"
    containers = workbook.findall(defined_names_tag)
    if len(containers) != 1 or containers[0].attrib or len(containers[0]) != 1:
        return None
    definition = containers[0][0]
    if definition.tag != defined_name_tag or set(definition.attrib) != {"name"}:
        return None
    name = definition.get("name")
    refers_to = definition.text
    if not isinstance(name, str) or not name or not isinstance(refers_to, str) or not refers_to:
        return None
    return {
        "workbook_member": "xl/workbook.xml",
        "name": name,
        "attributes": tuple(sorted(definition.attrib.items())),
        "refers_to": refers_to,
        "external_references_count": len(workbook.findall(external_references_tag)),
        "external_link_members": external_link_members,
    }


def _workbook_without_named_lambda_definition(path: Path) -> bytes | None:
    """Return raw workbook XML with WCAB's sole LAMBDA body erased."""

    if _raw_named_lambda_definition_state(path) is None:
        return None
    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    defined_names = workbook.find(f"{{{_SPREADSHEETML_NS}}}definedNames")
    if defined_names is None or len(defined_names) != 1:
        return None
    definition = defined_names[0]
    if definition.tag != f"{{{_SPREADSHEETML_NS}}}definedName" or definition.text is None:
        return None
    definition.text = None
    return ElementTree.tostring(workbook, encoding="utf-8", xml_declaration=True)


def _raw_table_calculated_column_formula_state(
    path: Path, sheet_name: str
) -> dict[str, Any] | None:
    """Read WCAB's local Table binding and its sole formula master.

    The reader follows only one internal worksheet-to-Table relationship and
    rejects any unexpected table shape. It establishes a stored formula
    declaration only; it never fills a column, evaluates a structured
    reference, or calculates a workbook.
    """

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
            table_parts = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}tableParts")
            if len(table_parts) != 1 or len(table_parts[0]) != 1:
                return None
            table_parts_container = table_parts[0]
            table_part = table_parts_container[0]
            relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
            if (
                table_parts_container.attrib != {"count": "1"}
                or table_part.tag != f"{{{_SPREADSHEETML_NS}}}tablePart"
                or set(table_part.attrib) != {relationship_id_attribute}
            ):
                return None
            table_relationship_id = table_part.get(relationship_id_attribute)
            if not isinstance(table_relationship_id, str):
                return None
            worksheet_directory, worksheet_filename = worksheet_member.rsplit("/", maxsplit=1)
            worksheet_relationships_member = (
                f"{worksheet_directory}/_rels/{worksheet_filename}.rels"
            )
            worksheet_relationships = ElementTree.fromstring(
                archive.read(worksheet_relationships_member)
            )
            table_member = _relationship_part_member(
                worksheet_relationships,
                table_relationship_id,
                worksheet_member,
                relationship_type=_TABLE_RELATIONSHIP,
            )
            if table_member is None:
                return None
            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            table_relationships = [
                relationship
                for relationship in worksheet_relationships.findall(relationship_tag)
                if relationship.get("Id") == table_relationship_id
                and relationship.get("Type") == _TABLE_RELATIONSHIP
                and relationship.get("TargetMode") is None
            ]
            if len(table_relationships) != 1:
                return None
            table = ElementTree.fromstring(archive.read(table_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError, ValueError):
        return None

    table_tag = f"{{{_SPREADSHEETML_NS}}}table"
    auto_filter_tag = f"{{{_SPREADSHEETML_NS}}}autoFilter"
    table_columns_tag = f"{{{_SPREADSHEETML_NS}}}tableColumns"
    table_column_tag = f"{{{_SPREADSHEETML_NS}}}tableColumn"
    calculated_formula_tag = f"{{{_SPREADSHEETML_NS}}}calculatedColumnFormula"
    if table.tag != table_tag or len(table) != 2:
        return None
    auto_filters = table.findall(auto_filter_tag)
    table_columns = table.findall(table_columns_tag)
    if (
        len(auto_filters) != 1
        or len(auto_filters[0])
        or len(table_columns) != 1
        or table_columns[0].get("count") != "3"
    ):
        return None
    columns = list(table_columns[0])
    if len(columns) != 3 or any(column.tag != table_column_tag for column in columns):
        return None
    calculated_columns = [
        column for column in columns if len(column) == 1 and column[0].tag == calculated_formula_tag
    ]
    if len(calculated_columns) != 1 or calculated_columns[0] is not columns[-1]:
        return None
    calculated_formula = calculated_columns[0][0]
    formula_text = calculated_formula.text
    if (
        calculated_formula.attrib
        or len(calculated_formula)
        or not isinstance(formula_text, str)
        or not formula_text
        or any(len(column) for column in columns[:-1])
    ):
        return None
    return {
        "worksheet_member": worksheet_member,
        "worksheet_relationships_member": worksheet_relationships_member,
        "worksheet_table_part_attributes": tuple(sorted(table_part.attrib.items())),
        "table_relationship_attributes": tuple(sorted(table_relationships[0].attrib.items())),
        "table_member": table_member,
        "table_attributes": tuple(sorted(table.attrib.items())),
        "auto_filter_attributes": tuple(sorted(auto_filters[0].attrib.items())),
        "table_columns_attributes": tuple(sorted(table_columns[0].attrib.items())),
        "table_column_attributes": tuple(
            tuple(sorted(column.attrib.items())) for column in columns
        ),
        "calculated_column_formula_attributes": tuple(sorted(calculated_formula.attrib.items())),
        "calculated_column_formula": formula_text,
    }


def _table_without_calculated_column_formula(path: Path, sheet_name: str) -> bytes | None:
    """Return WCAB's Table part with only its formula-master text erased."""

    state = _raw_table_calculated_column_formula_state(path, sheet_name)
    if state is None:
        return None
    try:
        with ZipFile(path) as archive:
            table = ElementTree.fromstring(archive.read(state["table_member"]))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    formula_tag = f"{{{_SPREADSHEETML_NS}}}calculatedColumnFormula"
    formulas = list(table.iter(formula_tag))
    if len(formulas) != 1 or formulas[0].text is None:
        return None
    formulas[0].text = None
    return ElementTree.tostring(table, encoding="utf-8", xml_declaration=True)


def _raw_power_pivot_data_model_relationship_state(path: Path) -> dict[str, Any] | None:
    """Read WCAB's one relationship-backed Power Pivot/Data Model declaration.

    ``xl/model/item.data`` is an opaque Analysis Services payload.  This
    reader establishes its bounded package identity and fixed digest without
    parsing it, then reads only the one explicit x15 relationship declaration.
    It never executes DAX, refreshes a model, or opens a client report.
    """

    workbook_member = "xl/workbook.xml"
    workbook_relationships_member = "xl/_rels/workbook.xml.rels"
    relationship_id = "rIdWCABPowerPivotData"
    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read(workbook_member))
            workbook_relationships = ElementTree.fromstring(
                archive.read(workbook_relationships_member)
            )
            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
            relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
            relationship_matches = [
                relationship
                for relationship in workbook_relationships.findall(relationship_tag)
                if relationship.get("Id") == relationship_id
                and relationship.get("Type") == _POWER_PIVOT_DATA_RELATIONSHIP
                and relationship.get("TargetMode") is None
            ]
            if (
                len(relationship_matches) != 1
                or sum(
                    relationship.get("Type") == _POWER_PIVOT_DATA_RELATIONSHIP
                    for relationship in workbook_relationships.findall(relationship_tag)
                )
                != 1
            ):
                return None
            data_model_member = _relationship_part_member(
                workbook_relationships,
                relationship_id,
                workbook_member,
                relationship_type=_POWER_PIVOT_DATA_RELATIONSHIP,
            )
            if data_model_member is None:
                return None
            payload_info = archive.getinfo(data_model_member)
            if payload_info.file_size != len(_POWER_PIVOT_DATA_PAYLOAD):
                return None
            payload = archive.read(data_model_member)
            data_model_members = tuple(
                sorted(
                    member
                    for member in archive.namelist()
                    if member.startswith("xl/model/")
                    and not member.startswith("xl/model/_rels/")
                    and not member.endswith("/")
                )
            )
            data_model_relationship_members = tuple(
                sorted(
                    member for member in archive.namelist() if member.startswith("xl/model/_rels/")
                )
            )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError, ValueError):
        return None

    ext_list_tag = f"{{{_SPREADSHEETML_NS}}}extLst"
    extension_tag = f"{{{_SPREADSHEETML_NS}}}ext"
    data_model_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}dataModel"
    model_tables_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}modelTables"
    model_table_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}modelTable"
    model_relationships_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}modelRelationships"
    model_relationship_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}modelRelationship"
    default_tag = f"{{{_CONTENT_TYPES_NS}}}Default"
    ext_lists = workbook.findall(ext_list_tag)
    data_models = list(workbook.iter(data_model_tag))
    if (
        workbook_relationships.tag != f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        or content_types.tag != f"{{{_CONTENT_TYPES_NS}}}Types"
        or len(ext_lists) != 1
        or ext_lists[0].attrib
        or len(ext_lists[0]) != 1
        or ext_lists[0][0].tag != extension_tag
        or ext_lists[0][0].attrib != {"uri": "{FCE2AD5D-F65C-4FA6-A056-5C36A1767C68}"}
        or len(ext_lists[0][0]) != 1
        or ext_lists[0][0][0].tag != data_model_tag
        or len(data_models) != 1
        or data_models[0] is not ext_lists[0][0][0]
    ):
        return None
    data_model = data_models[0]
    model_tables = data_model.findall(model_tables_tag)
    model_relationships = data_model.findall(model_relationships_tag)
    if (
        data_model.attrib != {"minVersionLoad": "5"}
        or len(data_model) != 2
        or len(model_tables) != 1
        or model_tables[0].attrib
        or len(model_tables[0]) != 2
        or any(table.tag != model_table_tag or len(table) for table in model_tables[0])
        or len(model_relationships) != 1
        or model_relationships[0].attrib
        or len(model_relationships[0]) != 1
        or model_relationships[0][0].tag != model_relationship_tag
        or len(model_relationships[0][0])
    ):
        return None
    data_model_content_types = [
        content_type
        for content_type in content_types.findall(default_tag)
        if content_type.get("Extension") == "data"
    ]
    if len(data_model_content_types) != 1:
        return None
    return {
        "workbook_member": workbook_member,
        "workbook_relationships_member": workbook_relationships_member,
        "data_model_member": data_model_member,
        "workbook_relationship_attributes": tuple(sorted(relationship_matches[0].attrib.items())),
        "data_model_content_type_attributes": tuple(
            sorted(data_model_content_types[0].attrib.items())
        ),
        "extension_attributes": tuple(sorted(ext_lists[0][0].attrib.items())),
        "data_model_attributes": tuple(sorted(data_model.attrib.items())),
        "model_table_attributes": tuple(
            tuple(sorted(table.attrib.items())) for table in model_tables[0]
        ),
        "model_relationship_attributes": tuple(sorted(model_relationships[0][0].attrib.items())),
        "data_model_members": data_model_members,
        "data_model_relationship_members": data_model_relationship_members,
        "data_model_payload_sha256": sha256(payload).hexdigest(),
        "data_model_payload_size": len(payload),
    }


def _workbook_without_power_pivot_data_model_to_column(path: Path) -> bytes | None:
    """Return WCAB's workbook XML with only its model target key removed."""

    if _raw_power_pivot_data_model_relationship_state(path) is None:
        return None
    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    relationship_tag = f"{{{_OFFICE_2013_SPREADSHEET_NS}}}modelRelationship"
    relationships = list(workbook.iter(relationship_tag))
    if len(relationships) != 1 or "toColumn" not in relationships[0].attrib:
        return None
    relationships[0].attrib.pop("toColumn")
    return ElementTree.tostring(workbook, encoding="utf-8", xml_declaration=True)


def _raw_xlm_auto_open_binding_state(path: Path) -> dict[str, Any] | None:
    """Read WCAB's one raw XLM Auto_Open dispatch declaration.

    The generated macro sheet has a fixed, tiny XML payload. This reader
    validates only its package wiring, direct A1 target, and exact inert
    formula cells; it never opens Excel, evaluates an XLM instruction, or
    follows a relationship target beyond the local package member.
    """

    workbook_member = "xl/workbook.xml"
    workbook_relationships_member = "xl/_rels/workbook.xml.rels"
    macro_sheet_member = "xl/macrosheets/sheet1.xml"
    relationship_id = "rIdWCABXlmMacroSheet"
    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read(workbook_member))
            workbook_relationships = ElementTree.fromstring(
                archive.read(workbook_relationships_member)
            )
            content_types = ElementTree.fromstring(archive.read("[Content_Types].xml"))
            macro_sheet_payload = archive.read(macro_sheet_member)
            macro_sheet = ElementTree.fromstring(macro_sheet_payload)
            macro_sheet_members = tuple(
                sorted(
                    member
                    for member in archive.namelist()
                    if member.startswith("xl/macrosheets/")
                    and not member.startswith("xl/macrosheets/_rels/")
                    and not member.endswith("/")
                )
            )
            macro_sheet_relationship_members = tuple(
                sorted(
                    member
                    for member in archive.namelist()
                    if member.startswith("xl/macrosheets/_rels/")
                )
            )
            vba_project_members = tuple(
                sorted(
                    member for member in archive.namelist() if member.startswith("xl/vbaProject")
                )
            )
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None

    relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
    sheets_tag = f"{{{_SPREADSHEETML_NS}}}sheets"
    sheet_tag = f"{{{_SPREADSHEETML_NS}}}sheet"
    defined_names_tag = f"{{{_SPREADSHEETML_NS}}}definedNames"
    defined_name_tag = f"{{{_SPREADSHEETML_NS}}}definedName"
    override_tag = f"{{{_CONTENT_TYPES_NS}}}Override"
    macro_sheet_tag = f"{{{_XLM_AUTO_OPEN_MACRO_SHEET_NS}}}macrosheet"
    sheet_data_tag = f"{{{_XLM_AUTO_OPEN_MACRO_SHEET_NS}}}sheetData"
    row_tag = f"{{{_XLM_AUTO_OPEN_MACRO_SHEET_NS}}}row"
    cell_tag = f"{{{_XLM_AUTO_OPEN_MACRO_SHEET_NS}}}c"
    formula_tag = f"{{{_XLM_AUTO_OPEN_MACRO_SHEET_NS}}}f"
    sheets = workbook.findall(sheets_tag)
    defined_names = workbook.findall(defined_names_tag)
    if (
        workbook_relationships.tag != f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationships"
        or content_types.tag != f"{{{_CONTENT_TYPES_NS}}}Types"
        or len(sheets) != 1
        or sheets[0].attrib
        or [sheet.get("name") for sheet in sheets[0]]
        != ["Inputs", "Model", "Dashboard", "Macro Automation"]
        or any(sheet.tag != sheet_tag for sheet in sheets[0])
        or len(defined_names) != 1
        or defined_names[0].attrib
        or len(defined_names[0]) != 1
        or defined_names[0][0].tag != defined_name_tag
        or defined_names[0][0].attrib != {"name": "_xlnm.Auto_Open"}
        or len(defined_names[0][0])
        or not isinstance(defined_names[0][0].text, str)
        or not defined_names[0][0].text
    ):
        return None
    macro_sheets = [sheet for sheet in sheets[0] if sheet.get("name") == "Macro Automation"]
    if len(macro_sheets) != 1:
        return None
    macro_sheet_declaration = macro_sheets[0]
    relationship_matches = [
        relationship
        for relationship in workbook_relationships.findall(relationship_tag)
        if relationship.get("Id") == relationship_id
        and relationship.get("Type") == _XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP
        and relationship.get("TargetMode") is None
    ]
    if (
        len(relationship_matches) != 1
        or sum(
            relationship.get("Type") == _XLM_AUTO_OPEN_MACRO_SHEET_RELATIONSHIP
            for relationship in workbook_relationships.findall(relationship_tag)
        )
        != 1
        or relationship_matches[0].get("Target") != "macrosheets/sheet1.xml"
        or macro_sheet_members != (macro_sheet_member,)
        or macro_sheet_relationship_members
        or vba_project_members
    ):
        return None
    workbook_content_types = [
        override
        for override in content_types.findall(override_tag)
        if override.get("PartName") == "/xl/workbook.xml"
    ]
    macro_sheet_content_types = [
        override
        for override in content_types.findall(override_tag)
        if override.get("PartName") == "/xl/macrosheets/sheet1.xml"
    ]
    if (
        len(workbook_content_types) != 1
        or len(macro_sheet_content_types) != 1
        or macro_sheet.tag != macro_sheet_tag
        or macro_sheet.attrib
        or len(macro_sheet) != 1
        or macro_sheet[0].tag != sheet_data_tag
        or macro_sheet[0].attrib
        or len(macro_sheet[0]) != 2
    ):
        return None
    macro_formulas: list[tuple[str, str]] = []
    for row_number, coordinate, row in zip(("1", "2"), ("A1", "A2"), macro_sheet[0], strict=True):
        if (
            row.tag != row_tag
            or row.attrib != {"r": row_number}
            or len(row) != 1
            or row[0].tag != cell_tag
            or row[0].attrib != {"r": coordinate}
            or len(row[0]) != 1
            or row[0][0].tag != formula_tag
            or row[0][0].attrib
            or len(row[0][0])
            or row[0][0].text != "HALT()"
        ):
            return None
        macro_formulas.append((coordinate, row[0][0].text))
    return {
        "workbook_member": workbook_member,
        "workbook_relationships_member": workbook_relationships_member,
        "macro_sheet_member": macro_sheet_member,
        "macro_sheet_relationship_attributes": tuple(
            sorted(relationship_matches[0].attrib.items())
        ),
        "workbook_content_type_attributes": tuple(sorted(workbook_content_types[0].attrib.items())),
        "macro_sheet_content_type_attributes": tuple(
            sorted(macro_sheet_content_types[0].attrib.items())
        ),
        "macro_sheet_declaration_attributes": tuple(sorted(macro_sheet_declaration.attrib.items())),
        "automatic_macro_attributes": tuple(sorted(defined_names[0][0].attrib.items())),
        "automatic_macro_target": defined_names[0][0].text,
        "macro_sheet_formulas": tuple(macro_formulas),
        "macro_sheet_members": macro_sheet_members,
        "macro_sheet_relationship_members": macro_sheet_relationship_members,
        "vba_project_members": vba_project_members,
        "macro_sheet_sha256": sha256(macro_sheet_payload).hexdigest(),
        "macro_sheet_size": len(macro_sheet_payload),
    }


def _workbook_without_xlm_auto_open_target(path: Path) -> bytes | None:
    """Return WCAB's workbook XML with its Auto_Open target text erased."""

    if _raw_xlm_auto_open_binding_state(path) is None:
        return None
    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    defined_name_tag = f"{{{_SPREADSHEETML_NS}}}definedName"
    automatic_names = [
        name
        for name in workbook.iter(defined_name_tag)
        if name.attrib == {"name": "_xlnm.Auto_Open"}
    ]
    if len(automatic_names) != 1 or not automatic_names[0].text:
        return None
    automatic_names[0].text = None
    return ElementTree.tostring(workbook, encoding="utf-8", xml_declaration=True)


def _sheet_protection_sort_state(
    path: Path, sheet_name: str
) -> tuple[str, bool, bool, tuple[tuple[str, str], ...]] | None:
    """Read one generated worksheet's explicit protection and sort locks."""

    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    protections = worksheet.findall(f"{{{_SPREADSHEETML_NS}}}sheetProtection")
    if len(protections) != 1:
        return None
    protection = protections[0]
    protected = _ooxml_boolean(protection.get("sheet"))
    sort_locked = _ooxml_boolean(protection.get("sort"))
    if protected is None or sort_locked is None:
        return None
    return (
        worksheet_member,
        protected,
        sort_locked,
        tuple(sorted(protection.attrib.items())),
    )


def _worksheet_without_sheet_protection_sort_control(path: Path, sheet_name: str) -> bytes | None:
    """Return one worksheet with its explicit ``sheetProtection/@sort`` erased."""

    if _sheet_protection_sort_state(path, sheet_name) is None:
        return None
    try:
        with ZipFile(path) as archive:
            worksheet_member = _worksheet_member_for_sheet(archive, sheet_name)
            if worksheet_member is None:
                return None
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    protection = worksheet.find(f"{{{_SPREADSHEETML_NS}}}sheetProtection")
    if protection is None or "sort" not in protection.attrib:
        return None
    protection.attrib.pop("sort")
    return ElementTree.tostring(worksheet, encoding="utf-8", xml_declaration=True)


def _xlsx_member_differences(baseline_path: Path, candidate_path: Path) -> set[str] | None:
    """Return changed member names when two small fixture archives align."""

    try:
        with ZipFile(baseline_path) as baseline, ZipFile(candidate_path) as candidate:
            baseline_members = {
                entry.filename: baseline.read(entry.filename) for entry in baseline.infolist()
            }
            candidate_members = {
                entry.filename: candidate.read(entry.filename) for entry in candidate.infolist()
            }
    except (BadZipFile, OSError):
        return None
    if set(baseline_members) != set(candidate_members):
        return None
    return {
        member
        for member in baseline_members
        if baseline_members[member] != candidate_members[member]
    }


def _dynamic_array_cell_metadata_indexes(archive: ZipFile) -> set[int] | None:
    """Read the generated package's cell-metadata bindings for dynamic arrays.

    This mirrors only the compact OOXML relationship used by WCAB's fixture:
    a workbook ``sheetMetadata`` relationship, an ``XLDAPR`` future-metadata
    record with ``fDynamic=true``, and one-based ``cm`` indices into
    ``cellMetadata``. It intentionally does not calculate formulas or infer a
    dynamic result extent.
    """

    try:
        relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    relationship_tag = f"{{{_PACKAGE_RELATIONSHIPS_NS}}}Relationship"
    targets = [
        relationship.get("Target")
        for relationship in relationships.findall(relationship_tag)
        if relationship.get("Type") == f"{_DOCUMENT_RELATIONSHIPS_NS}/sheetMetadata"
    ]
    if not targets:
        return set()
    if len(targets) != 1 or targets[0] != "metadata.xml":
        return None
    try:
        metadata = ElementTree.fromstring(archive.read("xl/metadata.xml"))
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None

    def tag(name: str) -> str:
        return f"{{{_SPREADSHEETML_NS}}}{name}"

    metadata_types = [
        item.get("name", "")
        for item in metadata.findall(f"./{tag('metadataTypes')}/{tag('metadataType')}")
    ]
    dynamic_future_indexes: dict[str, set[int]] = {}
    dynamic_properties_tag = f"{{{_DYNAMIC_ARRAY_NS}}}dynamicArrayProperties"
    for future_metadata in metadata.findall(tag("futureMetadata")):
        name = future_metadata.get("name")
        if not name:
            continue
        indexes = {
            index
            for index, block in enumerate(future_metadata.findall(tag("bk")))
            if any(
                item.tag == dynamic_properties_tag
                and item.get("fDynamic", "").casefold() in {"1", "true"}
                for item in block.iter()
            )
        }
        if indexes:
            dynamic_future_indexes[name] = indexes

    cell_metadata = metadata.find(tag("cellMetadata"))
    if cell_metadata is None:
        return set()
    dynamic_cells: set[int] = set()
    for cell_index, block in enumerate(cell_metadata.findall(tag("bk")), start=1):
        for record in block.findall(tag("rc")):
            try:
                type_index = int(record.get("t", ""))
                value_index = int(record.get("v", ""))
            except ValueError:
                continue
            if not 1 <= type_index <= len(metadata_types):
                continue
            if value_index in dynamic_future_indexes.get(metadata_types[type_index - 1], set()):
                dynamic_cells.add(cell_index)
                break
    return dynamic_cells


def _array_formula_state(
    path: Path, sheet_name: str, coordinate: str
) -> tuple[str, str | None, str | None] | None:
    """Classify a generated raw-OOXML array anchor as CSE or dynamic.

    The returned tuple is ``(mode, formula_text, output_range)``. This helper
    is deliberately limited to the generated fixture's anchor representation;
    any incomplete metadata becomes ``unclassified`` rather than a guess.
    """

    try:
        with ZipFile(path) as archive:
            workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            sheets = workbook.find(f"{{{_SPREADSHEETML_NS}}}sheets")
            if sheets is None:
                return None
            sheet_tag = f"{{{_SPREADSHEETML_NS}}}sheet"
            matching_sheets = [
                sheet for sheet in sheets.findall(sheet_tag) if sheet.get("name") == sheet_name
            ]
            if len(matching_sheets) != 1:
                return None
            relationship_id = matching_sheets[0].get(f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id")
            if not isinstance(relationship_id, str):
                return None
            relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            target = _relationship_target(
                relationships,
                relationship_id,
                relationship_type=f"{_DOCUMENT_RELATIONSHIPS_NS}/worksheet",
            )
            if target is None:
                return None
            worksheet_member = target if target.startswith("xl/") else f"xl/{target}"
            worksheet = ElementTree.fromstring(archive.read(worksheet_member))
            cell_tag = f"{{{_SPREADSHEETML_NS}}}c"
            cells = [cell for cell in worksheet.iter(cell_tag) if cell.get("r") == coordinate]
            if len(cells) != 1:
                return None
            formula = cells[0].find(f"{{{_SPREADSHEETML_NS}}}f")
            if formula is None:
                return None
            formula_text = f"={formula.text or ''}"
            output_range = formula.get("ref")
            if formula.get("t") != "array":
                return "ordinary", formula_text, output_range
            metadata_index = cells[0].get("cm")
            if metadata_index is None:
                return "legacy_cse", formula_text, output_range
            try:
                cell_metadata_index = int(metadata_index)
            except ValueError:
                return "unclassified", formula_text, output_range
            dynamic_cells = _dynamic_array_cell_metadata_indexes(archive)
            if dynamic_cells is None:
                return "unclassified", formula_text, output_range
    except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
        return None
    mode = "dynamic" if cell_metadata_index in dynamic_cells else "unclassified"
    return mode, formula_text, output_range


def _direct_graph(workbook: Workbook) -> dict[tuple[str, str], set[tuple[str, str]]]:
    """Return a small static local A1 dependency graph for the generated fixtures.

    This deliberately recognizes only ordinary direct A1 references.  The
    benchmark labels this as a lower bound rather than using it as an Excel
    calculation engine.
    """

    graph: dict[tuple[str, str], set[tuple[str, str]]] = {}
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    continue
                # External references are separate benchmark facts.  Avoid
                # treating an external sheet label as a local dependency.
                if "[" in cell.value:
                    continue
                dependent = (worksheet.title, cell.coordinate)
                for match in _CELL_REFERENCE.finditer(cell.value):
                    sheet = match.group("quoted") or match.group("plain") or worksheet.title
                    sheet = sheet.replace("''", "'")
                    if sheet not in workbook.sheetnames:
                        continue
                    source = (sheet, match.group("cell").replace("$", ""))
                    graph.setdefault(source, set()).add(dependent)
    return graph


def _reachable(
    graph: dict[tuple[str, str], set[tuple[str, str]]], source: tuple[str, str]
) -> set[tuple[str, str]]:
    discovered: set[tuple[str, str]] = set()
    pending = deque(graph.get(source, ()))
    while pending:
        current = pending.popleft()
        if current in discovered:
            continue
        discovered.add(current)
        pending.extend(graph.get(current, ()))
    return discovered


def _workbook_pair(
    case_dir: Path, truth: dict[str, Any], workbook: str | None = None
) -> tuple[Path, Path]:
    topology = truth.get("topology")
    if topology == "pair":
        if workbook is not None:
            raise FixtureValidationError(
                f"{case_dir}: pair fixture unexpectedly names workbook {workbook!r}"
            )
        pairs = [
            (case_dir / f"baseline.{extension}", case_dir / f"candidate.{extension}")
            for extension in _PAIR_WORKBOOK_EXTENSIONS
            if (case_dir / f"baseline.{extension}").is_file()
            and (case_dir / f"candidate.{extension}").is_file()
        ]
        if len(pairs) != 1:
            expected = " or ".join(
                f"baseline.{extension}/candidate.{extension}"
                for extension in _PAIR_WORKBOOK_EXTENSIONS
            )
            raise FixtureValidationError(
                f"{case_dir}: expected exactly one paired workbook set ({expected})"
            )
        return pairs[0]
    if topology == "portfolio":
        if not workbook:
            raise FixtureValidationError(f"{case_dir}: portfolio fact must name a workbook")
        return case_dir / "baseline" / workbook, case_dir / "candidate" / workbook
    raise FixtureValidationError(f"{case_dir}: unsupported topology {topology!r}")


def _assert(condition: bool, message: str, errors: list[str]) -> None:
    if not condition:
        errors.append(message)


def _validate_fact(
    case_dir: Path, truth: dict[str, Any], fact: dict[str, Any], errors: list[str]
) -> None:
    kind = fact.get("kind")
    baseline_path, candidate_path = _workbook_pair(case_dir, truth, fact.get("workbook"))
    baseline = _load_workbook(baseline_path)
    candidate = _load_workbook(candidate_path)

    if kind in {
        "formula_to_value",
        "formula_changed",
        "value_changed",
        "external_formula_added",
        "formula_cell_unlocked",
        "dynamic_formula_reference_added",
    }:
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        if sheet_name not in baseline.sheetnames or sheet_name not in candidate.sheetnames:
            errors.append(f"{truth['id']}: sheet {sheet_name!r} is absent")
            return
        before = baseline[sheet_name][coordinate]
        after = candidate[sheet_name][coordinate]
        if kind == "formula_to_value":
            _assert(
                _cell_kind(before) == "formula" and _cell_kind(after) == "value",
                f"{truth['id']}: expected {sheet_name}!{coordinate} formula-to-value change",
                errors,
            )
        elif kind == "formula_changed":
            _assert(
                _cell_kind(before) == "formula"
                and _cell_kind(after) == "formula"
                and before.value != after.value,
                f"{truth['id']}: expected {sheet_name}!{coordinate} changed formula",
                errors,
            )
        elif kind == "value_changed":
            _assert(
                _cell_kind(before) == "value"
                and _cell_kind(after) == "value"
                and before.value != after.value,
                f"{truth['id']}: expected {sheet_name}!{coordinate} changed literal value",
                errors,
            )
        elif kind == "external_formula_added":
            _assert(
                _cell_kind(after) == "formula" and "[" in str(after.value),
                f"{truth['id']}: expected external formula at {sheet_name}!{coordinate}",
                errors,
            )
        elif kind == "dynamic_formula_reference_added":
            expected_functions = fact.get("functions")
            before_functions = _dynamic_reference_functions(str(before.value))
            after_functions = _dynamic_reference_functions(str(after.value))
            _assert(
                _cell_kind(before) == "formula"
                and _cell_kind(after) == "formula"
                and before.value != after.value
                and not before_functions
                and isinstance(expected_functions, list)
                and all(isinstance(function, str) for function in expected_functions)
                and after_functions == tuple(expected_functions),
                f"{truth['id']}: expected dynamic reference functions {expected_functions!r} at {sheet_name}!{coordinate}",
                errors,
            )
        else:
            _assert(
                baseline[sheet_name].protection.sheet
                and candidate[sheet_name].protection.sheet
                and _cell_kind(before) == "formula"
                and bool(before.protection.locked)
                and not bool(after.protection.locked),
                f"{truth['id']}: expected protected formula {sheet_name}!{coordinate} to be explicitly unlocked",
                errors,
            )
        return

    if kind == "workbook_structure_lock_removed":
        hidden_sheet = fact.get("hidden_sheet")
        hidden_sheet_state = fact.get("hidden_sheet_state")
        formula_sheet = fact.get("formula_sheet")
        formula_cell = fact.get("formula_cell")
        formula = fact.get("formula")
        before_structure_state = _workbook_structure_protection_state(baseline_path)
        after_structure_state = _workbook_structure_protection_state(candidate_path)
        before_formula = (
            _raw_cell_state(baseline_path, formula_sheet, formula_cell)
            if isinstance(formula_sheet, str) and isinstance(formula_cell, str)
            else None
        )
        after_formula = (
            _raw_cell_state(candidate_path, formula_sheet, formula_cell)
            if isinstance(formula_sheet, str) and isinstance(formula_cell, str)
            else None
        )
        _assert(
            fact.get("baseline_lock_structure") is True
            and fact.get("candidate_lock_structure") is False
            and hidden_sheet == "ReviewControls"
            and hidden_sheet_state == "hidden"
            and formula_sheet == "Inputs"
            and formula_cell == "D2"
            and formula == "=B2*C2"
            and hidden_sheet in baseline.sheetnames
            and hidden_sheet in candidate.sheetnames
            and formula_sheet in baseline.sheetnames
            and formula_sheet in candidate.sheetnames
            and before_structure_state == (True, (("lockStructure", "1"),))
            and after_structure_state == (False, (("lockStructure", "0"),))
            and bool(baseline.security.lockStructure)
            and not bool(candidate.security.lockStructure)
            and baseline[hidden_sheet].sheet_state == hidden_sheet_state
            and candidate[hidden_sheet].sheet_state == hidden_sheet_state
            and before_formula is not None
            and after_formula is not None
            and before_formula == after_formula
            and before_formula[3] == formula
            and _workbook_without_structure_lock(baseline_path)
            == _workbook_without_structure_lock(candidate_path)
            and _xlsx_member_differences(baseline_path, candidate_path) == {"xl/workbook.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected only workbookProtection lockStructure true -> false with stable hidden-sheet and formula context",
            errors,
        )
        return

    if kind == "sheet_protection_sort_permission_enabled":
        sheet_name = fact.get("sheet")
        formula_cell = fact.get("formula_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = (
            _sheet_protection_sort_state(baseline_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        after_state = (
            _sheet_protection_sort_state(candidate_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        before_formula = (
            _raw_cell_state(baseline_path, sheet_name, formula_cell)
            if isinstance(sheet_name, str) and isinstance(formula_cell, str)
            else None
        )
        after_formula = (
            _raw_cell_state(candidate_path, sheet_name, formula_cell)
            if isinstance(sheet_name, str) and isinstance(formula_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        graph = _direct_graph(candidate)
        _assert(
            sheet_name == "Controls"
            and fact.get("worksheet_member") == "xl/worksheets/sheet1.xml"
            and fact.get("baseline_sort_locked") is True
            and fact.get("candidate_sort_locked") is False
            and formula_cell == "D2"
            and fact.get("formula") == "=B2*C2"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Controls!$D$2"
            and before_state is not None
            and after_state is not None
            and before_state[:3] == ("xl/worksheets/sheet1.xml", True, True)
            and after_state[:3] == ("xl/worksheets/sheet1.xml", True, False)
            and before_state[3] != after_state[3]
            and _ooxml_boolean(dict(before_state[3]).get("sort")) is True
            and _ooxml_boolean(dict(after_state[3]).get("sort")) is False
            and sheet_name in baseline.sheetnames
            and sheet_name in candidate.sheetnames
            and bool(baseline[sheet_name].protection.sheet)
            and bool(candidate[sheet_name].protection.sheet)
            and bool(baseline[sheet_name].protection.sort)
            and not bool(candidate[sheet_name].protection.sort)
            and before_formula is not None
            and after_formula is not None
            and before_formula == after_formula
            and before_formula[3] == fact.get("formula")
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == fact.get("dashboard_formula")
            and _worksheet_without_sheet_protection_sort_control(baseline_path, sheet_name)
            == _worksheet_without_sheet_protection_sort_control(candidate_path, sheet_name)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {"xl/worksheets/sheet1.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (dashboard_sheet, dashboard_cell) in _reachable(graph, (sheet_name, formula_cell)),
            f"{truth['id']}: expected only one protected-sheet sort lock true -> false with stable formula context",
            errors,
        )
        return

    if kind == "named_sheet_view_filter_criterion_changed":
        sheet_name = fact.get("sheet")
        subtotal_cell = fact.get("subtotal_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = (
            _raw_named_sheet_view_filter_state(baseline_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        after_state = (
            _raw_named_sheet_view_filter_state(candidate_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        before_subtotal = (
            _raw_cell_state(baseline_path, sheet_name, subtotal_cell)
            if isinstance(sheet_name, str) and isinstance(subtotal_cell, str)
            else None
        )
        after_subtotal = (
            _raw_cell_state(candidate_path, sheet_name, subtotal_cell)
            if isinstance(sheet_name, str) and isinstance(subtotal_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_base_auto_filter_attributes = (
            ("ref", "A1:B5"),
            (
                "{http://schemas.microsoft.com/office/spreadsheetml/2014/revision}uid",
                "{00000000-0001-0000-0000-000000000000}",
            ),
        )
        expected_relationship_attributes = (
            ("Id", "rIdWCABNamedSheetView"),
            ("Target", "../namedSheetViews/namedSheetView1.xml"),
            (
                "Type",
                "http://schemas.microsoft.com/office/2019/04/relationships/namedSheetView",
            ),
        )
        expected_content_type_attributes = (
            ("ContentType", "application/vnd.ms-excel.namedsheetviews+xml"),
            ("PartName", "/xl/namedSheetViews/namedSheetView1.xml"),
        )
        expected_named_view_attributes = (
            ("id", "{11111111-1111-1111-1111-111111111111}"),
            ("name", "WCAB regional review"),
        )
        expected_named_filter_attributes = (
            ("filterId", "{00000000-0001-0000-0000-000000000000}"),
            ("ref", "A1:B5"),
            ("tableId", "0"),
        )
        expected_column_filter_attributes = (
            ("colId", "0"),
            ("id", "{22222222-2222-2222-2222-222222222222}"),
        )
        _assert(
            sheet_name == "Report"
            and fact.get("view_member") == "xl/namedSheetViews/namedSheetView1.xml"
            and fact.get("base_filter_ref") == "A1:B5"
            and fact.get("filter_column_id") == 0
            and fact.get("baseline_filter_value") == "North"
            and fact.get("candidate_filter_value") == "South"
            and subtotal_cell == "D2"
            and fact.get("subtotal_formula") == "=SUBTOTAL(109,B2:B5)"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Report!$D$2"
            and sheet_name in baseline.sheetnames
            and sheet_name in candidate.sheetnames
            and dashboard_sheet in baseline.sheetnames
            and dashboard_sheet in candidate.sheetnames
            and before_state is not None
            and after_state is not None
            and before_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and after_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and before_state["relationships_member"] == "xl/worksheets/_rels/sheet1.xml.rels"
            and after_state["relationships_member"] == "xl/worksheets/_rels/sheet1.xml.rels"
            and before_state["view_member"] == "xl/namedSheetViews/namedSheetView1.xml"
            and after_state["view_member"] == "xl/namedSheetViews/namedSheetView1.xml"
            and before_state["base_auto_filter_attributes"] == expected_base_auto_filter_attributes
            and after_state["base_auto_filter_attributes"] == expected_base_auto_filter_attributes
            and before_state["relationship_attributes"] == expected_relationship_attributes
            and after_state["relationship_attributes"] == expected_relationship_attributes
            and before_state["content_type_attributes"] == expected_content_type_attributes
            and after_state["content_type_attributes"] == expected_content_type_attributes
            and before_state["named_view_attributes"] == expected_named_view_attributes
            and after_state["named_view_attributes"] == expected_named_view_attributes
            and before_state["named_filter_attributes"] == expected_named_filter_attributes
            and after_state["named_filter_attributes"] == expected_named_filter_attributes
            and before_state["column_filter_attributes"] == expected_column_filter_attributes
            and after_state["column_filter_attributes"] == expected_column_filter_attributes
            and before_state["filter_column_attributes"] == (("colId", "0"),)
            and after_state["filter_column_attributes"] == (("colId", "0"),)
            and before_state["filters_attributes"] == (("blank", "0"),)
            and after_state["filters_attributes"] == (("blank", "0"),)
            and before_state["criterion_attributes"] == (("val", "North"),)
            and after_state["criterion_attributes"] == (("val", "South"),)
            and before_state["criterion_value"] == "North"
            and after_state["criterion_value"] == "South"
            and before_subtotal is not None
            and after_subtotal is not None
            and before_subtotal == after_subtotal
            and before_subtotal[0] == "xl/worksheets/sheet1.xml"
            and before_subtotal[3] == "=SUBTOTAL(109,B2:B5)"
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[0] == "xl/worksheets/sheet2.xml"
            and before_dashboard[3] == "=Report!$D$2"
            and _named_sheet_view_without_filter_criterion(baseline_path, sheet_name)
            == _named_sheet_view_without_filter_criterion(candidate_path, sheet_name)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {"xl/namedSheetViews/namedSheetView1.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected only one relationship-backed Named Sheet View criterion North -> South with stable base filter, formulas, and package boundary",
            errors,
        )
        return

    if kind == "xml_map_table_column_xpath_retargeted":
        sheet_name = fact.get("sheet")
        total_cell = fact.get("total_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        single_cell = fact.get("single_cell")
        before_state = (
            _raw_xml_map_table_binding_state(baseline_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        after_state = (
            _raw_xml_map_table_binding_state(candidate_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        before_total = (
            _raw_cell_state(baseline_path, sheet_name, total_cell)
            if isinstance(sheet_name, str) and isinstance(total_cell, str)
            else None
        )
        after_total = (
            _raw_cell_state(candidate_path, sheet_name, total_cell)
            if isinstance(sheet_name, str) and isinstance(total_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        before_single_cell = (
            _raw_cell_state(baseline_path, sheet_name, single_cell)
            if isinstance(sheet_name, str) and isinstance(single_cell, str)
            else None
        )
        after_single_cell = (
            _raw_cell_state(candidate_path, sheet_name, single_cell)
            if isinstance(sheet_name, str) and isinstance(single_cell, str)
            else None
        )
        expected_table_part_attributes = (
            (
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id",
                "rId1",
            ),
        )
        expected_table_relationship_attributes = (
            ("Id", "rId1"),
            ("Target", "/xl/tables/table1.xml"),
            (
                "Type",
                "http://schemas.openxmlformats.org/officeDocument/2006/relationships/table",
            ),
        )
        expected_table_attributes = (
            ("connectionId", "7"),
            ("displayName", "InvoiceLines"),
            ("headerRowCount", "1"),
            ("id", "1"),
            ("name", "InvoiceLines"),
            ("ref", "A1:B3"),
            ("tableType", "xml"),
        )
        expected_map_relationship_attributes = (
            ("Id", "rIdWCABXmlMaps"),
            ("Target", "xmlMaps.xml"),
            (
                "Type",
                "http://schemas.openxmlformats.org/officeDocument/2006/relationships/xmlMaps",
            ),
        )
        expected_map_attributes = (
            ("Append", "false"),
            ("AutoFit", "true"),
            ("ID", "1"),
            ("Name", "WCAB invoice export"),
            ("PreserveFormat", "true"),
            ("PreserveSortAFLayout", "true"),
            ("RootElement", "Invoice"),
            ("SchemaID", "WCAB-INVOICE-EXPORT"),
            ("ShowImportExportValidationErrors", "false"),
        )
        expected_data_binding_attributes = (
            ("ConnectionID", "7"),
            ("DataBindingLoadMode", "1"),
            ("DataBindingName", "WCAB invoice export binding"),
            ("FileBinding", "true"),
            ("FileBindingName", "wcab-invoice-export.xml"),
        )
        expected_single_cell_relationship_attributes = (
            ("Id", "rIdWCABXmlMappingSingleCells"),
            ("Target", "../singleCellTables/singleCellTable1.xml"),
            (
                "Type",
                "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
                "tableSingleCells",
            ),
        )
        expected_single_cell_content_type_attributes = (
            ("ContentType", "application/vnd.ms-excel.tableSingleCells"),
            ("PartName", "/xl/singleCellTables/singleCellTable1.xml"),
        )
        expected_common_state = {
            "worksheet_member": "xl/worksheets/sheet1.xml",
            "worksheet_relationships_member": "xl/worksheets/_rels/sheet1.xml.rels",
            "worksheet_table_part_attributes": expected_table_part_attributes,
            "table_relationship_attributes": expected_table_relationship_attributes,
            "table_member": "xl/tables/table1.xml",
            "table_attributes": expected_table_attributes,
            "table_column_attributes": (("id", "2"), ("name", "Net amount")),
            "map_relationship_attributes": expected_map_relationship_attributes,
            "map_member": "xl/xmlMaps.xml",
            "map_info_attributes": (
                ("SelectionNamespaces", "xmlns:wcab='urn:wcab:invoice-export'"),
            ),
            "schema_attributes": (("ID", "WCAB-INVOICE-EXPORT"),),
            "schema_definition_attributes": (
                ("elementFormDefault", "qualified"),
                ("targetNamespace", "urn:wcab:invoice-export"),
            ),
            "schema_root_element_attributes": (("name", "Invoice"),),
            "schema_element_attributes": (
                (("name", "Invoice"),),
                (("minOccurs", "0"), ("name", "Header")),
                (("name", "AsOf"), ("type", "xs:date")),
                (
                    ("maxOccurs", "unbounded"),
                    ("minOccurs", "0"),
                    ("name", "Line"),
                ),
                (("name", "Item"), ("type", "xs:string")),
                (("name", "NetAmount"), ("type", "xs:decimal")),
                (("name", "TaxAmount"), ("type", "xs:decimal")),
            ),
            "map_attributes": expected_map_attributes,
            "data_binding_attributes": expected_data_binding_attributes,
            "single_cell_relationship_attributes": expected_single_cell_relationship_attributes,
            "single_cell_member": "xl/singleCellTables/singleCellTable1.xml",
            "single_cell_attributes": (("connectionId", "7"), ("id", "1"), ("r", "E2")),
            "single_cell_property_attributes": (
                ("id", "1"),
                ("uniqueName", "WCAB invoice export as of"),
            ),
            "single_cell_binding_attributes": (
                ("mapId", "1"),
                ("xmlDataType", "date"),
                ("xpath", "/wcab:Invoice/wcab:Header/wcab:AsOf"),
            ),
            "xml_default_attributes": (
                ("ContentType", "application/xml"),
                ("Extension", "xml"),
            ),
            "single_cell_content_type_attributes": expected_single_cell_content_type_attributes,
        }
        expected_before_xml_column_attributes = (
            ("denormalized", "false"),
            ("mapId", "1"),
            ("xmlDataType", "double"),
            ("xpath", "/wcab:Invoice/wcab:Line/wcab:NetAmount"),
        )
        expected_after_xml_column_attributes = (
            ("denormalized", "false"),
            ("mapId", "1"),
            ("xmlDataType", "double"),
            ("xpath", "/wcab:Invoice/wcab:Line/wcab:TaxAmount"),
        )
        _assert(
            sheet_name == "Export"
            and fact.get("table_member") == "xl/tables/table1.xml"
            and fact.get("table_name") == "InvoiceLines"
            and fact.get("table_ref") == "A1:B3"
            and fact.get("mapped_column_id") == 2
            and fact.get("mapped_column_name") == "Net amount"
            and fact.get("map_member") == "xl/xmlMaps.xml"
            and fact.get("map_id") == 1
            and fact.get("schema_id") == "WCAB-INVOICE-EXPORT"
            and fact.get("connection_id") == 7
            and fact.get("baseline_xpath") == "/wcab:Invoice/wcab:Line/wcab:NetAmount"
            and fact.get("candidate_xpath") == "/wcab:Invoice/wcab:Line/wcab:TaxAmount"
            and fact.get("single_cell_member") == "xl/singleCellTables/singleCellTable1.xml"
            and single_cell == "E2"
            and fact.get("single_cell_xpath") == "/wcab:Invoice/wcab:Header/wcab:AsOf"
            and total_cell == "D2"
            and fact.get("total_formula") == "=SUM(InvoiceLines[Net amount])"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Export!$D$2"
            and sheet_name in baseline.sheetnames
            and sheet_name in candidate.sheetnames
            and dashboard_sheet in baseline.sheetnames
            and dashboard_sheet in candidate.sheetnames
            and before_state is not None
            and after_state is not None
            and all(before_state.get(key) == value for key, value in expected_common_state.items())
            and all(after_state.get(key) == value for key, value in expected_common_state.items())
            and before_state["xml_column_attributes"] == expected_before_xml_column_attributes
            and after_state["xml_column_attributes"] == expected_after_xml_column_attributes
            and before_total is not None
            and before_total == after_total
            and before_total[0] == "xl/worksheets/sheet1.xml"
            and before_total[3] == "=SUM(InvoiceLines[Net amount])"
            and before_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[0] == "xl/worksheets/sheet2.xml"
            and before_dashboard[3] == "=Export!$D$2"
            and before_single_cell is not None
            and before_single_cell == after_single_cell
            and before_single_cell[0] == "xl/worksheets/sheet1.xml"
            and _xml_map_table_without_xpath(baseline_path, sheet_name)
            == _xml_map_table_without_xpath(candidate_path, sheet_name)
            and _xlsx_member_differences(baseline_path, candidate_path) == {"xl/tables/table1.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected only one XML Map table-column XPath NetAmount -> TaxAmount with stable map, schema, cells, formulas, and package boundary",
            errors,
        )
        return

    if kind == "office_web_addin_auto_show_enabled":
        input_sheet = fact.get("input_sheet")
        input_cell = fact.get("input_cell")
        model_sheet = fact.get("model_sheet")
        model_cell = fact.get("model_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = _raw_office_web_addin_auto_show_state(baseline_path)
        after_state = _raw_office_web_addin_auto_show_state(candidate_path)
        before_input = (
            _raw_cell_state(baseline_path, input_sheet, input_cell)
            if isinstance(input_sheet, str) and isinstance(input_cell, str)
            else None
        )
        after_input = (
            _raw_cell_state(candidate_path, input_sheet, input_cell)
            if isinstance(input_sheet, str) and isinstance(input_cell, str)
            else None
        )
        before_model = (
            _raw_cell_state(baseline_path, model_sheet, model_cell)
            if isinstance(model_sheet, str) and isinstance(model_cell, str)
            else None
        )
        after_model = (
            _raw_cell_state(candidate_path, model_sheet, model_cell)
            if isinstance(model_sheet, str) and isinstance(model_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_common_state = {
            "workbook_relationship_attributes": (
                ("Id", "rIdWCABOfficeWebAddinTaskpanes"),
                ("Target", "webextensions/taskpanes.xml"),
                (
                    "Type",
                    "http://schemas.microsoft.com/office/2011/relationships/webextensiontaskpanes",
                ),
            ),
            "taskpane_member": "xl/webextensions/taskpanes.xml",
            "taskpane_content_type_attributes": (
                (
                    "ContentType",
                    "application/vnd.ms-office.webextensiontaskpanes+xml",
                ),
                ("PartName", "/xl/webextensions/taskpanes.xml"),
            ),
            "taskpane_relationships_member": "xl/webextensions/_rels/taskpanes.xml.rels",
            "taskpane_to_extension_relationship_attributes": (
                ("Id", "rIdWCABOfficeWebAddin"),
                ("Target", "webextension1.xml"),
                (
                    "Type",
                    "http://schemas.microsoft.com/office/2011/relationships/webextension",
                ),
            ),
            "taskpane_attributes": (
                ("dockstate", "right"),
                ("locked", "1"),
                ("row", "4"),
                ("visibility", "0"),
                ("width", "350"),
            ),
            "taskpane_reference_attributes": (
                (
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id",
                    "rIdWCABOfficeWebAddin",
                ),
            ),
            "extension_member": "xl/webextensions/webextension1.xml",
            "extension_attributes": (("id", "{33333333-3333-3333-3333-333333333333}"),),
            "extension_content_type_attributes": (
                ("ContentType", "application/vnd.ms-office.webextension+xml"),
                ("PartName", "/xl/webextensions/webextension1.xml"),
            ),
            "reference_attributes": (
                ("id", "{44444444-4444-4444-4444-444444444444}"),
                ("store", "wcab-review-assistant.xml"),
                ("storeType", "Filesystem"),
                ("version", "1.0.0.0"),
            ),
            "properties_attributes": (),
        }
        _assert(
            fact.get("taskpane_member") == "xl/webextensions/taskpanes.xml"
            and fact.get("web_extension_member") == "xl/webextensions/webextension1.xml"
            and fact.get("addin_id") == "{33333333-3333-3333-3333-333333333333}"
            and fact.get("reference_id") == "{44444444-4444-4444-4444-444444444444}"
            and fact.get("reference_version") == "1.0.0.0"
            and fact.get("store") == "wcab-review-assistant.xml"
            and fact.get("store_type") == "Filesystem"
            and fact.get("baseline_auto_show") is False
            and fact.get("candidate_auto_show") is True
            and input_sheet == "Inputs"
            and input_cell == "B2"
            and fact.get("input_value") == 10
            and model_sheet == "Model"
            and model_cell == "B2"
            and fact.get("model_formula") == "=Inputs!$B$2*2"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Model!$B$2"
            and all(
                sheet in baseline.sheetnames and sheet in candidate.sheetnames
                for sheet in (input_sheet, model_sheet, dashboard_sheet)
            )
            and before_state is not None
            and after_state is not None
            and all(before_state.get(key) == value for key, value in expected_common_state.items())
            and all(after_state.get(key) == value for key, value in expected_common_state.items())
            and before_state["property_attributes"]
            == (
                ("name", "Office.AutoShowTaskpaneWithDocument"),
                ("value", "false"),
            )
            and after_state["property_attributes"]
            == (
                ("name", "Office.AutoShowTaskpaneWithDocument"),
                ("value", "true"),
            )
            and before_state["auto_show"] is False
            and after_state["auto_show"] is True
            and before_input is not None
            and before_input == after_input
            and before_input[0] == "xl/worksheets/sheet1.xml"
            and before_input[4] == "10"
            and before_model is not None
            and before_model == after_model
            and before_model[0] == "xl/worksheets/sheet2.xml"
            and before_model[3] == "=Inputs!$B$2*2"
            and before_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[0] == "xl/worksheets/sheet3.xml"
            and before_dashboard[3] == "=Model!$B$2"
            and _office_web_addin_without_auto_show(baseline_path)
            == _office_web_addin_without_auto_show(candidate_path)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {"xl/webextensions/webextension1.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected one Office Web Add-in auto-show property false -> true with stable task-pane declaration, cells, formulas, and package boundary",
            errors,
        )
        return

    if kind == "ole_object_auto_load_enabled":
        sheet_name = fact.get("sheet")
        input_sheet = fact.get("input_sheet")
        input_cell = fact.get("input_cell")
        model_sheet = fact.get("model_sheet")
        model_cell = fact.get("model_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = (
            _raw_ole_object_auto_load_state(baseline_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        after_state = (
            _raw_ole_object_auto_load_state(candidate_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        before_input = (
            _raw_cell_state(baseline_path, input_sheet, input_cell)
            if isinstance(input_sheet, str) and isinstance(input_cell, str)
            else None
        )
        after_input = (
            _raw_cell_state(candidate_path, input_sheet, input_cell)
            if isinstance(input_sheet, str) and isinstance(input_cell, str)
            else None
        )
        before_model = (
            _raw_cell_state(baseline_path, model_sheet, model_cell)
            if isinstance(model_sheet, str) and isinstance(model_cell, str)
            else None
        )
        after_model = (
            _raw_cell_state(candidate_path, model_sheet, model_cell)
            if isinstance(model_sheet, str) and isinstance(model_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_common_state = {
            "worksheet_member": "xl/worksheets/sheet1.xml",
            "worksheet_relationships_member": "xl/worksheets/_rels/sheet1.xml.rels",
            "ole_objects_attributes": (),
            "relationship_attributes": (
                ("Id", "rIdWCABEmbeddedOle"),
                ("Target", "../embeddings/wcab-review-embedded-object.bin"),
                (
                    "Type",
                    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/oleObject",
                ),
            ),
            "embedded_object_member": "xl/embeddings/wcab-review-embedded-object.bin",
            "content_type_attributes": (
                (
                    "ContentType",
                    "application/vnd.openxmlformats-officedocument.oleObject",
                ),
                ("Extension", "bin"),
            ),
            "payload_size": len(_OLE_OBJECT_PAYLOAD),
        }
        expected_before_attributes = (
            ("autoLoad", "false"),
            ("dvAspect", "DVASPECT_CONTENT"),
            ("progId", "WCAB.Review.Embedded.Object"),
            ("shapeId", "1026"),
            (
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id",
                "rIdWCABEmbeddedOle",
            ),
        )
        expected_after_attributes = (
            ("autoLoad", "true"),
            ("dvAspect", "DVASPECT_CONTENT"),
            ("progId", "WCAB.Review.Embedded.Object"),
            ("shapeId", "1026"),
            (
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id",
                "rIdWCABEmbeddedOle",
            ),
        )
        _assert(
            sheet_name == "Inputs"
            and fact.get("worksheet_member") == "xl/worksheets/sheet1.xml"
            and fact.get("worksheet_relationships_member") == "xl/worksheets/_rels/sheet1.xml.rels"
            and fact.get("relationship_id") == "rIdWCABEmbeddedOle"
            and fact.get("relationship_type")
            == "http://schemas.openxmlformats.org/officeDocument/2006/relationships/oleObject"
            and fact.get("target") == "../embeddings/wcab-review-embedded-object.bin"
            and fact.get("embedded_object_member")
            == "xl/embeddings/wcab-review-embedded-object.bin"
            and fact.get("content_type")
            == "application/vnd.openxmlformats-officedocument.oleObject"
            and fact.get("prog_id") == "WCAB.Review.Embedded.Object"
            and fact.get("dv_aspect") == "DVASPECT_CONTENT"
            and fact.get("shape_id") == 1026
            and fact.get("baseline_auto_load") is False
            and fact.get("candidate_auto_load") is True
            and input_sheet == "Inputs"
            and input_cell == "B2"
            and fact.get("input_value") == 10
            and model_sheet == "Model"
            and model_cell == "B2"
            and fact.get("model_formula") == "=Inputs!$B$2*2"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Model!$B$2"
            and all(
                sheet in baseline.sheetnames and sheet in candidate.sheetnames
                for sheet in (input_sheet, model_sheet, dashboard_sheet)
            )
            and before_state is not None
            and after_state is not None
            and all(before_state.get(key) == value for key, value in expected_common_state.items())
            and all(after_state.get(key) == value for key, value in expected_common_state.items())
            and before_state["ole_object_attributes"] == expected_before_attributes
            and after_state["ole_object_attributes"] == expected_after_attributes
            and before_state["auto_load"] is False
            and after_state["auto_load"] is True
            and before_input is not None
            and before_input == after_input
            and before_input[0] == "xl/worksheets/sheet1.xml"
            and before_input[4] == "10"
            and before_model is not None
            and before_model == after_model
            and before_model[0] == "xl/worksheets/sheet2.xml"
            and before_model[3] == "=Inputs!$B$2*2"
            and before_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[0] == "xl/worksheets/sheet3.xml"
            and before_dashboard[3] == "=Model!$B$2"
            and _ole_object_without_auto_load(baseline_path, sheet_name)
            == _ole_object_without_auto_load(candidate_path, sheet_name)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {"xl/worksheets/sheet1.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected one local embedded OLE auto-load flag false -> true with stable relationship, opaque bytes, cells, formulas, and package boundary",
            errors,
        )
        return

    if kind == "defined_name_changed":
        name = fact.get("name")
        _assert(
            _defined_name_text(baseline, name) != _defined_name_text(candidate, name),
            f"{truth['id']}: expected defined name {name!r} to change",
            errors,
        )
        return

    if kind in {"data_validation_count_changed", "conditional_formatting_count_changed"}:
        sheet_name = fact.get("sheet")
        if sheet_name not in baseline.sheetnames or sheet_name not in candidate.sheetnames:
            errors.append(f"{truth['id']}: sheet {sheet_name!r} is absent")
            return
        if kind == "data_validation_count_changed":
            before_count = len(baseline[sheet_name].data_validations.dataValidation)
            after_count = len(candidate[sheet_name].data_validations.dataValidation)
        else:
            before_count = len(baseline[sheet_name].conditional_formatting)
            after_count = len(candidate[sheet_name].conditional_formatting)
        _assert(
            before_count == fact.get("baseline_count")
            and after_count == fact.get("candidate_count"),
            f"{truth['id']}: expected {kind} {fact.get('baseline_count')} -> {fact.get('candidate_count')}, got {before_count} -> {after_count}",
            errors,
        )
        return

    if kind == "conditional_formatting_threshold_changed":
        sheet_name = fact.get("sheet")
        target_range = fact.get("target_range")

        def range_values(workbook: Workbook) -> tuple[Any, ...] | None:
            if (
                not isinstance(sheet_name, str)
                or not isinstance(target_range, str)
                or sheet_name not in workbook.sheetnames
            ):
                return None
            try:
                cells = workbook[sheet_name][target_range]
            except ValueError:
                return None
            if not isinstance(cells, tuple) or any(not isinstance(row, tuple) for row in cells):
                return None
            return tuple(cell.value for row in cells for cell in row)

        before_state = (
            _raw_conditional_formatting_threshold_state(baseline_path, sheet_name, target_range)
            if isinstance(sheet_name, str) and isinstance(target_range, str)
            else None
        )
        after_state = (
            _raw_conditional_formatting_threshold_state(candidate_path, sheet_name, target_range)
            if isinstance(sheet_name, str) and isinstance(target_range, str)
            else None
        )
        before_values = range_values(baseline)
        after_values = range_values(candidate)
        expected_values = (10, 75, 120)
        expected_rule_attributes = (
            ("dxfId", "0"),
            ("operator", "greaterThan"),
            ("priority", "1"),
            ("type", "cellIs"),
        )
        _assert(
            sheet_name == "Operations"
            and target_range == "B2:B4"
            and fact.get("priority") == 1
            and fact.get("rule_type") == "cellIs"
            and fact.get("operator") == "greaterThan"
            and fact.get("baseline_formula") == "100"
            and fact.get("candidate_formula") == "50"
            and fact.get("metric_values") == list(expected_values)
            and fact.get("fill_rgb") == "FFFFC7CE"
            and before_state is not None
            and after_state is not None
            and before_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and after_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and before_state["control_attributes"] == (("sqref", "B2:B4"),)
            and after_state["control_attributes"] == (("sqref", "B2:B4"),)
            and before_state["rule_attributes"] == expected_rule_attributes
            and after_state["rule_attributes"] == expected_rule_attributes
            and before_state["formula_attributes"] == ()
            and after_state["formula_attributes"] == ()
            and before_state["formula_child_count"] == 0
            and after_state["formula_child_count"] == 0
            and before_state["formula_text"] == "100"
            and after_state["formula_text"] == "50"
            and before_values == expected_values
            and after_values == expected_values
            and _conditional_formatting_differential_fill_state(baseline_path, 0)
            == ("solid", "FFFFC7CE")
            and _conditional_formatting_differential_fill_state(candidate_path, 0)
            == ("solid", "FFFFC7CE")
            and _conditional_formatting_worksheet_without_threshold_formula(
                baseline_path, sheet_name, target_range
            )
            == _conditional_formatting_worksheet_without_threshold_formula(
                candidate_path, sheet_name, target_range
            )
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {before_state["worksheet_member"]}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected only one cellIs conditional-formatting threshold formula change with stable control, fill, values, calculation properties, and package boundary",
            errors,
        )
        return

    if kind == "cell_number_format_changed":
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        formula_cell = fact.get("formula_cell")
        before_state = (
            _raw_custom_number_format_state(baseline_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        after_state = (
            _raw_custom_number_format_state(candidate_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        before_cell_state = (
            _raw_cell_state(baseline_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        after_cell_state = (
            _raw_cell_state(candidate_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        before_formula_state = (
            _raw_cell_state(baseline_path, sheet_name, formula_cell)
            if isinstance(sheet_name, str) and isinstance(formula_cell, str)
            else None
        )
        after_formula_state = (
            _raw_cell_state(candidate_path, sheet_name, formula_cell)
            if isinstance(sheet_name, str) and isinstance(formula_cell, str)
            else None
        )
        expected_baseline_format = "0.0%;[Red](0.0%);-"
        expected_candidate_format = ";;;"
        expected_custom_number_format_id = 164
        expected_num_fmts_attributes = (("count", "1"),)
        expected_baseline_num_fmt_attributes = (
            ("formatCode", expected_baseline_format),
            ("numFmtId", str(expected_custom_number_format_id)),
        )
        expected_candidate_num_fmt_attributes = (
            ("formatCode", expected_candidate_format),
            ("numFmtId", str(expected_custom_number_format_id)),
        )
        _assert(
            sheet_name == "Operations"
            and coordinate == "B2"
            and fact.get("value") == 0.125
            and fact.get("custom_number_format_id") == expected_custom_number_format_id
            and fact.get("baseline_format") == expected_baseline_format
            and fact.get("candidate_format") == expected_candidate_format
            and formula_cell == "B3"
            and fact.get("formula") == "=B2"
            and before_state is not None
            and after_state is not None
            and before_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and after_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and before_state["styles_member"] == "xl/styles.xml"
            and after_state["styles_member"] == "xl/styles.xml"
            and before_state["custom_number_format_id"] == expected_custom_number_format_id
            and after_state["custom_number_format_id"] == expected_custom_number_format_id
            and before_state["num_fmts_attributes"] == expected_num_fmts_attributes
            and after_state["num_fmts_attributes"] == expected_num_fmts_attributes
            and before_state["num_fmt_attributes"] == expected_baseline_num_fmt_attributes
            and after_state["num_fmt_attributes"] == expected_candidate_num_fmt_attributes
            and before_state["format_code"] == expected_baseline_format
            and after_state["format_code"] == expected_candidate_format
            and before_state["cell_style_index"] == after_state["cell_style_index"]
            and before_state["cell_xf_attributes"] == after_state["cell_xf_attributes"]
            and before_cell_state is not None
            and before_cell_state == after_cell_state
            and before_cell_state[0] == "xl/worksheets/sheet1.xml"
            and before_cell_state[3] is None
            and before_cell_state[4] == "0.125"
            and before_formula_state is not None
            and before_formula_state == after_formula_state
            and before_formula_state[0] == "xl/worksheets/sheet1.xml"
            and before_formula_state[3] == "=B2"
            and baseline[sheet_name][coordinate].value == 0.125
            and candidate[sheet_name][coordinate].value == 0.125
            and baseline[sheet_name][formula_cell].value == "=B2"
            and candidate[sheet_name][formula_cell].value == "=B2"
            and _styles_without_custom_number_format_code(
                baseline_path, expected_custom_number_format_id
            )
            == _styles_without_custom_number_format_code(
                candidate_path, expected_custom_number_format_id
            )
            and _xlsx_member_differences(baseline_path, candidate_path) == {"xl/styles.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected only one custom number-format code transition with stable target cells, formulas, styles, calculation properties, and package boundary",
            errors,
        )
        return

    if kind == "ignored_error_rule_added":
        sheet_name = fact.get("sheet")
        target_range = fact.get("target_range")
        downstream_formula_cell = fact.get("downstream_formula_cell")
        before_state = (
            _raw_ignored_error_formula_range_state(baseline_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        after_state = (
            _raw_ignored_error_formula_range_state(candidate_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        baseline_sheet = (
            baseline[sheet_name]
            if isinstance(sheet_name, str) and sheet_name in baseline.sheetnames
            else None
        )
        candidate_sheet = (
            candidate[sheet_name]
            if isinstance(sheet_name, str) and sheet_name in candidate.sheetnames
            else None
        )
        raw_cells = ("B2", "B3", "B4", "B5", "C5")
        before_cells = {
            coordinate: _raw_cell_state(baseline_path, sheet_name, coordinate)
            for coordinate in raw_cells
        }
        after_cells = {
            coordinate: _raw_cell_state(candidate_path, sheet_name, coordinate)
            for coordinate in raw_cells
        }
        expected_before_state = {
            "worksheet_member": "xl/worksheets/sheet1.xml",
            "present": False,
            "container_attributes": None,
            "container_child_count": 0,
            "container_text": None,
            "rule_attributes": None,
            "rule_child_count": 0,
            "rule_text": None,
        }
        expected_after_state = {
            "worksheet_member": "xl/worksheets/sheet1.xml",
            "present": True,
            "container_attributes": (),
            "container_child_count": 1,
            "container_text": None,
            "rule_attributes": (("formulaRange", "1"), ("sqref", "B5")),
            "rule_child_count": 0,
            "rule_text": None,
        }
        _assert(
            sheet_name == "Operations"
            and target_range == "B5"
            and fact.get("warning_flag") == "formulaRange"
            and fact.get("formula") == "=SUM(B2:B3)"
            and fact.get("adjacent_populated_cell") == "B4"
            and fact.get("adjacent_populated_value") == 30
            and downstream_formula_cell == "C5"
            and fact.get("downstream_formula") == "=B5"
            and before_state == expected_before_state
            and after_state == expected_after_state
            and all(state is not None for state in before_cells.values())
            and before_cells == after_cells
            and before_cells["B2"] is not None
            and before_cells["B2"][3:] == (None, "10")
            and before_cells["B3"] is not None
            and before_cells["B3"][3:] == (None, "20")
            and before_cells["B4"] is not None
            and before_cells["B4"][3:] == (None, "30")
            and before_cells["B5"] is not None
            and before_cells["B5"][3] == "=SUM(B2:B3)"
            and before_cells["C5"] is not None
            and before_cells["C5"][3] == "=B5"
            and baseline_sheet is not None
            and candidate_sheet is not None
            and baseline_sheet["B2"].value == 10
            and candidate_sheet["B2"].value == 10
            and baseline_sheet["B3"].value == 20
            and candidate_sheet["B3"].value == 20
            and baseline_sheet["B4"].value == 30
            and candidate_sheet["B4"].value == 30
            and baseline_sheet[target_range].value == "=SUM(B2:B3)"
            and candidate_sheet[target_range].value == "=SUM(B2:B3)"
            and baseline_sheet[downstream_formula_cell].value == "=B5"
            and candidate_sheet[downstream_formula_cell].value == "=B5"
            and _worksheet_without_ignored_error_controls(baseline_path, sheet_name)
            == _worksheet_without_ignored_error_controls(candidate_path, sheet_name)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {"xl/worksheets/sheet1.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected only one formula-range ignored-error rule addition with stable cells, formulas, calculation properties, and package boundary",
            errors,
        )
        return

    if kind == "data_validation_list_source_changed":
        validation_sheet = fact.get("validation_sheet")
        target_range = fact.get("target_range")
        source_sheet = fact.get("source_sheet")
        baseline_source_range = fact.get("baseline_source_range")
        candidate_source_range = fact.get("candidate_source_range")
        input_cell = fact.get("input_cell")
        model_sheet = fact.get("model_sheet")
        model_cell = fact.get("model_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")

        def validation_state(workbook: Workbook) -> dict[str, Any] | None:
            if not isinstance(validation_sheet, str) or validation_sheet not in workbook.sheetnames:
                return None
            validations = list(workbook[validation_sheet].data_validations.dataValidation)
            if len(validations) != 1:
                return None
            validation = validations[0]
            return {
                "type": validation.type,
                "sqref": str(validation.sqref),
                "formula1": validation.formula1,
                "formula2": validation.formula2,
                "allow_blank": validation.allowBlank,
                "dropdown_hidden": validation.showDropDown,
                "show_error_message": validation.showErrorMessage,
                "error_style": validation.errorStyle,
                "error_title": validation.errorTitle,
                "error": validation.error,
                "show_input_message": validation.showInputMessage,
                "prompt_title": validation.promptTitle,
                "prompt": validation.prompt,
            }

        def source_values(workbook: Workbook, source_range: Any) -> tuple[Any, ...] | None:
            if (
                not isinstance(source_sheet, str)
                or not isinstance(source_range, str)
                or source_sheet not in workbook.sheetnames
            ):
                return None
            try:
                cells = workbook[source_sheet][source_range]
            except ValueError:
                return None
            if not isinstance(cells, tuple) or any(not isinstance(row, tuple) for row in cells):
                return None
            return tuple(cell.value for row in cells for cell in row)

        def workbook_cell(workbook: Workbook, sheet_name: Any, coordinate: Any) -> Any | None:
            if (
                not isinstance(sheet_name, str)
                or not isinstance(coordinate, str)
                or sheet_name not in workbook.sheetnames
            ):
                return None
            try:
                return workbook[sheet_name][coordinate]
            except ValueError:
                return None

        before_state = (
            _raw_data_validation_list_source_state(baseline_path, validation_sheet, target_range)
            if isinstance(validation_sheet, str) and isinstance(target_range, str)
            else None
        )
        after_state = (
            _raw_data_validation_list_source_state(candidate_path, validation_sheet, target_range)
            if isinstance(validation_sheet, str) and isinstance(target_range, str)
            else None
        )
        before_validation = validation_state(baseline)
        after_validation = validation_state(candidate)
        before_baseline_source_values = source_values(baseline, baseline_source_range)
        after_baseline_source_values = source_values(candidate, baseline_source_range)
        before_candidate_source_values = source_values(baseline, candidate_source_range)
        after_candidate_source_values = source_values(candidate, candidate_source_range)
        before_input = workbook_cell(baseline, validation_sheet, input_cell)
        after_input = workbook_cell(candidate, validation_sheet, input_cell)
        before_model = workbook_cell(baseline, model_sheet, model_cell)
        after_model = workbook_cell(candidate, model_sheet, model_cell)
        before_dashboard = workbook_cell(baseline, dashboard_sheet, dashboard_cell)
        after_dashboard = workbook_cell(candidate, dashboard_sheet, dashboard_cell)
        expected_validation_attributes = tuple(
            sorted(
                {
                    "allowBlank": "0",
                    "error": "Choose an approved status.",
                    "errorStyle": "stop",
                    "errorTitle": "Invalid status",
                    "prompt": "Choose a documented status.",
                    "promptTitle": "Approved status",
                    "showDropDown": "0",
                    "showErrorMessage": "1",
                    "showInputMessage": "0",
                    "sqref": "B2",
                    "type": "list",
                }.items()
            )
        )
        expected_rule = {
            "type": "list",
            "sqref": "B2",
            "formula2": None,
            "allow_blank": False,
            "dropdown_hidden": False,
            "show_error_message": True,
            "error_style": "stop",
            "error_title": "Invalid status",
            "error": "Choose an approved status.",
            "show_input_message": False,
            "prompt_title": "Approved status",
            "prompt": "Choose a documented status.",
        }
        expected_baseline_values = ("Draft", "Review", "Approved")
        expected_candidate_values = ("Draft", "Suspended", "Rejected")
        graph = _direct_graph(candidate)
        _assert(
            validation_sheet == "Inputs"
            and target_range == "B2"
            and fact.get("validation_type") == "list"
            and fact.get("baseline_source_formula") == "=Lists!$A$2:$A$4"
            and fact.get("candidate_source_formula") == "=Lists!$B$2:$B$4"
            and fact.get("allow_blank") is False
            and fact.get("dropdown_hidden") is False
            and fact.get("show_error_message") is True
            and fact.get("error_style") == "stop"
            and fact.get("error_title") == "Invalid status"
            and fact.get("error") == "Choose an approved status."
            and fact.get("show_input_message") is False
            and fact.get("prompt_title") == "Approved status"
            and fact.get("prompt") == "Choose a documented status."
            and source_sheet == "Lists"
            and baseline_source_range == "A2:A4"
            and candidate_source_range == "B2:B4"
            and fact.get("baseline_source_values") == list(expected_baseline_values)
            and fact.get("candidate_source_values") == list(expected_candidate_values)
            and input_cell == "B2"
            and fact.get("input_value") == "Draft"
            and model_sheet == "Model"
            and model_cell == "B2"
            and fact.get("model_formula") == "=Inputs!$B$2"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Model!$B$2"
            and before_state is not None
            and after_state is not None
            and before_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and after_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and before_state["container_attributes"] == (("count", "1"),)
            and after_state["container_attributes"] == (("count", "1"),)
            and before_state["validation_attributes"] == expected_validation_attributes
            and after_state["validation_attributes"] == expected_validation_attributes
            and before_state["formula1_attributes"] == ()
            and after_state["formula1_attributes"] == ()
            and before_state["formula1_text"] == "=Lists!$A$2:$A$4"
            and after_state["formula1_text"] == "=Lists!$B$2:$B$4"
            and before_state["formula1_child_count"] == 0
            and after_state["formula1_child_count"] == 0
            and before_validation == {**expected_rule, "formula1": "=Lists!$A$2:$A$4"}
            and after_validation == {**expected_rule, "formula1": "=Lists!$B$2:$B$4"}
            and before_baseline_source_values == expected_baseline_values
            and after_baseline_source_values == expected_baseline_values
            and before_candidate_source_values == expected_candidate_values
            and after_candidate_source_values == expected_candidate_values
            and before_input is not None
            and after_input is not None
            and _cell_kind(before_input) == "value"
            and _cell_kind(after_input) == "value"
            and before_input.value == after_input.value == "Draft"
            and before_model is not None
            and after_model is not None
            and _cell_kind(before_model) == "formula"
            and _cell_kind(after_model) == "formula"
            and before_model.value == after_model.value == "=Inputs!$B$2"
            and before_dashboard is not None
            and after_dashboard is not None
            and _cell_kind(before_dashboard) == "formula"
            and _cell_kind(after_dashboard) == "formula"
            and before_dashboard.value == after_dashboard.value == "=Model!$B$2"
            and _data_validation_worksheet_without_source_formula(
                baseline_path, validation_sheet, target_range
            )
            == _data_validation_worksheet_without_source_formula(
                candidate_path, validation_sheet, target_range
            )
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {before_state["worksheet_member"]}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and {(model_sheet, model_cell), (dashboard_sheet, dashboard_cell)}
            <= _reachable(graph, (validation_sheet, input_cell)),
            f"{truth['id']}: expected only one list-validation formula1 source declaration to change with stable rule, source values, formulas, and package boundary",
            errors,
        )
        return

    if kind == "sheet_visibility_changed":
        sheet_name = fact.get("sheet")
        _assert(
            sheet_name in baseline.sheetnames
            and sheet_name in candidate.sheetnames
            and baseline[sheet_name].sheet_state == fact.get("baseline_state")
            and candidate[sheet_name].sheet_state == fact.get("candidate_state"),
            f"{truth['id']}: expected {sheet_name!r} visibility change",
            errors,
        )
        return

    if kind == "manual_calculation_incomplete":
        _assert(
            candidate.calculation.calcMode == "manual"
            and candidate.calculation.calcCompleted is False,
            f"{truth['id']}: expected manual, incomplete calculation metadata",
            errors,
        )
        return

    if kind == "iterative_calculation_enabled":
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        formula = fact.get("formula")
        before_state = _calculation_iteration_state(baseline_path)
        after_state = _calculation_iteration_state(candidate_path)
        before_properties = _calculation_properties(baseline_path)
        after_properties = _calculation_properties(candidate_path)
        before_formula = (
            baseline[sheet_name][coordinate]
            if isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and sheet_name in baseline.sheetnames
            else None
        )
        after_formula = (
            candidate[sheet_name][coordinate]
            if isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and sheet_name in candidate.sheetnames
            else None
        )
        expected_count = fact.get("iteration_count")
        expected_delta = fact.get("iteration_delta")
        try:
            declared_delta = Decimal(str(expected_delta))
        except InvalidOperation:
            declared_delta = None
        ignored_properties = {"iterate", "iterateCount", "iterateDelta"}
        before_non_iteration_properties = (
            {
                key: value
                for key, value in before_properties.items()
                if key not in ignored_properties
            }
            if before_properties is not None
            else None
        )
        after_non_iteration_properties = (
            {key: value for key, value in after_properties.items() if key not in ignored_properties}
            if after_properties is not None
            else None
        )
        source = (sheet_name, coordinate)
        graph = _direct_graph(candidate)
        _assert(
            isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and isinstance(formula, str)
            and isinstance(expected_count, int)
            and not isinstance(expected_count, bool)
            and declared_delta is not None
            and fact.get("baseline_iterate") is False
            and fact.get("candidate_iterate") is True
            and before_formula is not None
            and after_formula is not None
            and _cell_kind(before_formula) == "formula"
            and _cell_kind(after_formula) == "formula"
            and before_formula.value == formula
            and after_formula.value == formula
            and source in graph.get(source, set())
            and before_state == (False, expected_count, declared_delta)
            and after_state == (True, expected_count, declared_delta)
            and before_non_iteration_properties == after_non_iteration_properties,
            f"{truth['id']}: expected unchanged direct circular formula and calcPr iterate false -> true with declared bounds only",
            errors,
        )
        return

    if kind == "precision_as_displayed_enabled":
        input_sheet = fact.get("input_sheet")
        input_cell = fact.get("input_cell")
        formula_sheet = fact.get("formula_sheet")
        formula_cell = fact.get("formula_cell")
        formula = fact.get("formula")
        number_format = fact.get("number_format")
        expected_value = fact.get("input_value")
        try:
            declared_value = Decimal(str(expected_value))
        except InvalidOperation:
            declared_value = None
        before_properties = _calculation_properties(baseline_path)
        after_properties = _calculation_properties(candidate_path)
        before_non_precision_properties = (
            {key: value for key, value in before_properties.items() if key != "fullPrecision"}
            if before_properties is not None
            else None
        )
        after_non_precision_properties = (
            {key: value for key, value in after_properties.items() if key != "fullPrecision"}
            if after_properties is not None
            else None
        )
        if (
            not isinstance(input_sheet, str)
            or not isinstance(input_cell, str)
            or not isinstance(formula_sheet, str)
            or not isinstance(formula_cell, str)
            or input_sheet not in baseline.sheetnames
            or input_sheet not in candidate.sheetnames
            or formula_sheet not in baseline.sheetnames
            or formula_sheet not in candidate.sheetnames
        ):
            errors.append(f"{truth['id']}: precision-as-displayed locations are absent")
            return
        before_input = baseline[input_sheet][input_cell]
        after_input = candidate[input_sheet][input_cell]
        before_formula = baseline[formula_sheet][formula_cell]
        after_formula = candidate[formula_sheet][formula_cell]
        try:
            before_value = Decimal(str(before_input.value))
            after_value = Decimal(str(after_input.value))
        except InvalidOperation:
            before_value = after_value = None
        graph = _direct_graph(candidate)
        _assert(
            isinstance(formula, str)
            and isinstance(number_format, str)
            and declared_value is not None
            and declared_value.is_finite()
            and fact.get("baseline_full_precision") is True
            and fact.get("candidate_full_precision") is False
            and _calculation_full_precision(baseline_path) is True
            and _calculation_full_precision(candidate_path) is False
            and before_non_precision_properties == after_non_precision_properties
            and _cell_kind(before_input) == "value"
            and _cell_kind(after_input) == "value"
            and before_value == declared_value
            and after_value == declared_value
            and before_input.number_format == number_format
            and after_input.number_format == number_format
            and _cell_kind(before_formula) == "formula"
            and _cell_kind(after_formula) == "formula"
            and before_formula.value == formula
            and after_formula.value == formula
            and (formula_sheet, formula_cell) in graph.get((input_sheet, input_cell), set()),
            f"{truth['id']}: expected unchanged stored input/formula and calcPr fullPrecision true -> false only",
            errors,
        )
        return

    if kind == "auto_filter_criteria_changed":
        sheet_name = fact.get("sheet")
        filter_reference = fact.get("filter_ref")
        filter_column_id = fact.get("filter_column_id")
        baseline_filter_value = fact.get("baseline_filter_value")
        candidate_filter_value = fact.get("candidate_filter_value")
        subtotal_cell = fact.get("subtotal_cell")
        subtotal_formula = fact.get("subtotal_formula")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        dashboard_formula = fact.get("dashboard_formula")
        before_filter = (
            _raw_auto_filter_state(baseline_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        after_filter = (
            _raw_auto_filter_state(candidate_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        before_subtotal = (
            _raw_cell_state(baseline_path, sheet_name, subtotal_cell)
            if isinstance(sheet_name, str) and isinstance(subtotal_cell, str)
            else None
        )
        after_subtotal = (
            _raw_cell_state(candidate_path, sheet_name, subtotal_cell)
            if isinstance(sheet_name, str) and isinstance(subtotal_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        graph = _direct_graph(candidate)
        _assert(
            isinstance(sheet_name, str)
            and isinstance(filter_reference, str)
            and type(filter_column_id) is int
            and isinstance(baseline_filter_value, str)
            and isinstance(candidate_filter_value, str)
            and baseline_filter_value != candidate_filter_value
            and isinstance(subtotal_cell, str)
            and isinstance(subtotal_formula, str)
            and isinstance(dashboard_sheet, str)
            and isinstance(dashboard_cell, str)
            and isinstance(dashboard_formula, str)
            and before_filter is not None
            and after_filter is not None
            and before_filter[0] == after_filter[0]
            and before_filter[1] == (("ref", filter_reference),)
            and after_filter[1] == (("ref", filter_reference),)
            and before_filter[2] == filter_column_id
            and after_filter[2] == filter_column_id
            and before_filter[3] == after_filter[3]
            and before_filter[4] == after_filter[4]
            and before_filter[5] == (baseline_filter_value,)
            and after_filter[5] == (candidate_filter_value,)
            and before_subtotal is not None
            and after_subtotal is not None
            and before_subtotal == after_subtotal
            and before_subtotal[3] == subtotal_formula
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == dashboard_formula
            and _worksheet_without_auto_filter_criteria(baseline_path, sheet_name)
            == _worksheet_without_auto_filter_criteria(candidate_path, sheet_name)
            and _xlsx_member_differences(baseline_path, candidate_path) == {before_filter[0]}
            and (dashboard_sheet, dashboard_cell) in graph.get((sheet_name, subtotal_cell), set()),
            f"{truth['id']}: expected one raw AutoFilter criterion change only with stable formulas and dependency edge",
            errors,
        )
        return

    if kind == "workbook_date_system_changed":
        serial_sheet = fact.get("serial_sheet")
        serial_cell = fact.get("serial_cell")
        formula_sheet = fact.get("formula_sheet")
        formula_cell = fact.get("formula_cell")
        formula = fact.get("formula")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        dashboard_formula = fact.get("dashboard_formula")
        number_format = fact.get("number_format")
        try:
            declared_serial = Decimal(str(fact.get("serial_value")))
        except InvalidOperation:
            declared_serial = None
        before_date_system = _workbook_date_system_state(baseline_path)
        after_date_system = _workbook_date_system_state(candidate_path)
        before_input = (
            _raw_cell_state(baseline_path, serial_sheet, serial_cell)
            if isinstance(serial_sheet, str) and isinstance(serial_cell, str)
            else None
        )
        after_input = (
            _raw_cell_state(candidate_path, serial_sheet, serial_cell)
            if isinstance(serial_sheet, str) and isinstance(serial_cell, str)
            else None
        )
        before_formula = (
            _raw_cell_state(baseline_path, formula_sheet, formula_cell)
            if isinstance(formula_sheet, str) and isinstance(formula_cell, str)
            else None
        )
        after_formula = (
            _raw_cell_state(candidate_path, formula_sheet, formula_cell)
            if isinstance(formula_sheet, str) and isinstance(formula_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        try:
            before_serial = Decimal(before_input[4]) if before_input and before_input[4] else None
            after_serial = Decimal(after_input[4]) if after_input and after_input[4] else None
        except InvalidOperation:
            before_serial = after_serial = None
        graph = _direct_graph(candidate)
        _assert(
            fact.get("baseline_date_1904") is False
            and fact.get("candidate_date_1904") is True
            and fact.get("date_compatibility") is True
            and isinstance(serial_sheet, str)
            and isinstance(serial_cell, str)
            and isinstance(formula_sheet, str)
            and isinstance(formula_cell, str)
            and isinstance(formula, str)
            and isinstance(dashboard_sheet, str)
            and isinstance(dashboard_cell, str)
            and isinstance(dashboard_formula, str)
            and isinstance(number_format, str)
            and declared_serial is not None
            and declared_serial.is_finite()
            and before_date_system == (False, True)
            and after_date_system == (True, True)
            and before_input is not None
            and after_input is not None
            and before_input == after_input
            and before_input[1] in {None, "n"}
            and before_input[3] is None
            and before_serial == declared_serial
            and after_serial == declared_serial
            and before_formula is not None
            and after_formula is not None
            and before_formula == after_formula
            and before_formula[3] == formula
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == dashboard_formula
            and _raw_cell_number_format(baseline_path, serial_sheet, serial_cell) == number_format
            and _raw_cell_number_format(candidate_path, serial_sheet, serial_cell) == number_format
            and _workbook_without_date_system_controls(baseline_path)
            == _workbook_without_date_system_controls(candidate_path)
            and _xlsx_member_differences(baseline_path, candidate_path) == {"xl/workbook.xml"}
            and (formula_sheet, formula_cell) in graph.get((serial_sheet, serial_cell), set())
            and (dashboard_sheet, dashboard_cell)
            in graph.get((formula_sheet, formula_cell), set()),
            f"{truth['id']}: expected raw workbookPr date1904 false -> true only with stable serial, style, formulas, and dependency edges",
            errors,
        )
        return

    if kind == "formula_cached_result_changed":
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        formula = fact.get("formula")
        input_sheet = fact.get("input_sheet")
        input_cell = fact.get("input_cell")
        result_type = fact.get("result_type")
        try:
            declared_input = Decimal(str(fact.get("input_value")))
            declared_before_result = Decimal(str(fact.get("baseline_cached_result")))
            declared_after_result = Decimal(str(fact.get("candidate_cached_result")))
        except InvalidOperation:
            declared_input = declared_before_result = declared_after_result = None
        before_state = (
            _formula_cached_result_state(baseline_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        after_state = (
            _formula_cached_result_state(candidate_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        if before_state is None or after_state is None:
            errors.append(f"{truth['id']}: formula cached-result location is absent or malformed")
            return
        (
            before_member,
            before_formula_text,
            before_result_type,
            before_result_text,
            before_without_result,
        ) = before_state
        (
            after_member,
            after_formula_text,
            after_result_type,
            after_result_text,
            after_without_result,
        ) = after_state
        try:
            before_result = Decimal(before_result_text)
            after_result = Decimal(after_result_text)
        except InvalidOperation:
            before_result = after_result = None
        if (
            not isinstance(input_sheet, str)
            or not isinstance(input_cell, str)
            or input_sheet not in baseline.sheetnames
            or input_sheet not in candidate.sheetnames
        ):
            errors.append(f"{truth['id']}: formula cached-result input location is absent")
            return
        before_input = baseline[input_sheet][input_cell]
        after_input = candidate[input_sheet][input_cell]
        try:
            before_input_value = Decimal(str(before_input.value))
            after_input_value = Decimal(str(after_input.value))
        except InvalidOperation:
            before_input_value = after_input_value = None
        graph = _direct_graph(candidate)
        _assert(
            isinstance(formula, str)
            and result_type == "numeric"
            and declared_input is not None
            and declared_input.is_finite()
            and declared_before_result is not None
            and declared_before_result.is_finite()
            and declared_after_result is not None
            and declared_after_result.is_finite()
            and declared_before_result != declared_after_result
            and before_formula_text == formula
            and after_formula_text == formula
            and before_result_type == result_type
            and after_result_type == result_type
            and before_result == declared_before_result
            and after_result == declared_after_result
            and before_result is not None
            and after_result is not None
            and before_result.is_finite()
            and after_result.is_finite()
            and _cell_kind(before_input) == "value"
            and _cell_kind(after_input) == "value"
            and before_input_value == declared_input
            and after_input_value == declared_input
            and _cell_kind(baseline[sheet_name][coordinate]) == "formula"
            and _cell_kind(candidate[sheet_name][coordinate]) == "formula"
            and before_member == after_member
            and before_without_result == after_without_result
            and _xlsx_member_differences(baseline_path, candidate_path) == {before_member}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (sheet_name, coordinate) in graph.get((input_sheet, input_cell), set()),
            f"{truth['id']}: expected only raw numeric formula cache {sheet_name}!{coordinate} to change with formula, input, and controls unchanged",
            errors,
        )
        return

    if kind == "chart_series_value_reference_changed":
        chart_sheet = fact.get("chart_sheet")
        chart_anchor = fact.get("chart_anchor")
        source_sheet = fact.get("source_sheet")
        series_title_reference = fact.get("series_title_ref")
        category_reference = fact.get("category_ref")
        baseline_value_reference = fact.get("baseline_value_ref")
        candidate_value_reference = fact.get("candidate_value_ref")
        before_state = (
            _raw_chart_series_value_reference_state(baseline_path, chart_sheet)
            if isinstance(chart_sheet, str)
            else None
        )
        after_state = (
            _raw_chart_series_value_reference_state(candidate_path, chart_sheet)
            if isinstance(chart_sheet, str)
            else None
        )
        _assert(
            isinstance(chart_sheet, str)
            and isinstance(chart_anchor, str)
            and isinstance(source_sheet, str)
            and isinstance(series_title_reference, str)
            and isinstance(category_reference, str)
            and isinstance(baseline_value_reference, str)
            and isinstance(candidate_value_reference, str)
            and baseline_value_reference != candidate_value_reference
            and chart_sheet in baseline.sheetnames
            and chart_sheet in candidate.sheetnames
            and source_sheet in baseline.sheetnames
            and source_sheet in candidate.sheetnames
            and before_state is not None
            and after_state is not None
            and before_state["worksheet_member"] == after_state["worksheet_member"]
            and before_state["drawing_member"] == after_state["drawing_member"]
            and before_state["chart_member"] == after_state["chart_member"]
            and before_state["anchor"] == chart_anchor
            and after_state["anchor"] == chart_anchor
            and before_state["series_title_reference"] == series_title_reference
            and after_state["series_title_reference"] == series_title_reference
            and before_state["category_reference"] == category_reference
            and after_state["category_reference"] == category_reference
            and before_state["value_reference"] == baseline_value_reference
            and after_state["value_reference"] == candidate_value_reference
            and all(
                reference.startswith(f"'{source_sheet}'!")
                for reference in (
                    series_title_reference,
                    category_reference,
                    baseline_value_reference,
                    candidate_value_reference,
                )
            )
            and _chart_without_value_reference(baseline_path, before_state["chart_member"])
            == _chart_without_value_reference(candidate_path, after_state["chart_member"])
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {before_state["chart_member"]}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected one relationship-bound chart value-series reference to change with stable worksheets and chart bindings",
            errors,
        )
        return

    if kind == "power_query_m_filter_changed":
        data_mashup_part = fact.get("data_mashup_part")
        source_sheet = fact.get("source_sheet")
        source_table = fact.get("source_table")
        source_reference = fact.get("source_ref")
        query_section = fact.get("query_section")
        query_name = fact.get("query_name")
        filter_column = fact.get("filter_column")
        baseline_filter_value = fact.get("baseline_filter_value")
        candidate_filter_value = fact.get("candidate_filter_value")

        def source_table_state(
            workbook: Workbook,
        ) -> tuple[str, tuple[tuple[Any, ...], ...]] | None:
            if (
                not isinstance(source_sheet, str)
                or not isinstance(source_table, str)
                or not isinstance(source_reference, str)
                or source_sheet not in workbook.sheetnames
            ):
                return None
            worksheet = workbook[source_sheet]
            try:
                table = worksheet.tables[source_table]
                cells = worksheet[source_reference]
            except (KeyError, ValueError):
                return None
            if not isinstance(cells, tuple) or any(not isinstance(row, tuple) for row in cells):
                return None
            return table.ref, tuple(tuple(cell.value for cell in row) for row in cells)

        before_state = _raw_power_query_m_filter_state(baseline_path)
        after_state = _raw_power_query_m_filter_state(candidate_path)
        before_table_state = source_table_state(baseline)
        after_table_state = source_table_state(candidate)
        expected_root_relationship_attributes = tuple(
            sorted(
                {
                    "Id": "rIdWCABPowerQuery",
                    "Type": _POWER_QUERY_CUSTOM_XML_RELATIONSHIP,
                    "Target": "customXml/item1.xml",
                }.items()
            )
        )
        expected_package_members = (
            "Config/Package.xml",
            "Formulas/Section1.m",
            "[Content_Types].xml",
        )
        expected_source_rows = (
            ("Region", "Amount"),
            ("North", 10),
            ("South", 20),
            ("North", 15),
            ("South", 25),
        )

        def expected_formula(filter_value: str) -> str:
            return (
                f"section {query_section};\n\n"
                f"shared {query_name} = let\n"
                f'    Source = Excel.CurrentWorkbook(){{[Name="{source_table}"]}}[Content],\n'
                f'    FilteredRows = Table.SelectRows(Source, each [{filter_column}] = "{filter_value}")\n'
                "in\n"
                "    FilteredRows;\n"
            )

        mutable_fields = {"formula", "filter_value"}
        _assert(
            data_mashup_part == "customXml/item1.xml"
            and source_sheet == "Source"
            and source_table == "SourceData"
            and source_reference == "A1:B5"
            and query_section == "Section1"
            and query_name == "RegionQuery"
            and filter_column == "Region"
            and baseline_filter_value == "North"
            and candidate_filter_value == "South"
            and fact.get("fill_enabled") is False
            and fact.get("firewall_enabled") is True
            and fact.get("future_packages_allowed") is False
            and before_state is not None
            and after_state is not None
            and before_table_state == after_table_state
            and before_table_state == (source_reference, expected_source_rows)
            and {key: value for key, value in before_state.items() if key not in mutable_fields}
            == {key: value for key, value in after_state.items() if key not in mutable_fields}
            and before_state["data_mashup_member"] == data_mashup_part
            and after_state["data_mashup_member"] == data_mashup_part
            and before_state["root_relationship_attributes"]
            == expected_root_relationship_attributes
            and after_state["root_relationship_attributes"] == expected_root_relationship_attributes
            and before_state["stream_version"] == 0
            and after_state["stream_version"] == 0
            and before_state["package_members"] == expected_package_members
            and after_state["package_members"] == expected_package_members
            and before_state["package_content_types"] == b"<Types/>"
            and after_state["package_content_types"] == b"<Types/>"
            and before_state["package_configuration"]
            == b"<Package>WCAB connection-only local table query</Package>"
            and after_state["package_configuration"]
            == b"<Package>WCAB connection-only local table query</Package>"
            and before_state["formula"] == expected_formula(baseline_filter_value)
            and after_state["formula"] == expected_formula(candidate_filter_value)
            and before_state["query_section"] == query_section
            and after_state["query_section"] == query_section
            and before_state["query_name"] == query_name
            and after_state["query_name"] == query_name
            and before_state["source_table"] == source_table
            and after_state["source_table"] == source_table
            and before_state["filter_column"] == filter_column
            and after_state["filter_column"] == filter_column
            and before_state["filter_value"] == baseline_filter_value
            and after_state["filter_value"] == candidate_filter_value
            and before_state["metadata_version"] == 0
            and after_state["metadata_version"] == 0
            and before_state["metadata_item_path"] == f"{query_section}/{query_name}"
            and after_state["metadata_item_path"] == f"{query_section}/{query_name}"
            and before_state["fill_enabled"] is False
            and after_state["fill_enabled"] is False
            and before_state["firewall_enabled"] is True
            and after_state["firewall_enabled"] is True
            and before_state["future_packages_allowed"] is False
            and after_state["future_packages_allowed"] is False
            and before_state["permission_binding"] == b""
            and after_state["permission_binding"] == b""
            and _xlsx_member_differences(baseline_path, candidate_path) == {data_mashup_part}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected only one compact connection-only local-table Power Query M filter literal to change with stable source, metadata, permissions, and calculation controls",
            errors,
        )
        return

    if kind == "what_if_data_table_input_reference_changed":
        table_sheet = fact.get("table_sheet")
        master_cell = fact.get("master_cell")
        output_range = fact.get("output_range")
        baseline_input_cell = fact.get("baseline_input_cell")
        candidate_input_cell = fact.get("candidate_input_cell")
        input_value_range = fact.get("input_value_range")
        scale_cell = fact.get("scale_cell")
        output_formula_cell = fact.get("output_formula_cell")
        output_formula = fact.get("output_formula")
        model_sheet = fact.get("model_sheet")
        model_cell = fact.get("model_cell")
        model_formula = fact.get("model_formula")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        dashboard_formula = fact.get("dashboard_formula")
        before_state = (
            _raw_what_if_data_table_state(baseline_path, table_sheet, master_cell)
            if isinstance(table_sheet, str) and isinstance(master_cell, str)
            else None
        )
        after_state = (
            _raw_what_if_data_table_state(candidate_path, table_sheet, master_cell)
            if isinstance(table_sheet, str) and isinstance(master_cell, str)
            else None
        )
        try:
            declared_grid_values = tuple(
                Decimal(str(value)) for value in fact.get("input_values", [])
            )
            declared_primary_value = Decimal(str(fact.get("primary_input_value")))
            declared_alternate_value = Decimal(str(fact.get("alternate_input_value")))
            declared_scale_value = Decimal(str(fact.get("scale_value")))
            before_grid_values = tuple(
                Decimal(str(baseline[table_sheet][coordinate].value))
                for coordinate in ("C3", "C4", "C5")
            )
            after_grid_values = tuple(
                Decimal(str(candidate[table_sheet][coordinate].value))
                for coordinate in ("C3", "C4", "C5")
            )
            before_primary_value = Decimal(str(baseline[table_sheet][baseline_input_cell].value))
            after_primary_value = Decimal(str(candidate[table_sheet][baseline_input_cell].value))
            before_alternate_value = Decimal(str(baseline[table_sheet][candidate_input_cell].value))
            after_alternate_value = Decimal(str(candidate[table_sheet][candidate_input_cell].value))
            before_scale_value = Decimal(str(baseline[table_sheet][scale_cell].value))
            after_scale_value = Decimal(str(candidate[table_sheet][scale_cell].value))
        except (InvalidOperation, TypeError, ValueError, KeyError):
            declared_grid_values = ()
            declared_primary_value = declared_alternate_value = declared_scale_value = None
            before_grid_values = after_grid_values = ()
            before_primary_value = after_primary_value = None
            before_alternate_value = after_alternate_value = None
            before_scale_value = after_scale_value = None
        before_output = (
            baseline[table_sheet][output_formula_cell]
            if isinstance(table_sheet, str)
            and isinstance(output_formula_cell, str)
            and table_sheet in baseline.sheetnames
            else None
        )
        after_output = (
            candidate[table_sheet][output_formula_cell]
            if isinstance(table_sheet, str)
            and isinstance(output_formula_cell, str)
            and table_sheet in candidate.sheetnames
            else None
        )
        before_model = (
            baseline[model_sheet][model_cell]
            if isinstance(model_sheet, str)
            and isinstance(model_cell, str)
            and model_sheet in baseline.sheetnames
            else None
        )
        after_model = (
            candidate[model_sheet][model_cell]
            if isinstance(model_sheet, str)
            and isinstance(model_cell, str)
            and model_sheet in candidate.sheetnames
            else None
        )
        before_dashboard = (
            baseline[dashboard_sheet][dashboard_cell]
            if isinstance(dashboard_sheet, str)
            and isinstance(dashboard_cell, str)
            and dashboard_sheet in baseline.sheetnames
            else None
        )
        after_dashboard = (
            candidate[dashboard_sheet][dashboard_cell]
            if isinstance(dashboard_sheet, str)
            and isinstance(dashboard_cell, str)
            and dashboard_sheet in candidate.sheetnames
            else None
        )
        expected_baseline_attributes = (
            ("ca", "1"),
            ("r1", "B2"),
            ("ref", "D3:D5"),
            ("t", "dataTable"),
        )
        expected_candidate_attributes = (
            ("ca", "1"),
            ("r1", "B3"),
            ("ref", "D3:D5"),
            ("t", "dataTable"),
        )
        graph = _direct_graph(candidate)
        _assert(
            table_sheet == "Sensitivity"
            and master_cell == "D3"
            and output_range == "D3:D5"
            and baseline_input_cell == "B2"
            and candidate_input_cell == "B3"
            and fact.get("orientation") == "column"
            and fact.get("recalculation_requested") is True
            and input_value_range == "C3:C5"
            and declared_grid_values == (Decimal("0.04"), Decimal("0.08"), Decimal("0.12"))
            and declared_primary_value == Decimal("0.08")
            and declared_alternate_value == Decimal("0.12")
            and scale_cell == "B4"
            and declared_scale_value == Decimal("100")
            and output_formula_cell == "D2"
            and output_formula == "=Model!$B$2"
            and model_sheet == "Model"
            and model_cell == "B2"
            and model_formula == "=Sensitivity!$B$2*Sensitivity!$B$3*Sensitivity!$B$4"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and dashboard_formula == "=Model!$B$2"
            and before_state is not None
            and after_state is not None
            and before_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and after_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and before_state["master_cell_attributes"] == (("r", "D3"),)
            and after_state["master_cell_attributes"] == (("r", "D3"),)
            and before_state["formula_attributes"] == expected_baseline_attributes
            and after_state["formula_attributes"] == expected_candidate_attributes
            and before_state["formula_text"] is None
            and after_state["formula_text"] is None
            and before_state["formula_child_count"] == 0
            and after_state["formula_child_count"] == 0
            and before_state["master_value_text"] is None
            and after_state["master_value_text"] is None
            and before_grid_values == after_grid_values == declared_grid_values
            and before_primary_value == after_primary_value == declared_primary_value
            and before_alternate_value == after_alternate_value == declared_alternate_value
            and before_scale_value == after_scale_value == declared_scale_value
            and before_output is not None
            and after_output is not None
            and _cell_kind(before_output) == "formula"
            and _cell_kind(after_output) == "formula"
            and before_output.value == after_output.value == output_formula
            and before_model is not None
            and after_model is not None
            and _cell_kind(before_model) == "formula"
            and _cell_kind(after_model) == "formula"
            and before_model.value == after_model.value == model_formula
            and before_dashboard is not None
            and after_dashboard is not None
            and _cell_kind(before_dashboard) == "formula"
            and _cell_kind(after_dashboard) == "formula"
            and before_dashboard.value == after_dashboard.value == dashboard_formula
            and _what_if_data_table_worksheet_without_input_reference(
                baseline_path, table_sheet, master_cell
            )
            == _what_if_data_table_worksheet_without_input_reference(
                candidate_path, table_sheet, master_cell
            )
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {before_state["worksheet_member"]}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and {
                (model_sheet, model_cell),
                (dashboard_sheet, dashboard_cell),
            }
            <= _reachable(graph, (table_sheet, baseline_input_cell))
            and {
                (model_sheet, model_cell),
                (dashboard_sheet, dashboard_cell),
            }
            <= _reachable(graph, (table_sheet, candidate_input_cell)),
            f"{truth['id']}: expected only one What-If Data Table r1 input reference to change with stable grid, formulas, and package boundary",
            errors,
        )
        return

    if kind == "scenario_manager_stored_input_value_changed":
        scenario_sheet = fact.get("scenario_sheet")
        scenario_name = fact.get("scenario_name")
        changing_cell = fact.get("changing_cell")
        stable_input_cell = fact.get("stable_input_cell")
        baseline_stored_value = fact.get("baseline_stored_value")
        candidate_stored_value = fact.get("candidate_stored_value")
        stable_stored_value = fact.get("stable_stored_value")
        input_number_format_id = fact.get("input_number_format_id")
        summary_reference = fact.get("summary_ref")
        result_cell = fact.get("result_cell")
        result_formula = fact.get("result_formula")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        dashboard_formula = fact.get("dashboard_formula")
        before_state = (
            _raw_scenario_manager_stored_input_state(baseline_path, scenario_sheet)
            if isinstance(scenario_sheet, str)
            else None
        )
        after_state = (
            _raw_scenario_manager_stored_input_state(candidate_path, scenario_sheet)
            if isinstance(scenario_sheet, str)
            else None
        )
        before_result = (
            _raw_cell_state(baseline_path, scenario_sheet, result_cell)
            if isinstance(scenario_sheet, str) and isinstance(result_cell, str)
            else None
        )
        after_result = (
            _raw_cell_state(candidate_path, scenario_sheet, result_cell)
            if isinstance(scenario_sheet, str) and isinstance(result_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        before_input = (
            baseline[scenario_sheet][changing_cell]
            if isinstance(scenario_sheet, str)
            and isinstance(changing_cell, str)
            and scenario_sheet in baseline.sheetnames
            else None
        )
        after_input = (
            candidate[scenario_sheet][changing_cell]
            if isinstance(scenario_sheet, str)
            and isinstance(changing_cell, str)
            and scenario_sheet in candidate.sheetnames
            else None
        )
        before_stable_input = (
            baseline[scenario_sheet][stable_input_cell]
            if isinstance(scenario_sheet, str)
            and isinstance(stable_input_cell, str)
            and scenario_sheet in baseline.sheetnames
            else None
        )
        after_stable_input = (
            candidate[scenario_sheet][stable_input_cell]
            if isinstance(scenario_sheet, str)
            and isinstance(stable_input_cell, str)
            and scenario_sheet in candidate.sheetnames
            else None
        )
        try:
            worksheet_input_value = Decimal(str(fact.get("worksheet_input_value")))
            worksheet_stable_input_value = Decimal(str(fact.get("worksheet_stable_input_value")))
            before_input_value = Decimal(str(before_input.value)) if before_input else None
            after_input_value = Decimal(str(after_input.value)) if after_input else None
            before_stable_input_value = (
                Decimal(str(before_stable_input.value)) if before_stable_input else None
            )
            after_stable_input_value = (
                Decimal(str(after_stable_input.value)) if after_stable_input else None
            )
        except (InvalidOperation, TypeError, ValueError):
            worksheet_input_value = worksheet_stable_input_value = None
            before_input_value = after_input_value = None
            before_stable_input_value = after_stable_input_value = None
        expected_scenarios_attributes = (
            ("current", "0"),
            ("show", "0"),
            ("sqref", "D2"),
        )
        expected_scenario_attributes = (
            ("comment", "Synthetic downside assumption set"),
            ("count", "2"),
            ("locked", "1"),
            ("name", "WCAB downside"),
            ("user", "WCAB"),
        )
        expected_baseline_inputs = (
            (
                "B2",
                "0.08",
                (("numFmtId", "10"), ("r", "B2"), ("val", "0.08")),
            ),
            ("B3", "125", (("r", "B3"), ("val", "125"))),
        )
        expected_candidate_inputs = (
            (
                "B2",
                "0.16",
                (("numFmtId", "10"), ("r", "B2"), ("val", "0.16")),
            ),
            ("B3", "125", (("r", "B3"), ("val", "125"))),
        )
        graph = _direct_graph(candidate)
        _assert(
            scenario_sheet == "Inputs"
            and scenario_name == "WCAB downside"
            and changing_cell == "B2"
            and stable_input_cell == "B3"
            and baseline_stored_value == "0.08"
            and candidate_stored_value == "0.16"
            and baseline_stored_value != candidate_stored_value
            and stable_stored_value == "125"
            and input_number_format_id == 10
            and summary_reference == "D2"
            and result_cell == "D2"
            and result_formula == "=B2*B3"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and dashboard_formula == "=Inputs!$D$2"
            and worksheet_input_value == Decimal("0.1")
            and worksheet_stable_input_value == Decimal("125")
            and before_state is not None
            and after_state is not None
            and {key: value for key, value in before_state.items() if key != "input_cells"}
            == {key: value for key, value in after_state.items() if key != "input_cells"}
            and before_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and after_state["worksheet_member"] == "xl/worksheets/sheet1.xml"
            and before_state["scenarios_attributes"] == expected_scenarios_attributes
            and after_state["scenarios_attributes"] == expected_scenarios_attributes
            and before_state["scenario_attributes"] == expected_scenario_attributes
            and after_state["scenario_attributes"] == expected_scenario_attributes
            and before_state["input_cells"] == expected_baseline_inputs
            and after_state["input_cells"] == expected_candidate_inputs
            and before_input is not None
            and after_input is not None
            and before_stable_input is not None
            and after_stable_input is not None
            and _cell_kind(before_input) == "value"
            and _cell_kind(after_input) == "value"
            and _cell_kind(before_stable_input) == "value"
            and _cell_kind(after_stable_input) == "value"
            and before_input_value == worksheet_input_value
            and after_input_value == worksheet_input_value
            and before_stable_input_value == worksheet_stable_input_value
            and after_stable_input_value == worksheet_stable_input_value
            and before_result is not None
            and after_result is not None
            and before_result == after_result
            and before_result[3] == result_formula
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == dashboard_formula
            and _scenario_manager_worksheet_without_stored_input_value(
                baseline_path, scenario_sheet, changing_cell
            )
            == _scenario_manager_worksheet_without_stored_input_value(
                candidate_path, scenario_sheet, changing_cell
            )
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {before_state["worksheet_member"]}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and {
                (scenario_sheet, result_cell),
                (dashboard_sheet, dashboard_cell),
            }
            <= _reachable(graph, (scenario_sheet, changing_cell)),
            f"{truth['id']}: expected one stored Scenario Manager input value to change with stable worksheet cells, metadata, formulas, and package boundary",
            errors,
        )
        return

    if kind == "pivot_slicer_selection_changed":
        cache_id = fact.get("cache_id")
        source_type = fact.get("source_type")
        source_sheet = fact.get("source_sheet")
        source_reference = fact.get("source_ref")
        pivot_sheet = fact.get("pivot_sheet")
        pivot_reference = fact.get("pivot_ref")
        pivot_output_cell = fact.get("pivot_output_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        dashboard_formula = fact.get("dashboard_formula")
        slicer_name = fact.get("slicer_name")
        slicer_source_name = fact.get("slicer_source_name")
        slicer_pivot_table_name = fact.get("slicer_pivot_table_name")
        slicer_pivot_tab_id = fact.get("slicer_pivot_tab_id")
        item_count = fact.get("item_count")
        baseline_selected_item_index = fact.get("baseline_selected_item_index")
        candidate_selected_item_index = fact.get("candidate_selected_item_index")
        baseline_selected_value = fact.get("baseline_selected_value")
        candidate_selected_value = fact.get("candidate_selected_value")
        before_state = (
            _raw_pivot_slicer_selection_state(baseline_path, pivot_sheet)
            if isinstance(pivot_sheet, str)
            else None
        )
        after_state = (
            _raw_pivot_slicer_selection_state(candidate_path, pivot_sheet)
            if isinstance(pivot_sheet, str)
            else None
        )
        before_output = (
            _raw_cell_state(baseline_path, pivot_sheet, pivot_output_cell)
            if isinstance(pivot_sheet, str) and isinstance(pivot_output_cell, str)
            else None
        )
        after_output = (
            _raw_cell_state(candidate_path, pivot_sheet, pivot_output_cell)
            if isinstance(pivot_sheet, str) and isinstance(pivot_output_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_definition_attributes = tuple(
            sorted(
                {
                    "recordCount": "4",
                    "refreshOnLoad": "0",
                    "saveData": "1",
                    f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": "rIdWCABPivotRecords",
                }.items()
            )
        )
        expected_pivot_attributes = (
            tuple(
                sorted(
                    {
                        "cacheId": str(cache_id),
                        "dataCaption": "Amount",
                        "name": "WCAB Pivot Report",
                    }.items()
                )
            )
            if type(cache_id) is int
            else None
        )
        expected_location_attributes = (
            tuple(
                sorted(
                    {
                        "ref": pivot_reference,
                        "firstHeaderRow": "1",
                        "firstDataRow": "2",
                        "firstDataCol": "1",
                    }.items()
                )
            )
            if isinstance(pivot_reference, str)
            else None
        )
        expected_cache_source_attributes = (
            (("type", source_type),) if isinstance(source_type, str) else None
        )
        expected_worksheet_source_attributes = (
            tuple(sorted({"ref": source_reference, "sheet": source_sheet}.items()))
            if isinstance(source_reference, str) and isinstance(source_sheet, str)
            else None
        )
        expected_slicer_declaration_attributes = (
            (f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id", "rIdWCABSlicerCache"),
        )
        expected_slicer_extension_attributes = (("uri", "{BBE1A952-AA13-448E-AADC-164F8A28A991}"),)
        expected_slicer_content_type_attributes = (
            ("ContentType", "application/vnd.ms-excel.slicerCache+xml"),
            ("PartName", "/xl/slicerCaches/slicerCache1.xml"),
        )
        expected_slicer_attributes = (
            tuple(sorted({"name": slicer_name, "sourceName": slicer_source_name}.items()))
            if isinstance(slicer_name, str) and isinstance(slicer_source_name, str)
            else None
        )
        expected_slicer_pivot_table_attributes = (
            tuple(
                sorted({"name": slicer_pivot_table_name, "tabId": str(slicer_pivot_tab_id)}.items())
            )
            if isinstance(slicer_pivot_table_name, str) and type(slicer_pivot_tab_id) is int
            else None
        )
        expected_slicer_tabular_attributes = (
            (("pivotCacheId", str(cache_id)),) if type(cache_id) is int else None
        )
        expected_baseline_slicer_items = (
            (
                tuple(sorted({"x": "0", "s": "1"}.items())),
                tuple(sorted({"x": "1", "s": "0"}.items())),
            )
            if baseline_selected_item_index == 0 and candidate_selected_item_index == 1
            else None
        )
        expected_candidate_slicer_items = (
            (
                tuple(sorted({"x": "0", "s": "0"}.items())),
                tuple(sorted({"x": "1", "s": "1"}.items())),
            )
            if baseline_selected_item_index == 0 and candidate_selected_item_index == 1
            else None
        )
        graph = _direct_graph(candidate)
        selection_fields = {
            "slicer_item_attributes",
            "slicer_selected_item_index",
            "slicer_selected_value",
        }
        _assert(
            type(cache_id) is int
            and cache_id > 0
            and isinstance(source_type, str)
            and isinstance(source_sheet, str)
            and isinstance(source_reference, str)
            and isinstance(pivot_sheet, str)
            and isinstance(pivot_reference, str)
            and isinstance(pivot_output_cell, str)
            and isinstance(dashboard_sheet, str)
            and isinstance(dashboard_cell, str)
            and isinstance(dashboard_formula, str)
            and isinstance(slicer_name, str)
            and isinstance(slicer_source_name, str)
            and isinstance(slicer_pivot_table_name, str)
            and type(slicer_pivot_tab_id) is int
            and type(item_count) is int
            and item_count == 2
            and type(baseline_selected_item_index) is int
            and type(candidate_selected_item_index) is int
            and baseline_selected_item_index == 0
            and candidate_selected_item_index == 1
            and isinstance(baseline_selected_value, str)
            and isinstance(candidate_selected_value, str)
            and baseline_selected_value != candidate_selected_value
            and source_sheet in baseline.sheetnames
            and source_sheet in candidate.sheetnames
            and pivot_sheet in baseline.sheetnames
            and pivot_sheet in candidate.sheetnames
            and dashboard_sheet in baseline.sheetnames
            and dashboard_sheet in candidate.sheetnames
            and before_state is not None
            and after_state is not None
            and {key: value for key, value in before_state.items() if key not in selection_fields}
            == {key: value for key, value in after_state.items() if key not in selection_fields}
            and before_state["cache_id"] == cache_id
            and after_state["cache_id"] == cache_id
            and before_state["definition_attributes"] == expected_definition_attributes
            and after_state["definition_attributes"] == expected_definition_attributes
            and before_state["cache_source_attributes"] == expected_cache_source_attributes
            and after_state["cache_source_attributes"] == expected_cache_source_attributes
            and before_state["worksheet_source_attributes"] == expected_worksheet_source_attributes
            and after_state["worksheet_source_attributes"] == expected_worksheet_source_attributes
            and before_state["cache_records_attributes"] == (("count", "4"),)
            and after_state["cache_records_attributes"] == (("count", "4"),)
            and before_state["cache_record_indexes"] == ((0, 0), (0, 1), (1, 2), (1, 3))
            and after_state["cache_record_indexes"] == ((0, 0), (0, 1), (1, 2), (1, 3))
            and before_state["pivot_table_attributes"] == expected_pivot_attributes
            and after_state["pivot_table_attributes"] == expected_pivot_attributes
            and before_state["location_attributes"] == expected_location_attributes
            and after_state["location_attributes"] == expected_location_attributes
            and before_state["pivot_sheet_id"] == slicer_pivot_tab_id
            and after_state["pivot_sheet_id"] == slicer_pivot_tab_id
            and before_state["slicer_member"] == "xl/slicerCaches/slicerCache1.xml"
            and after_state["slicer_member"] == "xl/slicerCaches/slicerCache1.xml"
            and before_state["slicer_declaration_attributes"]
            == expected_slicer_declaration_attributes
            and after_state["slicer_declaration_attributes"]
            == expected_slicer_declaration_attributes
            and before_state["slicer_extension_attributes"] == expected_slicer_extension_attributes
            and after_state["slicer_extension_attributes"] == expected_slicer_extension_attributes
            and before_state["slicer_content_type_attributes"]
            == expected_slicer_content_type_attributes
            and after_state["slicer_content_type_attributes"]
            == expected_slicer_content_type_attributes
            and before_state["slicer_attributes"] == expected_slicer_attributes
            and after_state["slicer_attributes"] == expected_slicer_attributes
            and before_state["slicer_pivot_table_attributes"]
            == expected_slicer_pivot_table_attributes
            and after_state["slicer_pivot_table_attributes"]
            == expected_slicer_pivot_table_attributes
            and before_state["slicer_tabular_attributes"] == expected_slicer_tabular_attributes
            and after_state["slicer_tabular_attributes"] == expected_slicer_tabular_attributes
            and before_state["slicer_source_field_index"] == 0
            and after_state["slicer_source_field_index"] == 0
            and before_state["slicer_source_shared_item_values"]
            == (baseline_selected_value, candidate_selected_value)
            and after_state["slicer_source_shared_item_values"]
            == (baseline_selected_value, candidate_selected_value)
            and before_state["slicer_item_attributes"] == expected_baseline_slicer_items
            and after_state["slicer_item_attributes"] == expected_candidate_slicer_items
            and before_state["slicer_selected_item_index"] == baseline_selected_item_index
            and after_state["slicer_selected_item_index"] == candidate_selected_item_index
            and before_state["slicer_selected_value"] == baseline_selected_value
            and after_state["slicer_selected_value"] == candidate_selected_value
            and before_state["pivot_cache_slicer_extension_attributes"]
            == (("pivotCacheId", str(cache_id)),)
            and after_state["pivot_cache_slicer_extension_attributes"]
            == (("pivotCacheId", str(cache_id)),)
            and before_output is not None
            and after_output is not None
            and before_output == after_output
            and before_output[3] is None
            and before_output[4] is not None
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == dashboard_formula
            and _slicer_cache_without_selection_state(baseline_path, pivot_sheet)
            == _slicer_cache_without_selection_state(candidate_path, pivot_sheet)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {before_state["slicer_member"]}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (dashboard_sheet, dashboard_cell)
            in graph.get((pivot_sheet, pivot_output_cell), set()),
            f"{truth['id']}: expected one relationship-bound PivotTable Slicer selection to move with stable source, cache, stored output, and dashboard formula",
            errors,
        )
        return

    if kind == "pivot_data_field_aggregation_changed":
        cache_id = fact.get("cache_id")
        source_type = fact.get("source_type")
        source_sheet = fact.get("source_sheet")
        source_reference = fact.get("source_ref")
        pivot_sheet = fact.get("pivot_sheet")
        pivot_reference = fact.get("pivot_ref")
        pivot_output_cell = fact.get("pivot_output_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        dashboard_formula = fact.get("dashboard_formula")
        data_field_source_index = fact.get("data_field_source_index")
        baseline_subtotal = fact.get("baseline_subtotal")
        candidate_subtotal = fact.get("candidate_subtotal")
        before_state = (
            _raw_pivot_data_field_aggregation_state(baseline_path, pivot_sheet)
            if isinstance(pivot_sheet, str)
            else None
        )
        after_state = (
            _raw_pivot_data_field_aggregation_state(candidate_path, pivot_sheet)
            if isinstance(pivot_sheet, str)
            else None
        )
        before_output = (
            _raw_cell_state(baseline_path, pivot_sheet, pivot_output_cell)
            if isinstance(pivot_sheet, str) and isinstance(pivot_output_cell, str)
            else None
        )
        after_output = (
            _raw_cell_state(candidate_path, pivot_sheet, pivot_output_cell)
            if isinstance(pivot_sheet, str) and isinstance(pivot_output_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_definition_attributes = tuple(
            sorted(
                {
                    "recordCount": "4",
                    "refreshOnLoad": "0",
                    "saveData": "1",
                    f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": "rIdWCABPivotRecords",
                }.items()
            )
        )
        expected_pivot_attributes = (
            tuple(
                sorted(
                    {
                        "cacheId": str(cache_id),
                        "dataCaption": "Amount",
                        "name": "WCAB Pivot Report",
                    }.items()
                )
            )
            if type(cache_id) is int
            else None
        )
        expected_location_attributes = (
            tuple(
                sorted(
                    {
                        "ref": pivot_reference,
                        "firstHeaderRow": "1",
                        "firstDataRow": "2",
                        "firstDataCol": "1",
                    }.items()
                )
            )
            if isinstance(pivot_reference, str)
            else None
        )
        expected_cache_source_attributes = (
            (("type", source_type),) if isinstance(source_type, str) else None
        )
        expected_worksheet_source_attributes = (
            tuple(sorted({"ref": source_reference, "sheet": source_sheet}.items()))
            if isinstance(source_reference, str) and isinstance(source_sheet, str)
            else None
        )
        expected_baseline_data_field_attributes = (
            tuple(
                sorted(
                    {
                        "fld": str(data_field_source_index),
                        "subtotal": baseline_subtotal,
                    }.items()
                )
            )
            if type(data_field_source_index) is int and isinstance(baseline_subtotal, str)
            else None
        )
        expected_candidate_data_field_attributes = (
            tuple(
                sorted(
                    {
                        "fld": str(data_field_source_index),
                        "subtotal": candidate_subtotal,
                    }.items()
                )
            )
            if type(data_field_source_index) is int and isinstance(candidate_subtotal, str)
            else None
        )
        graph = _direct_graph(candidate)
        _assert(
            type(cache_id) is int
            and cache_id > 0
            and isinstance(source_type, str)
            and isinstance(source_sheet, str)
            and isinstance(source_reference, str)
            and isinstance(pivot_sheet, str)
            and isinstance(pivot_reference, str)
            and isinstance(pivot_output_cell, str)
            and isinstance(dashboard_sheet, str)
            and isinstance(dashboard_cell, str)
            and isinstance(dashboard_formula, str)
            and type(data_field_source_index) is int
            and data_field_source_index >= 0
            and isinstance(baseline_subtotal, str)
            and isinstance(candidate_subtotal, str)
            and baseline_subtotal != candidate_subtotal
            and source_sheet in baseline.sheetnames
            and source_sheet in candidate.sheetnames
            and pivot_sheet in baseline.sheetnames
            and pivot_sheet in candidate.sheetnames
            and dashboard_sheet in baseline.sheetnames
            and dashboard_sheet in candidate.sheetnames
            and before_state is not None
            and after_state is not None
            and {
                key: value for key, value in before_state.items() if key != "data_field_attributes"
            }
            == {key: value for key, value in after_state.items() if key != "data_field_attributes"}
            and before_state["cache_id"] == cache_id
            and after_state["cache_id"] == cache_id
            and before_state["definition_attributes"] == expected_definition_attributes
            and after_state["definition_attributes"] == expected_definition_attributes
            and before_state["cache_source_attributes"] == expected_cache_source_attributes
            and after_state["cache_source_attributes"] == expected_cache_source_attributes
            and before_state["worksheet_source_attributes"] == expected_worksheet_source_attributes
            and after_state["worksheet_source_attributes"] == expected_worksheet_source_attributes
            and before_state["cache_records_attributes"] == (("count", "4"),)
            and after_state["cache_records_attributes"] == (("count", "4"),)
            and before_state["cache_record_indexes"] == ((0, 0), (0, 1), (1, 2), (1, 3))
            and after_state["cache_record_indexes"] == ((0, 0), (0, 1), (1, 2), (1, 3))
            and before_state["pivot_table_attributes"] == expected_pivot_attributes
            and after_state["pivot_table_attributes"] == expected_pivot_attributes
            and before_state["location_attributes"] == expected_location_attributes
            and after_state["location_attributes"] == expected_location_attributes
            and before_state["data_field_source_index"] == data_field_source_index
            and after_state["data_field_source_index"] == data_field_source_index
            and before_state["data_field_attributes"] == expected_baseline_data_field_attributes
            and after_state["data_field_attributes"] == expected_candidate_data_field_attributes
            and before_output is not None
            and after_output is not None
            and before_output == after_output
            and before_output[3] is None
            and before_output[4] is not None
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == dashboard_formula
            and _pivot_table_without_data_field_subtotal(baseline_path, pivot_sheet)
            == _pivot_table_without_data_field_subtotal(candidate_path, pivot_sheet)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {before_state["pivot_table_member"]}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (dashboard_sheet, dashboard_cell)
            in graph.get((pivot_sheet, pivot_output_cell), set()),
            f"{truth['id']}: expected one relationship-bound PivotTable value aggregation to change with stable source, cache, stored output, and dashboard formula",
            errors,
        )
        return

    if kind == "pivot_cache_refresh_on_load_changed":
        cache_id = fact.get("cache_id")
        source_type = fact.get("source_type")
        source_sheet = fact.get("source_sheet")
        source_reference = fact.get("source_ref")
        pivot_sheet = fact.get("pivot_sheet")
        pivot_reference = fact.get("pivot_ref")
        pivot_output_cell = fact.get("pivot_output_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        dashboard_formula = fact.get("dashboard_formula")
        before_state = (
            _raw_pivot_cache_refresh_state(baseline_path, pivot_sheet)
            if isinstance(pivot_sheet, str)
            else None
        )
        after_state = (
            _raw_pivot_cache_refresh_state(candidate_path, pivot_sheet)
            if isinstance(pivot_sheet, str)
            else None
        )
        before_output = (
            _raw_cell_state(baseline_path, pivot_sheet, pivot_output_cell)
            if isinstance(pivot_sheet, str) and isinstance(pivot_output_cell, str)
            else None
        )
        after_output = (
            _raw_cell_state(candidate_path, pivot_sheet, pivot_output_cell)
            if isinstance(pivot_sheet, str) and isinstance(pivot_output_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_baseline_definition = tuple(
            sorted(
                {
                    "recordCount": "4",
                    "refreshOnLoad": "0",
                    "saveData": "1",
                    f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": "rIdWCABPivotRecords",
                }.items()
            )
        )
        expected_candidate_definition = tuple(
            sorted(
                {
                    "recordCount": "4",
                    "refreshOnLoad": "1",
                    "saveData": "1",
                    f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": "rIdWCABPivotRecords",
                }.items()
            )
        )
        expected_pivot_attributes = (
            tuple(
                sorted(
                    {
                        "cacheId": str(cache_id),
                        "dataCaption": "Amount",
                        "name": "WCAB Pivot Report",
                    }.items()
                )
            )
            if type(cache_id) is int
            else None
        )
        expected_location_attributes = (
            tuple(
                sorted(
                    {
                        "ref": pivot_reference,
                        "firstHeaderRow": "1",
                        "firstDataRow": "2",
                        "firstDataCol": "1",
                    }.items()
                )
            )
            if isinstance(pivot_reference, str)
            else None
        )
        expected_cache_source_attributes = (
            (("type", source_type),) if isinstance(source_type, str) else None
        )
        expected_worksheet_source_attributes = (
            tuple(sorted({"ref": source_reference, "sheet": source_sheet}.items()))
            if isinstance(source_reference, str) and isinstance(source_sheet, str)
            else None
        )
        graph = _direct_graph(candidate)
        _assert(
            type(cache_id) is int
            and cache_id > 0
            and isinstance(source_type, str)
            and isinstance(source_sheet, str)
            and isinstance(source_reference, str)
            and isinstance(pivot_sheet, str)
            and isinstance(pivot_reference, str)
            and isinstance(pivot_output_cell, str)
            and isinstance(dashboard_sheet, str)
            and isinstance(dashboard_cell, str)
            and isinstance(dashboard_formula, str)
            and fact.get("baseline_refresh_on_load") is False
            and fact.get("candidate_refresh_on_load") is True
            and before_state is not None
            and after_state is not None
            and before_state["cache_id"] == cache_id
            and after_state["cache_id"] == cache_id
            and before_state["definition_member"] == after_state["definition_member"]
            and before_state["definition_attributes"] == expected_baseline_definition
            and after_state["definition_attributes"] == expected_candidate_definition
            and before_state["cache_source_attributes"] == expected_cache_source_attributes
            and after_state["cache_source_attributes"] == expected_cache_source_attributes
            and before_state["worksheet_source_attributes"] == expected_worksheet_source_attributes
            and after_state["worksheet_source_attributes"] == expected_worksheet_source_attributes
            and before_state["cache_records_member"] == after_state["cache_records_member"]
            and before_state["cache_records_attributes"] == (("count", "4"),)
            and after_state["cache_records_attributes"] == (("count", "4"),)
            and before_state["cache_record_indexes"] == ((0, 0), (0, 1), (1, 2), (1, 3))
            and after_state["cache_record_indexes"] == ((0, 0), (0, 1), (1, 2), (1, 3))
            and before_state["report_member"] == after_state["report_member"]
            and before_state["pivot_table_member"] == after_state["pivot_table_member"]
            and before_state["pivot_table_attributes"] == expected_pivot_attributes
            and after_state["pivot_table_attributes"] == expected_pivot_attributes
            and before_state["location_attributes"] == expected_location_attributes
            and after_state["location_attributes"] == expected_location_attributes
            and before_output is not None
            and after_output is not None
            and before_output == after_output
            and before_output[3] is None
            and before_output[4] is not None
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == dashboard_formula
            and _pivot_cache_definition_without_refresh_on_load(baseline_path, pivot_sheet)
            == _pivot_cache_definition_without_refresh_on_load(candidate_path, pivot_sheet)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {before_state["definition_member"]}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (dashboard_sheet, dashboard_cell)
            in graph.get((pivot_sheet, pivot_output_cell), set()),
            f"{truth['id']}: expected one relationship-bound PivotCache refresh-on-open false -> true only with stable stored report/dashboard cells",
            errors,
        )
        return

    if kind == "external_data_connection_refresh_on_load_changed":
        connection_id = fact.get("connection_id")
        before_refresh_on_load = (
            _external_data_connection_refresh_on_load(baseline_path, connection_id)
            if isinstance(connection_id, int)
            else None
        )
        after_refresh_on_load = (
            _external_data_connection_refresh_on_load(candidate_path, connection_id)
            if isinstance(connection_id, int)
            else None
        )
        _assert(
            isinstance(connection_id, int)
            and fact.get("baseline_refresh_on_load") is False
            and fact.get("candidate_refresh_on_load") is True
            and before_refresh_on_load is fact.get("baseline_refresh_on_load")
            and after_refresh_on_load is fact.get("candidate_refresh_on_load"),
            f"{truth['id']}: expected connection {connection_id!r} refresh-on-open false -> true",
            errors,
        )
        return

    if kind == "external_data_connection_web_query_url_changed":
        connection_id = fact.get("connection_id")
        connection_member = fact.get("connection_member")
        workbook_relationships_member = fact.get("workbook_relationships_member")
        relationship_id = fact.get("relationship_id")
        relationship_type = fact.get("relationship_type")
        connection_content_type = fact.get("connection_content_type")
        baseline_url = fact.get("baseline_url")
        candidate_url = fact.get("candidate_url")
        saved_value_sheet = fact.get("saved_value_sheet")
        saved_value_cell = fact.get("saved_value_cell")
        saved_value = fact.get("saved_value")
        summary_sheet = fact.get("summary_sheet")
        summary_cell = fact.get("summary_cell")
        summary_formula = fact.get("summary_formula")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        dashboard_formula = fact.get("dashboard_formula")
        before_state = (
            _raw_external_data_connection_state(baseline_path, connection_id)
            if isinstance(connection_id, int)
            else None
        )
        after_state = (
            _raw_external_data_connection_state(candidate_path, connection_id)
            if isinstance(connection_id, int)
            else None
        )
        before_saved_value = (
            _raw_cell_state(baseline_path, saved_value_sheet, saved_value_cell)
            if isinstance(saved_value_sheet, str) and isinstance(saved_value_cell, str)
            else None
        )
        after_saved_value = (
            _raw_cell_state(candidate_path, saved_value_sheet, saved_value_cell)
            if isinstance(saved_value_sheet, str) and isinstance(saved_value_cell, str)
            else None
        )
        before_summary = (
            _raw_cell_state(baseline_path, summary_sheet, summary_cell)
            if isinstance(summary_sheet, str) and isinstance(summary_cell, str)
            else None
        )
        after_summary = (
            _raw_cell_state(candidate_path, summary_sheet, summary_cell)
            if isinstance(summary_sheet, str) and isinstance(summary_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_relationship_attributes = (
            tuple(
                sorted(
                    {
                        "Id": relationship_id,
                        "Type": relationship_type,
                        "Target": "connections.xml",
                    }.items()
                )
            )
            if isinstance(relationship_id, str) and isinstance(relationship_type, str)
            else None
        )
        expected_connection_attributes = tuple(
            sorted(
                {
                    "id": str(connection_id),
                    "name": "WCAB synthetic external-data connection",
                    "refreshedVersion": "1",
                    "refreshOnLoad": "0",
                    "type": "4",
                }.items()
            )
        )
        expected_content_type_attributes = (
            tuple(
                sorted(
                    {
                        "PartName": f"/{connection_member}",
                        "ContentType": connection_content_type,
                    }.items()
                )
            )
            if isinstance(connection_member, str) and isinstance(connection_content_type, str)
            else None
        )
        graph = _direct_graph(candidate)
        _assert(
            isinstance(connection_id, int)
            and isinstance(connection_member, str)
            and isinstance(workbook_relationships_member, str)
            and isinstance(baseline_url, str)
            and isinstance(candidate_url, str)
            and baseline_url != candidate_url
            and fact.get("refresh_on_load") is False
            and isinstance(saved_value_sheet, str)
            and isinstance(saved_value_cell, str)
            and isinstance(saved_value, int)
            and isinstance(summary_sheet, str)
            and isinstance(summary_cell, str)
            and isinstance(summary_formula, str)
            and isinstance(dashboard_sheet, str)
            and isinstance(dashboard_cell, str)
            and isinstance(dashboard_formula, str)
            and before_state is not None
            and after_state is not None
            and before_state["workbook_relationships_member"] == workbook_relationships_member
            and after_state["workbook_relationships_member"] == workbook_relationships_member
            and before_state["connection_member"] == connection_member
            and after_state["connection_member"] == connection_member
            and before_state["relationship_attributes"] == expected_relationship_attributes
            and after_state["relationship_attributes"] == expected_relationship_attributes
            and before_state["connection_attributes"] == expected_connection_attributes
            and after_state["connection_attributes"] == expected_connection_attributes
            and before_state["connection_content_type_attributes"]
            == expected_content_type_attributes
            and after_state["connection_content_type_attributes"]
            == expected_content_type_attributes
            and before_state["web_properties_attributes"] == (("url", baseline_url),)
            and after_state["web_properties_attributes"] == (("url", candidate_url),)
            and before_state["url"] == baseline_url
            and after_state["url"] == candidate_url
            and before_saved_value is not None
            and after_saved_value is not None
            and before_saved_value == after_saved_value
            and before_saved_value[4] == str(saved_value)
            and before_summary is not None
            and after_summary is not None
            and before_summary == after_summary
            and before_summary[3] == summary_formula
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == dashboard_formula
            and _external_data_connection_without_web_query_url(baseline_path, connection_id)
            == _external_data_connection_without_web_query_url(candidate_path, connection_id)
            and _xlsx_member_differences(baseline_path, candidate_path) == {connection_member}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (summary_sheet, summary_cell)
            in graph.get((saved_value_sheet, saved_value_cell), set())
            and (dashboard_sheet, dashboard_cell)
            in graph.get((summary_sheet, summary_cell), set()),
            f"{truth['id']}: expected one stable external-data web-query URL retarget with saved cells and local formulas unchanged",
            errors,
        )
        return

    if kind == "cell_hyperlink_target_changed":
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        summary_sheet = fact.get("summary_sheet")
        summary_cell = fact.get("summary_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = (
            _raw_cell_hyperlink_target_state(baseline_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        after_state = (
            _raw_cell_hyperlink_target_state(candidate_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        before_input = (
            _raw_cell_state(baseline_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        after_input = (
            _raw_cell_state(candidate_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        before_summary = (
            _raw_cell_state(baseline_path, summary_sheet, summary_cell)
            if isinstance(summary_sheet, str) and isinstance(summary_cell, str)
            else None
        )
        after_summary = (
            _raw_cell_state(candidate_path, summary_sheet, summary_cell)
            if isinstance(summary_sheet, str) and isinstance(summary_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        before_cell = (
            baseline[sheet_name][coordinate]
            if isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and sheet_name in baseline.sheetnames
            else None
        )
        after_cell = (
            candidate[sheet_name][coordinate]
            if isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and sheet_name in candidate.sheetnames
            else None
        )
        before_hyperlink = before_cell.hyperlink if before_cell is not None else None
        after_hyperlink = after_cell.hyperlink if after_cell is not None else None
        expected_hyperlink_attributes = (
            ("ref", "B2"),
            (
                f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id",
                "rIdWCABVendorPortal",
            ),
        )
        expected_before_relationship_attributes = (
            ("Id", "rIdWCABVendorPortal"),
            ("Target", "https://approved.example.invalid/wcab-vendor-portal"),
            ("TargetMode", "External"),
            (
                "Type",
                "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink",
            ),
        )
        expected_after_relationship_attributes = (
            ("Id", "rIdWCABVendorPortal"),
            ("Target", "https://review.example.invalid/wcab-vendor-portal"),
            ("TargetMode", "External"),
            (
                "Type",
                "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink",
            ),
        )
        graph = _direct_graph(candidate)
        _assert(
            sheet_name == "Inputs"
            and coordinate == "B2"
            and fact.get("cell_value") == "Open vendor portal"
            and fact.get("worksheet_member") == "xl/worksheets/sheet1.xml"
            and fact.get("worksheet_relationships_member") == "xl/worksheets/_rels/sheet1.xml.rels"
            and fact.get("relationship_id") == "rIdWCABVendorPortal"
            and fact.get("relationship_type")
            == "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink"
            and fact.get("target_mode") == "External"
            and fact.get("baseline_target") == "https://approved.example.invalid/wcab-vendor-portal"
            and fact.get("candidate_target") == "https://review.example.invalid/wcab-vendor-portal"
            and summary_sheet == "Summary"
            and summary_cell == "B2"
            and fact.get("summary_formula") == "=Inputs!$B$2"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Summary!$B$2"
            and all(
                sheet in baseline.sheetnames and sheet in candidate.sheetnames
                for sheet in (sheet_name, summary_sheet, dashboard_sheet)
            )
            and before_state is not None
            and after_state is not None
            and before_state["worksheet_member"] == fact.get("worksheet_member")
            and after_state["worksheet_member"] == fact.get("worksheet_member")
            and before_state["worksheet_relationships_member"]
            == fact.get("worksheet_relationships_member")
            and after_state["worksheet_relationships_member"]
            == fact.get("worksheet_relationships_member")
            and before_state["hyperlink_attributes"] == expected_hyperlink_attributes
            and after_state["hyperlink_attributes"] == expected_hyperlink_attributes
            and before_state["relationship_id"] == fact.get("relationship_id")
            and after_state["relationship_id"] == fact.get("relationship_id")
            and before_state["relationship_attributes"] == expected_before_relationship_attributes
            and after_state["relationship_attributes"] == expected_after_relationship_attributes
            and before_state["target_mode"] == fact.get("target_mode")
            and after_state["target_mode"] == fact.get("target_mode")
            and before_state["target"] == fact.get("baseline_target")
            and after_state["target"] == fact.get("candidate_target")
            and before_cell is not None
            and after_cell is not None
            and before_cell.value == fact.get("cell_value")
            and after_cell.value == fact.get("cell_value")
            and before_hyperlink is not None
            and after_hyperlink is not None
            and before_hyperlink.id == fact.get("relationship_id")
            and after_hyperlink.id == fact.get("relationship_id")
            and before_hyperlink.target == fact.get("baseline_target")
            and after_hyperlink.target == fact.get("candidate_target")
            and all(
                hyperlink.location is None
                and hyperlink.display is None
                and hyperlink.tooltip is None
                for hyperlink in (before_hyperlink, after_hyperlink)
            )
            and before_input is not None
            and after_input is not None
            and before_input == after_input
            and before_input[0] == "xl/worksheets/sheet1.xml"
            and before_input[1] == "inlineStr"
            and before_input[3:] == (None, None)
            and before_summary is not None
            and after_summary is not None
            and before_summary == after_summary
            and before_summary[3] == fact.get("summary_formula")
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == fact.get("dashboard_formula")
            and _cell_hyperlink_relationship_without_target(baseline_path, sheet_name, coordinate)
            == _cell_hyperlink_relationship_without_target(candidate_path, sheet_name, coordinate)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {"xl/worksheets/_rels/sheet1.xml.rels"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (summary_sheet, summary_cell) in _reachable(graph, (sheet_name, coordinate))
            and (dashboard_sheet, dashboard_cell) in _reachable(graph, (sheet_name, coordinate)),
            f"{truth['id']}: expected one external worksheet cell hyperlink target change only with stable visible text, formulas, and package boundary",
            errors,
        )
        return

    if kind == "query_table_refresh_on_load_changed":
        sheet_name = fact.get("sheet")
        connection_id = fact.get("connection_id")
        saved_value_cell = fact.get("saved_value_cell")
        summary_sheet = fact.get("summary_sheet")
        summary_cell = fact.get("summary_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = (
            _raw_query_table_refresh_state(baseline_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        after_state = (
            _raw_query_table_refresh_state(candidate_path, sheet_name)
            if isinstance(sheet_name, str)
            else None
        )
        before_saved_value = (
            _raw_cell_state(baseline_path, sheet_name, saved_value_cell)
            if isinstance(sheet_name, str) and isinstance(saved_value_cell, str)
            else None
        )
        after_saved_value = (
            _raw_cell_state(candidate_path, sheet_name, saved_value_cell)
            if isinstance(sheet_name, str) and isinstance(saved_value_cell, str)
            else None
        )
        before_summary = (
            _raw_cell_state(baseline_path, summary_sheet, summary_cell)
            if isinstance(summary_sheet, str) and isinstance(summary_cell, str)
            else None
        )
        after_summary = (
            _raw_cell_state(candidate_path, summary_sheet, summary_cell)
            if isinstance(summary_sheet, str) and isinstance(summary_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_common_state = {
            "worksheet_member": "xl/worksheets/sheet1.xml",
            "worksheet_relationships_member": "xl/worksheets/_rels/sheet1.xml.rels",
            "worksheet_relationship_attributes": (
                ("Id", "rIdWCABQueryTable"),
                ("Target", "../queryTables/queryTable1.xml"),
                (
                    "Type",
                    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/queryTable",
                ),
            ),
            "query_table_member": "xl/queryTables/queryTable1.xml",
            "connection_relationship_attributes": (
                ("Id", "rIdWCABQueryTableConnections"),
                ("Target", "connections.xml"),
                (
                    "Type",
                    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/connections",
                ),
            ),
            "connection_member": "xl/connections.xml",
            "connection_attributes": (
                ("background", "0"),
                ("id", "1"),
                ("name", "WCAB synthetic query-table connection"),
                ("refreshOnLoad", "0"),
                ("refreshedVersion", "1"),
                ("type", "4"),
            ),
            "web_properties_attributes": (
                ("url", "https://example.invalid/wcab-query-table-refresh"),
            ),
            "query_table_content_type_attributes": (
                (
                    "ContentType",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.queryTable+xml",
                ),
                ("PartName", "/xl/queryTables/queryTable1.xml"),
            ),
            "connection_content_type_attributes": (
                (
                    "ContentType",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml",
                ),
                ("PartName", "/xl/connections.xml"),
            ),
        }
        expected_before_query_table_attributes = (
            ("backgroundRefresh", "0"),
            ("connectionId", "1"),
            ("disableEdit", "1"),
            ("disableRefresh", "0"),
            ("fillFormulas", "0"),
            ("growShrinkType", "insertClear"),
            ("name", "WCAB synthetic query table"),
            ("refreshOnLoad", "0"),
            ("removeDataOnSave", "0"),
        )
        expected_after_query_table_attributes = (
            ("backgroundRefresh", "0"),
            ("connectionId", "1"),
            ("disableEdit", "1"),
            ("disableRefresh", "0"),
            ("fillFormulas", "0"),
            ("growShrinkType", "insertClear"),
            ("name", "WCAB synthetic query table"),
            ("refreshOnLoad", "1"),
            ("removeDataOnSave", "0"),
        )
        graph = _direct_graph(candidate)
        _assert(
            sheet_name == "ImportedData"
            and connection_id == 1
            and fact.get("connection_member") == "xl/connections.xml"
            and fact.get("connection_url") == "https://example.invalid/wcab-query-table-refresh"
            and fact.get("query_table_member") == "xl/queryTables/queryTable1.xml"
            and fact.get("worksheet_member") == "xl/worksheets/sheet1.xml"
            and fact.get("worksheet_relationships_member") == "xl/worksheets/_rels/sheet1.xml.rels"
            and fact.get("relationship_id") == "rIdWCABQueryTable"
            and fact.get("relationship_type")
            == "http://schemas.openxmlformats.org/officeDocument/2006/relationships/queryTable"
            and fact.get("baseline_refresh_on_load") is False
            and fact.get("candidate_refresh_on_load") is True
            and fact.get("background_refresh") is False
            and fact.get("refresh_disabled") is False
            and fact.get("remove_data_on_save") is False
            and fact.get("fill_formulas") is False
            and fact.get("connection_edit_disabled") is True
            and fact.get("growth_behavior") == "insertClear"
            and saved_value_cell == "B2"
            and fact.get("saved_value") == 100
            and summary_sheet == "Summary"
            and summary_cell == "B2"
            and fact.get("summary_formula") == "=ImportedData!$B$2"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Summary!$B$2"
            and all(
                sheet in baseline.sheetnames and sheet in candidate.sheetnames
                for sheet in (sheet_name, summary_sheet, dashboard_sheet)
            )
            and before_state is not None
            and after_state is not None
            and all(before_state.get(key) == value for key, value in expected_common_state.items())
            and all(after_state.get(key) == value for key, value in expected_common_state.items())
            and before_state["query_table_attributes"] == expected_before_query_table_attributes
            and after_state["query_table_attributes"] == expected_after_query_table_attributes
            and before_state["refresh_on_load"] is False
            and after_state["refresh_on_load"] is True
            and before_saved_value is not None
            and before_saved_value == after_saved_value
            and before_saved_value[0] == "xl/worksheets/sheet1.xml"
            and before_saved_value[3] is None
            and before_saved_value[4] == "100"
            and before_summary is not None
            and before_summary == after_summary
            and before_summary[0] == "xl/worksheets/sheet2.xml"
            and before_summary[3] == "=ImportedData!$B$2"
            and before_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[0] == "xl/worksheets/sheet3.xml"
            and before_dashboard[3] == "=Summary!$B$2"
            and _query_table_without_refresh_on_load(baseline_path, sheet_name)
            == _query_table_without_refresh_on_load(candidate_path, sheet_name)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {"xl/queryTables/queryTable1.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (summary_sheet, summary_cell) in _reachable(graph, (sheet_name, saved_value_cell))
            and (dashboard_sheet, dashboard_cell)
            in _reachable(graph, (sheet_name, saved_value_cell)),
            f"{truth['id']}: expected one relationship-backed QueryTable refresh-on-open false -> true with stable connection, saved cells, formulas, and package boundary",
            errors,
        )
        return

    if kind == "external_workbook_link_update_policy_changed":
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        formula = fact.get("formula")
        before_properties = _workbook_properties(baseline_path)
        after_properties = _workbook_properties(candidate_path)
        before_formula = (
            baseline[sheet_name][coordinate]
            if isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and sheet_name in baseline.sheetnames
            else None
        )
        after_formula = (
            candidate[sheet_name][coordinate]
            if isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and sheet_name in candidate.sheetnames
            else None
        )
        before_without_policy = (
            {key: value for key, value in before_properties.items() if key != "updateLinks"}
            if before_properties is not None
            else None
        )
        after_without_policy = (
            {key: value for key, value in after_properties.items() if key != "updateLinks"}
            if after_properties is not None
            else None
        )
        _assert(
            isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and isinstance(formula, str)
            and "[" in formula
            and before_formula is not None
            and after_formula is not None
            and _cell_kind(before_formula) == "formula"
            and _cell_kind(after_formula) == "formula"
            and before_formula.value == formula
            and after_formula.value == formula
            and fact.get("baseline_update_links") == "never"
            and fact.get("candidate_update_links") == "always"
            and before_properties is not None
            and after_properties is not None
            and before_properties.get("updateLinks") == fact.get("baseline_update_links")
            and after_properties.get("updateLinks") == fact.get("candidate_update_links")
            and before_without_policy == after_without_policy,
            f"{truth['id']}: expected unchanged {sheet_name}!{coordinate} external formula and workbook updateLinks never -> always only",
            errors,
        )
        return

    if kind == "external_workbook_link_source_changed":
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = _raw_external_workbook_link_source_state(baseline_path)
        after_state = _raw_external_workbook_link_source_state(candidate_path)
        before_formula = (
            baseline[sheet_name][coordinate]
            if isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and sheet_name in baseline.sheetnames
            else None
        )
        after_formula = (
            candidate[sheet_name][coordinate]
            if isinstance(sheet_name, str)
            and isinstance(coordinate, str)
            and sheet_name in candidate.sheetnames
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_common_state = {
            "workbook_member": "xl/workbook.xml",
            "workbook_relationships_member": "xl/_rels/workbook.xml.rels",
            "workbook_relationship_id": "rIdWCABExternalLink",
            "workbook_relationship_attributes": (
                ("Id", "rIdWCABExternalLink"),
                ("Target", "externalLinks/externalLink1.xml"),
                (
                    "Type",
                    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink",
                ),
            ),
            "external_link_member": "xl/externalLinks/externalLink1.xml",
            "external_link_relationships_member": ("xl/externalLinks/_rels/externalLink1.xml.rels"),
            "external_book_attributes": (
                (
                    f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id",
                    "rIdWCABExternalLinkPath",
                ),
            ),
            "external_sheet": "Inputs",
            "external_link_relationship_id": "rIdWCABExternalLinkPath",
            "content_type_attributes": (
                (
                    "ContentType",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml",
                ),
                ("PartName", "/xl/externalLinks/externalLink1.xml"),
            ),
        }
        expected_before_state = {
            **expected_common_state,
            "external_link_relationship_attributes": (
                ("Id", "rIdWCABExternalLinkPath"),
                (
                    "Target",
                    "https://approved.example.invalid/wcab-external-workbook/WCABSource.xlsx",
                ),
                ("TargetMode", "External"),
                (
                    "Type",
                    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath",
                ),
            ),
            "target": "https://approved.example.invalid/wcab-external-workbook/WCABSource.xlsx",
        }
        expected_after_state = {
            **expected_common_state,
            "external_link_relationship_attributes": (
                ("Id", "rIdWCABExternalLinkPath"),
                (
                    "Target",
                    "https://review.example.invalid/wcab-external-workbook/WCABSource.xlsx",
                ),
                ("TargetMode", "External"),
                (
                    "Type",
                    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath",
                ),
            ),
            "target": "https://review.example.invalid/wcab-external-workbook/WCABSource.xlsx",
        }
        graph = _direct_graph(candidate)
        _assert(
            sheet_name == "LinkedModel"
            and coordinate == "B2"
            and fact.get("formula") == "='[WCABSource.xlsx]Inputs'!$B$2"
            and fact.get("workbook_member") == "xl/workbook.xml"
            and fact.get("workbook_relationships_member") == "xl/_rels/workbook.xml.rels"
            and fact.get("workbook_relationship_id") == "rIdWCABExternalLink"
            and fact.get("workbook_relationship_type")
            == "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink"
            and fact.get("external_link_member") == "xl/externalLinks/externalLink1.xml"
            and fact.get("external_link_relationships_member")
            == "xl/externalLinks/_rels/externalLink1.xml.rels"
            and fact.get("external_link_content_type")
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"
            and fact.get("external_link_relationship_id") == "rIdWCABExternalLinkPath"
            and fact.get("external_link_relationship_type")
            == "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath"
            and fact.get("external_sheet") == "Inputs"
            and fact.get("target_mode") == "External"
            and fact.get("baseline_target")
            == "https://approved.example.invalid/wcab-external-workbook/WCABSource.xlsx"
            and fact.get("candidate_target")
            == "https://review.example.invalid/wcab-external-workbook/WCABSource.xlsx"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=LinkedModel!$B$2"
            and before_state == expected_before_state
            and after_state == expected_after_state
            and before_formula is not None
            and after_formula is not None
            and _cell_kind(before_formula) == "formula"
            and _cell_kind(after_formula) == "formula"
            and before_formula.value == fact.get("formula")
            and after_formula.value == fact.get("formula")
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == fact.get("dashboard_formula")
            and _external_workbook_link_relationship_without_target(baseline_path)
            == _external_workbook_link_relationship_without_target(candidate_path)
            and _xlsx_member_differences(baseline_path, candidate_path)
            == {"xl/externalLinks/_rels/externalLink1.xml.rels"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (dashboard_sheet, dashboard_cell) in _reachable(graph, (sheet_name, coordinate)),
            f"{truth['id']}: expected one external-workbook source target change only with stable formulas and package graph",
            errors,
        )
        return

    if kind == "external_defined_name_source_changed":
        name = fact.get("name")
        formula_sheet = fact.get("formula_sheet")
        formula_cell = fact.get("formula_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = _raw_external_defined_name_source_state(baseline_path)
        after_state = _raw_external_defined_name_source_state(candidate_path)
        before_formula = (
            _raw_cell_state(baseline_path, formula_sheet, formula_cell)
            if isinstance(formula_sheet, str) and isinstance(formula_cell, str)
            else None
        )
        after_formula = (
            _raw_cell_state(candidate_path, formula_sheet, formula_cell)
            if isinstance(formula_sheet, str) and isinstance(formula_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_common_state = {
            "workbook_member": "xl/workbook.xml",
            "name": "ScenarioRate",
            "attributes": (("name", "ScenarioRate"),),
            "external_references_count": 0,
            "external_link_members": (),
        }
        graph = _direct_graph(candidate)
        _assert(
            name == "ScenarioRate"
            and fact.get("workbook_member") == "xl/workbook.xml"
            and fact.get("baseline_refers_to") == "'[WCABApprovedSource.xlsx]Inputs'!$B$2"
            and fact.get("candidate_refers_to") == "'[WCABReviewSource.xlsx]Inputs'!$B$2"
            and formula_sheet == "Model"
            and formula_cell == "B2"
            and fact.get("formula") == "=ScenarioRate*2"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Model!$B$2"
            and before_state
            == {
                **expected_common_state,
                "refers_to": "'[WCABApprovedSource.xlsx]Inputs'!$B$2",
            }
            and after_state
            == {
                **expected_common_state,
                "refers_to": "'[WCABReviewSource.xlsx]Inputs'!$B$2",
            }
            and _defined_name_text(baseline, name) == fact.get("baseline_refers_to")
            and _defined_name_text(candidate, name) == fact.get("candidate_refers_to")
            and formula_sheet in baseline.sheetnames
            and formula_sheet in candidate.sheetnames
            and before_formula is not None
            and after_formula is not None
            and before_formula == after_formula
            and before_formula[3] == fact.get("formula")
            and before_dashboard is not None
            and after_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[3] == fact.get("dashboard_formula")
            and _workbook_without_external_defined_name_source(baseline_path)
            == _workbook_without_external_defined_name_source(candidate_path)
            and _xlsx_member_differences(baseline_path, candidate_path) == {"xl/workbook.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (dashboard_sheet, dashboard_cell)
            in _reachable(graph, (formula_sheet, formula_cell)),
            f"{truth['id']}: expected only one local defined-name external source change with stable formula context",
            errors,
        )
        return

    if kind == "named_lambda_definition_changed":
        name = fact.get("name")
        input_sheet = fact.get("input_sheet")
        rate_cell = fact.get("rate_cell")
        amount_cell = fact.get("amount_cell")
        formula_sheet = fact.get("formula_sheet")
        formula_cell = fact.get("formula_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = _raw_named_lambda_definition_state(baseline_path)
        after_state = _raw_named_lambda_definition_state(candidate_path)
        before_rate = (
            _raw_cell_state(baseline_path, input_sheet, rate_cell)
            if isinstance(input_sheet, str) and isinstance(rate_cell, str)
            else None
        )
        after_rate = (
            _raw_cell_state(candidate_path, input_sheet, rate_cell)
            if isinstance(input_sheet, str) and isinstance(rate_cell, str)
            else None
        )
        before_amount = (
            _raw_cell_state(baseline_path, input_sheet, amount_cell)
            if isinstance(input_sheet, str) and isinstance(amount_cell, str)
            else None
        )
        after_amount = (
            _raw_cell_state(candidate_path, input_sheet, amount_cell)
            if isinstance(input_sheet, str) and isinstance(amount_cell, str)
            else None
        )
        before_formula = (
            _raw_cell_state(baseline_path, formula_sheet, formula_cell)
            if isinstance(formula_sheet, str) and isinstance(formula_cell, str)
            else None
        )
        after_formula = (
            _raw_cell_state(candidate_path, formula_sheet, formula_cell)
            if isinstance(formula_sheet, str) and isinstance(formula_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_common_state = {
            "workbook_member": "xl/workbook.xml",
            "name": "ScenarioValue",
            "attributes": (("name", "ScenarioValue"),),
            "external_references_count": 0,
            "external_link_members": (),
        }
        graph = _direct_graph(candidate)
        _assert(
            name == "ScenarioValue"
            and fact.get("workbook_member") == "xl/workbook.xml"
            and fact.get("parameters") == ["rate", "amount"]
            and fact.get("baseline_refers_to") == "=LAMBDA(rate,amount,rate*amount)"
            and fact.get("candidate_refers_to") == "=LAMBDA(rate,amount,rate*(amount+10))"
            and input_sheet == "Inputs"
            and rate_cell == "B2"
            and fact.get("rate_value") == 0.08
            and amount_cell == "B3"
            and fact.get("amount_value") == 100
            and formula_sheet == "Model"
            and formula_cell == "B2"
            and fact.get("formula") == "=ScenarioValue(Inputs!B2,Inputs!B3)"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Model!$B$2"
            and before_state
            == {
                **expected_common_state,
                "refers_to": "=LAMBDA(rate,amount,rate*amount)",
            }
            and after_state
            == {
                **expected_common_state,
                "refers_to": "=LAMBDA(rate,amount,rate*(amount+10))",
            }
            and _defined_name_text(baseline, name) == fact.get("baseline_refers_to")
            and _defined_name_text(candidate, name) == fact.get("candidate_refers_to")
            and all(
                sheet in baseline.sheetnames and sheet in candidate.sheetnames
                for sheet in (input_sheet, formula_sheet, dashboard_sheet)
            )
            and before_rate is not None
            and before_rate == after_rate
            and before_rate[0] == "xl/worksheets/sheet1.xml"
            and before_rate[4] == "0.08"
            and before_amount is not None
            and before_amount == after_amount
            and before_amount[0] == "xl/worksheets/sheet1.xml"
            and before_amount[4] == "100"
            and before_formula is not None
            and before_formula == after_formula
            and before_formula[0] == "xl/worksheets/sheet2.xml"
            and before_formula[3] == fact.get("formula")
            and before_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[0] == "xl/worksheets/sheet3.xml"
            and before_dashboard[3] == fact.get("dashboard_formula")
            and _workbook_without_named_lambda_definition(baseline_path)
            == _workbook_without_named_lambda_definition(candidate_path)
            and _xlsx_member_differences(baseline_path, candidate_path) == {"xl/workbook.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (dashboard_sheet, dashboard_cell)
            in _reachable(graph, (formula_sheet, formula_cell))
            and (formula_sheet, formula_cell) in _reachable(graph, (input_sheet, rate_cell))
            and (dashboard_sheet, dashboard_cell) in _reachable(graph, (input_sheet, amount_cell)),
            f"{truth['id']}: expected one local named LAMBDA body change with stable inputs, formulas, and package boundary",
            errors,
        )
        return

    if kind == "power_pivot_data_model_relationship_changed":
        before_state = _raw_power_pivot_data_model_relationship_state(baseline_path)
        after_state = _raw_power_pivot_data_model_relationship_state(candidate_path)
        expected_common_state = {
            "workbook_member": "xl/workbook.xml",
            "workbook_relationships_member": "xl/_rels/workbook.xml.rels",
            "data_model_member": "xl/model/item.data",
            "workbook_relationship_attributes": (
                ("Id", "rIdWCABPowerPivotData"),
                ("Target", "model/item.data"),
                ("Type", _POWER_PIVOT_DATA_RELATIONSHIP),
            ),
            "data_model_content_type_attributes": (
                ("ContentType", _POWER_PIVOT_DATA_CONTENT_TYPE),
                ("Extension", "data"),
            ),
            "extension_attributes": (("uri", "{FCE2AD5D-F65C-4FA6-A056-5C36A1767C68}"),),
            "data_model_attributes": (("minVersionLoad", "5"),),
            "model_table_attributes": (
                (
                    ("connection", "SalesModel"),
                    ("id", "SalesModel_{11111111-1111-1111-1111-111111111111}"),
                    ("name", "SalesModel"),
                ),
                (
                    ("connection", "CalendarModel"),
                    ("id", "CalendarModel_{22222222-2222-2222-2222-222222222222}"),
                    ("name", "CalendarModel"),
                ),
            ),
            "data_model_members": ("xl/model/item.data",),
            "data_model_relationship_members": (),
            "data_model_payload_sha256": sha256(_POWER_PIVOT_DATA_PAYLOAD).hexdigest(),
            "data_model_payload_size": len(_POWER_PIVOT_DATA_PAYLOAD),
        }
        expected_before_relationship = (
            ("fromColumn", "CalendarKey"),
            ("fromTable", "SalesModel"),
            ("toColumn", "DateKey"),
            ("toTable", "CalendarModel"),
        )
        expected_after_relationship = (
            ("fromColumn", "CalendarKey"),
            ("fromTable", "SalesModel"),
            ("toColumn", "FiscalDateKey"),
            ("toTable", "CalendarModel"),
        )
        expected_sheets = ["Sales", "Calendar", "Dashboard"]
        _assert(
            fact.get("workbook_member") == "xl/workbook.xml"
            and fact.get("workbook_relationships_member") == "xl/_rels/workbook.xml.rels"
            and fact.get("data_model_member") == "xl/model/item.data"
            and fact.get("workbook_relationship_id") == "rIdWCABPowerPivotData"
            and fact.get("workbook_relationship_type") == _POWER_PIVOT_DATA_RELATIONSHIP
            and fact.get("workbook_relationship_target") == "model/item.data"
            and fact.get("data_model_content_type") == _POWER_PIVOT_DATA_CONTENT_TYPE
            and fact.get("extension_uri") == "{FCE2AD5D-F65C-4FA6-A056-5C36A1767C68}"
            and fact.get("min_version_load") == "5"
            and fact.get("model_tables") == ["SalesModel", "CalendarModel"]
            and fact.get("from_table") == "SalesModel"
            and fact.get("from_column") == "CalendarKey"
            and fact.get("to_table") == "CalendarModel"
            and fact.get("baseline_to_column") == "DateKey"
            and fact.get("candidate_to_column") == "FiscalDateKey"
            and fact.get("data_model_payload_sha256")
            == sha256(_POWER_PIVOT_DATA_PAYLOAD).hexdigest()
            and fact.get("data_model_payload_size") == len(_POWER_PIVOT_DATA_PAYLOAD)
            and baseline.sheetnames == expected_sheets
            and candidate.sheetnames == expected_sheets
            and "SalesModel" in baseline["Sales"].tables
            and "SalesModel" in candidate["Sales"].tables
            and baseline["Sales"].tables["SalesModel"].ref == "A1:B3"
            and candidate["Sales"].tables["SalesModel"].ref == "A1:B3"
            and baseline["Sales"]["A1"].value == "CalendarKey"
            and candidate["Sales"]["A1"].value == "CalendarKey"
            and "CalendarModel" in baseline["Calendar"].tables
            and "CalendarModel" in candidate["Calendar"].tables
            and baseline["Calendar"].tables["CalendarModel"].ref == "A1:B3"
            and candidate["Calendar"].tables["CalendarModel"].ref == "A1:B3"
            and baseline["Calendar"]["A1"].value == "DateKey"
            and candidate["Calendar"]["A1"].value == "DateKey"
            and baseline["Calendar"]["B1"].value == "FiscalDateKey"
            and candidate["Calendar"]["B1"].value == "FiscalDateKey"
            and before_state
            == {
                **expected_common_state,
                "model_relationship_attributes": expected_before_relationship,
            }
            and after_state
            == {
                **expected_common_state,
                "model_relationship_attributes": expected_after_relationship,
            }
            and _workbook_without_power_pivot_data_model_to_column(baseline_path)
            == _workbook_without_power_pivot_data_model_to_column(candidate_path)
            and _xlsx_member_differences(baseline_path, candidate_path) == {"xl/workbook.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected one Power Pivot Data Model relationship target-key change with a fixed opaque payload and package boundary",
            errors,
        )
        return

    if kind == "xlm_auto_open_binding_retargeted":
        input_sheet = fact.get("input_sheet")
        input_cell = fact.get("input_cell")
        model_sheet = fact.get("model_sheet")
        model_cell = fact.get("model_cell")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = _raw_xlm_auto_open_binding_state(baseline_path)
        after_state = _raw_xlm_auto_open_binding_state(candidate_path)
        before_input = (
            _raw_cell_state(baseline_path, input_sheet, input_cell)
            if isinstance(input_sheet, str) and isinstance(input_cell, str)
            else None
        )
        after_input = (
            _raw_cell_state(candidate_path, input_sheet, input_cell)
            if isinstance(input_sheet, str) and isinstance(input_cell, str)
            else None
        )
        before_model = (
            _raw_cell_state(baseline_path, model_sheet, model_cell)
            if isinstance(model_sheet, str) and isinstance(model_cell, str)
            else None
        )
        after_model = (
            _raw_cell_state(candidate_path, model_sheet, model_cell)
            if isinstance(model_sheet, str) and isinstance(model_cell, str)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        expected_common_state = {
            "workbook_member": "xl/workbook.xml",
            "workbook_relationships_member": "xl/_rels/workbook.xml.rels",
            "macro_sheet_member": "xl/macrosheets/sheet1.xml",
            "macro_sheet_relationship_attributes": (
                ("Id", "rIdWCABXlmMacroSheet"),
                ("Target", "macrosheets/sheet1.xml"),
                (
                    "Type",
                    "http://schemas.microsoft.com/office/2006/relationships/xlMacrosheet",
                ),
            ),
            "workbook_content_type_attributes": (
                (
                    "ContentType",
                    "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                ),
                ("PartName", "/xl/workbook.xml"),
            ),
            "macro_sheet_content_type_attributes": (
                ("ContentType", "application/vnd.ms-excel.macrosheet+xml"),
                ("PartName", "/xl/macrosheets/sheet1.xml"),
            ),
            "macro_sheet_declaration_attributes": tuple(
                sorted(
                    {
                        "name": "Macro Automation",
                        "sheetId": "4",
                        "state": "veryHidden",
                        f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id": "rIdWCABXlmMacroSheet",
                    }.items()
                )
            ),
            "automatic_macro_attributes": (("name", "_xlnm.Auto_Open"),),
            "macro_sheet_formulas": (("A1", "HALT()"), ("A2", "HALT()")),
            "macro_sheet_members": ("xl/macrosheets/sheet1.xml",),
            "macro_sheet_relationship_members": (),
            "vba_project_members": (),
            "macro_sheet_sha256": sha256(_XLM_AUTO_OPEN_MACRO_SHEET_PAYLOAD).hexdigest(),
            "macro_sheet_size": len(_XLM_AUTO_OPEN_MACRO_SHEET_PAYLOAD),
        }
        expected_sheets = ["Inputs", "Model", "Dashboard", "Macro Automation"]
        graph = _direct_graph(candidate)
        _assert(
            fact.get("workbook_member") == "xl/workbook.xml"
            and fact.get("workbook_relationships_member") == "xl/_rels/workbook.xml.rels"
            and fact.get("macro_sheet_member") == "xl/macrosheets/sheet1.xml"
            and fact.get("macro_sheet_relationship_id") == "rIdWCABXlmMacroSheet"
            and fact.get("macro_sheet_relationship_type")
            == "http://schemas.microsoft.com/office/2006/relationships/xlMacrosheet"
            and fact.get("macro_sheet_relationship_target") == "macrosheets/sheet1.xml"
            and fact.get("macro_sheet_content_type") == "application/vnd.ms-excel.macrosheet+xml"
            and fact.get("workbook_content_type")
            == "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
            and fact.get("macro_sheet_name") == "Macro Automation"
            and fact.get("macro_sheet_sheet_id") == "4"
            and fact.get("macro_sheet_state") == "veryHidden"
            and fact.get("automatic_macro_name") == "_xlnm.Auto_Open"
            and fact.get("automatic_macro_event") == "Auto_Open"
            and fact.get("baseline_target") == "'Macro Automation'!$A$1"
            and fact.get("candidate_target") == "'Macro Automation'!$A$2"
            and fact.get("macro_sheet_formula") == "HALT()"
            and fact.get("macro_sheet_formula_cells") == ["A1", "A2"]
            and fact.get("macro_sheet_sha256")
            == sha256(_XLM_AUTO_OPEN_MACRO_SHEET_PAYLOAD).hexdigest()
            and fact.get("macro_sheet_size") == len(_XLM_AUTO_OPEN_MACRO_SHEET_PAYLOAD)
            and input_sheet == "Inputs"
            and input_cell == "B2"
            and fact.get("input_value") == 10
            and model_sheet == "Model"
            and model_cell == "B2"
            and fact.get("model_formula") == "=Inputs!$B$2*2"
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=Model!$B$2"
            and baseline.sheetnames == expected_sheets
            and candidate.sheetnames == expected_sheets
            and before_state
            == {
                **expected_common_state,
                "automatic_macro_target": "'Macro Automation'!$A$1",
            }
            and after_state
            == {
                **expected_common_state,
                "automatic_macro_target": "'Macro Automation'!$A$2",
            }
            and before_input is not None
            and before_input == after_input
            and before_input[0] == "xl/worksheets/sheet1.xml"
            and before_input[4] == "10"
            and before_model is not None
            and before_model == after_model
            and before_model[0] == "xl/worksheets/sheet2.xml"
            and before_model[3] == "=Inputs!$B$2*2"
            and before_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[0] == "xl/worksheets/sheet3.xml"
            and before_dashboard[3] == "=Model!$B$2"
            and _workbook_without_xlm_auto_open_target(baseline_path)
            == _workbook_without_xlm_auto_open_target(candidate_path)
            and _xlsx_member_differences(baseline_path, candidate_path) == {"xl/workbook.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path)
            and (model_sheet, model_cell) in _reachable(graph, (input_sheet, input_cell))
            and (dashboard_sheet, dashboard_cell) in _reachable(graph, (input_sheet, input_cell))
            and (dashboard_sheet, dashboard_cell) in _reachable(graph, (model_sheet, model_cell)),
            f"{truth['id']}: expected one raw XLM Auto_Open target change with fixed macro-sheet and ordinary-formula context",
            errors,
        )
        return

    if kind == "array_formula_mode_changed":
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        before = (
            _array_formula_state(baseline_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        after = (
            _array_formula_state(candidate_path, sheet_name, coordinate)
            if isinstance(sheet_name, str) and isinstance(coordinate, str)
            else None
        )
        expected_before = (
            fact.get("baseline_mode"),
            fact.get("formula"),
            fact.get("baseline_output_range"),
        )
        expected_after = (
            fact.get("candidate_mode"),
            fact.get("formula"),
            fact.get("candidate_output_range"),
        )
        _assert(
            before == expected_before
            and after == expected_after
            and fact.get("baseline_mode") == "legacy_cse"
            and fact.get("candidate_mode") == "dynamic",
            f"{truth['id']}: expected unchanged {sheet_name}!{coordinate} formula to switch legacy CSE -> dynamic array mode",
            errors,
        )
        return

    if kind == "static_cycle_introduced":
        graph = _direct_graph(candidate)
        cells = [(cell["sheet"], cell["cell"]) for cell in fact.get("cells", [])]
        for cell in cells:
            _assert(
                cell in _reachable(graph, cell),
                f"{truth['id']}: expected a static cycle through {cell[0]}!{cell[1]}",
                errors,
            )
        return

    if kind == "three_d_scope_changed":
        formula_sheet = fact.get("formula_sheet")
        formula_cell = fact.get("formula_cell")
        _assert(
            baseline[formula_sheet][formula_cell].value
            == candidate[formula_sheet][formula_cell].value,
            f"{truth['id']}: 3-D formula text should remain unchanged",
            errors,
        )
        order = candidate.sheetnames
        try:
            after_index = order.index(fact["after_sheet"])
            inserted_index = order.index(fact["inserted_sheet"])
            before_index = order.index(fact["before_sheet"])
        except ValueError:
            errors.append(f"{truth['id']}: missing declared 3-D scope sheet")
            return
        _assert(
            after_index < inserted_index < before_index,
            f"{truth['id']}: inserted sheet is not inside the declared 3-D span",
            errors,
        )
        return

    if kind == "structural_formula_rewrite":
        before = fact.get("baseline", {})
        after = fact.get("candidate", {})
        _assert(
            baseline[before["sheet"]][before["cell"]].value == before["formula"]
            and candidate[after["sheet"]][after["cell"]].value == after["formula"],
            f"{truth['id']}: declared structural formula rewrite is not present",
            errors,
        )
        return

    if kind == "structured_table_scope_changed":
        table_sheet = fact.get("table_sheet")
        table_name = fact.get("table")
        formula_sheet = fact.get("formula_sheet")
        formula_cell = fact.get("formula_cell")
        if (
            table_sheet not in baseline.sheetnames
            or table_sheet not in candidate.sheetnames
            or formula_sheet not in baseline.sheetnames
            or formula_sheet not in candidate.sheetnames
        ):
            errors.append(f"{truth['id']}: declared table or formula sheet is absent")
            return
        try:
            before_table = baseline[table_sheet].tables[table_name]
            after_table = candidate[table_sheet].tables[table_name]
        except KeyError:
            errors.append(f"{truth['id']}: declared Excel Table {table_name!r} is absent")
            return
        before_formula = baseline[formula_sheet][formula_cell].value
        after_formula = candidate[formula_sheet][formula_cell].value
        _assert(
            before_table.ref == fact.get("baseline_ref")
            and after_table.ref == fact.get("candidate_ref")
            and before_formula == after_formula
            and f"{table_name}[" in str(after_formula),
            f"{truth['id']}: expected structured-reference formula to retain text while table scope changes",
            errors,
        )
        return

    if kind == "table_calculated_column_formula_changed":
        table_sheet = fact.get("table_sheet")
        table_name = fact.get("table")
        stable_formula_cells = fact.get("stable_formula_cells")
        dashboard_sheet = fact.get("dashboard_sheet")
        dashboard_cell = fact.get("dashboard_cell")
        before_state = (
            _raw_table_calculated_column_formula_state(baseline_path, table_sheet)
            if isinstance(table_sheet, str)
            else None
        )
        after_state = (
            _raw_table_calculated_column_formula_state(candidate_path, table_sheet)
            if isinstance(table_sheet, str)
            else None
        )
        before_formula_cells = (
            tuple(
                _raw_cell_state(baseline_path, table_sheet, cell) for cell in stable_formula_cells
            )
            if isinstance(table_sheet, str)
            and isinstance(stable_formula_cells, list)
            and all(isinstance(cell, str) for cell in stable_formula_cells)
            else None
        )
        after_formula_cells = (
            tuple(
                _raw_cell_state(candidate_path, table_sheet, cell) for cell in stable_formula_cells
            )
            if isinstance(table_sheet, str)
            and isinstance(stable_formula_cells, list)
            and all(isinstance(cell, str) for cell in stable_formula_cells)
            else None
        )
        before_dashboard = (
            _raw_cell_state(baseline_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        after_dashboard = (
            _raw_cell_state(candidate_path, dashboard_sheet, dashboard_cell)
            if isinstance(dashboard_sheet, str) and isinstance(dashboard_cell, str)
            else None
        )
        relationship_id_attribute = f"{{{_DOCUMENT_RELATIONSHIPS_NS}}}id"
        expected_common_state = {
            "worksheet_member": "xl/worksheets/sheet1.xml",
            "worksheet_relationships_member": "xl/worksheets/_rels/sheet1.xml.rels",
            "worksheet_table_part_attributes": ((relationship_id_attribute, "rId1"),),
            "table_relationship_attributes": (
                ("Id", "rId1"),
                ("Target", "/xl/tables/table1.xml"),
                ("Type", _TABLE_RELATIONSHIP),
            ),
            "table_member": "xl/tables/table1.xml",
            "table_attributes": (
                ("displayName", "ScenarioLedger"),
                ("headerRowCount", "1"),
                ("id", "1"),
                ("name", "ScenarioLedger"),
                ("ref", "A1:C4"),
            ),
            "auto_filter_attributes": (("ref", "A1:C4"),),
            "table_columns_attributes": (("count", "3"),),
            "table_column_attributes": (
                (("id", "1"), ("name", "Units")),
                (("id", "2"), ("name", "Rate")),
                (("id", "3"), ("name", "Calculated amount")),
            ),
            "calculated_column_formula_attributes": (),
        }
        expected_formula_cells = (
            ("C2", "=A2*B2"),
            ("C3", "=A3*B3"),
            ("C4", "=A4*B4"),
        )
        _assert(
            table_sheet == "Ledger"
            and table_name == "ScenarioLedger"
            and fact.get("table_member") == "xl/tables/table1.xml"
            and fact.get("table_ref") == "A1:C4"
            and fact.get("calculated_column_id") == 3
            and fact.get("calculated_column_name") == "Calculated amount"
            and fact.get("baseline_formula") == "A2*B2"
            and fact.get("candidate_formula") == "A2*(B2+1)"
            and stable_formula_cells == [cell for cell, _formula in expected_formula_cells]
            and dashboard_sheet == "Dashboard"
            and dashboard_cell == "B4"
            and fact.get("dashboard_formula") == "=SUM(ScenarioLedger[Calculated amount])"
            and table_sheet in baseline.sheetnames
            and table_sheet in candidate.sheetnames
            and dashboard_sheet in baseline.sheetnames
            and dashboard_sheet in candidate.sheetnames
            and table_name in baseline[table_sheet].tables
            and table_name in candidate[table_sheet].tables
            and baseline[table_sheet].tables[table_name].ref == fact.get("table_ref")
            and candidate[table_sheet].tables[table_name].ref == fact.get("table_ref")
            and before_state
            == {
                **expected_common_state,
                "calculated_column_formula": "A2*B2",
            }
            and after_state
            == {
                **expected_common_state,
                "calculated_column_formula": "A2*(B2+1)",
            }
            and before_formula_cells is not None
            and after_formula_cells == before_formula_cells
            and all(cell_state is not None for cell_state in before_formula_cells)
            and all(
                cell_state is not None
                and cell_state[0] == "xl/worksheets/sheet1.xml"
                and cell_state[3] == formula
                for cell_state, (_cell, formula) in zip(
                    before_formula_cells, expected_formula_cells, strict=True
                )
            )
            and before_dashboard is not None
            and before_dashboard == after_dashboard
            and before_dashboard[0] == "xl/worksheets/sheet2.xml"
            and before_dashboard[3] == fact.get("dashboard_formula")
            and _table_without_calculated_column_formula(baseline_path, table_sheet)
            == _table_without_calculated_column_formula(candidate_path, table_sheet)
            and _xlsx_member_differences(baseline_path, candidate_path) == {"xl/tables/table1.xml"}
            and _calculation_properties(baseline_path) == _calculation_properties(candidate_path),
            f"{truth['id']}: expected only one Table calculated-column master formula change with stable cells and package boundary",
            errors,
        )
        return

    if kind == "portfolio_value_changed":
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        before = baseline[sheet_name][coordinate]
        after = candidate[sheet_name][coordinate]
        _assert(
            _cell_kind(before) == "value"
            and _cell_kind(after) == "value"
            and before.value != after.value,
            f"{truth['id']}: expected portfolio value change at {sheet_name}!{coordinate}",
            errors,
        )
        return

    if kind == "portfolio_external_reference":
        sheet_name = fact.get("sheet")
        coordinate = fact.get("cell")
        value = candidate[sheet_name][coordinate].value
        target = fact.get("target_workbook")
        _assert(
            _cell_kind(candidate[sheet_name][coordinate]) == "formula"
            and f"[{target}]" in str(value),
            f"{truth['id']}: expected local external reference to {target!r}",
            errors,
        )
        return

    errors.append(f"{truth['id']}: unsupported fact kind {kind!r}")


def _validate_coverage_expectations(
    case_dir: Path, truth: dict[str, Any], errors: list[str]
) -> None:
    expectations = truth.get("coverage_expectations")
    if not isinstance(expectations, list):
        errors.append(f"{truth['id']}: coverage_expectations must be an array")
        return
    for expectation in expectations:
        if not isinstance(expectation, dict):
            errors.append(f"{truth['id']}: coverage expectation must be an object")
            continue
        kind = expectation.get("kind")
        if kind not in {
            "dynamic_reference_static_coverage",
            "dynamic_reference_driver_changed",
        }:
            errors.append(f"{truth['id']}: unsupported coverage expectation kind {kind!r}")
            continue
        try:
            baseline_path, candidate_path = _workbook_pair(case_dir, truth)
        except FixtureValidationError as error:
            errors.append(str(error))
            continue
        baseline = _load_workbook(baseline_path)
        candidate = _load_workbook(candidate_path)
        expected_functions = expectation.get("functions")
        if kind == "dynamic_reference_static_coverage":
            sheet_name = expectation.get("sheet")
            coordinate = expectation.get("cell")
            if (
                not isinstance(sheet_name, str)
                or not isinstance(coordinate, str)
                or sheet_name not in baseline.sheetnames
                or sheet_name not in candidate.sheetnames
            ):
                errors.append(f"{truth['id']}: dynamic-reference coverage location is absent")
                continue
            before = baseline[sheet_name][coordinate]
            after = candidate[sheet_name][coordinate]
            _assert(
                _cell_kind(before) == "formula"
                and _cell_kind(after) == "formula"
                and not _dynamic_reference_functions(str(before.value))
                and isinstance(expected_functions, list)
                and all(isinstance(function, str) for function in expected_functions)
                and _dynamic_reference_functions(str(after.value)) == tuple(expected_functions),
                f"{truth['id']}: expected static-dependency coverage boundary at {sheet_name}!{coordinate}",
                errors,
            )
            continue

        driver = expectation.get("driver")
        formula = expectation.get("formula")
        if not isinstance(driver, dict) or not isinstance(formula, dict):
            errors.append(
                f"{truth['id']}: dynamic-reference driver coverage needs driver and formula"
            )
            continue
        driver_sheet = driver.get("sheet")
        driver_cell = driver.get("cell")
        formula_sheet = formula.get("sheet")
        formula_cell = formula.get("cell")
        locations = (driver_sheet, driver_cell, formula_sheet, formula_cell)
        if (
            not all(isinstance(location, str) for location in locations)
            or driver_sheet not in baseline.sheetnames
            or driver_sheet not in candidate.sheetnames
            or formula_sheet not in baseline.sheetnames
            or formula_sheet not in candidate.sheetnames
        ):
            errors.append(f"{truth['id']}: dynamic-reference driver coverage location is absent")
            continue
        before_driver = baseline[driver_sheet][driver_cell]
        after_driver = candidate[driver_sheet][driver_cell]
        before_formula = baseline[formula_sheet][formula_cell]
        after_formula = candidate[formula_sheet][formula_cell]
        graph = _direct_graph(candidate)
        _assert(
            _cell_kind(before_driver) == "value"
            and _cell_kind(after_driver) == "value"
            and before_driver.value != after_driver.value
            and _cell_kind(before_formula) == "formula"
            and _cell_kind(after_formula) == "formula"
            and before_formula.value == after_formula.value
            and isinstance(expected_functions, list)
            and all(isinstance(function, str) for function in expected_functions)
            and _dynamic_reference_functions(str(after_formula.value)) == tuple(expected_functions)
            and (formula_sheet, formula_cell) in graph.get((driver_sheet, driver_cell), set()),
            f"{truth['id']}: expected changed dynamic-reference driver {driver_sheet}!{driver_cell} to feed unchanged formula {formula_sheet}!{formula_cell}",
            errors,
        )


def _validate_impacts(case_dir: Path, truth: dict[str, Any], errors: list[str]) -> None:
    if truth.get("topology") != "pair":
        return
    _baseline_path, candidate_path = _workbook_pair(case_dir, truth)
    candidate = _load_workbook(candidate_path)
    graph = _direct_graph(candidate)
    for impact in truth.get("must_reach", []):
        source = impact.get("source", {})
        source_key = (source.get("sheet"), source.get("cell"))
        reachable = _reachable(graph, source_key)
        for target in impact.get("targets", []):
            target_key = (target.get("sheet"), target.get("cell"))
            _assert(
                target_key in reachable,
                f"{truth['id']}: static impact {source_key[0]}!{source_key[1]} -> {target_key[0]}!{target_key[1]} is absent",
                errors,
            )


def validate_case(case_dir: str | Path) -> list[str]:
    """Validate one case directory and return its errors without raising."""

    directory = Path(case_dir)
    truth_path = directory / "truth.json"
    truth = _load_truth(truth_path)
    errors: list[str] = []
    _assert(
        truth.get("schema_version") == FIXTURE_SCHEMA_VERSION,
        f"{directory}: unsupported schema version {truth.get('schema_version')!r}",
        errors,
    )
    _assert(
        truth.get("id") in CASE_IDS,
        f"{directory}: unknown case id {truth.get('id')!r}",
        errors,
    )
    _assert(
        truth.get("review_expectation") in {"allow", "review", "block"},
        f"{truth.get('id', directory)}: invalid review expectation",
        errors,
    )
    _assert(
        isinstance(truth.get("coverage_expectations"), list),
        f"{truth.get('id', directory)}: coverage_expectations must be an array",
        errors,
    )
    if truth.get("topology") == "pair":
        try:
            _workbook_pair(directory, truth)
        except FixtureValidationError as error:
            errors.append(str(error))
            return errors
    elif truth.get("topology") == "portfolio":
        _assert(
            (directory / "baseline").is_dir(),
            f"{truth.get('id')}: missing baseline portfolio",
            errors,
        )
        _assert(
            (directory / "candidate").is_dir(),
            f"{truth.get('id')}: missing candidate portfolio",
            errors,
        )
    else:
        errors.append(f"{truth.get('id', directory)}: invalid topology {truth.get('topology')!r}")
        return errors

    for fact in truth.get("facts", []):
        if not isinstance(fact, dict):
            errors.append(f"{truth.get('id')}: fact must be an object")
            continue
        _validate_fact(directory, truth, fact, errors)
    _validate_coverage_expectations(directory, truth, errors)
    _validate_impacts(directory, truth, errors)
    return errors


def validate_all(root: str | Path) -> dict[str, list[str]]:
    """Validate all fixture manifests below *root*.

    Returns a mapping of case IDs to validation errors.  An empty mapping means
    the entire tree satisfies its documented truth contract.
    """

    fixture_root = Path(root)
    manifests = sorted(fixture_root.rglob("truth.json"))
    if len(manifests) != len(CASE_IDS):
        return {"_catalog": [f"expected {len(CASE_IDS)} truth manifests, found {len(manifests)}"]}
    failures: dict[str, list[str]] = {}
    seen: set[str] = set()
    for manifest in manifests:
        truth = _load_truth(manifest)
        case_id = str(truth.get("id", manifest.parent))
        seen.add(case_id)
        errors = validate_case(manifest.parent)
        if errors:
            failures[case_id] = errors
    missing = sorted(set(CASE_IDS) - seen)
    if missing:
        failures["_catalog"] = [f"missing case IDs: {', '.join(missing)}"]
    return failures
